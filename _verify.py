import openpyxl

print('=== activity (3050) ===')
wb = openpyxl.load_workbook(r'resources\activity.xlsx', read_only=True)
ws = wb['Activity']
for r in ws.iter_rows(min_row=3):
    if ws.cell(r[0].row, 1).value == 3050:
        print([c.value for c in r][:8])
        break

print('\n=== reward (7条) ===')
wb = openpyxl.load_workbook(r'resources\reward.xlsx', read_only=True)
ws = wb['Reward']
for r in ws.iter_rows(min_row=3):
    rid = ws.cell(r[0].row, 1).value
    if rid in (100609, 100610, 100611, 100612, 100613, 100614, 100615):
        name = ws.cell(r[0].row, 2).value
        day_limit = ws.cell(r[0].row, 4).value
        exp_p = ws.cell(r[0].row, 6).value
        exp_f = ws.cell(r[0].row, 7).value
        gold_p = ws.cell(r[0].row, 8).value
        gold_f = ws.cell(r[0].row, 9).value
        m1 = ws.cell(r[0].row, 11).value
        m1n = ws.cell(r[0].row, 12).value
        m2 = ws.cell(r[0].row, 13).value
        m2n = ws.cell(r[0].row, 14).value
        w2 = ws.cell(r[0].row, 27).value
        print(f'{rid} {name} daylimit={day_limit} exp={exp_p}/{exp_f} gold={gold_p}/{gold_f} must1={m1}x{m1n} must2={m2}x{m2n} weight2={w2}')

print('\n=== mail template (99003) + global (100) ===')
wb = openpyxl.load_workbook(r'resources\mail.xlsx', read_only=True)
ws = wb['MailTemplate']
for r in ws.iter_rows(min_row=3):
    if ws.cell(r[0].row, 1).value == 99003:
        print('tpl', [c.value for c in r][:3])
        break
ws = wb['GlobalMail']
for r in ws.iter_rows(min_row=3):
    if ws.cell(r[0].row, 1).value == 100:
        print('global', [c.value for c in r][:9])
        break

print('\n=== tips (7条) ===')
wb = openpyxl.load_workbook(r'resources\tips.xlsx', read_only=True)
ws = wb['ui_txt']
for r in ws.iter_rows(min_row=ws.max_row - 6):
    key = ws.cell(r[0].row, 4).value
    if key and 'QRCX' in str(key):
        print(ws.cell(r[0].row, 2).value, '|', ws.cell(r[0].row, 3).value, '|', ws.cell(r[0].row, 5).value)
