"""制造"ID 重映射 → 跨表（不同 xlsx 文件间）外键引用同步更新"的完整演示场景。

背景：engine/cross_table_fk.py::sync_cross_table_refs 是"第三遍"全局扫描——
merge_stages.py::_build_group 只能处理"同一张表文件内"的外键同步（同 sheet /
跨 sheet），不同 xlsx 文件之间的外键（如 item.xlsx·Chest."reward ID" →
reward.xlsx·Reward.reward_id）各自独立 _build_group，彼此的 id_mapping 互不
可见，必须等一次 compare 请求内所有表都比对完成后，才能拿到"reward.xlsx 的
重映射记录"去同步"item.xlsx 里引用它的外键列"。这条链路挂在
merge_branch.py::branch_compare（/api/merge/branch/compare）里，
group_names 需**同时包含** "item/item" 和 "reward" 才会生效（sync_cross_table_
refs 对未同时被 compare 到的表对会直接跳过，不误报不同步）。

场景设计（比照 seed_ref_integrity_sync.py 的同 sheet 场景，往前多一步——
demo 里默认没有 reward.xlsx 的共享 PK 插入冲突，本脚本自己造）：
  1. dev1 / dev2 的 reward.xlsx·Reward 各自新增一行 reward_id=88888，
     内容（名称列）不同 → 合并 dev1<->dev2 时触发"多分支同 PK 新增冲突"，
     dev2 那条会被 id_resolver 重映射到一个新编号（下方注释 88889 只是示例，
     实际取决于当前已占用的最大编号）。
  2. dev2 的 item.xlsx·Chest 里 item_id=11001 这一行的 "reward ID" 列改成
     88888——引用同分支 reward.xlsx 里正在被重映射的那一行。

幂等：目标单元格已是预期值则跳过对应分支的写入。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"

REWARD_REL = Path("reward.xlsx")
REWARD_SHEET = "Reward"
SHARED_REWARD_ID = 88888

ITEM_REL = Path("item") / "item.xlsx"
CHEST_SHEET = "Chest"
CHEST_ITEM_ID = 11001   # item.xlsx·Chest 里已存在的一行（物品编号）
FK_COL = 2               # B列："reward ID"


def _find_row_by_pk(ws, pk, col: int = 1, max_row: int = None) -> int:
    # max_row 缺省用 ws.max_row（不能硬编码固定上限）：reward.xlsx 这类模板文件
    # 常带大段预格式化的空行（无值但有样式，openpyxl 仍计入 max_row），本表首次
    # 追加共享 PK 那一行落到了 2000 开外——固定 max_row=2000 扫描不到它，第二次
    # 运行时误判"不存在"又 append 一次，产生重复行（已在 demo 数据里踩过这个坑，
    # 手工清理过；这里改成动态取 ws.max_row 避免复发）。
    limit = max_row if max_row is not None else ws.max_row
    for r in range(1, limit + 1):
        if ws.cell(r, col).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def _ensure_shared_reward(wc: Path, name_suffix: str) -> bool:
    """在 wc/reward.xlsx·Reward 插入/校正共享 PK=88888 的行，内容按分支区分。

    返回 True 表示本次做了修改（需要 commit）。
    """
    fp = wc / REWARD_REL
    wb = load_workbook(fp, data_only=False)
    ws = wb[REWARD_SHEET]
    row = _find_row_by_pk(ws, SHARED_REWARD_ID)
    name_val = f"跨表外键同步测试奖励_{name_suffix}"
    changed = False
    if row < 0:
        ws.append([SHARED_REWARD_ID, name_val])
        changed = True
    elif ws.cell(row, 2).value != name_val:
        ws.cell(row, 2).value = name_val
        changed = True
    if changed:
        wb.save(fp)
    wb.close()
    return changed


def main() -> None:
    print("=== 制造 ID 重映射 -> 跨表外键引用同步更新 演示场景 ===")

    changed1 = _ensure_shared_reward(DEV1, "dev1")
    if changed1:
        _commit(DEV1, f"dev1: reward.xlsx 新增共享 PK={SHARED_REWARD_ID}（制造跨分支同 PK 插入冲突）")
    else:
        print(f"  [skip] dev1 reward.xlsx 行 {SHARED_REWARD_ID} 已是预期内容")

    changed2 = _ensure_shared_reward(DEV2, "dev2")

    item_fp = DEV2 / ITEM_REL
    wb = load_workbook(item_fp, data_only=False)
    ws = wb[CHEST_SHEET]
    row = _find_row_by_pk(ws, CHEST_ITEM_ID, max_row=730)
    if row < 0:
        wb.close()
        print(f"  [skip] dev2 item.xlsx·Chest 未找到 物品编号={CHEST_ITEM_ID}")
    elif ws.cell(row, FK_COL).value == SHARED_REWARD_ID:
        wb.close()
        print(f"  [skip] dev2 item.xlsx·Chest 第{row}行 reward ID 已是 {SHARED_REWARD_ID}（幂等）")
    else:
        ws.cell(row, FK_COL).value = SHARED_REWARD_ID
        wb.save(item_fp)
        wb.close()
        print(f"  dev2: item.xlsx·Chest 第{row}行（物品编号={CHEST_ITEM_ID}）"
              f"reward ID -> {SHARED_REWARD_ID}（引用 reward.xlsx 里正在被重映射的那一行）")
        changed2 = True

    if changed2:
        _commit(DEV2, f"dev2: reward.xlsx 新增共享 PK={SHARED_REWARD_ID}（内容冲突）"
                      f" + item.xlsx·Chest 引用该 PK（制造跨表外键同步场景）")
    else:
        print("  [skip] dev2 无需提交（内容与 FK 引用均已是预期状态）")

    print(f"\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2，"
          f"table_names 需**同时勾选** item/item 与 reward 两张表（缺一个 sync_cross_"
          f"table_refs 会跳过这条规则，不会同步）。reward 表里 reward_id={SHARED_REWARD_ID} "
          f"那一行应能看到「重编号」徽标（dev2 的插入被重映射成新编号）；item 表·Chest "
          f"sheet 物品编号={CHEST_ITEM_ID} 那一行的「reward ID」列应自动同步显示成同样的"
          f"新编号（不是 {SHARED_REWARD_ID}，也不悬空）。")


if __name__ == "__main__":
    main()
