"""制造"ID 重映射 → 同表外键引用同步更新"的完整演示场景。

背景：id_resolver.py 的 id_mapping（多分支同 PK 冲突重映射记录）会被
ref_integrity.py::validate_sheet_references 消费，对同一张表/同一 sheet 内被
merge_strategies.yaml 标为 base_priority 的"外键候选列"做同步改写——如果某分支
自己的一行用了一个后来被重映射的旧 PK，改写后这一列会自动变成新 PK，而不是
悬空引用。这条链路已经接在 merge_stages.py::_build_group（/api/merge/branch/
compare、/api/merge/subdir/compare 都走它）里，不需要额外触发 apply，纯 compare
阶段就能看到。

demo 里已经有 dev1/dev2 都插入 id=9999（不同内容）的同 PK 冲突（build_svn_small_
branches.py 的 RICH_INSERT_SHARED_PK），dev2 那条会被重映射成一个新编号（比如
10810，具体值取决于当前已占用的最大编号）。本脚本额外让 dev2 里**另一行**的
"被动id"外键列指向 9999——重映射生效后，这一列应该自动同步成新编号，而不是
指向一个已经不存在/被改号的旧 9999。

目标：ability.xlsx · Ability sheet
  dev2：第5行（神通id=1001）的"被动id"列（G列）设为 9999（引用同分支正在被
        重映射的那一行）
  dev1/trunk：不动（该行被动id 保持原样，不参与这个场景）

幂等：dev2 该单元格已是 9999 则跳过。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV2 = WC_ROOT / "branches" / "dev2"
TABLE_REL = Path("ability.xlsx")
SHEET = "Ability"
TARGET_PK = 1001  # 引用来源行：神通id=1001（A列）
FK_COL = 7  # G列：被动id
REF_PK = 9999  # 引用目标：同分支正在被重映射的那个共享 PK


def _find_row_by_pk(ws, pk: int, max_row: int = 730) -> int:
    for r in range(1, max_row + 1):
        if ws.cell(r, 1).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def main() -> None:
    print("=== 制造 ID 重映射 -> 外键引用同步更新 演示场景 ===")
    fp = DEV2 / TABLE_REL
    wb = load_workbook(fp, data_only=False)
    ws = wb[SHEET]
    row = _find_row_by_pk(ws, TARGET_PK)
    if row < 0:
        wb.close()
        print(f"  [skip] dev2 未找到 神通id={TARGET_PK}")
        return
    if ws.cell(row, FK_COL).value == REF_PK:
        wb.close()
        print(f"  [skip] dev2 第{row}行 被动id 已是 {REF_PK}（幂等）")
        return
    ws.cell(row, FK_COL).value = REF_PK
    wb.save(fp)
    wb.close()
    print(f"  dev2: 第{row}行（神通id={TARGET_PK}）被动id -> {REF_PK}"
          f"（引用同分支正在被重映射的那一行）")
    _commit(DEV2, f"dev2: 制造外键引用同步场景（被动id 引用即将被重映射的 id={REF_PK}）")
    print(f"\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2，"
          f"表 ability / sheet Ability，神通id={TARGET_PK} 那一行的"
          f"「被动id」列——正常情况下它应该显示成 dev2 的 id=9999 被重映射后的新编号"
          f"（不是 9999，也不是悬空），因为 ref_integrity 会在 compare 阶段就把这个"
          f"引用同步改写。可以对照 id=9999/新编号那一行的「重编号」徽章核实两者是否"
          f"一致。")


if __name__ == "__main__":
    main()
