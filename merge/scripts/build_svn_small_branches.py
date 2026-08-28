"""扩展 demo_svn fixture，新增 4 个小表分支/子目录，并保证 dev1-4 任两两都有冲突。

布局（位置严格区分）：
  平行分支（branches/ 下，与 dev1/dev2 同级；subdev_1 不再作分支，改 trunk 子目录）：
    /branches/dev3     svn copy /trunk@R，删 3 大表，改 ability/reward/tips
    /branches/dev4     svn copy /trunk@R，删 3 大表，改 const/fabao/guild
  trunk 下子目录（私有表集合，目录合并 source，冲突规模按需差异化）：
    /trunk/subdev_1/   8 张小私有表 → 16 冲突（10-19）
    /trunk/subdev_2/   21 张小私有表 → 31 冲突（30-40）
    /trunk/subdev_3/   15 张小私有表 → 23 冲突（20-30）
  全配对冲突锚点：dev1/dev2/dev3/dev4 的 tips.xlsx 首个 sheet B2 各写不同值，
    使任两分支合并都在该单元格冲突。

排除大表 monster(5.6MB)/skill_level(2.8MB)/item_drop(2.4MB)，仅留小表测合并时间。
依赖：build_svn_real.py 已建好 repo + trunk + dev1/dev2（subdev_1 随 trunk 导入）。

合并引导可见：
  /api/merge/branch/dirs 列出 svn/demo_svn/wc/trunk + branches/{dev1..4}（不含 subdev_1）
  /api/merge/subdir/dirs 递归扫描列出 svn/demo_svn/wc/trunk/subdev_{1,2,3}

幂等：分支/子目录已存在则跳过，仅补改表。
用法：
    python merge/scripts/build_svn_small_branches.py
    python merge/scripts/build_svn_small_branches.py --clean-branches  # 删旧 dev3/dev4 + trunk/subdev_2|3 + 遗留 subdev_1 再重建
"""
import argparse
import os
import shutil
import stat
import subprocess
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
MERGE_DIR = SCRIPT_DIR.parent
SVN_DEMO_DIR = MERGE_DIR / "svn" / "demo_svn"
REPO_DIR = SVN_DEMO_DIR / "repo"
WC_DIR = SVN_DEMO_DIR / "wc"
TRUNK_WC = WC_DIR / "trunk"
DEFAULT_SRC = MERGE_DIR / "_seed_data"
# 不可变原始种子 trunk（各 stage 读 base 一律从这里取，避免工作副本被前置 stage 污染叠加）
SEED_TRUNK = DEFAULT_SRC / "trunk"

# 排除的 3 张大表（均 >2MB，compare 阶段 O(n^2) 卡顿）
BIG_TABLES = {"monster.xlsx", "skill_level.xlsx", "item_drop.xlsx"}

# 平行分支（branches/ 下）
NEW_BRANCHES = ["dev3", "dev4"]
DEV3_MODIFY = ["ability.xlsx", "reward.xlsx", "tips.xlsx"]
DEV4_MODIFY = ["const.xlsx", "fabao.xlsx", "guild.xlsx"]

# trunk 下子目录（目录合并 source）：表集合与冲突规模按需差异化。
#  - subdev_1：8 张表 → 16 冲突（10-19）
#  - subdev_2：21 张表 → 31 冲突（30-40）
#  - subdev_3：15 张表 → 23 冲突（20-30）
NEW_SUBDIRS = ["subdev_2", "subdev_3"]
SUBDEV1_KEEP = ["ability.xlsx", "const.xlsx", "reward.xlsx", "fabao.xlsx",
                "mail.xlsx", "tips.xlsx", "world_buff.xlsx", "interaction.xlsx"]
SUBDEV2_KEEP = [
    "ability.xlsx", "activity.xlsx", "const.xlsx", "entity_prefab.xlsx",
    "exclusive_state.xlsx", "fabao.xlsx", "gameplay_tags.xlsx", "guild.xlsx",
    "id_mgr.xlsx", "interaction.xlsx", "mail.xlsx", "map.xlsx",
    "material_effect.xlsx", "model_prefab.xlsx", "player_common.xlsx",
    "realm_info.xlsx", "reward.xlsx", "space.xlsx", "spawn_world_entity.xlsx",
    "tips.xlsx", "world_buff.xlsx",
]
SUBDEV3_KEEP = ["ability.xlsx", "activity.xlsx", "const.xlsx", "entity_prefab.xlsx",
                "exclusive_state.xlsx", "fabao.xlsx", "gameplay_tags.xlsx", "guild.xlsx",
                "interaction.xlsx", "mail.xlsx", "map.xlsx", "reward.xlsx",
                "space.xlsx", "tips.xlsx", "world_buff.xlsx"]

# 目录合并冲突规模：(子目录, 表集合, 每表冲突数)。冲突数列表与表集合等长，
# 每表 1 或 2 个 name 列数据行冲突；subdev 侧与 trunk 侧写不同语义值构成真冲突。
SUBDIR_PLANS = [
    ("subdev_1", SUBDEV1_KEEP, [2] * 8),               # 8 表 → 16 冲突
    ("subdev_2", SUBDEV2_KEEP, [2] * 10 + [1] * 11),   # 21 表 → 31 冲突
    ("subdev_3", SUBDEV3_KEEP, [2] * 8 + [1] * 7),     # 15 表 → 23 冲突
]

# dev1-4 全面冲突场景：分散多表 + 同表不同 sheet + 新增 sheet(结构差异) + 新增行(含同 PK 冲突)。
# 每个 dev 都参与每场景（写本分支值），故任两 dev 合并在每场景都冲突/差异。用户要求
# "不应集中在一个表格，要分散多表、同表不同 sheet、新增 sheet 和行"。
# 1) content 冲突：5 个 表/sheet 组合（4 张表，fabao 跨 2 sheet 展示同表不同 sheet）
RICH_CONTENT = [
    ("ability.xlsx", "Ability", 2, (2, 3)),     # 2 数据行 × (B,C) = 4 格
    ("reward.xlsx", "Reward", 2, (2,)),          # 2 格
    ("const.xlsx", "Const", 2, (2,)),            # 2 格
    ("fabao.xlsx", "Fabao", 2, (2,)),            # 2 格（同表 sheet1）
    ("fabao.xlsx", "FabaoLevel", 2, (2,)),      # 2 格（同表 sheet2 → 同表不同 sheet 冲突）
]
# 2) 新增 sheet（结构差异：每分支加 DevNote_{branch}，合并不在 common → extra/missing sheet）
RICH_NEW_SHEET_TABLE = "const.xlsx"
# 3) 新增行（插入）：每分支独有 PK（单向 inserted）+ 共享 PK 9999（不同内容 → 同 PK 插入冲突）
RICH_INSERT_TABLE = "ability.xlsx"
RICH_INSERT_SHEET = "Ability"
RICH_INSERT_SHARED_PK = 9999  # 所有分支都插此 PK，写不同值 → 合并时插入行 PK 冲突
ALL_BRANCHES = ["dev1", "dev2", "dev3", "dev4"]

# 目录合并（subdev_N → trunk）冲突制造：subdev_N/table 与 trunk/table 同名表同单元格
# 双方都改、值不同即冲突。stage4 只改 subdev 一侧（B2 私有标记）→ 单向变更非冲突。
# 用户反馈"subdev_2 和 trunk 没有冲突、数量太少无参考价值"，要求增至 10-20 个。
# 3 单元格 × 5 表 × 2 子目录 = 30 个双方冲突（每子目录 15，落在 10-20）。
SUBDIR_COLS = (2,)          # 冲突列固定为名称列(B)


def _name_variant(base_val, branch: str) -> str:
    """名称列(string)相对 base 的语义替换：换成语义相似但真实不同的词，非后缀占位符。

    覆盖种子数据真实值（item/ability/reward/monster/item_drop 名称列），
    每个 (原值, 分支) 固定映射保证幂等；monster/掉落 ID 式名称做前缀替换；
    未命中时兜底后缀（避免 None 空值破坏表语义）。
    """
    _SEMANTIC = {
        ("可罗", "dev1"): "可罗",  # 占位，避免误伤（下面按真实值覆盖）
        ("三味真火", "dev1"): "重水",
        ("三味真火", "dev2"): "南明离火",
        ("三味真火", "dev3"): "九天玄火",
        ("三味真火", "dev4"): "红莲业火",
        ("三味真火", "trunk"): "弱水",
        ("三味真火", "subdev_2"): "玄冥重水",
        ("三味真火", "subdev_3"): "幽冥鬼火",
        ("蛮牛狂击", "dev1"): "莽牛冲撞",
        ("蛮牛狂击", "dev2"): "狂暴牛袭",
        ("蛮牛狂击", "dev3"): "蛮牛践踏",
        ("蛮牛狂击", "dev4"): "牛魔冲撞",
        ("蛮牛狂击", "trunk"): "狂奔牛突",
        ("蛮牛狂击", "subdev_2"): "莽牛撞",
        ("蛮牛狂击", "subdev_3"): "疯牛顶",
        ("雨露均沾", "dev1"): "甘霖普降",
        ("雨露均沾", "dev2"): "春风化雨",
        ("雨露均沾", "dev3"): "泽被苍生",
        ("雨露均沾", "dev4"): "久旱逢霖",
        ("雨露均沾", "trunk"): "雨过天青",
        ("雨露均沾", "subdev_2"): "普降甘霖",
        ("雨露均沾", "subdev_3"): "沾衣欲湿",
        ("测试道具", "dev1"): "测试法器",
        ("测试道具", "dev2"): "试炼道具",
        ("测试道具", "dev3"): "测试灵器",
        ("测试道具", "dev4"): "试炼符",
        ("测试道具", "trunk"): "试炼符",
        ("测试道具", "subdev_2"): "测试灵器",
        ("测试道具", "subdev_3"): "试炼道具",
        ("金流露", "dev1"): "金露膏",
        ("金流露", "dev2"): "玉露散",
        ("金流露", "dev3"): "金露液",
        ("金流露", "dev4"): "玉露膏",
        ("金流露", "trunk"): "金露液",
        ("金流露", "subdev_2"): "玉露膏",
        ("金流露", "subdev_3"): "金露膏",
        ("进化石", "dev1"): "进化结晶",
        ("进化石", "dev2"): "突破石",
        ("进化石", "dev3"): "演化石",
        ("进化石", "dev4"): "进阶石",
        ("进化石", "trunk"): "演化石",
        ("进化石", "subdev_2"): "进阶石",
        ("进化石", "subdev_3"): "突破石",
        ("宝石原石", "dev1"): "宝石矿母",
        ("宝石原石", "dev2"): "璞玉原矿",
        ("宝石原石", "dev3"): "灵石原胚",
        ("宝石原石", "dev4"): "原石胚",
        ("宝石原石", "trunk"): "灵石原胚",
        ("宝石原石", "subdev_2"): "原石胚",
        ("宝石原石", "subdev_3"): "璞玉原矿",
        ("宝石精华", "dev1"): "宝石精髓",
        ("宝石精华", "dev2"): "灵石精华",
        ("宝石精华", "dev3"): "玉髓精华",
        ("宝石精华", "dev4"): "晶核精华",
        ("宝石精华", "trunk"): "玉髓精华",
        ("宝石精华", "subdev_2"): "晶核精华",
        ("宝石精华", "subdev_3"): "灵石精华",
        ("测试奖励1", "dev1"): "试炼奖励一",
        ("测试奖励1", "dev2"): "历练奖励壹",
        ("测试奖励1", "dev3"): "挑战奖励甲",
        ("测试奖励1", "dev4"): "任务奖励壹",
        ("测试奖励1", "trunk"): "挑战奖励甲",
        ("测试奖励1", "subdev_2"): "任务奖励壹",
        ("测试奖励1", "subdev_3"): "历练奖励壹",
        ("境界1药物", "dev1"): "筑基灵药",
        ("境界1药物", "dev2"): "练气妙药",
        ("境界1药物", "dev3"): "凝气灵液",
        ("境界1药物", "dev4"): "炼体灵汤",
        ("境界1药物", "trunk"): "凝气灵液",
        ("境界1药物", "subdev_2"): "炼体灵汤",
        ("境界1药物", "subdev_3"): "练气妙药",
    }
    sv = str(base_val).strip() if base_val is not None else ""
    key = (sv, branch)
    if key in _SEMANTIC and _SEMANTIC[key]:
        return _SEMANTIC[key]
    # monster/掉落 ID 式名称前缀替换
    id_prefix = {
        "怪物_": {"dev1": "妖灵_", "dev2": "精怪_", "dev3": "异兽_", "dev4": "灵兽_",
                  "trunk": "魔兽_", "subdev_2": "妖灵_", "subdev_3": "精怪_"},
        "掉落_": {"dev1": "秘宝_", "dev2": "奇珍_", "dev3": "灵物_", "dev4": "异材_",
                  "trunk": "灵物_", "subdev_2": "异材_", "subdev_3": "奇珍_"},
    }
    for prefix, mapping in id_prefix.items():
        if sv.startswith(prefix):
            return mapping.get(branch, "变异_") + sv[len(prefix):]
    # 未命中兜底后缀
    suffix = {
        "dev1": "·精", "dev2": "·极", "dev3": "·异", "dev4": "·变",
        "subdev_1": "·异", "subdev_2": "·私", "subdev_3": "·独", "trunk": "·改",
    }.get(branch, "·改")
    return sv + suffix


def _int_variant(base_val, branch: str, step: int = 1) -> int:
    """int 列相对 base +N 微调（等级/品质等），分支不同 step 保证冲突。

    例：base=5 → dev1=6, dev2=7, dev3=8, dev4=9, trunk=10。
    """
    try:
        b = int(float(base_val))
    except (TypeError, ValueError):
        b = 0
    order = {"dev1": 1, "dev2": 2, "dev3": 3, "dev4": 4,
             "subdev_2": 1, "subdev_3": 2, "trunk": 5}.get(branch, 1)
    return b + step * order


def _cell_val(xlsx: Path, sheet: str, row: int, col: int):
    """读 base 单元格当前值（read_only）。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    v = wb[sheet].cell(row, col).value
    wb.close()
    return v


def _run(cmd, check=True, timeout=600, cwd=None):
    r = subprocess.run(cmd, capture_output=True, text=True,
                       encoding="utf-8", errors="replace", timeout=timeout, cwd=cwd)
    if check and r.returncode != 0:
        err = (r.stderr or "").strip()
        out = (r.stdout or "").strip()
        raise RuntimeError(f"命令失败 [{' '.join(cmd)}]\n  stdout: {out[:800]}\n  stderr: {err[:800]}")
    return r


def _wc_root(p: Path):
    """给定工作副本内某路径，返回其所属 SVN 工作副本根目录（含 .svn 的最近祖先）。"""
    cur = p if p.is_dir() else p.parent
    while cur != cur.parent:
        if (cur / ".svn").exists():
            return cur
        cur = cur.parent
    return None


def _svn(*args, check=True, timeout=600, cwd=None):
    """执行 svn，自动把绝对文件系统路径改写为 (cwd=工作副本根, 相对路径)。

    本机 TortoiseSVN/unisvn 对绝对路径做大小写解析时偶发 E720005，但相对路径 + cwd
    稳定。URL（file:// 等）与非路径参数原样透传，不受影响。
    """
    new_args = []
    wc_cwd = cwd
    for a in args:
        s = str(a)
        pa = Path(s)
        if pa.is_absolute():
            root = _wc_root(pa)
            if root is not None:
                if wc_cwd is None:
                    wc_cwd = str(root)
                rel = "." if pa == root else pa.relative_to(root).as_posix()
                new_args.append(rel)
                continue
        new_args.append(s)
    return _run(["svn", *new_args], check=check, timeout=timeout, cwd=wc_cwd)


def _repo_url() -> str:
    p = str(REPO_DIR).replace("\\", "/")
    return f"file:///{p}"


def _head_rev() -> int:
    r = _svn("info", "--show-item", "revision", _repo_url())
    return int(r.stdout.strip())


def _branch_exists(branch: str) -> bool:
    r = _svn("info", f"{_repo_url()}/branches/{branch}", check=False)
    return r.returncode == 0


def _svn_path_exists(url_path: str) -> bool:
    r = _svn("info", url_path, check=False)
    return r.returncode == 0


def _set_cell(xlsx: Path, sheet: str, row: int, col: int, value) -> bool:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[sheet]
    if ws.cell(row, col).value == value:
        wb.close()
        return False
    ws.cell(row, col).value = value
    wb.save(xlsx)
    wb.close()
    return True


def _first_sheet(xlsx: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    name = wb.sheetnames[0]
    wb.close()
    return name


# 非数据 sheet 关键词（与 server/engine/id_scope._NON_DATA_SHEET_KEYWORDS 对齐）：
# 合并 compare 跳过这些 sheet，故冲突标记必须写在数据 sheet 上才可见。
# tips.xlsx 等"首个 sheet 是 CONFIG"的表，_first_sheet 会取到 CONFIG（被跳过）→ 标记不可见。
_NON_DATA_SHEET_KEYWORDS = ("CONFIG", "PATCH_CONFIG", "说明", "SETTING", "INDEX")


def _first_data_sheet(xlsx: Path) -> str:
    """首个数据 sheet（跳过 CONFIG/PATCH_CONFIG/说明 等被合并 compare 忽略的 sheet）。
    全是非数据 sheet 时回退首个（保旧行为）。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    names = wb.sheetnames
    wb.close()
    for n in names:
        if not any(kw in n for kw in _NON_DATA_SHEET_KEYWORDS):
            return n
    return names[0]


def _svn_commit_as(wc_path: Path, msg: str, username: str):
    """svn commit 用指定作者。file:// repo 用 --username 设 svn:author，使两侧不同作者
    → compare 的 commit_authors 判定为真冲突（不同作者）而非同作者自动合并（D3 规则
    会把同作者同格改动自动合并、不算冲突）。--non-interactive --no-auth-cache 防 prompt。
    """
    return _svn("commit", str(wc_path), "--username", username, "--password", "",
                "--non-interactive", "--no-auth-cache", "-m", msg)


def _find_data_rows(xlsx: Path, sheet: str, count: int) -> list[int]:
    """返回前 count 个"A 列非空且唯一"的行号（数据行）。compare 按主键(A 列)行匹配，
    空/重复 A 列的行会折叠成 1 行 → 15 格只出 3 冲突（tips.xlsx 的坑）。故冲突单元格
    必须写在 A 列唯一的数据行上。从 row 3 起跳过表头(1)+类型(2)。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows: list[int] = []
    seen: set[str] = set()
    for r in range(3, (ws.max_row or 0) + 1):
        v = ws.cell(r, 1).value
        if v is None or v == "":
            continue
        key = str(v)
        if key in seen:
            continue
        seen.add(key)
        rows.append(r)
        if len(rows) >= count:
            break
    wb.close()
    return rows


def _edit_xlsx(xlsx: Path, fn) -> bool:
    """可写打开 xlsx，应用 fn(wb)（多编辑批量），保存。返回是否有写盘（fn 返回 True 表改了）。
    比 _set_cell 逐格 open/save 高效，且支持 create_sheet/insert_rows 等结构编辑。
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    changed = bool(fn(wb))
    if changed:
        wb.save(xlsx)
    wb.close()
    return changed


def _ensure_sheet(wb, sheet: str) -> bool:
    """确保 sheet 存在（不存在则建空 sheet）。返回是否新建（True=本次新建）。"""
    if sheet in wb.sheetnames:
        return False
    wb.create_sheet(sheet)
    return True


def _append_row(ws, values: list) -> bool:
    """在 ws 末尾追加一行（values[0]→A, [1]→B…）。返回 True。"""
    ws.append(values)
    return True


def _data_rows_ws(ws, count: int) -> list[int]:
    """同 _find_data_rows，但直接在已打开的 ws 上找（供 _edit_xlsx 内批量编辑用）。"""
    rows: list[int] = []
    seen: set[str] = set()
    for r in range(3, (ws.max_row or 0) + 1):
        v = ws.cell(r, 1).value
        if v is None or v == "":
            continue
        key = str(v)
        if key in seen:
            continue
        seen.add(key)
        rows.append(r)
        if len(rows) >= count:
            break
    return rows


def _pk_row(ws, pk) -> int:
    """返回 A 列等于 pk 的行号，无则 0。"""
    for r in range(3, (ws.max_row or 0) + 1):
        if ws.cell(r, 1).value == pk:
            return r
    return 0


def stage0_check_base():
    print("=== [0/5] 检查基础 repo 是否就绪 ===")
    if not REPO_DIR.exists() or not TRUNK_WC.is_dir():
        print(f" {TRUNK_WC} 不存在。", file=sys.stderr)
        print("  请先运行：python merge/scripts/setup_svn_demo.py", file=sys.stderr)
        sys.exit(2)
    rev = _head_rev()
    print(f"  repo 就绪，HEAD r{rev}")
    return rev


def stage1_branch_off(trunk_rev: int, url: str):
    """从 /trunk@r 切出 dev3/dev4 平行分支（已存在则跳过）。"""
    print(f"=== [1/5] 从 /trunk@r{trunk_rev} 切出平行分支 dev3/dev4 ===")
    for name in NEW_BRANCHES:
        if _branch_exists(name):
            print(f"  branches/{name}: 已存在，跳过 copy")
            continue
        _svn("copy", f"{url}/trunk@{trunk_rev}", f"{url}/branches/{name}",
             "-m", f"{name}: 从 trunk@r{trunk_rev} 切出（小表分支，排除 {', '.join(sorted(BIG_TABLES))}）")
        print(f"  branches/{name}: 已切出")


def stage2_checkout(url: str):
    print("=== [2/5] checkout dev3/dev4 工作副本 ===")
    for name in NEW_BRANCHES:
        wc = WC_DIR / "branches" / name
        if (wc / ".svn").is_dir():
            print(f"  wc/branches/{name}: 已存在，revert+update")
            _svn("revert", "-R", str(wc), check=False)
            _svn("update", str(wc))
            continue
        wc.parent.mkdir(parents=True, exist_ok=True)
        # 清掉 clean 阶段留下的空目录（存在空目录会让 svn checkout 触发大小写解析失败）
        if wc.exists() and not (wc / ".svn").is_dir():
            shutil.rmtree(wc, onerror=lambda f, p, e: (os.chmod(p, stat.S_IWRITE), f(p)))
        # 相对路径 + cwd=父目录 检出，规避 Windows 上绝对路径大小写不一致导致的 E720005
        _svn("checkout", f"{url}/branches/{name}", name, cwd=str(wc.parent))
        print(f"  wc/branches/{name}")


def stage3_dev_branch(branch: str, modify_tables: list[str]):
    """dev3/dev4：删 3 大表，改若干小表（与 trunk 形成差异）。

    改动相对 trunk 基准真实差异化：B2 名称列做语义替换（如"狂刀"→"狂刀·异"），
    非 {branch}_mark 占位符。分支不同后缀 → dev3/dev4 同表同行冲突。
    """
    print(f"=== [3/5] 整理 branches/{branch}：删大表 + 改小表 ===")
    wc = WC_DIR / "branches" / branch
    for big in BIG_TABLES:
        p = wc / big
        if p.exists():
            _svn("rm", str(p))
            print(f"  rm {big}")
    for name in modify_tables:
        fp = wc / name
        if not fp.exists():
            print(f"  跳过 {name}（不存在）")
            continue
        sheet = _first_sheet(fp)
        # base = trunk 同名表同格当前值
        trunk_fp = TRUNK_WC / name
        base_val = _cell_val(trunk_fp, sheet, 2, 2) if trunk_fp.exists() else None
        marker = _name_variant(base_val, branch)
        if _set_cell(fp, sheet, 2, 2, marker):
            print(f"  {name}!{sheet}!B2 = {marker}")
        else:
            print(f"  {name}!{sheet}!B2 已是 {marker}，跳过")
    _svn("commit", str(wc), "-m",
         f"{branch}: 删除 3 大表({', '.join(sorted(BIG_TABLES))})，改 {len(modify_tables)} 张小表"
         f"({', '.join(modify_tables)})，相对基准真实差异化")
    print(f"  {branch} 提交完成")


def stage4_trunk_subdirs(url: str):
    """在 trunk 工作副本下创建/整理 subdev_1/2/3 子目录（目录合并 source）。

    每子目录内放入其私有表集合（SUBDIR_PLANS 定义）：
      - subdev_1 由 build_svn_real 随 trunk 导入，此处归一化为 8 张小表（去掉大表）
      - subdev_2/3 新建，各放目标张数小表
    从 trunk 顶层拷入同名表，svn add/rm 后随 trunk 一次提交。实际冲突制造在
    stage_subdir_trunk_conflicts 中统一完成（按 SUBDIR_PLANS 目标规模）。
    """
    print("=== [4/5] 在 trunk 下整理 subdev_1/2/3 子目录 ===")
    for subname, keep, _conf in SUBDIR_PLANS:
        sdir = TRUNK_WC / subname
        existed = sdir.is_dir() and (sdir / ".svn").is_dir()
        sdir.mkdir(parents=True, exist_ok=True)
        keep_set = set(keep)
        # 删除不在目标集合中的旧表（如 subdev_1 的 monster/item_drop 大表）
        for fp in list(sdir.iterdir()):
            if fp.is_file() and fp.suffix == ".xlsx" and fp.name not in keep_set:
                _svn("rm", str(fp), check=False)
                print(f"  rm {subname}/{fp.name}")
        # 拷入缺失的目标表
        for name in keep:
            src_fp = TRUNK_WC / name
            dst = sdir / name
            if dst.exists():
                continue
            if not src_fp.exists():
                print(f"  警告：trunk/{name} 不存在，跳过", file=sys.stderr)
                continue
            shutil.copy2(src_fp, dst)
        # 新拷入的文件在已版本化子目录下需显式 add（svn add 子目录对新目录递归；已有目录只加新文件）
        for name in keep:
            dst = sdir / name
            if not dst.exists():
                continue
            st = _svn("status", str(dst), check=False).stdout or ""
            if st.startswith("?") or st.startswith("I"):
                _svn("add", str(dst), check=False)
        if not existed:
            _svn("add", str(sdir), check=False)
        msg = (f"trunk/{subname}: 私有子目录归一化（{len(keep)} 张小表），目录合并 source"
               if not existed else f"trunk/{subname}: 刷新私有表集合为 {len(keep)} 张")
        _svn("commit", str(TRUNK_WC), "-m", msg)
        print(f"  trunk/{subname} 提交完成（{len(keep)} 张表）")


def stage_subdir_trunk_conflicts():
    """按 SUBDIR_PLANS 为 subdev_1/2/3 vs trunk 制造目标数量的真实冲突。

    目录合并 source=subdev_N（trunk 子目录）→ target=trunk，base=subdev 创建时 trunk 快照。
    冲突 = subdev_N/table 与 trunk/table 同名表同单元格都改且值不同。本函数对每张表取
    指定数量的数据行，在名称列(B)写不同语义值（subdev 侧 vs trunk 侧），构成双方都改的
    真冲突。subdev 子目录与 trunk 同名表都在 trunk 工作副本下，一次 commit trunk wc 覆盖。

    两侧分作者提交：commit_authors 同作者自动合并会把同作者同格改动算自动合并、不算冲突。
    故 subdev 侧用作者 subname 提交，trunk 侧用作者 trunk 提交 → 真冲突。demo_svn 单用户
    仓库在比对时已禁用同作者合并（use_authors={}），但仍保持分作者以贴近真实场景。

    每个冲突格对应一个 (subname, tname, sheet, row)：subdev 侧与 trunk 侧各写一次，
    总共 2 次提交（先全部 subdev 子树，后全部 trunk 顶层）。
    """
    print("=== [5/7] subdev_1/2/3 vs trunk 冲突制造（按 SUBDIR_PLANS 目标规模）===")
    targets = []  # (subname, tname, sheet, row) — 每个元素 = 1 个冲突单元格
    for subname, keep, conflict_counts in SUBDIR_PLANS:
        for tname, nconf in zip(keep, conflict_counts):
            sub_fp = TRUNK_WC / subname / tname
            trunk_fp = TRUNK_WC / tname
            if not sub_fp.exists() or not trunk_fp.exists():
                print(f"  警告：{subname}/{tname} 或 trunk/{tname} 不存在，跳过", file=sys.stderr)
                continue
            sheet = _first_data_sheet(sub_fp)
            rows = _find_data_rows(sub_fp, sheet, nconf)
            if len(rows) < nconf:
                print(f"  警告：{subname}/{tname}!{sheet} 数据行不足（需 {nconf}，得 {len(rows)}），跳过", file=sys.stderr)
                continue
            for r in rows:
                targets.append((subname, tname, sheet, r))

    # 阶段一：改并提交 subdev 子树（作者=subname）
    changed = 0
    for subname, tname, sheet, r in targets:
        sub_fp = TRUNK_WC / subname / tname
        seed_fp = SEED_TRUNK / tname
        base_val = _cell_val(seed_fp, sheet, r, 2) if seed_fp.exists() else _cell_val(sub_fp, sheet, r, 2)
        if _set_cell(sub_fp, sheet, r, 2, _name_variant(base_val, subname)):
            changed += 1
    for subname, _keep, _conf in SUBDIR_PLANS:
        _svn_commit_as(TRUNK_WC / subname,
             f"{subname}: 目录合并 source 侧制造冲突（作者 {subname}，相对基准真实差异化）",
             subname)
    print(f"  subdev 侧共改 {changed} 格")

    # 阶段二：改并提交 trunk 顶层（作者=trunk）
    changed = 0
    for subname, tname, sheet, r in targets:
        trunk_fp = TRUNK_WC / tname
        seed_fp = SEED_TRUNK / tname
        base_val = _cell_val(seed_fp, sheet, r, 2) if seed_fp.exists() else _cell_val(trunk_fp, sheet, r, 2)
        if _set_cell(trunk_fp, sheet, r, 2, _name_variant(base_val, "trunk")):
            changed += 1
    if changed:
        _svn_commit_as(TRUNK_WC,
             f"trunk: 目录合并 target 侧制造冲突（作者 trunk，相对基准真实差异化）",
             "trunk")
    print(f"  trunk 侧共改 {changed} 格")
    print(f"  冲突规模：{[(p[0], len(p[1]), sum(p[2])) for p in SUBDIR_PLANS]}")


def stage_rich_dev_conflicts():
    """dev1-4 全面冲突：分散多表 + 同表不同 sheet + 新增 sheet + 新增行(含同 PK 冲突)。

    用户要求"不应集中在一个表格，要分散多表、同表不同 sheet、新增 sheet 和行"。
    每个 dev 都参与每场景（写本分支值），任两 dev 合并在每场景都冲突/差异：
      1. content：5 个 表/sheet 组合（ability/reward/const/fabao×2 sheet），每表 2 数据行
         值相对 base（dev 分支 fork 时的 trunk 同名表）做语义替换（名称列）或 int+1（数值列）
      2. 新增 sheet：每分支在 const.xlsx 加 DevNote_{branch}（合并不在 common → 结构差异）
      3. 新增行：ability!Ability 插本分支独有 PK（单向 inserted）+ 共享 PK 9999（不同值 → 同 PK 插入冲突）
    分作者提交（--username {branch}）避同作者自动合并。幂等：值/sheet/行已存在则跳过。
    """
    print("=== [6/7] dev1-4 全面冲突(多表/多sheet/新sheet/新行) ===")
    for bi, branch in enumerate(ALL_BRANCHES):
        wc = WC_DIR / "branches" / branch
        edits = 0

        # 1+3) ability.xlsx：content(2 行×2 格) + 新增行(2 独有 PK + 共享 PK 9999)
        # base = 原始种子 trunk/ability.xlsx（不可变，避免被前置 stage 污染叠加）
        def _abl(wb, b=branch, i=bi):
            ch = 0
            ws = wb["Ability"] if "Ability" in wb.sheetnames else wb.active
            abl_base_fp = SEED_TRUNK / "ability.xlsx"
            abl_base_sheet = _first_data_sheet(abl_base_fp) if abl_base_fp.exists() else ws.title
            for r in _data_rows_ws(ws, 2):
                for c in (2, 3):
                    base_val = _cell_val(abl_base_fp, abl_base_sheet, r, c) if abl_base_fp.exists() else None
                    v = _name_variant(base_val, b)
                    if ws.cell(r, c).value != v:
                        ws.cell(r, c).value = v; ch += 1
            for pk in (9000 + i * 10 + 1, 9000 + i * 10 + 2, RICH_INSERT_SHARED_PK):
                # 共享 PK 9999 各分支写不同语义值 → 同 PK 插入冲突
                v = _name_variant(f"新增神通{pk}", b)
                rr = _pk_row(ws, pk)
                if rr == 0:
                    ws.append([pk, v, v, v]); ch += 1
                elif ws.cell(rr, 2).value != v:
                    ws.cell(rr, 2).value = v; ch += 1
            return ch > 0
        if (wc / "ability.xlsx").exists() and _edit_xlsx(wc / "ability.xlsx", _abl):
            edits += 1

        # reward.xlsx：content(2 行×B) 名称列语义替换
        def _rwd(wb, b=branch):
            ch = 0
            ws = wb["Reward"] if "Reward" in wb.sheetnames else wb.active
            rwd_base_fp = SEED_TRUNK / "reward.xlsx"
            rwd_base_sheet = _first_data_sheet(rwd_base_fp) if rwd_base_fp.exists() else ws.title
            for r in _data_rows_ws(ws, 2):
                base_val = _cell_val(rwd_base_fp, rwd_base_sheet, r, 2) if rwd_base_fp.exists() else None
                v = _name_variant(base_val, b)
                if ws.cell(r, 2).value != v:
                    ws.cell(r, 2).value = v; ch += 1
            return ch > 0
        if (wc / "reward.xlsx").exists() and _edit_xlsx(wc / "reward.xlsx", _rwd):
            edits += 1

        # const.xlsx：content(2 行×B) 名称列语义替换 + 新增 DevNote_{branch} sheet（结构差异）
        def _cst(wb, b=branch):
            ch = 0
            ws = wb["Const"] if "Const" in wb.sheetnames else wb.active
            cst_base_fp = SEED_TRUNK / "const.xlsx"
            cst_base_sheet = _first_data_sheet(cst_base_fp) if cst_base_fp.exists() else ws.title
            for r in _data_rows_ws(ws, 2):
                base_val = _cell_val(cst_base_fp, cst_base_sheet, r, 2) if cst_base_fp.exists() else None
                v = _name_variant(base_val, b)
                if ws.cell(r, 2).value != v:
                    ws.cell(r, 2).value = v; ch += 1
            sn = f"DevNote_{b}"
            if sn not in wb.sheetnames:
                nws = wb.create_sheet(sn)
                nws.append(["key", "note"]); nws.append([1, f"{b} 分支备注"])
                ch += 1
            return ch > 0
        if (wc / "const.xlsx").exists() and _edit_xlsx(wc / "const.xlsx", _cst):
            edits += 1

        # fabao.xlsx：content 跨 2 sheet(Fabao + FabaoLevel) → 同表不同 sheet 冲突，名称列语义替换
        def _fab(wb, b=branch):
            ch = 0
            fab_base_fp = SEED_TRUNK / "fabao.xlsx"
            for sn in ("Fabao", "FabaoLevel"):
                if sn not in wb.sheetnames:
                    continue
                ws = wb[sn]
                for r in _data_rows_ws(ws, 2):
                    base_val = _cell_val(fab_base_fp, sn, r, 2) if fab_base_fp.exists() else None
                    v = _name_variant(base_val, b)
                    if ws.cell(r, 2).value != v:
                        ws.cell(r, 2).value = v; ch += 1
            return ch > 0
        if (wc / "fabao.xlsx").exists() and _edit_xlsx(wc / "fabao.xlsx", _fab):
            edits += 1

        if edits:
            _svn_commit_as(wc,
                 f"{branch}: 全面冲突(5 表/sheet content + DevNote_{branch} 新 sheet + ability 新行含同 PK 冲突)，"
                 f"相对基准真实差异化，分作者提交",
                 branch)
            print(f"  branches/{branch}: content(ability/reward/const/fabao×2) + DevNote_{branch} sheet + ability 新行")
        else:
            print(f"  branches/{branch}: 已是本分支值，跳过")


def stage5_verify(url: str, trunk_rev: int):
    print(f"=== [7/7] 验证布局 ===")
    import re
    ok_all = True
    # dev3/dev4：copyfrom=/trunk 且 rev <= trunk_rev（分支切点可能早于当前 HEAD）+ 大表已删
    for name in NEW_BRANCHES:
        wc = WC_DIR / "branches" / name
        r = _svn("log", "-v", "--stop-on-copy", "--xml", str(wc), check=False)
        out = r.stdout or ""
        cf_path = re.search(r'copyfrom-path="([^"]+)"', out)
        cf_rev = re.search(r'copyfrom-rev="([^"]+)"', out)
        cf_ok = (cf_path and cf_path.group(1).endswith("/trunk") and
                 cf_rev and int(cf_rev.group(1)) <= trunk_rev)
        big_missing = all(not (wc / b).exists() for b in BIG_TABLES)
        status = "OK" if (cf_ok and big_missing) else "WARN"
        if status != "OK":
            ok_all = False
        print(f"  branches/{name}: copyfrom={'/trunk@r'+cf_rev.group(1) if cf_rev else '?'} "
              f"大表已删={big_missing} [{status}]")
    # trunk/subdev_1|2|3：子目录存在 + 含目标张数小表
    for subname, keep, _conf in SUBDIR_PLANS:
        sdir = TRUNK_WC / subname
        has_all = all((sdir / n).exists() for n in keep)
        status = "OK" if has_all else "WARN"
        if status != "OK":
            ok_all = False
        print(f"  trunk/{subname}: 含 {len(keep)} 张小表={has_all} [{status}]")
    return ok_all


def _clean(url: str):
    """删除已存在的 dev3/dev4 分支 + trunk/subdev_2|3 子目录 + 遗留的 branches/subdev_2|3 脏分支。"""
    print("=== [clean] 删除旧 dev3/dev4 + trunk/subdev_2|3 + 遗留 branches/subdev_2|3 ===")
    for name in NEW_BRANCHES:
        if _branch_exists(name):
            _svn("rm", f"{url}/branches/{name}", "-m", f"clean: 删除旧 {name}")
            print(f"  删除 /branches/{name}")
        wc = WC_DIR / "branches" / name
        if (wc / ".svn").is_dir():
            shutil.rmtree(wc, onerror=lambda f, p, e: (os.chmod(p, stat.S_IWRITE), f(p)))
            print(f"  删除 wc/branches/{name}")
    # 遗留：旧脚本曾把 subdev_2|3 建成 branches/ 下分支，此处一并清掉
    for subname in NEW_SUBDIRS:
        if _branch_exists(subname):
            _svn("rm", f"{url}/branches/{subname}", "-m", f"clean: 删除遗留 branches/{subname}（应作 trunk 子目录）")
            print(f"  删除遗留 /branches/{subname}")
        wc = WC_DIR / "branches" / subname
        if (wc / ".svn").is_dir():
            shutil.rmtree(wc, onerror=lambda f, p, e: (os.chmod(p, stat.S_IWRITE), f(p)))
            print(f"  删除遗留 wc/branches/{subname}")
    # 遗留：旧脚本曾把 subdev_1 建成 branches/ 下分支（现改作 trunk 子目录），一并清掉
    if _branch_exists("subdev_1"):
        _svn("rm", f"{url}/branches/subdev_1", "-m",
             "clean: 删除遗留 branches/subdev_1（subdev_1 改作 trunk 子目录，不再作分支）")
        print(f"  删除遗留 /branches/subdev_1")
    wc_sub1 = WC_DIR / "branches" / "subdev_1"
    if (wc_sub1 / ".svn").is_dir() or wc_sub1.exists():
        shutil.rmtree(wc_sub1, onerror=lambda f, p, e: (os.chmod(p, stat.S_IWRITE), f(p)))
        print(f"  删除遗留 wc/branches/subdev_1")
    # trunk 下子目录：svn rm 后随 trunk commit 删除
    for subname in NEW_SUBDIRS:
        sdir = TRUNK_WC / subname
        if sdir.is_dir() and (sdir / ".svn").is_dir():
            _svn("rm", "-R", str(sdir), check=False)
            print(f"  svn rm trunk/{subname}")
    status_out = _svn("status", str(TRUNK_WC), check=False).stdout
    if status_out.strip():
        _svn("commit", str(TRUNK_WC), "-m", "clean: 删除 trunk/subdev_2|3 子目录")
        print("  提交 trunk 子目录删除")


def main():
    ap = argparse.ArgumentParser(description="扩展 demo_svn：新增 dev3/dev4 平行分支 + trunk/subdev_2|3 子目录")
    ap.add_argument("--src", default=str(DEFAULT_SRC), help=f"源数据目录（默认 {DEFAULT_SRC}）")
    ap.add_argument("--clean-branches", action="store_true",
                    help="先删除 dev3/dev4 分支 + trunk/subdev_2|3 子目录再重建")
    args = ap.parse_args()

    src = Path(args.src)
    if not src.is_dir():
        print(f"错误：源目录不存在: {src}", file=sys.stderr)
        sys.exit(2)

    trunk_rev = stage0_check_base()
    url = _repo_url()

    if args.clean_branches:
        _clean(url)

    stage1_branch_off(trunk_rev, url)
    stage2_checkout(url)
    stage3_dev_branch("dev3", DEV3_MODIFY)
    stage3_dev_branch("dev4", DEV4_MODIFY)
    stage4_trunk_subdirs(url)
    stage_subdir_trunk_conflicts()
    stage_rich_dev_conflicts()
    ok = stage5_verify(url, trunk_rev)

    print("\n" + "=" * 60)
    print("小表分支/子目录扩展完成。")
    print("  平行分支（/api/merge/branch/dirs）：")
    print("    svn/demo_svn/wc/branches/dev3  ← 跨分支合并（小表，已删 monster/skill_level/item_drop）")
    print("    svn/demo_svn/wc/branches/dev4  ← 跨分支合并（小表）")
    print("  trunk 下子目录（/api/merge/subdir/dirs 递归扫描，目录合并 source）：")
    print("    svn/demo_svn/wc/trunk/subdev_1  ← 8 张表，16 冲突")
    print("    svn/demo_svn/wc/trunk/subdev_2  ← 21 张表，31 冲突")
    print("    svn/demo_svn/wc/trunk/subdev_3  ← 15 张表，23 冲突")
    print("  排除大表：", ", ".join(sorted(BIG_TABLES)))
    if not ok:
        print("\n注意：部分校验 [WARN]，请人工核对。", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
