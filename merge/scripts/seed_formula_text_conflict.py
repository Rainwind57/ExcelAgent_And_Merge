"""制造一个真正的"公式文本冲突"（diff_type=formula_conflict），供合并引导页截图验证。

之前几个 seed 脚本造的冲突都是"值冲突"（普通列），seed_formula_row_drift.py 造的是
"公式行漂移"（非冲突，公式文本没变）。目前 demo 里所有含公式的表（guild.xlsx /
assistant_weapon.xlsx / residence_putuan.xlsx）dev1/dev2 的公式**文本**都跟 trunk
一样，没有一处真正的公式文本冲突——本脚本补上这个场景。

目标：guild.xlsx · Const sheet · E27（GUILD_APPLY_VALIDITY 行，E 列是独立于 C 列
"配置值"的公式注释列，不影响 A 列主键，改它不会破坏 PK 匹配）。
  trunk 保持 "=1 * 86400"
  dev1  改成 "=2 * 86400"
  dev2  改成 "=3 * 86400"
三边公式文本互不相同 → compare_sheet 判定 diff_type="formula_conflict"，
conflict=True，前端会弹"选择采纳的公式版本"。

幂等：目标单元格已是期望值则跳过对应分支的提交。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
TABLE_REL = Path("guild.xlsx")
SHEET = "Const"
PK_COL = 1
TARGET_PK = "GUILD_APPLY_VALIDITY"
FORMULA_COL = 5  # E 列


def _find_row_by_pk(ws, pk: str, max_row: int = 40) -> int:
    for r in range(1, max_row + 1):
        if ws.cell(r, PK_COL).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def _set_formula(branch_dir: Path, new_formula: str, msg: str) -> None:
    fp = branch_dir / TABLE_REL
    if not fp.exists():
        print(f"  [skip] {branch_dir.name} 缺少 {TABLE_REL}")
        return
    wb = load_workbook(fp, data_only=False)
    if SHEET not in wb.sheetnames:
        wb.close()
        print(f"  [skip] {branch_dir.name} {TABLE_REL} 缺少 sheet {SHEET}")
        return
    ws = wb[SHEET]
    row = _find_row_by_pk(ws, TARGET_PK)
    if row < 0:
        wb.close()
        print(f"  [skip] {branch_dir.name} 未找到 PK={TARGET_PK}")
        return
    if ws.cell(row, FORMULA_COL).value == new_formula:
        wb.close()
        print(f"  [skip] {branch_dir.name} 已是目标公式（幂等）")
        return
    ws.cell(row, FORMULA_COL).value = new_formula
    wb.save(fp)
    wb.close()
    print(f"  {branch_dir.name}: 第{row}行 E列 -> {new_formula}")
    _commit(branch_dir, msg)


def main() -> None:
    print("=== 制造公式文本冲突（formula_conflict）===")
    _set_formula(DEV1, "=2 * 86400", "dev1: 制造公式文本冲突（guild.xlsx Const E27）")
    _set_formula(DEV2, "=3 * 86400", "dev2: 制造公式文本冲突（guild.xlsx Const E27）")
    print("\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2，"
          f"表 guild / sheet Const，PK={TARGET_PK}（第27行）E 列——三边公式文本互不"
          "相同，格子标红，点开会弹出选择采纳哪个公式版本的冲突框。")


if __name__ == "__main__":
    main()
