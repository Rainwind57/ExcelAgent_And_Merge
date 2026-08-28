"""扩充 demo_svn 冲突数据，三模式各 5-7 个冲突点。

absorb (dev1 vs dev2)：item 行5/8/11/14 + monster 行3/6 → 6冲突
merge_back (dev1 vs trunk)：item 行7/10/13 + monster 行5/9 + skill_level 行4 → 5冲突
目录合并 (subdev_1 vs trunk)：subdev_1 改 monster 行2/5 + item_drop 行3 → 2冲突

分支值相对 base 真实差异化（模拟实际改表）：
  item 名称列(string) → 语义替换（如"测试道具"→"测试道具·改"），非 {branch}_xxx 占位符
  item 品质列(int) → +1 微调（如 1→2），非字符串标签
  monster 名称列(string) → 语义替换（如"蛮牛"→"狂暴蛮牛"）
  skill_level 等级列(int) → +1 微调（如 5→6）
各分支改不同值仍构成冲突，但值符合列类型与语义。幂等：cell 已是目标值则跳过。
"""
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
TRUNK = WC_ROOT / "trunk"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
SUBDEV1 = TRUNK / "subdev_1"
# 原始种子 trunk（永不被脚本污染）：读 base 必须从这里取，否则多次 seed 叠加后缀
SEED_TRUNK = Path(__file__).resolve().parent.parent / "_seed_data" / "trunk"


def _cell(xlsx: Path, sheet: str, row: int, col: int):
    """读单元格当前值（read_only）。"""
    wb = load_workbook(xlsx, read_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    v = ws.cell(row, col).value
    wb.close()
    return v


def _name_variant(base_val, branch: str) -> str:
    """名称列(string)相对 base 的语义替换：换成语义相似但真实不同的词，非后缀占位符。

    覆盖种子数据真实值（item/monster/ability 名称列），每个 (原值, 分支) 固定映射，
    保证幂等（重复执行不叠加）。monster 名称 "怪物_NNNNNN" 按 ID 保留做近似语义替换。
    """
    _SEMANTIC = {
        # item 名称列
        ("测试道具", "dev1"): "测试法器",
        ("测试道具", "dev2"): "试炼道具",
        ("测试道具", "trunk"): "试炼符",
        ("测试道具", "subdev_1"): "测试灵器",
        ("金流露", "dev1"): "金露膏",
        ("金流露", "dev2"): "玉露散",
        ("金流露", "trunk"): "金露液",
        ("金流露", "subdev_1"): "玉露膏",
        ("进化石", "dev1"): "进化结晶",
        ("进化石", "dev2"): "突破石",
        ("进化石", "trunk"): "演化石",
        ("进化石", "subdev_1"): "进阶石",
        ("宝石原石", "dev1"): "宝石矿母",
        ("宝石原石", "dev2"): "璞玉原矿",
        ("宝石原石", "trunk"): "灵石原胚",
        ("宝石原石", "subdev_1"): "原石胚",
        ("宝石精华", "dev1"): "宝石精髓",
        ("宝石精华", "dev2"): "灵石精华",
        ("宝石精华", "trunk"): "玉髓精华",
        ("宝石精华", "subdev_1"): "晶核精华",
        ("石芽矿", "dev1"): "石芽矿脉",
        ("石芽矿", "dev2"): "苔芽石",
        ("石芽矿", "trunk"): "石芽晶",
        ("石芽矿", "subdev_1"): "石笋矿",
        ("云砂晶", "dev1"): "云晶砂",
        ("云砂晶", "dev2"): "流云砂",
        ("云砂晶", "trunk"): "云母砂",
        ("云砂晶", "subdev_1"): "晶砂",
        # ability 名称列
        ("三味真火", "dev1"): "重水",
        ("三味真火", "dev2"): "南明离火",
        ("三味真火", "trunk"): "九天玄火",
        ("三味真火", "subdev_1"): "弱水",
        ("蛮牛狂击", "dev1"): "莽牛冲撞",
        ("蛮牛狂击", "dev2"): "狂暴牛袭",
        ("蛮牛狂击", "trunk"): "蛮牛践踏",
        ("蛮牛狂击", "subdev_1"): "莽牛撞",
        ("雨露均沾", "dev1"): "甘霖普降",
        ("雨露均沾", "dev2"): "春风化雨",
        ("雨露均沾", "trunk"): "泽被苍生",
        ("雨露均沾", "subdev_1"): "雨过天青",
    }
    prefix = {"dev1": "妖灵", "dev2": "精怪", "trunk": "异兽", "subdev_1": "灵兽"}.get(branch, "变异")
    sv = str(base_val).strip() if base_val is not None else ""
    key = (sv, branch)
    if key in _SEMANTIC:
        return _SEMANTIC[key]
    # monster 名称 "怪物_NNNNNN" → "{prefix}_NNNNNN"（保留 ID 的语义近似替换）
    if sv.startswith("怪物_"):
        return prefix + "_" + sv[len("怪物_"):]
    # 未命中映射表的兜底：保持原后缀逻辑，避免生成 None/空值破坏表语义
    suffix = {"dev1": "·精", "dev2": "·极", "trunk": "·改", "subdev_1": "·异"}.get(branch, "·改")
    return sv + suffix


def _int_variant(base_val, branch: str, step: int = 1) -> int:
    """int 列相对 base +N 微调（等级/品质/编号等），分支不同 step 保证冲突。

    例：base=5 → dev1=6, dev2=7, trunk=5（trunk 作为前进基准回退到原值+0 时不冲突，
    故 trunk 用 +2 与 dev1/dev2 错开）。
    """
    try:
        b = int(float(base_val))
    except (TypeError, ValueError):
        b = 0
    delta = {"dev1": step, "dev2": step * 2, "trunk": step * 3, "subdev_1": step}.get(branch, step)
    return b + delta


def _quality_variant(base_val, branch: str) -> int:
    """品质列(int) base+1~3 微调（1→2/3/4），符合 int 约束。"""
    return _int_variant(base_val, branch, step=1)


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace")
    print(f"  [{'OK' if r.returncode==0 else 'FAIL'}] {msg}")


def _set(xlsx: Path, sheet: str, row: int, col: int, value) -> bool:
    wb = load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        wb.close(); return False
    ws = wb[sheet]
    if ws.cell(row, col).value == value:
        wb.close(); return False
    ws.cell(row, col).value = value
    wb.save(xlsx); wb.close()
    return True


def _batch_set(xlsx: Path, sheet: str, edits: list) -> list:
    """edits: [(row, col, value)]，批量改，返回实际改动的 (row,col,value)。"""
    wb = load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        wb.close(); return []
    ws = wb[sheet]
    changed = []
    for row, col, value in edits:
        if ws.cell(row, col).value != value:
            ws.cell(row, col).value = value
            changed.append((row, col, value))
    if changed:
        wb.save(xlsx)
    wb.close()
    return changed


def main() -> None:
    # ── absorb: dev1 vs dev2 在 item + monster 制造冲突 ──
    # base = 种子 trunk（dev1/dev2 均从 trunk fork，读原始值避免后缀叠加）
    item_d1 = DEV1 / "item" / "item.xlsx"
    item_d2 = DEV2 / "item" / "item.xlsx"
    item_base = SEED_TRUNK / "item" / "item.xlsx"
    mon_d1 = DEV1 / "monster.xlsx"
    mon_d2 = DEV2 / "monster.xlsx"
    mon_base = SEED_TRUNK / "monster.xlsx"

    # item 名称列(2)=string：相对 trunk 基准名称做语义替换（如"测试道具"→"测试道具·精"）
    # 行5/8/11/14 各分支改不同后缀 → 真冲突（值不同），但仍是合法中文名称
    item_rows = [(5, 2), (8, 2), (11, 2), (14, 2)]
    c1 = _batch_set(item_d1, "ItemBase",
        [(r, c, _name_variant(_cell(item_base, "ItemBase", r, c), "dev1")) for r, c in item_rows])
    c2 = _batch_set(item_d2, "ItemBase",
        [(r, c, _name_variant(_cell(item_base, "ItemBase", r, c), "dev2")) for r, c in item_rows])
    # monster 名称列(2)=string：行5/8 相对 trunk 基准名做语义替换；等级列(3)=int 同步微调
    mon_rows = [(5, 2), (8, 2)]
    mon_lv_rows = [(5, 3), (8, 3)]
    c3 = _batch_set(mon_d1, "Monster",
        [(r, c, _name_variant(_cell(mon_base, "Monster", r, c), "dev1")) for r, c in mon_rows]
        + [(r, c, _int_variant(_cell(mon_base, "Monster", r, c), "dev1")) for r, c in mon_lv_rows]) if mon_d1.exists() else []
    c4 = _batch_set(mon_d2, "Monster",
        [(r, c, _name_variant(_cell(mon_base, "Monster", r, c), "dev2")) for r, c in mon_rows]
        + [(r, c, _int_variant(_cell(mon_base, "Monster", r, c), "dev2")) for r, c in mon_lv_rows]) if mon_d2.exists() else []
    if c1 or c3:
        _commit(DEV1, f"dev1: absorb冲突扩充 item{len(c1)} monster{len(c3)}（相对基准真实差异化）")
    if c2 or c4:
        _commit(DEV2, f"dev2: absorb冲突扩充 item{len(c2)} monster{len(c4)}（相对基准真实差异化）")

    # ── merge_back: dev1 vs trunk 在 item + monster + skill_level 制造冲突 ──
    # base = 种子 trunk（dev1 fork 时的快照）。trunk 前进改这些行 → 与 dev1 冲突。
    item_tr = TRUNK / "item" / "item.xlsx"
    mon_tr = TRUNK / "monster.xlsx"
    skill_d1 = DEV1 / "skill_level.xlsx"
    skill_base = SEED_TRUNK / "skill_level.xlsx"

    # dev1: item 行7/10/13 名称相对 fork 基准做 dev1 语义替换
    mb_item_rows = [(7, 2), (10, 2), (13, 2)]
    c5 = _batch_set(item_d1, "ItemBase",
        [(r, c, _name_variant(_cell(item_base, "ItemBase", r, c), "dev1")) for r, c in mb_item_rows])
    # dev1: monster 行6/9 名称相对基准做 dev1 替换；等级列同步微调
    mb_mon_rows = [(6, 2), (9, 2)]
    mb_mon_lv_rows = [(6, 3), (9, 3)]
    c6 = _batch_set(mon_d1, "Monster",
        [(r, c, _name_variant(_cell(mon_base, "Monster", r, c), "dev1")) for r, c in mb_mon_rows]
        + [(r, c, _int_variant(_cell(mon_base, "Monster", r, c), "dev1")) for r, c in mb_mon_lv_rows]) if mon_d1.exists() else []
    # trunk 前进：同行改 trunk 语义替换值（与 dev1 不同 → 真冲突）
    c7 = _batch_set(item_tr, "ItemBase",
        [(r, c, _name_variant(_cell(item_base, "ItemBase", r, c), "trunk")) for r, c in mb_item_rows])
    c8 = _batch_set(mon_tr, "Monster",
        [(r, c, _name_variant(_cell(mon_base, "Monster", r, c), "trunk")) for r, c in mb_mon_rows]
        + [(r, c, _int_variant(_cell(mon_base, "Monster", r, c), "trunk")) for r, c in mb_mon_lv_rows]) if mon_tr.exists() else []
    # skill_level 等级列(int)：相对基准等级 +1~3 微调（5→6/7/8），符合 int 约束
    c9 = []
    if skill_d1.exists() and skill_base.exists():
        # skill_level 列4 为"等级"（int），行4 为首条数据
        sl_rows = [(4, 4)]
        c9 = _batch_set(skill_d1, skill_level_sheet(skill_d1),
            [(r, c, _int_variant(_cell(skill_base, skill_level_sheet(skill_base), r, c), "dev1"))
             for r, c in sl_rows])
    if c5 or c6 or c9:
        _commit(DEV1, f"dev1: merge_back冲突扩充 item{len(c5)} monster{len(c6)} skill{len(c9)}（相对基准真实差异化）")
    if c7 or c8:
        _commit(TRUNK, f"trunk: merge_back冲突扩充 item{len(c7)} monster{len(c8)}（相对基准真实差异化）")

    # 目录合并的 subdev_1/2/3 冲突已由 build_svn_small_branches.stage_subdir_trunk_conflicts
    # 统一按目标规模制造，此处不再重复注入 subdev_1 冲突，避免污染计数。
    print("\n冲突扩充完成（分支值相对基准真实差异化，非 {branch}_xxx 占位符）。")


def skill_level_sheet(xlsx: Path) -> str:
    """skill_level.xlsx 首个数据 sheet 名（跳过 CONFIG 等非数据 sheet）。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    names = wb.sheetnames
    wb.close()
    for n in names:
        if not any(kw in n for kw in ("CONFIG", "说明", "SETTING", "INDEX")):
            return n
    return names[0]


if __name__ == "__main__":
    main()
