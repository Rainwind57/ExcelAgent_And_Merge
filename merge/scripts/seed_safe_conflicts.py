"""补充冲突到 15+，仅用小表（item/reward/ability），避免大表 openpyxl 丢行。

absorb: dev1/dev2 item 行5/8/11/14/17/20 + reward 行5/7 = 8
merge_back: dev1/trunk item 行7/10/13 + reward 行3/5 = 5
目录合并: subdev_1/trunk ability CONFIG 行 + item_drop 行 = 3

分支值相对 trunk 基准真实差异化（名称列语义替换，非 dev1_i7/trunk_ab1b 占位符）。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC / "branches" / "dev1"
DEV2 = WC / "branches" / "dev2"
TRUNK = WC / "trunk"
SUBDEV1 = TRUNK / "subdev_1"
# 原始种子 trunk（永不被脚本污染）：读 base 必须从这里取，否则多次 seed 叠加后缀
SEED_TRUNK = Path(__file__).resolve().parent.parent / "_seed_data" / "trunk"


def _name_variant(base_val, branch: str) -> str:
    """名称列(string)相对 base 的语义替换：换成语义相似但真实不同的词，非后缀占位符。"""
    _SEMANTIC = {
        ("金流露", "dev1"): "金露膏", ("金流露", "dev2"): "玉露散", ("金流露", "trunk"): "金露液", ("金流露", "subdev_1"): "玉露膏",
        ("宝石精华", "dev1"): "宝石精髓", ("宝石精华", "dev2"): "灵石精华", ("宝石精华", "trunk"): "玉髓精华", ("宝石精华", "subdev_1"): "晶核精华",
        ("石芽矿", "dev1"): "石芽矿脉", ("石芽矿", "dev2"): "苔芽石", ("石芽矿", "trunk"): "石芽晶", ("石芽矿", "subdev_1"): "石笋矿",
        ("测试奖励1", "dev1"): "试炼奖励一", ("测试奖励1", "dev2"): "历练奖励壹", ("测试奖励1", "trunk"): "挑战奖励甲",
        ("测试奖励3", "dev1"): "试炼奖励三", ("测试奖励3", "dev2"): "历练奖励叁", ("测试奖励3", "trunk"): "挑战奖励丙",
        ("能力配置", "trunk"): "能力设置", ("能力配置", "subdev_1"): "能力配表",
    }
    prefix = {"dev1": "秘宝", "dev2": "奇珍", "trunk": "灵物", "subdev_1": "异材"}.get(branch, "变异")
    sv = str(base_val).strip() if base_val is not None else ""
    key = (sv, branch)
    if key in _SEMANTIC:
        return _SEMANTIC[key]
    # 掉落名称 "掉落_NNNNNN" → "{prefix}_NNNNNN"（保留 ID 的语义近似替换）
    if sv.startswith("掉落_"):
        return prefix + "_" + sv[len("掉落_"):]
    suffix = {"dev1": "·精", "dev2": "·极", "trunk": "·改", "subdev_1": "·异"}.get(branch, "·改")
    return sv + suffix


def _cell(xlsx: Path, sheet: str, row: int, col: int):
    wb = load_workbook(xlsx, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); return None
    v = wb[sheet].cell(row, col).value
    wb.close()
    return v


def _set_commit(path: Path, sheet: str, edits, wc: Path, msg: str):
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        print(f"  skip {msg}: sheet {sheet} not found"); return
    ws = wb[sheet]
    n = 0
    for r, c, v in edits:
        if ws.cell(r, c).value != v:
            ws.cell(r, c).value = v; n += 1
    if n:
        wb.save(path); wb.close()
        subprocess.run(["svn", "commit", "-m", msg, str(wc)],
                       capture_output=True, text=True, encoding="utf-8")
        print(f"  [OK] {msg} ({n} cells)")
    else:
        wb.close(); print(f"  skip {msg} (no change)")


item_base = SEED_TRUNK / "item" / "item.xlsx"
# merge_back: dev1 与 trunk item 行7/10/13 名称相对基准语义替换
_set_commit(DEV1 / "item" / "item.xlsx", "ItemBase",
           [(7, 2, _name_variant(_cell(item_base, "ItemBase", 7, 2), "dev1")),
            (10, 2, _name_variant(_cell(item_base, "ItemBase", 10, 2), "dev1")),
            (13, 2, _name_variant(_cell(item_base, "ItemBase", 13, 2), "dev1"))], DEV1, "dev1 merge_back item")
_set_commit(TRUNK / "item" / "item.xlsx", "ItemBase",
           [(7, 2, _name_variant(_cell(item_base, "ItemBase", 7, 2), "trunk")),
            (10, 2, _name_variant(_cell(item_base, "ItemBase", 10, 2), "trunk")),
            (13, 2, _name_variant(_cell(item_base, "ItemBase", 13, 2), "trunk"))], TRUNK, "trunk merge_back item")
# reward merge_back（相对种子 trunk reward 基准）
rw1 = DEV1 / "reward.xlsx"; rwt = TRUNK / "reward.xlsx"
rw_base = SEED_TRUNK / "reward.xlsx"
if rw1.exists() and rwt.exists():
    wb = load_workbook(rw1); sn = wb.sheetnames[0]; wb.close()
    _set_commit(rw1, sn, [(3, 2, _name_variant(_cell(rw_base, sn, 3, 2), "dev1")),
                          (5, 2, _name_variant(_cell(rw_base, sn, 5, 2), "dev1"))], DEV1, "dev1 merge_back reward")
    _set_commit(rwt, sn, [(3, 2, _name_variant(_cell(rw_base, sn, 3, 2), "trunk")),
                          (5, 2, _name_variant(_cell(rw_base, sn, 5, 2), "trunk"))], TRUNK, "trunk merge_back reward")

# 目录合并的 subdev_1/2/3 冲突已由 build_svn_small_branches.stage_subdir_trunk_conflicts
# 统一按目标规模制造，此处不再重复注入 subdev_1 冲突，避免污染计数。
print("\n冲突补充完成。")
