"""制造"一边删除、另一边修改"的场景（区别于"两边都删"的普通 deleted 行）。

背景：现有的 row_type='deleted' 语义是"base 有，两个衍生分支都没有"——这种没有
归因歧义（两边都删了）。真正需要人工判断的是**一边删了、另一边还在改**这种情况：
compare.py 目前会把它归类成普通 matched 行，被删的那侧在冲突格里显示成空值，容易
被误读成"对方就是填了个空值"而不是"对方把整行删了"。

R25 已经在 compare.py 补上了 presence 字段的正确透传（之前被 id_resolver 静默丢弃），
前端也加了"⊘ 已删"徽章消费 presence，本脚本负责在 demo 里造出这个场景。

目标：reward.xlsx · Reward sheet · id=10005（"测试奖励5"）
  dev1：删除这一行
  dev2：保留这一行，把 name 改成"测试奖励5·紧急修复"（模拟"我方还在改这行，
        没想到对面已经把它删了"的真实冲突场景）
  trunk：保持不动，作为基准

幂等：dev1 已删除 / dev2 已是目标值则跳过对应分支。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
TABLE_REL = Path("reward.xlsx")
SHEET = "Reward"
PK_COL = 1
TARGET_PK = 10005
NEW_NAME = "测试奖励5·紧急修复"


def _find_row_by_pk(ws, pk, max_row: int = 110) -> int:
    for r in range(1, max_row + 1):
        if ws.cell(r, PK_COL).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def seed_dev1_delete() -> None:
    fp = DEV1 / TABLE_REL
    wb = load_workbook(fp, data_only=False)
    ws = wb[SHEET]
    row = _find_row_by_pk(ws, TARGET_PK)
    if row < 0:
        wb.close()
        print(f"  [skip] dev1 未找到 id={TARGET_PK}（可能已删过，幂等跳过）")
        return
    ws.delete_rows(row, amount=1)
    wb.save(fp)
    wb.close()
    print(f"  dev1: 删除 id={TARGET_PK}（原第{row}行）")
    _commit(DEV1, f"dev1: 制造删除/修改冲突（删除 id={TARGET_PK}，dev2 同时在改这行）")


def seed_dev2_modify() -> None:
    fp = DEV2 / TABLE_REL
    wb = load_workbook(fp, data_only=False)
    ws = wb[SHEET]
    row = _find_row_by_pk(ws, TARGET_PK)
    if row < 0:
        wb.close()
        print(f"  [skip] dev2 未找到 id={TARGET_PK}")
        return
    if ws.cell(row, 2).value == NEW_NAME:
        wb.close()
        print(f"  [skip] dev2 已是目标值（幂等）")
        return
    ws.cell(row, 2).value = NEW_NAME
    wb.save(fp)
    wb.close()
    print(f"  dev2: 第{row}行 name -> {NEW_NAME}")
    _commit(DEV2, f"dev2: 制造删除/修改冲突（修改 id={TARGET_PK} name，dev1 同时删了这行）")


def main() -> None:
    print("=== 制造删除/修改冲突（一边删行，另一边改行）===")
    seed_dev1_delete()
    seed_dev2_modify()
    print(f"\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2，"
          f"表 reward / sheet Reward，id={TARGET_PK}——该行会显示为 matched 行，"
          "name 列标红冲突，行首多一个红色「已删」徽章（图标是个带斜线的圈），"
          "提示这不是普通值变化，是单侧删除。")


if __name__ == "__main__":
    main()
