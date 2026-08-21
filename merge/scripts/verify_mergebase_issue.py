"""最小验证：阶段2 三方合并（merge-base）修复单向改动静默丢失。

模式A（两方 base=trunk，无 merge-base）：复现 bug —— 生产者单向改动静默丢失。
模式B（三方 base=fork + merge_base_file）：修复 —— 真三方合并，单向改动正确采纳。

自包含：openpyxl 造极简表，直接调 compare_sheet + _apply_edits_to_workbook，绕过 HTTP。
"""
from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

SERVER = Path(__file__).resolve().parent.parent.parent / "server"
sys.path.insert(0, str(SERVER))

from engine.parser import read_excel  # noqa: E402
from engine.compare import compare_sheet  # noqa: E402
from routers.diff import _apply_edits_to_workbook  # noqa: E402

TMP = Path(__file__).resolve().parent.parent / "_verify_tmp"
SHEET = "Data"
HEADERS = ["ID", "Name", "Value"]


def _build(path: Path, overrides: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(HEADERS)
    for i in range(1, 6):
        name, val = overrides.get(i, (f"a{i}", i * 10))
        ws.append([i, name, val])
    wb.save(path)


def build_samples():
    TMP.mkdir(exist_ok=True)
    fork = TMP / "fork.xlsx"
    trunk = TMP / "trunk.xlsx"
    inter = TMP / "inter.xlsx"
    _build(fork, {})
    _build(trunk, {3: ("a3", 300)})   # 他人改 trunk ID3
    _build(inter, {2: ("a2", 200)})   # 生产者改 ID2
    return fork, trunk, inter


def _wrap_rows(rows):
    out = []
    for r in rows:
        cells = [
            SimpleNamespace(
                col=c["col"],
                value=c["value"],
                versions=c["versions"],
                diff_type=c.get("diff_type", ""),
                conflict=c.get("conflict"),
                changed=c.get("changed"),
            )
            for c in r["cells"]
        ]
        out.append(SimpleNamespace(key=r["key"], row_type=r["row_type"], cells=cells))
    return out


def run_export(file_sheets, compare_base, export_base, src_path, out_path, merge_base_file=None):
    """compare_base=判定基准, export_base=写回目标(导出 base_file), merge_base_file=三方标记。"""
    merged = compare_sheet(file_sheets, compare_base, SHEET, merge_base_file=merge_base_file)
    rows = merged["rows"]
    req = SimpleNamespace(
        base_file=export_base,
        sheets=[SimpleNamespace(name=SHEET, rows=_wrap_rows(rows))],
    )
    wb = load_workbook(src_path, data_only=False)
    _apply_edits_to_workbook(wb, req)
    wb.save(out_path)
    wb.close()

    wb2 = load_workbook(out_path, data_only=True)
    ws = wb2[SHEET]
    result = {}
    for r in range(2, ws.max_row + 1):
        result[str(ws.cell(row=r, column=1).value)] = ws.cell(row=r, column=3).value
    wb2.close()
    return result, rows


def cell_summary(rows, pk):
    for r in rows:
        if str(r["key"]) == str(pk):
            for c in r["cells"]:
                if c["col"] == 2:
                    return c
    return None


def main():
    fork, trunk, inter = build_samples()
    fs = {
        "fork.xlsx": read_excel(str(fork)),
        "trunk.xlsx": read_excel(str(trunk)),
        "inter.xlsx": read_excel(str(inter)),
    }
    print("=" * 72)
    print("样本: fork=公共祖先, trunk=他人改ID3(30→300), inter=生产者改ID2(20→200)")
    print("期望: ID2=200(采纳生产者), ID3=300(保留trunk)")
    print("=" * 72)

    # 模式A: 两方 base=trunk 无 merge-base（复现 bug）
    print("\n[模式A] 两方 base=trunk 无 merge-base（复现 bug）")
    sheets_a = {"trunk.xlsx": fs["trunk.xlsx"], "inter.xlsx": fs["inter.xlsx"]}
    out_a = TMP / "out_a.xlsx"
    ra, rows_a = run_export(sheets_a, "trunk.xlsx", "trunk.xlsx", trunk, out_a)
    for pk in ("2", "3"):
        c = cell_summary(rows_a, pk)
        print(f"  ID={pk}: changed={c['changed']} conflict={c['conflict']} "
              f"cell.value={c['value']} versions={c['versions']}")
    print(f"  导出: ID2={ra.get('2')} ID3={ra.get('3')}")
    print(f"  >>> ID2 生产者改动(200) {'丢失 ❌ (bug 复现)' if ra.get('2') != 200 else '保留 ✓'}")

    # 模式B: 三方 base=fork + merge_base_file（修复）
    print("\n[模式B] 三方 base=fork + merge_base_file（修复）")
    out_b = TMP / "out_b.xlsx"
    rb, rows_b = run_export(fs, "fork.xlsx", "trunk.xlsx", trunk, out_b, merge_base_file="fork.xlsx")
    for pk in ("2", "3"):
        c = cell_summary(rows_b, pk)
        print(f"  ID={pk}: changed={c['changed']} conflict={c['conflict']} "
              f"cell.value={c['value']} versions={c['versions']}")
    print(f"  导出: ID2={rb.get('2')} ID3={rb.get('3')}")

    ok2 = rb.get("2") == 200
    ok3 = rb.get("3") == 300
    print(f"  >>> ID2={'200 采纳生产者 ✓' if ok2 else '丢失 ❌'}")
    print(f"  >>> ID3={'300 保留trunk ✓' if ok3 else '丢失 ❌'}")

    print("\n" + "=" * 72)
    if ok2 and ok3:
        print("结论: 三方合并修复成功 —— ID2=200 ID3=300，单向改动不再丢失 ✓")
    else:
        print("结论: 修复未生效 ❌")
    print("=" * 72)
    return 0 if (ok2 and ok3) else 1


if __name__ == "__main__":
    sys.exit(main())
