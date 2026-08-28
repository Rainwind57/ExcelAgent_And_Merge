"""向 merge/svn/demo_svn 真实 SVN 仓库提交冲突数据，供三种合并模式验证。

制造冲突点：
  absorb (dev1 → dev2)：item.xlsx 行5 name 双方改不同值 → 真冲突
  merge_back (dev1 → trunk)：trunk 前进改 item 行7，与 dev1 改动冲突
  目录合并 (trunk/subdev_1 → trunk)：subdev_1 改 ability.xlsx 制造冲突

结构增删：
  dev1 新增 sheet "NewSheet_Dev1" 到 item.xlsx
  dev2 新增表格 new_table_dev2.xlsx
  subdev_1 新增 sheet 到 ability.xlsx

分支值相对 base 真实差异化（模拟实际改表）：
  item 名称列(string) → 语义替换（如"测试道具"→"测试道具·精"），非 dev1_改row5 占位符
  subdev_1 ability 名称 → 语义替换
各分支改不同值仍构成冲突，但值符合列类型与语义。
幂等：重复执行只产生一次提交（检测目标 cell 已是目标值则跳过）。
"""
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
TRUNK = WC_ROOT / "trunk"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
SUBDEV1 = TRUNK / "subdev_1"
# 原始种子 trunk（永不被脚本污染）：读 base 必须从这里取，否则多次 seed 叠加后缀
SEED_TRUNK = Path(__file__).resolve().parent.parent / "_seed_data" / "trunk"


def _name_variant(base_val, branch: str) -> str:
    """名称列(string)相对 base 的语义替换：换成语义相似但真实不同的词，避免后缀占位符。

    例：base="测试道具" → dev1="测试法器", dev2="试炼道具"；base="三味真火" → "重水"。
    每个 (原值, 分支) 固定映射，保证幂等（重复执行不叠加）。
    """
    # 逐 (原值, 分支) 的语义替换表（不同分支给不同"新值"，仍构成冲突）
    SEMANTIC_MAP = {
        ("测试道具", "dev1"): "测试法器",
        ("测试道具", "dev2"): "试炼道具",
        ("测试道具", "trunk"): "试炼符",
        ("测试道具", "subdev_1"): "测试灵器",
        ("金流露", "dev1"): "金露膏",
        ("金流露", "dev2"): "玉露散",
        ("金流露", "trunk"): "金露液",
        ("金流露", "subdev_1"): "玉露膏",
        ("三味真火", "dev1"): "重水",
        ("三味真火", "dev2"): "南明离火",
        ("三味真火", "trunk"): "九天玄火",
        ("三味真火", "subdev_1"): "弱水",
        ("蛮牛狂击", "subdev_1"): "莽牛冲撞",
        ("雨露均沾", "subdev_1"): "甘霖普降",
    }
    key = (str(base_val).strip() if isinstance(base_val, str) else str(base_val), branch)
    if key in SEMANTIC_MAP:
        return SEMANTIC_MAP[key]
    # 未命中映射表的兜底：保持原后缀逻辑，避免生成 None/空值破坏表语义
    suffix = {"dev1": "·精", "dev2": "·极", "trunk": "·改", "subdev_1": "·异"}.get(branch, "·改")
    return (str(base_val) if base_val else "") + suffix


def _svn_commit(wc_path: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc_path)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    if r.returncode != 0:
        print(f"  svn commit 失败: {r.stderr.strip()}", file=sys.stderr)
    else:
        print(f"  committed: {msg}")


def _set_cell(xlsx: Path, sheet: str, row: int, col: int, value) -> bool:
    """若 cell 当前值 != value 则改并返回 True（需提交），否则 False（跳过）。"""
    wb = load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[sheet]
    if ws.cell(row, col).value == value:
        wb.close()
        return False
    ws.cell(row, col).value = value
    wb.save(xlsx)
    wb.close()
    return True


def _read_cell(xlsx: Path, sheet: str, row: int, col: int):
    """读 base 单元格当前值（read_only）。"""
    wb = load_workbook(xlsx, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    v = wb[sheet].cell(row, col).value
    wb.close()
    return v


def _add_sheet(xlsx: Path, sheet: str, header: list, row: list) -> bool:
    wb = load_workbook(xlsx)
    if sheet in wb.sheetnames:
        wb.close()
        return False
    ws = wb.create_sheet(sheet)
    ws.append(header)
    ws.append(row)
    wb.save(xlsx)
    wb.close()
    return True


def _new_table(dir_path: Path, name: str) -> bool:
    fp = dir_path / f"{name}.xlsx"
    if fp.exists():
        return False
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "name"])
    ws.append([1, "新表测试"])
    wb.save(fp)
    return True


def _svn_add(path: Path) -> None:
    subprocess.run(
        ["svn", "add", str(path), "--force"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )


def main() -> None:
    # base = 种子 trunk（dev1/dev2 均从 trunk fork，读原始值避免后缀叠加）。
    # item 名称列(string) 相对种子 trunk 基准做语义替换。
    item_base = SEED_TRUNK / "item" / "item.xlsx"

    # ── dev1: item 行5 name 相对基准做 dev1 语义替换（"测试道具"→"测试道具·精"），新增 sheet ──
    item_d1 = DEV1 / "item" / "item.xlsx"
    changed = []
    v1 = _name_variant(_read_cell(item_base, "ItemBase", 5, 2), "dev1")
    if _set_cell(item_d1, "ItemBase", 5, 2, v1):
        changed.append("item 行5 name")
    if _add_sheet(item_d1, "NewSheet_Dev1", ["id", "info"], [1, "dev1新增sheet"]):
        changed.append("item 新增 sheet NewSheet_Dev1")
    if changed:
        print(f"dev1 改动: {changed}")
        _svn_commit(DEV1, f"dev1: 制造冲突 {', '.join(changed)}（相对基准真实差异化）")

    # ── dev2: item 行5 name 改不同语义值（冲突），行6 单向变更，新增表格 ──
    item_d2 = DEV2 / "item" / "item.xlsx"
    changed = []
    v5d2 = _name_variant(_read_cell(item_base, "ItemBase", 5, 2), "dev2")
    v6d2 = _name_variant(_read_cell(item_base, "ItemBase", 6, 2), "dev2")
    if _set_cell(item_d2, "ItemBase", 5, 2, v5d2):
        changed.append("item 行5 name（与 dev1 冲突）")
    if _set_cell(item_d2, "ItemBase", 6, 2, v6d2):
        changed.append("item 行6 name（单向）")
    if _new_table(DEV2, "new_table_dev2"):
        _svn_add(DEV2 / "new_table_dev2.xlsx")
        changed.append("新增表格 new_table_dev2.xlsx（结构增删）")
    if changed:
        print(f"dev2 改动: {changed}")
        _svn_commit(DEV2, f"dev2: 制造冲突 {', '.join(changed)}（相对基准真实差异化）")

    # ── trunk: item 行7 name 相对基准做 trunk 语义替换（merge_back 时与 dev1 冲突）──
    item_tr = TRUNK / "item" / "item.xlsx"
    v7tr = _name_variant(_read_cell(item_base, "ItemBase", 7, 2), "trunk")
    if _set_cell(item_tr, "ItemBase", 7, 2, v7tr):
        print("trunk 改动: item 行7 name（merge_back 冲突点）")
        _svn_commit(TRUNK, "trunk: 前进改 item 行7（merge_back 冲突源，相对基准真实差异化）")

    # 目录合并的 subdev_1/2/3 冲突已由 build_svn_small_branches.stage_subdir_trunk_conflicts
    # 统一按目标规模制造，此处不再重复注入 subdev_1 冲突，避免污染计数。
    print("\n完成。冲突数据已提交到 demo_svn 仓库（分支值相对基准真实差异化，非占位符）。")


if __name__ == "__main__":
    main()
