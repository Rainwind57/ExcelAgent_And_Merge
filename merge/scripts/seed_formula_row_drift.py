"""制造"公式行漂移"风险场景（注意：这不是 formula_conflict 冲突，是非阻塞提示）。

背景：现有 seed_svn_conflicts*.py 制造的都是"同一 PK 行、同一单元格，各分支写了
不同的值/公式文本"这类真冲突（compare.py 靠逐字比对公式文本即可发现，且各版本
文本不同，用户确实需要选一个）。但还有一类更隐蔽的真实 bug：某分支在表格中间
插入/删除了行，而表格里那些"硬编码绝对行号"的公式（如 `=D9-(E9/2)`）在插删后
**文本没有跟着调整**（openpyxl/脚本编辑表格时不会像人在 Excel 里操作那样自动
重写公式引用），导致公式物理位置变了但引用的还是旧行号——**两分支的公式文本
仍然逐字相同**（都没被人为改过），所以这不是"合并冲突"（没有版本分歧可选），
而是一种"数据本身可能已经错了"的风险提示。

server/engine/compare.py 已新增非阻塞标记 `formula_row_drift=True`（挂在
diff_type 仍为 "formula" 的单元格上，不计入 conflict/stats.conflicts，不会弹出
选版本框），对比同一 PK 在各版本中的物理行号（row_idx_maps）是否一致来判定。
本脚本负责在 demo_svn 里造出这个场景，供核实/截图该风险提示。

目标表：residence/residence_putuan.xlsx · PutuanAcupoint sheet
  - F 列 `=D{row}-(E{row}/2)`、I 列 `=(H{row}/2)-G{row}`：同行绝对行号引用公式。

制造方式：
  dev1：在 id=301 行（原物理第 9 行）前插入 1 行新数据（id=999），
        原第 9~19 行整体下移 1 行，公式文本保留原行号 → 与新的物理行错位 1 行。
  dev2：删除 id=204 行（原物理第 16 行），
        原第 17~19 行整体上移 1 行，公式文本同理错位。
  trunk：保持不动，作为三方比对基准（与 dev1/dev2 比较时都能看到漂移）。

幂等：dev1 已存在 id=999 / dev2 已不存在 id=204 时跳过对应分支，重复执行不叠加。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
TABLE_REL = Path("residence") / "residence_putuan.xlsx"
SHEET = "PutuanAcupoint"

# 插入锚点：新行插在该 PK 之前（原物理行号，1-based，含表头共 4 行）
INSERT_BEFORE_PK = 301
NEW_ROW_PK = 999
# 删除目标：删掉该 PK 所在整行
DELETE_PK = 204


def _find_row_by_pk(ws, pk: int, pk_col: int = 1, max_row: int = 40) -> int:
    for r in range(1, max_row + 1):
        if ws.cell(r, pk_col).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def seed_dev1_insert() -> None:
    """dev1：id=301 行前插入新行 id=999，下方行整体下移但公式文本不跟着改行号。"""
    fp = DEV1 / TABLE_REL
    if not fp.exists():
        print(f"  [skip] dev1 缺少 {TABLE_REL}")
        return
    wb = load_workbook(fp, data_only=False)
    if SHEET not in wb.sheetnames:
        wb.close()
        print(f"  [skip] dev1 {TABLE_REL} 缺少 sheet {SHEET}")
        return
    ws = wb[SHEET]
    if _find_row_by_pk(ws, NEW_ROW_PK) > 0:
        wb.close()
        print(f"  [skip] dev1 已存在 id={NEW_ROW_PK}，跳过（幂等）")
        return
    anchor = _find_row_by_pk(ws, INSERT_BEFORE_PK)
    if anchor < 0:
        wb.close()
        print(f"  [skip] dev1 未找到 id={INSERT_BEFORE_PK} 锚点行")
        return
    # 插入新行：不重写下方任何已有公式文本（模拟"插行忘了调公式"的真实 bug）
    ws.insert_rows(anchor, amount=1)
    ws.cell(anchor, 1).value = NEW_ROW_PK
    ws.cell(anchor, 2).value = "dev1新增穴位"
    ws.cell(anchor, 3).value = "0,0"
    ws.cell(anchor, 4).value = 500
    ws.cell(anchor, 5).value = 1000
    ws.cell(anchor, 6).value = f"=D{anchor}-(E{anchor}/2)"   # 新行公式，行号是对的
    ws.cell(anchor, 7).value = 100
    ws.cell(anchor, 8).value = 900
    ws.cell(anchor, 9).value = f"=(H{anchor}/2)-G{anchor}"   # 新行公式，行号是对的
    wb.save(fp)
    wb.close()
    print(f"  dev1: id={INSERT_BEFORE_PK} 前插入新行 id={NEW_ROW_PK}（原第{anchor}行起下移1行，"
          f"下方行公式文本未随行号调整 → 漂移）")
    _commit(DEV1, f"dev1: 制造公式行漂移（插入 id={NEW_ROW_PK}，下方行公式未随行号调整）")


def seed_dev2_delete() -> None:
    """dev2：删除 id=204 行，上方行整体上移但公式文本不跟着改行号。"""
    fp = DEV2 / TABLE_REL
    if not fp.exists():
        print(f"  [skip] dev2 缺少 {TABLE_REL}")
        return
    wb = load_workbook(fp, data_only=False)
    if SHEET not in wb.sheetnames:
        wb.close()
        print(f"  [skip] dev2 {TABLE_REL} 缺少 sheet {SHEET}")
        return
    ws = wb[SHEET]
    target = _find_row_by_pk(ws, DELETE_PK)
    if target < 0:
        wb.close()
        print(f"  [skip] dev2 未找到 id={DELETE_PK}（可能已被删过，幂等跳过）")
        return
    ws.delete_rows(target, amount=1)
    wb.save(fp)
    wb.close()
    print(f"  dev2: 删除 id={DELETE_PK}（原第{target}行），下方行整体上移1行，"
          f"公式文本未随行号调整 → 漂移")
    _commit(DEV2, f"dev2: 制造公式行漂移（删除 id={DELETE_PK}，下方行公式未随行号调整）")


def main() -> None:
    print("=== 制造公式行漂移风险场景（formula_row_drift，非冲突）===")
    seed_dev1_insert()
    seed_dev2_delete()
    print("\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2 或"
          "任一分支 vs trunk，表 residence_putuan / sheet PutuanAcupoint，"
          "F 列、I 列在 id=301/302/303/304/201/202/203/205/206/207 附近的行——"
          "这些格子 diff_type 仍是 'formula'（不是红色冲突），但 API 返回里"
          "formula_row_drift=True，可在 devtools/接口响应里核实；前端目前未对"
          "这个字段做专门的视觉呈现（属于非阻塞提示，未来如需要可再加轻量图标）。")


if __name__ == "__main__":
    main()
