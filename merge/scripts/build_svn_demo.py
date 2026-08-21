"""文件夹模拟 SVN 快照的样例数据生成脚本（方式一：demo 快照，merge/demo/）。

在 merge/demo/ 下生成 trunk_r1~r3 / A_r1~r3 / B_r1~r3 版本快照目录，
每个目录内含 ability.xlsx/item.xlsx/match_stat.xlsx 等表，并配 _meta.json
记录该快照的 copyfrom 信息（模拟 SVN copyfrom-path/copyfrom-rev）。

与真实 SVN（方式二，merge/svn/ 下 svnadmin 初始化的仓库与工作副本）物理隔离：
本脚本产物仅依赖文件夹快照 + openpyxl，无需 svn CLI。

提交历史设计（版本号仅作设计标识，实际由目录名后缀 _rN 表示）：
    trunk_r1  trunk 基准（从 merge/trunk 复制现有 3 表）
    trunk_r2  trunk 前进 1：改 ability/match_stat/big_data 若干行
    trunk_r3  trunk 前进 2：改 ability/item/match_stat/big_data 若干行（与 A/B 冲突点交错）
    A_r1      A 分支从 trunk_r1 切出（copyfrom=trunk_r1），改 ability/item 若干行
    A_r2      A 分支继续提交（copyfrom=A_r1）：多行多列修改 + 新增 1 行
    A_r3      A 分支再提交（copyfrom=A_r2）：新增一批与 B_r3/trunk_r3 交错冲突的修改
    B_r1      B 分支从 trunk_r2 切出（copyfrom=trunk_r2）
    B_r2      B 分支继续提交（copyfrom=B_r1）：多行多列修改 + 新增 1 行
    B_r3      B 分支再提交（copyfrom=B_r2）：与 A_r3 形成新的冲突点

A vs B（absorb A_r3 → B_r3）合并时 merge_base = trunk_r1（LCA）：
    ability：既有 5 处真冲突（行5/6/7 name + 行6 desc + 行8 name）+ 1 处单向改动（行9，仅 A）
             + 新增冲突 行10 name（A_r3 vs B_r3）+ A/B 各新增 1 行（inserted，不冲突）
    item：   既有 3 处真冲突（行6/7 name + 行6 desc）+ 1 处单向改动（行8，仅 B）
             + 新增冲突 行9 name（A_r3 vs B_r3）
    match_stat：既有 3 处真冲突（行3/4 c1 + 行5 col6）+ 1 处单向改动（行6，仅 A）
             + 新增冲突 行5 c2（A_r3 vs B_r3，base=35）
             + 1 处公式文本冲突（行7 total 公式：A =B7+C7+D7+E7 / B =(B7+C7+D7+E7)*2，base=SUM）
    config/skill（嵌套）：既有 行3/4 name 冲突 + 行5 单向（仅 A_r2）；A_r3/B_r3 在行5 name 形成新冲突
    big_data（10w 行）：既有 4 处真冲突（行10001/25000/50000/99999）+ 单向变更 ×4
             + 新增冲突 行77777（A_r3 vs B_r3 vs trunk_r3）
数量刻意做多，确保跨分支合并界面（冲突/变更/新增着色、搜索、批量、AI建议、公式重算等）有足够真实素材可测试。

嵌套子文件夹样例（跨分支场景）：
    trunk_r1~r3 / A_r1~r3 / B_r1~r3 各快照均含对称 config/skill.xlsx（一层嵌套）。
    A_r3 / B_r3 在 config/skill.xlsx 行3/4/5 name 制造冲突，LCA=trunk_r1。

大数据量 merge 样例（big_data.xlsx，10w 行级）：
    trunk_r1~r3/A_r1~r3/trunk_r2/B_r1~r3 各快照均含 big_data.xlsx（BigData sheet），
    基准 10 万行数据（id=1..100000，name/value/desc/flag 四列）。用 write_only 逐快照
    生成（内容等价于 copy+改单元格，避免 openpyxl 全量 load 10w 行的慢路径）。
    冲突设计（base=trunk_r1，A_r3 absorb → B_r3）：
        行10001/25000/50000/99999：A 与 B 各改不同值 → 4 处真冲突（分散在头/中/尾）
        行77777：A_r3 / B_r3 / trunk_r3 各改不同值 → 新增冲突点（三向交错）
        行501（仅A_r1）/ 601（仅trunk_r2→B）/ 12345（仅A_r2）/ 23456（仅B_r2）：单向变更 ×4
        行100001/100002：A/B 各追加 1 行（inserted，不冲突）
"""
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

DEMO_ROOT = Path(__file__).resolve().parent.parent / "demo"
TRUNK_SRC = Path(__file__).resolve().parent.parent / "trunk"


def _copy_table(src_dir: Path, name: str, dst_dir: Path) -> None:
    shutil.copy2(src_dir / f"{name}.xlsx", dst_dir / f"{name}.xlsx")


def _init_dir(name: str) -> Path:
    d = DEMO_ROOT / name
    if d.exists():
        shutil.rmtree(d)
    d.mkdir(parents=True)
    return d


def _write_meta(d: Path, copyfrom_path: str, copyfrom_rev: str, note: str) -> None:
    (d / "_meta.json").write_text(
        json.dumps({
            "copyfrom_path": copyfrom_path,
            "copyfrom_rev": copyfrom_rev,
            "note": note,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_meta_ext(d: Path, meta: dict) -> None:
    """写入扩展 _meta.json（子目录场景需额外记录 base_snapshot / inferred）。"""
    (d / "_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8",
    )


def _make_table(fp: Path, sheet: str, headers, rows) -> None:
    """新建一张简单表（单 sheet + 表头 + 数据行），用于演示结构增删。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(fp)


def _add_sheet(fp: Path, sheet: str, headers, rows) -> None:
    """给已有 xlsx 追加一个 sheet（用于演示 sheet 级结构增删）。"""
    wb = load_workbook(fp)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        for r in rows:
            ws.append(r)
        wb.save(fp)


def _set_cells(path: Path, sheet: str, edits: list) -> None:
    """edits: [(row, col, value), ...] 批量设置单元格，避免多次打开/保存文件。"""
    wb = load_workbook(path)
    ws = wb[sheet]
    for row, col, value in edits:
        ws.cell(row=row, column=col, value=value)
    wb.save(path)
    wb.close()


def _append_row(path: Path, sheet: str, row: int, values: list) -> None:
    """在指定行位置写入一整行新数据（用于模拟新增行，主键取 values[0]）。"""
    wb = load_workbook(path)
    ws = wb[sheet]
    for i, v in enumerate(values):
        ws.cell(row=row, column=i + 1, value=v)
    wb.save(path)
    wb.close()


# ── 大数据量 merge 样例（10w 行级 big_data.xlsx）──
# 基准 10 万行（id=1..100000），A/B 分支在分散行造冲突。逐快照用 write_only 全量生成
# （内容等价于 copy+改单元格，但避免 openpyxl 全量 load 10w 行的慢路径，生成秒级）。
_BIG_TABLE = "big_data"
_BIG_SHEET = "BigData"
_BIG_HEADERS = ["id:int", "name:string", "value:int", "desc:string", "flag:string"]
_BIG_ROWS = 100_000  # 基准 10 万行数据


def _make_big_table(fp: Path, overrides: dict = None, appends: list = None,
                    nrows: int = _BIG_ROWS) -> None:
    """write_only 生成大表。

    overrides: {数据行号(1..nrows): {列号(1..5): 值}}，把指定单元格改成分支自己的值；
    appends:   [[id, name, value, desc, flag], ...] 追加在数据末尾的新行（模拟分支新增行）。
    数据行号 = sheet 行号 - 1（第1行是表头），与前端比对展示的行号一致。
    """
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(_BIG_SHEET)
    ws.append(_BIG_HEADERS)
    overrides = overrides or {}
    for i in range(1, nrows + 1):
        ov = overrides.get(i, {})
        ws.append([
            i,
            ov.get(2, f"data_{i}"),
            ov.get(3, i * 10),
            ov.get(4, f"desc_{i}"),
            ov.get(5, "Y" if i % 2 == 0 else "N"),
        ])
    for r in appends or []:
        ws.append(r)
    wb.save(fp)


# ── 嵌套子文件夹样例（验证多层目录合并）──
# 跨分支与目录合并两个场景都加对称嵌套结构：分支目录下除平铺表外，还有子文件夹，
# 子文件夹里再放表（可多层）。这里统一用 config/ 子文件夹 + skill.xlsx 演示一层嵌套，
# 结构对称（trunk/A/B 各快照都有 config/skill.xlsx），制造跨分支嵌套层冲突供验证。
_NESTED_SUBDIR = "config"
_NESTED_TABLE = "skill"
_NESTED_SHEET = "Skill"
_NESTED_HEADERS = ["skill_id:int", "name:string", "desc:string"]
_NESTED_BASE_ROWS = [
    [1, "普攻", "普通攻击"],
    [2, "重击", "强力一击"],
    [3, "连击", "连续两次攻击"],
    [4, "格挡", "减少受到伤害"],
    [5, "闪避", "概率规避伤害"],
]


def _make_nested_table(parent_dir: Path, subdir: str = _NESTED_SUBDIR,
                       table: str = _NESTED_TABLE) -> Path:
    """在 parent_dir/subdir 下新建一张基准 skill.xlsx，返回该文件路径。"""
    sub = parent_dir / subdir
    sub.mkdir(parents=True, exist_ok=True)
    fp = sub / f"{table}.xlsx"
    _make_table(fp, _NESTED_SHEET, _NESTED_HEADERS, _NESTED_BASE_ROWS)
    return fp


def _copy_nested_table(src_dir: Path, dst_dir: Path,
                       subdir: str = _NESTED_SUBDIR,
                       table: str = _NESTED_TABLE) -> Path:
    """把 src_dir/subdir/table.xlsx 复制到 dst_dir/subdir/table.xlsx，返回目标路径。"""
    src_fp = src_dir / subdir / f"{table}.xlsx"
    dst_fp = dst_dir / subdir / f"{table}.xlsx"
    dst_fp.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_fp, dst_fp)
    return dst_fp


def main() -> None:
    print(f"=== 生成 SVN demo 样例: {DEMO_ROOT} ===")
    if DEMO_ROOT.exists():
        shutil.rmtree(DEMO_ROOT)
    DEMO_ROOT.mkdir(parents=True)

    # ── big_data.xlsx（10w 行）分支改动设计：数据行号 = 前端比对展示行号 ──
    a1_big = {501: {2: "A_r1改的数据名501"}}            # 行501 name：仅 A 侧，单向变更
    t2_big = {601: {2: "trunk_r2改的数据名601"}}        # 行601 name：仅 B 侧（经 trunk_r2→B），单向变更
    t3_big = {77777: {2: "trunk_r3改的数据名77777"}}    # 行77777 name：trunk_r3 改，与 A_r3/B_r3 交错
    a2_big = {
        10001: {2: "A_r2改的数据名10001"},              # 行10001 name：与 B_r2 冲突
        25000: {5: "A_r2改的flag25000"},                # 行25000 flag：与 B_r2 冲突
        50000: {3: 500001},                             # 行50000 value：与 B_r2 冲突
        99999: {4: "A_r2改的描述99999"},                # 行99999 desc：与 B_r2 冲突
        12345: {2: "A_r2单方改的数据名12345"},          # 行12345 name：仅 A 改，单向变更
    }
    a2_big_appends = [[100001, "A新增数据行", 1000010, "A分支新增的描述", "N"]]
    a3_big = {77777: {2: "A_r3改的数据名77777"}}         # 行77777 name：与 B_r3/trunk_r3 冲突
    b2_big = {
        10001: {2: "B_r2改的数据名10001"},              # 行10001 name：与 A_r2 冲突
        25000: {5: "B_r2改的flag25000"},                # 行25000 flag：与 A_r2 冲突
        50000: {3: 555555},                             # 行50000 value：与 A_r2 冲突
        99999: {4: "B_r2改的描述99999"},                # 行99999 desc：与 A_r2 冲突
        23456: {2: "B_r2单方改的数据名23456"},          # 行23456 name：仅 B 改，单向变更
    }
    b2_big_appends = [[100002, "B新增数据行", 1000020, "B分支新增的描述", "N"]]
    b3_big = {77777: {2: "B_r3改的数据名77777"}}         # 行77777 name：与 A_r3/trunk_r3 冲突

    # trunk_r1：基准
    trunk_r1 = _init_dir("trunk_r1")
    for n in ("ability", "item", "match_stat"):
        _copy_table(TRUNK_SRC, n, trunk_r1)
    # 嵌套子文件夹样例（config/skill.xlsx，基准态）
    _make_nested_table(trunk_r1)
    _make_big_table(trunk_r1 / f"{_BIG_TABLE}.xlsx")
    _write_meta(trunk_r1, "", "", "trunk 基准版本 r1（含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # A_r1：从 trunk_r1 切出，先做一批基础修改
    a_r1 = _init_dir("A_r1")
    for n in ("ability", "item", "match_stat"):
        _copy_table(trunk_r1, n, a_r1)
    _copy_nested_table(trunk_r1, a_r1)
    _make_big_table(a_r1 / f"{_BIG_TABLE}.xlsx", overrides=a1_big)
    _set_cells(a_r1 / "ability.xlsx", "Ability", [(5, 2, "A分支改的神通名")])
    _set_cells(a_r1 / "item.xlsx", "ItemBase", [(5, 2, "A分支改的道具名")])
    _write_meta(a_r1, "trunk_r1", "r1", "A 分支从 trunk_r1 切出（含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # trunk_r2：trunk 前进 1
    trunk_r2 = _init_dir("trunk_r2")
    for n in ("ability", "item", "match_stat"):
        _copy_table(trunk_r1, n, trunk_r2)
    _copy_nested_table(trunk_r1, trunk_r2)
    _make_big_table(trunk_r2 / f"{_BIG_TABLE}.xlsx", overrides=t2_big)
    _set_cells(trunk_r2 / "ability.xlsx", "Ability", [(5, 2, "trunk改的神通名")])
    _set_cells(trunk_r2 / "match_stat.xlsx", "SeasonStat", [(3, 2, 66)])
    _write_meta(trunk_r2, "trunk_r1", "r1", "trunk r2 前进版本（含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # trunk_r3：trunk 前进 2（与 A_r3/B_r3 的提交交错，制造 merge_back 冲突素材）
    trunk_r3 = _init_dir("trunk_r3")
    for n in ("ability", "item", "match_stat"):
        _copy_table(trunk_r2, n, trunk_r3)
    _copy_nested_table(trunk_r2, trunk_r3)
    _make_big_table(trunk_r3 / f"{_BIG_TABLE}.xlsx", overrides={**t2_big, **t3_big})
    _set_cells(trunk_r3 / "ability.xlsx", "Ability", [(10, 2, "trunk_r3改的星落")])
    _set_cells(trunk_r3 / "item.xlsx", "ItemBase", [(9, 2, "trunk_r3改的宝石原石")])
    _set_cells(trunk_r3 / "match_stat.xlsx", "SeasonStat", [(5, 3, 350)])
    _write_meta(trunk_r3, "trunk_r2", "r2", "trunk r3 前进版本（含嵌套 config/skill.xlsx + 10w行 big_data.xlsx，merge_back 默认 To 目标）")

    # B_r1：从 trunk_r2 切出
    b_r1 = _init_dir("B_r1")
    for n in ("ability", "item", "match_stat"):
        _copy_table(trunk_r2, n, b_r1)
    _copy_nested_table(trunk_r2, b_r1)
    _make_big_table(b_r1 / f"{_BIG_TABLE}.xlsx", overrides=t2_big)  # B_r1 = trunk_r2 拷贝态
    _write_meta(b_r1, "trunk_r2", "r2", "B 分支从 trunk_r2 切出（含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # A_r2：A 继续提交，多行多列修改 + 新增 1 行（与 B_r2 在多处形成真冲突）
    a_r2 = _init_dir("A_r2")
    for n in ("ability", "item", "match_stat"):
        _copy_table(a_r1, n, a_r2)
    _copy_nested_table(a_r1, a_r2)
    _make_big_table(a_r2 / f"{_BIG_TABLE}.xlsx", overrides={**a1_big, **a2_big}, appends=a2_big_appends)
    _set_cells(a_r2 / "ability.xlsx", "Ability", [
        (5, 2, "A_r2再改的神通名"),          # 行5 name：与 B_r2 冲突
        (6, 2, "A_r2改的蛮牛技能名"),        # 行6 name：与 B_r2 冲突
        (6, 3, "A_r2改的蛮牛描述"),          # 行6 desc：与 B_r2 冲突
        (7, 2, "A_r2改的雨露均沾"),          # 行7 name：与 B_r2 冲突
        (8, 2, "A_r2改的群体治疗"),          # 行8 name：与 B_r2 冲突
        (9, 2, "A_r2单方面改的复活技能"),     # 行9 name：仅 A 改，单向变更（非冲突）
    ])
    _append_row(a_r2 / "ability.xlsx", "Ability", 715, [9001, 'A分支新增技能', 'A分支新增的技能描述', 'icon_spell_a01', 100901, 1, None, None])
    _set_cells(a_r2 / "item.xlsx", "ItemBase", [
        (6, 2, "A_r2改的道具名6"),           # 行6 name：与 B_r2 冲突
        (6, 7, "A_r2改的道具描述6"),          # 行6 desc：与 B_r2 冲突
        (7, 2, "A_r2改的道具名7"),           # 行7 name：与 B_r2 冲突
    ])
    _set_cells(a_r2 / "match_stat.xlsx", "SeasonStat", [
        (3, 2, "A_r2改的赛季统计"),          # 行3 c1：与 B_r2 冲突（trunk_r2 已改过一次）
        (4, 2, 999),                        # 行4 c1：与 B_r2 冲突
        (5, 6, 5001),                       # 行5 total：与 B_r2 冲突
        (6, 2, 888),                        # 行6 c1：仅 A 改，单向变更
        (7, 6, "=B7+C7+D7+E7"),             # 行7 total 公式文本：与 B_r2 冲突（base=trunk_r1 的 =SUM(B7:E7)）→ formula_conflict
    ])
    # 嵌套 config/skill.xlsx：A_r2 改行3/4 name（与 B_r2 冲突）+ 行5 单向变更
    _set_cells(a_r2 / "config" / "skill.xlsx", "Skill", [
        (3, 2, "A_r2改的连击"),              # 行3 name：与 B_r2 冲突
        (4, 2, "A_r2改的格挡"),              # 行4 name：与 B_r2 冲突
        (5, 2, "A_r2单方面改的闪避"),         # 行5 name：仅 A 改，单向变更
    ])
    _write_meta(a_r2, "A_r1", "r1", "A 分支 r2 版本（copyfrom=A_r1，含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # B_r2：B 继续提交，多行多列修改 + 新增 1 行（与 A_r2 在多处形成真冲突）
    b_r2 = _init_dir("B_r2")
    for n in ("ability", "item", "match_stat"):
        _copy_table(b_r1, n, b_r2)
    _copy_nested_table(b_r1, b_r2)
    _make_big_table(b_r2 / f"{_BIG_TABLE}.xlsx", overrides={**t2_big, **b2_big}, appends=b2_big_appends)
    _set_cells(b_r2 / "ability.xlsx", "Ability", [
        (5, 2, "B_r2改的神通名"),
        (6, 2, "B_r2改的蛮牛技能名"),
        (6, 3, "B_r2改的蛮牛描述"),
        (7, 2, "B_r2改的雨露均沾"),
        (8, 2, "B_r2改的群体治疗"),
    ])
    _append_row(b_r2 / "ability.xlsx", "Ability", 716, [9002, 'B分支新增技能', 'B分支新增的技能描述', 'icon_spell_b01', 100902, 1, None, None])
    _set_cells(b_r2 / "item.xlsx", "ItemBase", [
        (5, 2, "B_r2改的道具名"),
        (6, 2, "B_r2改的道具名6"),
        (6, 7, "B_r2改的道具描述6"),
        (7, 2, "B_r2改的道具名7"),
        (8, 2, "B_r2单方面改的道具名8"),      # 行8 name：仅 B 改，单向变更
    ])
    _set_cells(b_r2 / "match_stat.xlsx", "SeasonStat", [
        (3, 2, "B_r2改的赛季统计"),
        (4, 2, 111),
        (5, 6, 5002),
        (7, 6, "=(B7+C7+D7+E7)*2"),         # 行7 total 公式文本：与 A_r2 冲突（base=trunk_r1 的 =SUM(B7:E7)）→ formula_conflict
    ])
    # 嵌套 config/skill.xlsx：B_r2 改行3/4 name（与 A_r2 冲突），无单向变更
    _set_cells(b_r2 / "config" / "skill.xlsx", "Skill", [
        (3, 2, "B_r2改的连击"),              # 行3 name：与 A_r2 冲突
        (4, 2, "B_r2改的格挡"),              # 行4 name：与 A_r2 冲突
    ])
    _write_meta(b_r2, "B_r1", "r1", "B 分支 r2 版本（copyfrom=B_r1，含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # A_r3：A 再提交（copyfrom=A_r2），新增与 B_r3/trunk_r3 交错的冲突点
    a_r3 = _init_dir("A_r3")
    for n in ("ability", "item", "match_stat"):
        _copy_table(a_r2, n, a_r3)
    _copy_nested_table(a_r2, a_r3)
    _make_big_table(a_r3 / f"{_BIG_TABLE}.xlsx",
                    overrides={**a1_big, **a2_big, **a3_big}, appends=a2_big_appends)
    _set_cells(a_r3 / "ability.xlsx", "Ability", [
        (10, 2, "A_r3改的星落"),             # 行10 name：与 B_r3/trunk_r3 冲突
    ])
    _set_cells(a_r3 / "item.xlsx", "ItemBase", [
        (9, 2, "A_r3改的宝石原石"),           # 行9 name：与 B_r3/trunk_r3 冲突
    ])
    _set_cells(a_r3 / "match_stat.xlsx", "SeasonStat", [
        (5, 3, 360),                        # 行5 c2：与 B_r3(370)/trunk_r3(350) 冲突
    ])
    _set_cells(a_r3 / "config" / "skill.xlsx", "Skill", [
        (5, 2, "A_r3改的闪避"),              # 行5 name：与 B_r3 冲突
    ])
    _write_meta(a_r3, "A_r2", "r2", "A 分支 r3 版本（copyfrom=A_r2，含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    # B_r3：B 再提交（copyfrom=B_r2），与 A_r3 形成新冲突点
    b_r3 = _init_dir("B_r3")
    for n in ("ability", "item", "match_stat"):
        _copy_table(b_r2, n, b_r3)
    _copy_nested_table(b_r2, b_r3)
    _make_big_table(b_r3 / f"{_BIG_TABLE}.xlsx",
                    overrides={**t2_big, **b2_big, **b3_big}, appends=b2_big_appends)
    _set_cells(b_r3 / "ability.xlsx", "Ability", [
        (10, 2, "B_r3改的星落"),             # 行10 name：与 A_r3/trunk_r3 冲突
    ])
    _set_cells(b_r3 / "item.xlsx", "ItemBase", [
        (9, 2, "B_r3改的宝石原石"),           # 行9 name：与 A_r3/trunk_r3 冲突
    ])
    _set_cells(b_r3 / "match_stat.xlsx", "SeasonStat", [
        (5, 3, 370),                        # 行5 c2：与 A_r3(360)/trunk_r3(350) 冲突
    ])
    _set_cells(b_r3 / "config" / "skill.xlsx", "Skill", [
        (5, 2, "B_r3改的闪避"),              # 行5 name：与 A_r3 冲突
    ])
    _write_meta(b_r3, "B_r2", "r2", "B 分支 r3 版本（copyfrom=B_r2，含嵌套 config/skill.xlsx + 10w行 big_data.xlsx）")

    print("生成目录：")
    for d in sorted(DEMO_ROOT.iterdir()):
        if d.is_dir():
            tables = [p.name for p in d.glob("*.xlsx")]
            nested = [str(p.relative_to(d)).replace("\\", "/")
                      for p in d.rglob("*.xlsx")
                      if p.parent != d and not p.name.startswith("~$")]
            print(f"  {d.name}: tables={tables}" + (f"  nested={nested}" if nested else ""))


if __name__ == "__main__":
    main()
