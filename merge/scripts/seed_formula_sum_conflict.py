"""制造一个基于 SUM(...) 聚合函数的公式文本冲突，补充 seed_formula_text_conflict.py
（那个是简单算术公式 =N*86400），这里用真正的 SUM 区间聚合公式，覆盖不同的公式形态。

目标：guild.xlsx · Const sheet · 第43行 PK=GUILD_CREATE_COST_NUM（trunk/dev1/dev2
三边都有这一行，是 matched 行，不涉及新增行的 ID 冲突/重映射逻辑，专注验证公式
文本冲突判定）。用 F 列（该 sheet 里本来没声明用途的空列，之前 E 列已被少数行
占用做零散公式注释，这里换一列避免互相干扰）新增一个 SUM 公式：

  trunk：F 列不写（无公式，走"从衍生版本取模板"识别公式列）
  dev1 的 F 列公式：=SUM(C5:C14)
  dev2 的 F 列公式：=SUM(C5:C20)

两边区间不同 → 公式文本不同 → compare_sheet 判定 diff_type="formula_conflict"。

注：最初尝试是把这个 SUM 公式放在一个"trunk 没有、dev1/dev2 各自新增"的新 PK 行上
（inserted 行），结果触发了 id_resolver 的"多分支同 PK 冲突"重映射逻辑，把 dev2 那
行的字符串主键 GUILD_CONST_SUM_TOTAL 重映射成了数字 "1"（_next_pk 对非数字 PK 的
兜底逻辑不适合字符串主键表，属于另一个应该跟进的边界 bug）。为了让这个 SUM 场景
干净地演示 formula_conflict（而不是被 id 重映射路径吞掉/转移），改用已存在的
matched 行 + 新增一个原本没人用的公式列来触发。

幂等：目标单元格已是期望公式则跳过对应分支的提交。
"""
import subprocess
from pathlib import Path
from openpyxl import load_workbook

WC_ROOT = Path(__file__).resolve().parent.parent / "svn" / "demo_svn" / "wc"
DEV1 = WC_ROOT / "branches" / "dev1"
DEV2 = WC_ROOT / "branches" / "dev2"
TABLE_REL = Path("guild.xlsx")
SHEET = "Const"
PK_COL = 1
TARGET_PK = "GUILD_CREATE_COST_NUM"
FORMULA_COL = 6  # F 列（未被占用）


def _find_row_by_pk(ws, pk: str, max_row: int = 130) -> int:
    for r in range(1, max_row + 1):
        if ws.cell(r, PK_COL).value == pk:
            return r
    return -1


def _commit(wc: Path, msg: str) -> None:
    r = subprocess.run(
        ["svn", "commit", "-m", msg, str(wc)],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    print(f"  [{'OK' if r.returncode == 0 else 'FAIL'}] {msg}")
    if r.returncode != 0:
        print(f"    stderr: {r.stderr.strip()}")


def _set_sum_formula(branch_dir: Path, formula: str, msg: str) -> None:
    fp = branch_dir / TABLE_REL
    if not fp.exists():
        print(f"  [skip] {branch_dir.name} 缺少 {TABLE_REL}")
        return
    wb = load_workbook(fp, data_only=False)
    if SHEET not in wb.sheetnames:
        wb.close()
        print(f"  [skip] {branch_dir.name} {TABLE_REL} 缺少 sheet {SHEET}")
        return
    ws = wb[SHEET]
    row = _find_row_by_pk(ws, TARGET_PK)
    if row < 0:
        wb.close()
        print(f"  [skip] {branch_dir.name} 未找到 PK={TARGET_PK}")
        return
    if ws.cell(row, FORMULA_COL).value == formula:
        wb.close()
        print(f"  [skip] {branch_dir.name} 已是目标公式（幂等）")
        return
    ws.cell(row, FORMULA_COL).value = formula
    wb.save(fp)
    wb.close()
    print(f"  {branch_dir.name}: 第{row}行 F列 -> {formula}")
    _commit(branch_dir, msg)


def main() -> None:
    print("=== 制造 SUM 公式文本冲突（formula_conflict，matched 行）===")
    _set_sum_formula(DEV1, "=SUM(C5:C14)", "dev1: 新增 SUM 公式（制造公式冲突，guild.xlsx Const F列）")
    _set_sum_formula(DEV2, "=SUM(C5:C20)", "dev2: 新增 SUM 公式（制造公式冲突，guild.xlsx Const F列）")
    print(f"\n完成。查看方式：合并引导页 /merge-guide?mode=branch，选择 dev1<->dev2，"
          f"表 guild / sheet Const，PK={TARGET_PK}，F 列——dev1 是 =SUM(C5:C14)，"
          "dev2 是 =SUM(C5:C20)，格子标红，点开会弹出选择采纳哪个公式版本的冲突框。")


if __name__ == "__main__":
    main()
