"""表格元信息索引（表格注册中心 · 层1）：扫描 resources 下所有 xlsx，生成/加载语义索引。

索引结构（每个表条目）:
  - path:        相对路径，如 school/school_ability.xlsx
  - stem:        文件名无后缀
  - md5:         文件内容 MD5，用于变更检测（外部修改后增量刷新）
  - sheets:      每个 sheet 的 {name, headers, header_names, header_row, data_start_row, row_count, row_index, samples}
                 samples = 前 3 行样本数据（list[list[str]]），供 LLM 上下文与人工预览

索引产物保存在 agent/_table_index.json，供 TableAgent / CodemakerNLParser / TableLocator 检索使用。
重新生成调用 build_index(workspace)；增量刷新调用 refresh_if_changed(workspace)。
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional


# 样本单元格长度上限：过长样本会撑大索引体积，截断保留语义即可
_SAMPLE_CELL_MAX = 80
# 每个 sheet 保留的样本行数
_SAMPLE_ROW_COUNT = 3
# T1: search_blob 每个 sheet 收集的去重值上限（控制索引体积）
_SEARCH_BLOB_CAP = 3000
# T1: 单个值超过此长度不纳入 blob（长文本误判 + 膨胀）
_SEARCH_VALUE_MAX_LEN = 60


@dataclass
class SheetMeta:
    name: str
    headers: list[str]
    header_names: list[str]
    header_row: int = 1
    data_start_row: int = 5
    row_count: int = 0
    # 行级倒排索引：{列名: {值: [行号, ...]}}，仅对"名称"和"id"类高频定位列建立
    row_index: dict[str, dict[str, list[int]]] = field(default_factory=dict)
    # 前 3 行样本数据（每行为 list[str]，单元格超长截断），供 LLM 上下文与人工预览
    samples: list[list[str]] = field(default_factory=list)
    # T30：per-col 非空计数（每列实际有值的行数），供 derive_required_fields 派生必填列。
    # 非空率 ≥ 0.9 的列 → required_fields.yaml。旧索引无此字段 → load 时默认空 list（兼容）。
    col_non_empty: list[int] = field(default_factory=list)
    # T1 搜索预检 blob：非长文本列的去重值小写拼接（"\n" 分隔，上限 3000 条），
    # search_rows 加载 workbook 前先做 keyword in blob 子串预判，不命中则跳过该表，
    # 避免 65+ 表全量扫描。空串表示未建（旧索引兼容），预检放行走原扫描流程。
    search_blob: str = ""


# ── 全局反向列名索引：列名 → [(stem, sheet)] ──────────────────────
# 模块级单例，build_index / refresh_if_changed 时重建。
# 供 Step1 列名提取阶段：用户输入命中某列名 → 反查该列出现在哪些 (stem, sheet)，
# 用 topK 收敛候选表/sheet，避免 LLM 黑盒选表选错（案例三 spirit 误路由根因）。
_COLUMN_REVERSE_INDEX: dict[str, list[tuple[str, str]]] = {}


def get_column_reverse_index() -> dict[str, list[tuple[str, str]]]:
    """返回全局列名→[(stem,sheet)] 反查索引。

    build_index/refresh_if_changed 后填充。空表示索引未建（旧索引/首次加载），
    调用方应回退到 TableLocator._level5_column 内存扫描。
    """
    return _COLUMN_REVERSE_INDEX


def _rebuild_column_reverse_index(tables: list[TableMeta]) -> None:
    """重建全局反向列名索引：遍历所有表所有 sheet 的 header_names，聚合 {列名: [(stem,sheet)]}。

    列名清洗后入库（_clean_header 去类型后缀/换行/括号），空串与过短（<2）跳过。
    同名列出现在多表多 sheet → 全部收录，供 topK 打分时区分度归并。
    """
    global _COLUMN_REVERSE_INDEX
    idx: dict[str, list[tuple[str, str]]] = {}
    for tm in tables:
        for s in tm.sheets:
            for hn in s.header_names:
                if not hn or len(hn) < 2:
                    continue
                idx.setdefault(hn, []).append((tm.stem, s.name))
    _COLUMN_REVERSE_INDEX = idx


@dataclass
class TableMeta:
    path: str                      # 相对 workspace 的路径
    stem: str                      # 文件名无后缀，如 school_ability
    md5: str = ""                  # 文件内容 MD5，变更检测用
    sheets: list[SheetMeta] = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class RefreshResult:
    """refresh_if_changed 的返回结果，描述本次增量刷新的变更明细。"""
    changed: list[str] = field(default_factory=list)   # 内容变化的文件 stem
    added: list[str] = field(default_factory=list)     # 新增文件 stem
    removed: list[str] = field(default_factory=list)   # 被删除文件 stem
    total: int = 0                                     # 刷新后总表数
    refreshed: bool = False                            # 是否实际发生刷新


def compute_md5(path: Path) -> str:
    """分块计算文件 MD5，避免大文件一次性读入内存。"""
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _sample_cell(v) -> str:
    """把单元格值转成字符串样本，超长截断。None → 空串。"""
    if v is None:
        return ""
    s = str(v)
    if len(s) > _SAMPLE_CELL_MAX:
        return s[:_SAMPLE_CELL_MAX] + "…"
    return s


def _clean_header(h) -> str:
    """去除表头里的换行、括号注释、类型后缀，保留纯列名。"""
    if h is None:
        return ""
    s = str(h)
    # 去换行
    s = s.replace("\n", "").replace("\r", "")
    # 去括号注释（中英文括号）
    s = re.sub(r"[（(].*?[）)]", "", s)
    # 去类型后缀 :int / :string 等
    s = s.split(":")[0]
    return s.strip()


def _is_type_cell(v) -> bool:
    """判断单元格是否为类型标注（如 'dan_id:int'、'string'、'base_score:float'）。

    严格匹配 name:type 或纯类型名，避免把公式（如 '=SUM(B3:E3)'，含冒号）
    误判为类型标注；纯类型名仅匹配已知原始类型词，避免把数据值
    （如 'attack'、'gold'、'defense'）误判为类型行。
    """
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s or s.startswith("="):
        return False  # 公式串不是类型标注
    # name:type 形式（冒号分隔，左侧为标识符，右侧为类型名）
    if ":" in s:
        name, _, type_part = s.partition(":")
        name = name.strip()
        type_part = type_part.strip()
        # 左侧 name 允许中英文标识符（如 灵兽id:int、ability_id:int、dan_id:int）。
        # 原 _IDENT_RE 仅 ASCII，把 perf 表/中文前缀类型行（灵兽id:int）误判为数据行，
        # 导致 data_start_row 错标为类型行行号（2 而非 3）。
        name_ok = bool(name) and bool(
            re.fullmatch(r"[a-zA-Z_][a-zA-Z0-9_]*", name)
            or re.fullmatch(r"[\u4e00-\u9fffA-Za-z_][\u4e00-\u9fffA-Za-z0-9_]*", name))
        type_ok = bool(re.fullmatch(r"[a-zA-Z_][a-zA-Z0-9_\[\], ]*", type_part))
        return name_ok and type_ok
    # 纯类型名：仅匹配已知原始类型词，避免 attack/gold/defense 等数据值误判
    _PRIMITIVE_TYPES = {
        "int", "int8", "int16", "int32", "int64", "uint", "uint8", "uint16",
        "uint32", "uint64", "float", "float32", "float64", "double",
        "string", "str", "bool", "boolean", "list", "dict", "any", "number",
        "void", "object", "bytes", "long", "short", "char",
    }
    return s.lower() in _PRIMITIVE_TYPES


def _detect_rows(ws) -> tuple[int, int]:
    """推断表头行与数据起始行。

    约定（参考现有表格）:
      - 第1行表头、第2行类型行；类型行后可能紧跟空/示例行，第5行起才是真实数据。
      - 兼容个别表：类型行后紧接真实数据（无空/示例行），此时数据从第3行起。
      - 兼容说明/标题类 sheet：第1行空，真正表头在后续行（向下扫30行
        找首个≥2非空单元格行作表头，data_start=表头+1）。

    数据起始行判定（有类型行时）：从类型行下一行起扫描，首个既非类型标注、
    又含≥2非空单元格的行即为 data_start；这样空/示例表跳到第5行，紧接数据的
    测试表正确落到第3行。
    """
    upper = min(30, ws.max_row)
    top = list(ws.iter_rows(min_row=1, max_row=upper, values_only=True))
    # 首个≥2非空单元格行作为表头
    header_row = 1
    for idx, cells in enumerate(top, start=1):
        if sum(1 for v in (cells or []) if v is not None) >= 2:
            header_row = idx
            break

    if header_row != 1:
        # 表头不在第1行（说明/标题类 sheet），表头下一行起即数据
        return header_row, header_row + 1

    # 表头在第1行：判断第2行是否为类型行
    type_row_cells = list(top[1]) if len(top) > 1 else []
    is_type_row = any(_is_type_cell(v) for v in type_row_cells if v)
    if not is_type_row:
        # 无类型行，数据从第2行起
        return 1, 2

    # 有类型行：从类型行下一行起扫描首个真实数据行
    # （跳过空行、示例行、以及仍为类型标注的行）
    scan_end = min(upper, ws.max_row)
    for r in range(header_row + 2, scan_end + 1):
        cells = top[r - 1] if r - 1 < len(top) else []
        non_empty = [v for v in (cells or []) if v is not None]
        if len(non_empty) < 2:
            continue  # 空行或仅1个值的示例行，跳过
        # 整行非空值必须全部是类型标注才算类型行，跳过；
        # 一旦出现数字/中文/公式等非类型值即为数据行。
        str_vals = [v for v in non_empty if isinstance(v, str)]
        if str_vals and all(_is_type_cell(v) for v in str_vals) and len(str_vals) == len(non_empty):
            continue  # 全是类型标注，跳过
        return header_row, r
    # 扫描完仍未确定（极端空表），回退表头+1
    return header_row, header_row + 1


def _scan_sheet(ws) -> Optional[SheetMeta]:
    if ws.title.lower() == "config":
        return None
    header_row, data_start = _detect_rows(ws)
    hr = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    if not hr:
        return None
    raw_headers = list(hr[0])
    while raw_headers and raw_headers[-1] is None:
        raw_headers.pop()
    if not raw_headers or all(v is None for v in raw_headers):
        return None

    # 类型行（仅当 data_start > header_row+1 时存在）；空表头列回退用类型行前缀翻译
    # （如 ability 表第1列表头空、类型行 ability_id:int → 神通id），让无表头主键列可被寻址。
    type_row = header_row + 1 if data_start > header_row + 1 else None
    type_vals: list = []
    if type_row:
        tr = list(ws.iter_rows(min_row=type_row, max_row=type_row, values_only=True))
        type_vals = list(tr[0]) if tr else []
    from .column_name_resolver import resolve_header_cell
    headers: list[str] = []
    header_names: list[str] = []
    for i in range(len(raw_headers)):
        hv = raw_headers[i]
        tv = type_vals[i] if i < len(type_vals) else None
        resolved = resolve_header_cell(hv, tv)
        headers.append(str(resolved) if resolved is not None else "")
        header_names.append(_clean_header(headers[-1]))

    # 精确计数：只统计有实际数据的行（非全空行）
    actual_rows = 0
    row_index: dict[str, dict[str, list[int]]] = {}
    # 收集名称/id类列索引，用于后续行级倒排
    id_cols = _find_id_columns(header_names)
    # T1: 收集可搜索列（非长文本/资源路径类），用于 search_blob 预检
    searchable_cols = _find_searchable_columns(header_names)
    value_set: set[str] = set()
    # 前 N 行样本数据（仅取有实际数据的前 _SAMPLE_ROW_COUNT 行）
    samples: list[list[str]] = []
    ncols = len(headers)
    # T30：per-col 非空计数（非 None 且 str(v).strip() 非空），供派生必填列
    col_non_empty: list[int] = [0] * ncols

    # read_only 模式下 ws.cell(r,c) 是 O(n) 随机访问，range(data_start, max_row+1)
    # 整体 O(n²)；且 max_row 可能被报成超大值（如 1048576）导致空转卡死。
    # 改用 iter_rows 前向迭代：只产出 XML 中真实存在的行，天然跳过空尾行，O(n)。
    for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True)):
        r = data_start + i
        # 对齐到表头列数：超长截断，不足补 None
        if row:
            row_vals = [row[c] if c < len(row) else None for c in range(ncols)]
        else:
            row_vals = [None] * ncols
        if any(v is not None for v in row_vals):
            actual_rows += 1
            # T30：per-col 非空计数（非 None 且 strip 非空）
            for c in range(ncols):
                v = row_vals[c] if c < len(row_vals) else None
                if v is not None and str(v).strip():
                    col_non_empty[c] += 1
            # 对名称/id类列建立倒排索引
            for col_idx, col_name in id_cols.items():
                v = row_vals[col_idx] if col_idx < len(row_vals) else None
                if v is not None:
                    key = str(v).strip()
                    if key:
                        row_index.setdefault(col_name, {}).setdefault(key, []).append(r)
            # T1: 收集可搜索列的去重值（小写），供 search_blob 预检
            if len(value_set) < _SEARCH_BLOB_CAP:
                for col_idx in searchable_cols:
                    v = row_vals[col_idx] if col_idx < len(row_vals) else None
                    if v is not None:
                        s = str(v).strip().lower()
                        if s and len(s) <= _SEARCH_VALUE_MAX_LEN:
                            value_set.add(s)
            # 收集前几行作为样本
            if len(samples) < _SAMPLE_ROW_COUNT:
                samples.append([_sample_cell(v) for v in row_vals])

    search_blob = "\n".join(sorted(value_set))

    return SheetMeta(
        name=ws.title,
        headers=headers,
        header_names=header_names,
        header_row=header_row,
        data_start_row=data_start,
        row_count=actual_rows,
        row_index=row_index,
        samples=samples,
        col_non_empty=col_non_empty,
        search_blob=search_blob,
    )


def _find_id_columns(header_names: list[str]) -> dict[int, str]:
    """找出名称/id类高频定位列的索引（0-based）及其清理后列名。

    匹配"名称""名字""id""编号"类列名，这些列适合建立行级倒排索引。
    """
    result: dict[int, str] = {}
    id_keywords = ("名称", "名字", "id", "编号", "名")
    for i, hn in enumerate(header_names):
        hn_l = hn.lower()
        for kw in id_keywords:
            if kw in hn_l or hn_l in kw:
                # 排除非定位列（如 "名称描述" 不是纯名称列，但 "物品名称" 是）
                if not any(bad in hn_l for bad in ("描述", "icon", "图标")):
                    result[i] = hn
                    break
    return result


# T1: search_blob 排除的长文本/资源路径类列关键字（值过长或为路径，纳入会膨胀且误判）
_SEARCH_EXCLUDE_KEYWORDS = (
    "描述", "说明", "comment", "备注", "icon", "图标", "path", "路径",
    "texture", "贴图", "tips", "提示", "url", "link", "配置", "config",
)


def _find_searchable_columns(header_names: list[str]) -> list[int]:
    """找出可纳入 search_blob 的列索引（0-based），排除长文本/资源路径类列。

    与 _find_id_columns 互补：id/名称列进 row_index 倒排（精准定位），
    其余非长文本列进 search_blob（子串预检，跳过不可能命中的表）。
    """
    out: list[int] = []
    for i, hn in enumerate(header_names):
        if not hn:
            continue
        hn_l = hn.lower()
        if any(bad in hn_l for bad in _SEARCH_EXCLUDE_KEYWORDS):
            continue
        out.append(i)
    return out


def build_index(workspace: Path) -> list[TableMeta]:
    """扫描 workspace 下所有 xlsx，返回 TableMeta 列表并写 JSON。

    每个文件计算 MD5 一并写入索引，供 refresh_if_changed 做变更检测。
    """
    workspace = Path(workspace)
    tables: list[TableMeta] = []
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("需要 openpyxl：python -m pip install openpyxl") from e

    for p in sorted(workspace.rglob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception:
            continue
        sheets = [s for s in (_scan_sheet(ws) for ws in wb.worksheets) if s]
        if not sheets:
            continue
        tables.append(TableMeta(
            path=str(p.relative_to(workspace)).replace("\\", "/"),
            stem=p.stem,
            md5=compute_md5(p),
            sheets=sheets,
        ))
    _write_index(tables)
    _rebuild_column_reverse_index(tables)
    return tables


def _idx_path() -> Path:
    """索引 JSON 文件路径（excel/_table_index.json）。

    重构把本模块挪进 excel/locator/ 子包，但 76MB _table_index.json 仍在
    excel/ 父级生成/存储。原 `Path(__file__).parent` 会指向
    locator/_table_index.json（不存在）→ load_index 抛 FileNotFoundError →
    TableLocator.index 为空 → 表定位全 miss、is_cross_table 永假、DecomposeAgent
    永不触发（quest/reward 全丢的根因之一）。
    逐级向上（本目录→父 excel/→祖父）找首个存在处，找不到回退父级以触发重建。
    """
    here = Path(__file__).resolve().parent
    for cand in (here, here.parent, here.parent.parent):
        p = cand / "_table_index.json"
        if p.exists():
            return p
    return here.parent / "_table_index.json"


def _write_index(tables: list[TableMeta]) -> None:
    """把 TableMeta 列表序列化写到索引 JSON（原子写）。

    先写 .tmp 再 os.replace：读者要么看到旧完整文件要么看到新完整文件，
    不会读到 truncate 后未写完的空文件。TableFileWatcher 后台线程与主线程
    均会触发 refresh_if_changed，非原子 write_text 在并发下会留下瞬时空文件，
    导致 TableResolver.__init__ 的 json.loads 抛 JSONDecodeError 崩溃。
    异常时清理临时文件，避免遗留 .tmp 残留。
    """
    import os
    import tempfile
    p = _idx_path()
    data = json.dumps([t.to_dict() for t in tables], ensure_ascii=False, indent=2)
    # 原子写：mkstemp 生成临时文件再 os.replace，避免 eval 多 case 并发读时读到半截 JSON
    tmp_fd, tmp_path = tempfile.mkstemp(dir=str(p.parent), suffix=".tmp",
                                         prefix="_table_index_")
    try:
        with os.fdopen(tmp_fd, "w", encoding="utf-8") as f:
            f.write(data)
        os.replace(tmp_path, p)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def load_index() -> list[TableMeta]:
    """加载已生成的索引（不重新扫描）。

    兼容旧索引：md5/samples 字段缺失时回退默认值。
    索引缺失或损坏时抛异常——refresh_if_changed 据此触发全量重建。
    只读场景（TableResolver/get_tables 等）请用 load_index_safe()。
    """
    p = _idx_path()
    data = json.loads(p.read_text(encoding="utf-8"))
    out = []
    for d in data:
        sheets = []
        for s in d["sheets"]:
            row_index = s.get("row_index", {})
            sheets.append(SheetMeta(
                name=s["name"], headers=s["headers"], header_names=s["header_names"],
                header_row=s.get("header_row", 1), data_start_row=s.get("data_start_row", 5),
                row_count=s.get("row_count", 0), row_index=row_index,
                samples=s.get("samples", []),
                search_blob=s.get("search_blob", ""),
                col_non_empty=s.get("col_non_empty", []),
            ))
        out.append(TableMeta(
            path=d["path"], stem=d["stem"], md5=d.get("md5", ""), sheets=sheets,
        ))
    _rebuild_column_reverse_index(out)
    return out


def load_index_safe() -> list[TableMeta]:
    """防御式加载：索引缺失/损坏/并发写截断时返回 [] 而非抛异常。

    供 TableResolver/get_tables 等只读场景使用——避免与 refresh_if_changed
    的原子写竞争瞬间读到空文件导致 JSONDecodeError 崩溃整个 Agent。
    返回 [] 时路由/列名匹配退化为空，Agent 报"未匹配到表"而非崩溃。
    """
    try:
        return load_index()
    except (json.JSONDecodeError, ValueError, OSError, FileNotFoundError):
        return []


def refresh_if_changed(workspace: Path) -> RefreshResult:
    """增量刷新：比对已存索引的 MD5 与磁盘文件当前 MD5，仅重扫变更/新增/删除的文件。

    流程：
      1. 加载已存索引（无则全量 build_index）
      2. 扫描 workspace 现有 xlsx 列表
      3. 逐文件比对 MD5：变化 → 重扫该文件；新增 → 加入；删除 → 移除
      4. 写回索引，返回 RefreshResult

    若无任何变化，refreshed=False，不写盘。
    """
    workspace = Path(workspace)
    idx_path = _idx_path()
    if not idx_path.exists():
        tables = build_index(workspace)
        return RefreshResult(
            added=[t.stem for t in tables], total=len(tables), refreshed=True,
        )

    try:
        existing = load_index()
    except Exception:
        tables = build_index(workspace)
        return RefreshResult(
            added=[t.stem for t in tables], total=len(tables), refreshed=True,
        )

    # 已存索引按 path 建表，便于按文件定位
    by_path: dict[str, TableMeta] = {t.path: t for t in existing}
    # 仅当文件需要 openpyxl 时才 import，避免无依赖环境报错
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("需要 openpyxl：python -m pip install openpyxl") from e

    result = RefreshResult(total=0)
    changed_map: dict[str, TableMeta] = {}

    # 扫描现有文件：检测变化与新增
    current_paths: set[str] = set()
    for p in sorted(workspace.rglob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        rel = str(p.relative_to(workspace)).replace("\\", "/")
        current_paths.add(rel)
        cur_md5 = compute_md5(p)
        old = by_path.get(rel)
        if old is None:
            # 新增文件
            try:
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                sheets = [s for s in (_scan_sheet(ws) for ws in wb.worksheets) if s]
                if sheets:
                    changed_map[rel] = TableMeta(
                        path=rel, stem=p.stem, md5=cur_md5, sheets=sheets,
                    )
                    result.added.append(p.stem)
            except Exception:
                continue
        elif old.md5 != cur_md5:
            # 内容变化 → 重扫
            try:
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                sheets = [s for s in (_scan_sheet(ws) for ws in wb.worksheets) if s]
                if sheets:
                    changed_map[rel] = TableMeta(
                        path=rel, stem=p.stem, md5=cur_md5, sheets=sheets,
                    )
                    result.changed.append(p.stem)
                else:
                    # 重扫后无有效 sheet，视为删除
                    result.removed.append(p.stem)
            except Exception:
                # 文件损坏 → 跳过，保留旧索引条目
                changed_map[rel] = old
        else:
            # 未变化，保留原条目
            changed_map[rel] = old

    # 检测删除：旧索引中存在但磁盘已无的文件
    for rel, old in by_path.items():
        if rel not in current_paths:
            result.removed.append(old.stem)

    result.refreshed = bool(result.changed or result.added or result.removed)
    if result.refreshed:
        tables = list(changed_map.values())
        tables.sort(key=lambda t: t.path)
        try:
            _write_index(tables)
        except (PermissionError, OSError) as e:
            # Windows 并发写索引文件锁冲突（eval 多 case 串行但文件句柄残留）：
            # 写盘失败不阻断流程，旧索引仍可用（load_index 有 JSON 损坏 fallback）。
            import warnings
            warnings.warn(f"写盘失败（{e}），保留旧索引", RuntimeWarning)
        _rebuild_column_reverse_index(tables)
        result.total = len(tables)
    else:
        _rebuild_column_reverse_index(existing)
        result.total = len(existing)
    return result


if __name__ == "__main__":
    import sys
    # 解析参数：--refresh 触发增量刷新，其余非 flag 参数为 workspace 路径
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    do_refresh = "--refresh" in sys.argv
    ws = Path(args[0]) if args else Path("resources")
    if do_refresh:
        r = refresh_if_changed(ws)
        print(f"增量刷新完成：变更 {len(r.changed)}，新增 {len(r.added)}，"
              f"删除 {len(r.removed)}，总表数 {r.total}，实际写盘 {r.refreshed}")
        if r.changed:
            print(f"  变更: {r.changed}")
        if r.added:
            print(f"  新增: {r.added}")
        if r.removed:
            print(f"  删除: {r.removed}")
    else:
        tables = build_index(ws)
        print(f"索引完成：{len(tables)} 个表，{sum(len(t.sheets) for t in tables)} 个 sheet")
        for t in tables[:5]:
            print(f"  {t.path}: {[s.name for s in t.sheets]}")
