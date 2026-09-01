"""CodeMaker CLI 抽象接口与 stub 实现。

真实 CLI 对接前先使用 StubCodeMakerCLI 跑通端到端流程。
后续接入真实 CLI 时，只需继承 CodeMakerCLI 实现一个新类即可，
上层调用方无需改动。
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from ..formula.formula_ref_shifter import permute_formula_rows, shift_workbook_formulas


def _serialize_cell_value(v):
    """复合值序列化兜底，防 openpyxl 直接写非标量报错。

    与 agent 侧规则一致：list 逗号分隔不写 []，tuple 用 ()，dict 用
    JSON 文本。嵌套子项用 () 包裹。非复合值原样返回。CLI 层兜底覆盖所有调用方
    （含未走 agent._coerce_value 的直调路径）。
    """
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
    if isinstance(v, list):
        return ",".join(_serialize_cell_item(x) for x in v)
    if isinstance(v, tuple):
        return "(" + ",".join(_serialize_cell_item(x) for x in v) + ")"
    return v


def _serialize_cell_item(x):
    if isinstance(x, (list, tuple)):
        return "(" + ",".join(str(e) for e in x) + ")"
    return str(x)


@dataclass
class CLICallResult:
    """CLI 单次调用结果。"""

    ok: bool
    stdout: str = ""
    stderr: str = ""
    data: Any = None
    error: Optional[str] = None
    needs_manual_fix: bool = False
    cache_message: str = ""
    # 批注守门：{replayed: bool, still_lost: int}。save 后检测批注丢失 → 回写 → 二次 save，
    # 仍丢则 still_lost 记数。replayed=True 表示触发了一次回写补救。默认空 dict 兼容老调用方。
    comment_replay: dict = field(default_factory=dict)
    # AD1/AD2：pre_commit_hold 事件（kind=formula_loss / comment_loss）。CLI 层构造 +
    # record_hold_audit 留痕，agent 层读此触发 emit_hold_sse + hold 阻断。默认空 list。
    hold_events: list = field(default_factory=list)


@dataclass
class SearchResult:
    """单条搜索结果。

    row: 绝对 Excel 行号（1-based，供 write/delete 等操作定位）。
    data_row: 浏览页相对行号（1-based，相对 read_browse 的数据起始行），
              与前端表格 "#" 列一致，供搜索跳转高亮定位。无法换算时为 0。
    """
    sheet: str
    row: int
    col: int
    col_name: str
    cell_value: Any
    row_data: list
    data_row: int = 0


class CodeMakerCLI:
    """CodeMaker CLI 抽象接口。

    约定的最小操作集合：
      - list_tables()          列出工作环境下所有 Excel 表格路径
      - get_sheets(path)       读取某表格的 sheet 名列表
      - read_sheet(path, sheet) 读取某个 sheet 的二维数据（含表头与类型行）
      - search_rows(path, sheet, keyword)  按关键词搜索行
      - locate_row(path, sheet, col, value, mode) 按值定位行号
      - write_cell(path, sheet, row, col, value) 写入单元格
      - read_cell(path, sheet, row, col)        读取单元格
      - append_row(path, sheet, values)         追加新行
      - delete_row(path, sheet, row)            删除行

    真实实现应通过 subprocess 调用 CLI 二进制；stub 实现直接用 openpyxl 操作文件，
    便于在无 CLI 环境下验证流程。
    """

    name: str = "base"

    def read_header(self, path: Path, sheet: str) -> list:
        raise NotImplementedError

    def list_tables(self) -> list[Path]:
        raise NotImplementedError

    def get_sheets(self, path: Path) -> list[str]:
        raise NotImplementedError

    def read_sheet(self, path: Path, sheet: str,
                   offset: int = 0, limit: int | None = None) -> list[list]:
        raise NotImplementedError

    def search_rows(self, path: Path, keyword: str, sheet: str = "",
                    col: str = "") -> list[SearchResult]:
        raise NotImplementedError

    def locate_row(self, path: Path, sheet: str, col: int, value: str,
                   mode: str = "contains") -> int | None:
        raise NotImplementedError

    def write_cell(self, path: Path, sheet: str, row: int, col: int, value: Any,
                   number_format: str | None = None) -> CLICallResult:
        raise NotImplementedError

    def read_cell(self, path: Path, sheet: str, row: int, col: int) -> CLICallResult:
        raise NotImplementedError

    def append_row(self, path: Path, sheet: str, values: dict[int, Any]) -> CLICallResult:
        raise NotImplementedError

    def delete_row(self, path: Path, sheet: str, row: int) -> CLICallResult:
        raise NotImplementedError

    def insert_row(self, path: Path, sheet: str, row: int,
                   values: dict[int, Any] | None = None) -> CLICallResult:
        """在指定行上方插入新行（1-based）。

        AI 想在有序编号段中间插入新配置行时使用，区别于末尾追加的 append_row。
        内部 ws.insert_rows(row) 后调 shift_workbook_formulas(row, +1) 重写公式引用。
        """
        raise NotImplementedError

    # ---- L2 公式语义原语（agent 化公式处理，判断权在 AI）----
    def interpret_formula(self, path: Path, sheet: str, cell: str):
        """解析单格公式语义。纯读。详见 formula_semantics.interpret_formula。"""
        raise NotImplementedError

    def scan_sheet_formulas(self, path: Path, sheet: str):
        """扫描全 sheet 公式语义。纯读。详见 formula_semantics.scan_sheet_formulas。"""
        raise NotImplementedError

    def preview_formula_impact(self, path: Path, sheet: str, op: dict):
        """dry-run 模拟增删改，返回机械位移结果 + 语义缺口标注。纯读。"""
        raise NotImplementedError

    def rewrite_formula(self, path: Path, sheet: str, cell: str,
                        new_formula: str) -> CLICallResult:
        """按 AI 决策写入目标公式 + 缓存校验。详见 formula_semantics.rewrite_formula。"""
        raise NotImplementedError

    # ---- 列级操作 ----
    def list_columns(self, path: Path, sheet: str) -> list[tuple[int, str]]:
        """列出 sheet 的所有列，返回 [(col_idx, col_name), ...]。"""
        raise NotImplementedError

    def insert_column(self, path: Path, sheet: str, name: str,
                      after: int | None = None, type_str: str | None = None,
                      default: Any = None) -> CLICallResult:
        """新增列。after 为列索引时在其右侧插入，None 时追加到末尾。

        Args:
            name: 新列表头名
            after: 插入位置（某列索引），None 追加末尾
            type_str: 类型标注（写入表头下一行，即类型行），可选
            default: 数据区默认值，None 不填充
        """
        raise NotImplementedError

    def delete_column(self, path: Path, sheet: str, col: int | str) -> CLICallResult:
        """删除列。col 可为列索引(1-based)或列名。"""
        raise NotImplementedError

    def rename_column(self, path: Path, sheet: str, col: int | str,
                      new_name: str) -> CLICallResult:
        """重命名列。col 可为列索引(1-based)或列名。"""
        raise NotImplementedError

    def sort_sheet(self, path: Path, sheet: str, key_col: int = 1,
                   ascending: bool = True) -> CLICallResult:
        """按指定列对数据区行重排序（表头/类型行不动）。

        整行搬移（值 + 单元格样式），公式单元格按文本原样搬移、不重写引用
        ——遵循"写操作→排序→最后更新公式"工作流，公式引用更新留待独立步骤。

        Args:
            key_col: 排序依据列（1-based），默认第1列（编号/主键列）。
            ascending: True 升序，False 降序。
        """
        raise NotImplementedError


class StubCodeMakerCLI(CodeMakerCLI):
    """基于 openpyxl 的 stub 实现，模拟 CodeMaker CLI 行为。

    行约定：
      - 默认数据行从第 5 行开始（第1行表头、第2行类型、第3-4行空行/示例）。
      - 可通过 _resolve_data_start 从索引动态获取每个表的数据起始行。
      - read_sheet 返回从数据起始行起的数据行，不含表头。
    """

    name = "stub"

    def __init__(self, workspace: Path, header_row: int = 1, data_start_row: int = 5):
        self.workspace = Path(workspace)
        self.header_row = header_row
        self.data_start_row = data_start_row
        self._cache: dict[tuple[str, str], Any] = {}
        # 5.6：read_sheet 全量结果缓存，键 (path, sheet)，写操作经 _invalidate 清除。
        self._row_cache: dict[tuple[str, str], list[list]] = {}
        # 公式检测缓存：键 (path, mtime)，避免每次 read_sheet 都读 zip 扫描 <f> 标签
        self._formula_check_cache: dict[tuple[str, float], bool] = {}
        # xlsx 特性缓存：键 (path, mtime) -> (has_formula, has_comment, has_merge)
        self._xlsx_features_cache: dict[tuple[str, float], tuple[bool, bool, bool]] = {}
        self._index: dict[str, dict[str, int]] = {}
        self._load_data_start_from_index()
        # T1: 搜索预检用的表索引缓存（含 search_blob），懒加载
        self._table_idx_cache: Optional[list] = None
        # 索引文件 mtime（缓存失效判定：refresh_if_changed 重写后自动重载）
        self._table_idx_mtime: float = 0.0
        # 公式缓存校验器（延迟导入避免循环依赖）
        from ..formula.formula_cache_validator import FormulaCacheValidator
        self._formula_validator = FormulaCacheValidator()
        # 方法 A：BackupAuditor 懒加载，用于 needs_manual_fix 的 audit 留痕
        self._auditor = None

    def _get_auditor(self):
        """懒加载 BackupAuditor（用于 needs_manual_fix=True 的 audit 留痕）。

        复用 server/backups/audit_log.jsonl 路径约定（BackupAuditor 默认）。
        加载失败返回 None（化为不记 audit）。
        """
        if self._auditor is None:
            try:
                from ..core.backup_audit import BackupAuditor
                self._auditor = BackupAuditor(workspace=self.workspace)
            except Exception:
                return None
        return self._auditor

    def _save_with_cache_check(self, wb, path: Path, sheet: Optional[str] = None) -> dict:
        """带公式缓存保护 + 批注守门的 save。

        流程：
          公式：save 前快照公式缓存 → wb.save → save 后校验 → 丢失则 LibreOffice 重算。
          批注：save 前快照批注 → wb.save → save 后 reload 做差 → 丢失则原 wb 回写 Comment
                → 二次 save → 二次做差记 still_lost（openpyxl save 偶发丢批注的精细化保护层）。
        fast-path：无公式 + 无批注零开销。

        缓存清除范围(3.4):无公式 fast-path + 已知 sheet → 收敛到受影响 sheet
        (wb 缓存仍按文件全清,整文件已重写;row 缓存仅清该 sheet,他表数据未变保留);
        含公式(可能触发 LibreOffice 重算触及他表)/缺 sheet → 全文件清(保守)。

        环境开关 CODEMAKER_COMMENT_GUARD=on|off（默认 on，off 退化仅 save 不守门）。

        Returns:
            dict: {needs_manual_fix, cache_message, comment_replay}，供写表方法填充 CLICallResult。
            comment_replay = {replayed: bool, still_lost: int}。
        """
        import logging
        import os
        logger = logging.getLogger(__name__)
        path = Path(path)
        # save 前快照（fast-path 内部判断无公式返回空 dict，零开销）
        before = self._formula_validator.snapshot_before(path)
        # 批注快照（环境开关控制，off 则跳过）
        comment_guard = os.environ.get("CODEMAKER_COMMENT_GUARD", "on").lower() != "off"
        # 无批注表跳过 _comment_snapshot 全量遍历（大表 10w 格遍历秒级开销省掉）
        _has_comment = self._xlsx_features(path)[1]
        before_comments = self._comment_snapshot(wb) if (comment_guard and _has_comment) else {}
        # 执行 save
        wb.save(path)
        # 批注守门：reload 做差 → 丢失则原 wb 回写 Comment → 二次 save → 二次做差记数
        comment_replay = {"replayed": False, "still_lost": 0}
        if before_comments:
            lost = self._detect_comment_loss(path, before_comments)
            if lost:
                self._replay_comments(wb, lost)
                wb.save(path)
                comment_replay["replayed"] = True
                lost2 = self._detect_comment_loss(path, before_comments)
                comment_replay["still_lost"] = len(lost2)
                if comment_replay["still_lost"]:
                    logger.warning(
                        "[CommentGuard] %s: %d comments still lost after replay",
                        path.name, comment_replay["still_lost"])
                    # D4：二次做差仍丢 → 记 audit_log operation=comment_replay_partial
                    # （AD2：同时产 pre_commit_hold 事件 kind=comment_loss，见下方统一注入）
                    auditor = self._get_auditor()
                    if auditor is not None:
                        try:
                            auditor.record(
                                operation="comment_replay_partial",
                                path=str(path),
                                sheet=sheet or "",
                                extra={
                                    "still_lost": comment_replay["still_lost"],
                                    "lost_coords": [list(k) for k in lost2][:50],
                                },
                            )
                        except Exception:
                            pass
        if not before and sheet:
            # 无公式 + 已知 sheet → 缓存清除收敛到受影响 sheet
            self._invalidate(path, sheet)
        else:
            self._invalidate(path)
        if not before:
            # 无公式单元格，fast-path 通过
            result = None
        else:
            # 含公式 → 校验 + 必要时重算
            result = self._formula_validator.validate_and_fix(path, before)
            if result.needs_manual_fix:
                # 方法 A：环境开关 CODEMAKER_FORMULA_GATE=on|off|hold
                # off=完全静默；on/hold=warning+audit（hold 阻断留第二波接 pre_commit_hold）
                gate = os.environ.get("CODEMAKER_FORMULA_GATE", "on").lower()
                if gate != "off":
                    logger.warning("[FormulaCache] %s: %s", path.name, result.message)
                    auditor = self._get_auditor()
                    if auditor is not None:
                        try:
                            auditor.record(
                                operation="formula_loss_detected",
                                path=str(path),
                                sheet=sheet or "",
                                extra={
                                    "cache_message": result.message,
                                    "gate": gate,
                                    "replayed_comments": comment_replay.get("replayed", False),
                                    "still_lost_comments": comment_replay.get("still_lost", 0),
                                },
                            )
                        except Exception:
                            pass
        # AD1/AD2：pre_commit_hold 事件构造（kind=formula_loss / comment_loss）。
        # CLI 层无 SSE task 上下文 → 仅 record_hold_audit 留痕 + 事件附返回 dict，
        # agent 层（_run_add/_run_set）读 hold_events 触发 emit_hold_sse + hold 阻断。
        hold_events: list = []
        try:
            from routers.precommit_hold import PreCommitHoldEvent, record_hold_audit
            _aud = self._get_auditor()
            # AD1：gate=hold + needs_manual_fix → formula_loss hold 事件
            if result is not None and result.needs_manual_fix and os.environ.get(
                    "CODEMAKER_FORMULA_GATE", "on").lower() == "hold":
                _ev = PreCommitHoldEvent(
                    kind="formula_loss", severity="hold", count=1,
                    sheets={sheet or "": {"cache_message": result.message}},
                    message=f"公式重算后仍需手动修复：{result.message}",
                    recommendation="manual_fix")
                record_hold_audit(_aud, _ev, str(path), sheet or "",
                                  extra={"replayed_comments": comment_replay.get("replayed", False)})
                hold_events.append(_ev)
            # AD2：still_lost>0 → comment_loss hold 事件
            if comment_replay.get("still_lost", 0) > 0:
                _ev2 = PreCommitHoldEvent(
                    kind="comment_loss", severity="hold",
                    count=comment_replay["still_lost"],
                    sheets={sheet or "": {"lost": comment_replay["still_lost"]}},
                    message=f"批注二次回写后仍丢 {comment_replay['still_lost']} 条",
                    recommendation="manual_fix")
                record_hold_audit(_aud, _ev2, str(path), sheet or "")
                hold_events.append(_ev2)
        except Exception:
            pass
        if result is None:
            return {"needs_manual_fix": False, "cache_message": "",
                    "comment_replay": comment_replay, "hold_events": hold_events}
        return {
            "needs_manual_fix": result.needs_manual_fix,
            "cache_message": result.message,
            "comment_replay": comment_replay,
            "hold_events": hold_events,
        }

    # ── 批注守门（方法 D）─────────────────────────────────────
    def _comment_snapshot(self, wb) -> dict:
        """遍历 wb 所有 sheet 所有 cell，收集批注快照。

        返回 {(sheet_name, coord): (text, author)}。无批注返回 {}（fast-path）。
        用于 save 前快照，save 后 reload 做差定位丢失批注。
        """
        snap: dict = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cm = cell.comment
                    if cm is not None:
                        snap[(ws.title, cell.coordinate)] = (cm.text, cm.author or "")
        return snap

    def _detect_comment_loss(self, path, before_comments: dict) -> dict:
        """reload 文件读批注，返回丢失/文本变更的 {(sheet, coord): (text, author)}。

        用独立 openpyxl.load_workbook 不走 _load，避免污染 wb 缓存且不与写 wb 冲突。
        before_comments 为空直接返回 {}（fast-path）。
        """
        if not before_comments:
            return {}
        import openpyxl
        wb2 = openpyxl.load_workbook(Path(path), data_only=False)
        after = self._comment_snapshot(wb2)
        wb2.close()
        lost: dict = {}
        for k, v in before_comments.items():
            cur = after.get(k)
            if cur is None or cur[0] != v[0]:
                lost[k] = v
        return lost

    def _replay_comments(self, wb, lost: dict) -> None:
        """在原 wb 上对 lost 的 coord 重新写回 Comment（text+author）。

        openpyxl save 偶发丢批注，原 wb 内存里 Comment 对象仍在，但序列化时丢失。
        重新构造新 Comment 对象赋值，触发二次 save 重新序列化。
        """
        from openpyxl.comments import Comment
        for (sheet, coord), (text, author) in lost.items():
            if sheet in wb.sheetnames:
                cell = wb[sheet][coord]
                cell.comment = Comment(text, author)


    def _load_data_start_from_index(self):
        """从 _table_index.json 加载每个表/sheet 的 data_start_row。"""
        try:
            idx_path = Path(__file__).resolve().parent.parent / "_table_index.json"
            if idx_path.exists():
                import json
                data = json.loads(idx_path.read_text(encoding="utf-8"))
                for t in data:
                    for s in t.get("sheets", []):
                        key = f"{t['stem']}|{s['name']}"
                        self._index[key] = s.get("data_start_row", self.data_start_row)
        except Exception:
            pass

    def _resolve_data_start(self, path: Path, sheet: str) -> int:
        """获取指定表/sheet 的数据起始行（优先从索引读取，回退默认值）。"""
        key = f"{path.stem}|{sheet}"
        return self._index.get(key, self.data_start_row)

    # ---- 基础工具 ----
    def _load(self, path: Path, data_only: bool = False):
        """加载 Excel 工作簿，带缓存。

        Args:
            path: Excel 文件路径。
            data_only: True 时只读取公式的计算结果，不读取公式文本。
                用于 _read_cell_value 的公式回退逻辑。

        Returns:
            openpyxl.Workbook 对象（缓存命中直接返回，避免重复 IO）。
        """
        import openpyxl

        key = ("wb", str(path), data_only)
        if key not in self._cache:
            self._cache[key] = openpyxl.load_workbook(path, data_only=data_only)
        return self._cache[key]

    def _invalidate(self, path: Path, sheet: Optional[str] = None):
        """清除指定路径的缓存变体。

        写操作后调用，确保下次 _load 能读到最新文件内容。

        wb 缓存按文件全清(整文件已重写,wb 对象整体过期)。row 缓存:
          sheet 给定 → 仅清该 (path, sheet),其余 sheet 数据未变保留命中;
          sheet 缺省 → 清该文件全部 sheet(保守,公式重算可能触及他表)。
        """
        for key in list(self._cache.keys()):
            if key[1] == str(path):
                del self._cache[key]
        # 5.6：清除该文件行缓存（sheet 给定时收敛到受影响 sheet）
        for key in list(self._row_cache.keys()):
            if key[0] != str(path):
                continue
            if sheet is None or key[1] == sheet:
                del self._row_cache[key]

    def _xlsx_features(self, path: Path) -> tuple[bool, bool, bool]:
        """读 zip 一次检测 (has_formula, has_comment, has_merge)，按 (path, mtime) 缓存。

        读取失败保守返回 (True, True, True)（全走慢路径，不破坏语义）。
        用于 read_sheet 快读分流 与 写路径快路径 eligibility 判定。
        """
        try:
            mtime = path.stat().st_mtime
        except OSError:
            return (True, True, True)
        key = (str(path), mtime)
        cached = self._xlsx_features_cache.get(key)
        if cached is not None:
            return cached
        try:
            import zipfile
            has_f = has_c = has_m = False
            with zipfile.ZipFile(path) as zf:
                for n in zf.namelist():
                    if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                        data = zf.read(n)
                        if not has_f and (b"<f>" in data or b"<f " in data or b"<f/>" in data):
                            has_f = True
                        if not has_m and b"<mergeCells" in data:
                            has_m = True
                    elif not has_c and n.startswith("xl/") and n.endswith(".xml") and "comments" in n:
                        has_c = True
            result = (has_f, has_c, has_m)
            self._xlsx_features_cache[key] = result
            return result
        except Exception:
            return (True, True, True)

    def _sheet_has_formula(self, path: Path) -> bool:
        """快速检测文件是否含公式（读 zip 的 worksheet xml 是否含 <f> 标签）。

        结果按 (path, mtime) 缓存。读取失败保守返回 True（走慢路径，不破坏公式语义）。
        """
        return self._xlsx_features(path)[0]

    @staticmethod
    def _normalize_calamine(v):
        """calamine 读出的值归一对齐 openpyxl 语义。

        - 整数统一为 float（如 100001.0）→ 归一回 int（100001）
        - 空单元格 calamine 返回 ""（openpyxl 返回 None）→ 归一回 None
        - 非整数值（如 1.49）保留 float；bool/str 原样返回
        """
        if isinstance(v, float) and v.is_integer():
            return int(v)
        if v == "":
            return None
        return v

    def _read_sheet_calamine(self, path: Path, sheet: str) -> list[list]:
        """calamine 快读整表：绕过 openpyxl load_workbook（10w 行约 22s），
        从 1.4s 级降到亚秒级。读后按 _normalize_calamine 归一，语义对齐 openpyxl。

        dsr 从索引取（1-based），切片 rows[dsr-1:] 后过滤全 None 空行。
        calamine 读取失败时抛异常，由调用方回退 openpyxl 路径。
        """
        from python_calamine import CalamineWorkbook
        cw = CalamineWorkbook.from_path(str(path))
        raw_rows = cw.get_sheet_by_name(sheet).to_python()
        dsr = self._resolve_data_start(path, sheet)
        result: list[list] = []
        for row in raw_rows[dsr - 1:]:
            nrow = [self._normalize_calamine(v) for v in row]
            if any(v is not None for v in nrow):
                result.append(nrow)
        return result

    @staticmethod
    def _calamine_rows(path, sheet) -> list[list]:
        """用 calamine to_python() 读 sheet 全部行。

        不用 iter_rows()：python-calamine 0.8.2 对空 sheet（height=0）调 iter_rows()
        触发 Rust `Option::unwrap()` panic（见 pet_evolve.xlsx「灵兽进化表说明」），
        而 to_python() 对空 sheet 安全返回 []。空 sheet / 异常统一返回 []，
        调用方据此走 openpyxl 回退或空表头。
        """
        from python_calamine import CalamineWorkbook
        cw = CalamineWorkbook.from_path(str(path))
        sh = cw.get_sheet_by_name(sheet)
        try:
            if sh.height == 0:
                return []
            return sh.to_python()
        except BaseException:
            return []

    def _detect_formula(self, path: Path, sheet: str, row: int, col: int) -> str | None:
        """检测单元格是否为公式。

        以 data_only=False 模式打开工作簿，直接读取单元格原始值。
        若原始值为以 ``=`` 开头的字符串，则判定为公式。

        Args:
            path: Excel 文件路径。
            sheet: Sheet 名称。
            row: 行号（1-based）。
            col: 列号（1-based）。

        Returns:
            公式文本（如 ``=VLOOKUP(A1,B:C,2,FALSE)``），非公式单元格返回 None。
            读取异常时静默返回 None。
        """
        try:
            wb = self._load(path, data_only=False)
            ws = wb[sheet]
            val = ws.cell(row, col).value
            if isinstance(val, str) and val.startswith("="):
                return val
        except Exception:
            pass
        return None

    def _read_cell_value(self, path: Path, sheet: str, row: int, col: int):
        """读取单元格值，自动处理公式单元格。

        逻辑：
        1. 以 data_only=False 读取原始值。
        2. 若原始值是公式（以 ``=`` 开头），切换到 data_only=True 模式重新加载，
           读取缓存的公式计算结果。openpyxl 的 data_only 模式依赖 Excel 上次保存
           时写入的缓存值——若文件尚未被 Excel 打开保存过，计算结果可能为 None，
           此时回退返回公式文本。
        3. 非公式单元格直接返回原始值。

        Returns:
            单元格值（公式单元格优先返回计算结果，回退返回公式文本）。
        """
        wb = self._load(path, data_only=False)
        ws = wb[sheet]
        val = ws.cell(row, col).value
        if isinstance(val, str) and val.startswith("="):
            # 尝试 data_only 模式读取缓存的公式计算结果
            try:
                wb_data = self._load(path, data_only=True)
                computed = wb_data[sheet].cell(row, col).value
                if computed is not None:
                    return computed
            except Exception:
                pass
        return val

    @staticmethod
    def _last_data_row(ws, data_start_row: int) -> int:
        """找到实际最后一个有数据的行号。

        从 data_start_row 遍历到 ws.max_row，遇到任一列有值即更新 last。
        跳过仅含格式残留（无实际值）的空行。

        Args:
            ws: openpyxl Worksheet 对象。
            data_start_row: 数据起始行号。

        Returns:
            最后一个有数据的行号；若 data_start_row 起完全无数据，返回 data_start_row - 1。
        """
        last = data_start_row - 1
        # 缓存 max_row/max_column：openpyxl 的 max_row/max_column 是动态 property，
        # 每次求值都 O(n_cells) 遍历 _cells 字典；循环内反复求值会退化为 O(N^2)，
        # 10w 行大表在此直接卡死（实测 >60s）。提到循环外一次求值即可。
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(data_start_row, max_row + 1):
            if any(ws.cell(r, c).value is not None for c in range(1, max_col + 1)):
                last = r
        return last

    # ---- 接口实现 ----
    def list_tables(self) -> list[Path]:
        return sorted(p for p in self.workspace.rglob("*.xlsx") if not p.name.startswith("~$"))

    def get_sheets(self, path: Path) -> list[str]:
        wb = self._load(path)
        return [ws.title for ws in wb.worksheets]

    def _header_of(self, ws) -> list:
        # 表头空列回退用类型行(header_row+1)前缀翻译成中文列名
        # （如 ability 表第1列表头空、类型行 ability_id:int → 神通id）。
        # resolve_header_cell 仅接受 name:type 形式，无类型行时数据行不会被误用。
        from ..locator.column_name_resolver import build_headers
        type_row = self.header_row + 1 if (ws.max_row or 0) >= self.header_row + 1 else None
        return build_headers(ws, self.header_row, type_row)

    def read_header(self, path: Path, sheet: str) -> list:
        """读取指定 sheet 的表头行。

        Args:
            path: Excel 文件路径。
            sheet: Sheet 名称。

        Returns:
            表头行各列的字符串值列表（由 _header_of 提取）。
        """
        # 无公式表：calamine 读真实表头（第1行 + 第2行类型行回退翻译），
        # 摆脱索引缓存过期脏读（watchdog 未安装/外部改表时索引可能滞后），
        # 同时免 openpyxl 大表 load（~22s）。表头是元数据，读前2行毫秒级。
        if not self._sheet_has_formula(path):
            try:
                from ..locator.column_name_resolver import resolve_header_cell
                rows = self._calamine_rows(path, sheet)
                first = rows[0] if rows else None
                second = rows[1] if len(rows) > 1 else None
                if first is None:
                    return []
                header_vals = [self._normalize_calamine(v) for v in first]
                type_vals = ([self._normalize_calamine(v) for v in second]
                             if second is not None else [])
                result = []
                n = max(len(header_vals), len(type_vals))
                for c in range(n):
                    hv = header_vals[c] if c < len(header_vals) else None
                    tv = type_vals[c] if c < len(type_vals) else None
                    resolved = resolve_header_cell(hv, tv)
                    result.append(str(resolved) if resolved is not None else "")
                # 去尾部全空列（与 read_browse 一致）
                while result and result[-1] == "":
                    result.pop()
                return result
            except BaseException:
                # calamine 内部 Rust panic（pyo3 PanicException 继承 BaseException，
                # `except Exception` 接不住会杀死线程）或解析异常：一律回退 openpyxl。
                pass
        # 含公式表或 calamine 失败：openpyxl 实读（不读索引缓存，避免过期脏读）
        ws = self._load(path)[sheet]
        return self._header_of(ws)

    def read_type_row(self, path: Path, sheet: str) -> list:
        """读取表头行下一行（类型行/规范名行），供 matcher 作 canonical 别名兜底。

        row1=中文展示名（read_header 返回），row2=规范名（如 option_function.function_type:int）。
        splitter 产出的点分规范名（option_function.data.1.conv_id）需对 row2 整名匹配，
        否则末段"conv_id"对中文表头"1:新对话ID"匹配失败（原则9）。
        返回行与 read_header 列对齐；无类型行返回空列表。
        """
        # 无公式表：calamine 读前 2 行取类型行，绕过 openpyxl load（大表 ~22s → ~1.4s）
        if not self._sheet_has_formula(path):
            try:
                rows = self._calamine_rows(path, sheet)
                first = rows[0] if rows else None
                second = rows[1] if len(rows) > 1 else None
                if first is None:
                    return []
                type_row = second if second is not None else []
                return [self._normalize_calamine(v) for v in type_row]
            except BaseException:
                # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退 openpyxl
                pass
        ws = self._load(path)[sheet]
        type_row_idx = self.header_row + 1 if (ws.max_row or 0) >= self.header_row + 1 else None
        if type_row_idx is None:
            return []
        max_col = ws.max_column or 0
        return [ws.cell(type_row_idx, c).value for c in range(1, max_col + 1)]

    def read_sheet(self, path: Path, sheet: str,
                   offset: int = 0, limit: int | None = None) -> list[list]:
        """读取指定 sheet 的全部数据行（不含表头）。

        使用动态 data_start_row（优先从索引读取），过滤全为 None 的空行。

        5.6：全量结果按 (path, sheet) 缓存，写操作经 `_invalidate` 清除，避免
        同一请求内重复逐格读取。`offset`/`limit` 在缓存结果上做分页切片（默认全量，
        不影响既有调用方）。
        """
        key = (str(path), sheet)
        rows = self._row_cache.get(key)
        if rows is None:
            if not self._sheet_has_formula(path):
                # 无公式表：calamine 快读，绕过 openpyxl load_workbook
                # （10w 行 load 约 22s，calamine 全表 ~1.4s）。失败回退 openpyxl。
                try:
                    rows = self._read_sheet_calamine(path, sheet)
                except BaseException:
                    # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退 openpyxl
                    rows = None
            if rows is None:
                # 含公式表 或 calamine 失败：openpyxl 路径（公式格需 data_only=True 缓存值回退）
                ws = self._load(path)[sheet]
                dsr = self._resolve_data_start(path, sheet)
                last_row = self._last_data_row(ws, dsr)
                # 缓存 max_column：动态 property 每次求值 O(n_cells)，循环内反复求值退化 O(N^2)
                max_col = ws.max_column
                rows = []
                for r in range(dsr, last_row + 1):
                    row = [self._read_cell_value(path, sheet, r, c) for c in range(1, max_col + 1)]
                    if any(v is not None for v in row):
                        rows.append(row)
            self._row_cache[key] = rows
        if offset or limit is not None:
            end = None if limit is None else offset + limit
            return rows[offset:end]
        return rows

    def _detect_header_row(self, ws, max_scan: int = 30) -> int:
        """探测表头所在行号：从第1行起向下找第一个至少2个非空单元格的行。

        某些 sheet 第1行为标题/说明，真正表头在后续行（如"道具表说明"表头在第12行）。
        扫描范围默认30行，覆盖常见说明区。浏览专用，不影响 CRUD。
        """
        upper = min(max_scan, ws.max_row)
        max_col = ws.max_column
        for r in range(1, upper + 1):
            non_empty = sum(
                1 for c in range(1, max_col + 1)
                if ws.cell(r, c).value is not None
            )
            if non_empty >= 2:
                return r
        return self.header_row

    def _browse_data_start(self, ws, path: Path, sheet: str) -> int:
        """计算浏览页的数据起始行。逻辑必须与 read_browse 一致，
        保证搜索命中的行号换算后能对齐前端 browse 表格的 "#" 列。

        - 探测表头行 → data_start = header_row + 1
        - 索引记录的 data_start_row 更大（存在类型行/示例行）→ 取索引值
        - 索引错位导致无数据 → 回退到 header_row + 1
        - 表头全空（说明类 sheet，合成表头）→ data_start = 1
        """
        header_row = self._detect_header_row(ws)
        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]
        while headers and headers[-1] is None:
            headers.pop()
        if len(headers) == 0:
            return 1
        data_start = header_row + 1
        idx_key = f"{path.stem}|{sheet}"
        if idx_key in self._index and self._index[idx_key] > header_row + 1:
            data_start = self._index[idx_key]
        last_row = self._last_data_row(ws, data_start)
        if last_row < data_start:
            data_start = header_row + 1
        return data_start

    def read_browse(self, path: Path, sheet: str,
                    page: int = 1, page_size: int = 100) -> tuple[list, list[list], int]:
        """表格浏览专用读取：自动探测表头行，返回 (headers, page_rows, total)。

        与 read_sheet 不同，不依赖索引的 data_start_row，从探测到的表头行下一行起
        读全部非空行，避免索引 data_start 错位导致浏览页空数据。
        """
        # 无公式表：calamine 快读，绕过 openpyxl load + 逐格读取（大表 ~22s+ → ~1.4s）
        if not self._sheet_has_formula(path):
            try:
                from python_calamine import CalamineWorkbook
                cw = CalamineWorkbook.from_path(str(path))
                sh = cw.get_sheet_by_name(sheet)
                raw = sh.to_python()
                # 表头探测：与 _detect_header_row 同规则（首个 ≥2 非空单元格行）
                header_row = 1
                for i, row in enumerate(raw[:30], start=1):
                    if sum(1 for v in row if v is not None) >= 2:
                        header_row = i
                        break
                headers = [self._normalize_calamine(v) for v in raw[header_row - 1]]
                while headers and headers[-1] is None:
                    headers.pop()
                if not headers:
                    return [], [], 0
                data_start = header_row + 1
                idx_key = f"{path.stem}|{sheet}"
                if idx_key in self._index and self._index[idx_key] > header_row + 1:
                    data_start = self._index[idx_key]
                # 过滤全 None 空行
                all_rows = []
                for i in range(data_start - 1, len(raw)):
                    nrow = [self._normalize_calamine(v) for v in raw[i]]
                    if any(v is not None for v in nrow):
                        all_rows.append(nrow)
                total = len(all_rows)
                start = (page - 1) * page_size
                end = start + page_size
                return headers, all_rows[start:end], total
            except BaseException:
                # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退 openpyxl
                pass
        ws = self._load(path)[sheet]
        header_row = self._detect_header_row(ws)
        headers = [ws.cell(header_row, c).value
                   for c in range(1, ws.max_column + 1)]
        # 去除尾部全空列，避免渲染大量空列
        while headers and headers[-1] is None:
            headers.pop()

        data_start = header_row + 1
        # 索引中明确记录了 data_start_row 且大于表头下一行（说明存在类型行/示例行区间）
        # 时优先采用，避免把类型行（如 "ability_id:int"）当作数据展示。
        # 仅当索引条目存在时才用，未知 sheet 仍走表头下一行。
        idx_key = f"{path.stem}|{sheet}"
        if idx_key in self._index and self._index[idx_key] > header_row + 1:
            data_start = self._index[idx_key]
        last_row = self._last_data_row(ws, data_start)
        # 索引 data_start 错位（文件变更后数据前移）导致空 → 回退到表头下一行
        if last_row < data_start:
            data_start = header_row + 1
            last_row = self._last_data_row(ws, data_start)
        col_count = len(headers)
        synthetic_header = False

        # 表头全空（说明/标题类 sheet 无标准表头行）→ 原始 dump：用列字母作表头，
        # 从第1行起读全部非空行，确保内容可见，避免前端只渲染出行号列 "#"。
        if col_count == 0:
            from openpyxl.utils import get_column_letter
            col_count = ws.max_column or 1
            headers = [get_column_letter(c) for c in range(1, col_count + 1)]
            data_start = 1
            last_row = self._last_data_row(ws, data_start)
            synthetic_header = True

        # 合成表头且实际无任何数据行 → 返回空，让前端显示"无可显示数据"而非裸"#"
        if synthetic_header and last_row < data_start:
            return [], [], 0

        all_rows: list[list] = []
        for r in range(data_start, last_row + 1):
            row = [self._read_cell_value(path, sheet, r, c) for c in range(1, col_count + 1)]
            if any(v is not None for v in row):
                all_rows.append(row)

        total = len(all_rows)
        start = (page - 1) * page_size
        end = start + page_size
        page_rows = all_rows[start:end]
        return headers, page_rows, total

    def _can_fast_write(self, path: Path) -> bool:
        """纯数据大表（无公式/批注/合并单元格 且 >512KB）可走 zip+XML 直改快路径。"""
        try:
            if path.stat().st_size < 512 * 1024:
                return False
        except OSError:
            return False
        has_f, has_c, has_m = self._xlsx_features(path)
        return not (has_f or has_c or has_m)

    def _fast_write_cell(self, path: Path, sheet: str, row: int, col: int,
                         value: Any) -> dict | None:
        """纯数据大表 zip+XML 直改单单元格（绕过 openpyxl load+save，大表各 ~10s+）。

        复用 engine.fast_apply 的 XML 辅助函数。仅支持 None/bool/int/float/str 基本值，
        datetime/date 等需 number_format 的类型回退 openpyxl（返回 None）。
        成功返回 cache_info dict（与 _save_with_cache_check 同构），失败返回 None。
        """
        from datetime import datetime, date
        if isinstance(value, (datetime, date)):
            return None
        try:
            import engine.fast_apply as _fa
        except Exception:
            return None
        import zipfile
        import tempfile
        import os
        _xet = _fa._xet
        path = Path(path)
        try:
            with zipfile.ZipFile(path) as zf:
                sheet_map = _fa._sheet_map(zf)
                xml_path = sheet_map.get(sheet)
                if not xml_path:
                    return None
                shared = _fa._load_shared_strings(zf)
                xml_bytes = zf.read(xml_path)
            root = _xet.fromstring(xml_bytes)
            sheet_data = root.find(f"{{{_fa._NS}}}sheetData")
            if sheet_data is None:
                return None
            rows = list(sheet_data)
            if row < 1 or row > len(rows):
                return None
            target_row = rows[row - 1]
            col0 = col - 1
            cell = _fa._find_cell(target_row, col0)
            if cell is not None:
                _fa._replace_cell(target_row, cell, col0, value)
            else:
                _fa._insert_cell_sorted(target_row, col0, value, style=None)
            new_xml = _xet.tostring(root, encoding="utf-8", xml_declaration=True)
            fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
            os.close(fd)
            try:
                with zipfile.ZipFile(path) as zf, \
                        zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
                    for item in zf.infolist():
                        data = new_xml if item.filename == xml_path else zf.read(item.filename)
                        out.writestr(item, data)
                os.replace(tmp, str(path))
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            return {
                "needs_manual_fix": False,
                "cache_message": "",
                "comment_replay": {"replayed": False, "still_lost": 0},
                "hold_events": [],
            }
        except Exception:
            return None

    def _fast_append_row(self, path: Path, sheet: str,
                         values: dict[int, Any]) -> dict | None:
        """纯数据大表 zip+XML 尾部追加整行，绕过 openpyxl load+save。

        values 为 {col_idx(1-based): value}。样式继承自最后一数据行同列（与慢路径
        copy_row_style 对齐：无样式格则不设 s 属性）。失败返回 None 回退 openpyxl。
        """
        try:
            import engine.fast_apply as _fa
        except Exception:
            return None
        import zipfile
        import tempfile
        import os
        _xet = _fa._xet
        path = Path(path)
        try:
            with zipfile.ZipFile(path) as zf:
                sheet_map = _fa._sheet_map(zf)
                xml_path = sheet_map.get(sheet)
                if not xml_path:
                    return None
                xml_bytes = zf.read(xml_path)
            root = _xet.fromstring(xml_bytes)
            sheet_data = root.find(f"{{{_fa._NS}}}sheetData")
            if sheet_data is None:
                return None
            rows = list(sheet_data)
            new_row_num = len(rows) + 1
            ref_row = rows[-1] if rows else None
            new_row = _xet.Element(f"{{{_fa._NS}}}row", {"r": str(new_row_num)})
            for col_idx in sorted(values):
                col0 = col_idx - 1
                style = _fa._cell_style(ref_row, col0) if ref_row is not None else None
                new_row.append(_fa._make_cell(_fa._col_letter(col0) + str(new_row_num),
                                              values[col_idx], style=style))
            sheet_data.append(new_row)
            # dimension 只扩展末行（列数不变，取原 dimension 列字母）
            import re
            dim = root.find(f"{{{_fa._NS}}}dimension")
            if dim is not None:
                m = re.match(r"^[^:]*:?([A-Z]+)", dim.get("ref") or "")
                col = m.group(1) if m else "A"
                dim.set("ref", f"A1:{col}{new_row_num}")
            new_xml = _xet.tostring(root, encoding="utf-8", xml_declaration=True)
            fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
            os.close(fd)
            try:
                with zipfile.ZipFile(path) as zf, \
                        zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
                    for item in zf.infolist():
                        data = new_xml if item.filename == xml_path else zf.read(item.filename)
                        out.writestr(item, data)
                os.replace(tmp, str(path))
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            return {
                "needs_manual_fix": False,
                "cache_message": "",
                "comment_replay": {"replayed": False, "still_lost": 0},
                "hold_events": [],
                "data": {"row": new_row_num, "values": values},
            }
        except Exception:
            return None

    def _fast_delete_row(self, path: Path, sheet: str, row: int) -> dict | None:
        """纯数据大表 zip+XML 删除整行并重排行号，绕过 openpyxl load+save。

        纯数据表无公式/批注/合并单元格，无需 shift_workbook_formulas（公式引用位移
        只对含公式表有意义）。行号经 _renumber_rows 重排（1-based 连续）。
        失败返回 None 回退 openpyxl 路径。
        """
        try:
            import engine.fast_apply as _fa
        except Exception:
            return None
        import zipfile
        import tempfile
        import os
        _xet = _fa._xet
        path = Path(path)
        try:
            with zipfile.ZipFile(path) as zf:
                sheet_map = _fa._sheet_map(zf)
                xml_path = sheet_map.get(sheet)
                if not xml_path:
                    return None
                xml_bytes = zf.read(xml_path)
            root = _xet.fromstring(xml_bytes)
            sheet_data = root.find(f"{{{_fa._NS}}}sheetData")
            if sheet_data is None:
                return None
            rows = list(sheet_data)
            idx = row - 1
            if idx < 0 or idx >= len(rows):
                return None
            del rows[idx]
            _fa._renumber_rows(rows, root)
            sheet_data[:] = rows
            new_xml = _xet.tostring(root, encoding="utf-8", xml_declaration=True)
            fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
            os.close(fd)
            try:
                with zipfile.ZipFile(path) as zf, \
                        zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
                    for item in zf.infolist():
                        data = new_xml if item.filename == xml_path else zf.read(item.filename)
                        out.writestr(item, data)
                os.replace(tmp, str(path))
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            return {
                "needs_manual_fix": False,
                "cache_message": "",
                "comment_replay": {"replayed": False, "still_lost": 0},
                "hold_events": [],
            }
        except Exception:
            return None

    def _fast_insert_row(self, path: Path, sheet: str, row: int,
                         values: dict[int, Any] | None = None) -> dict | None:
        """纯数据大表 zip+XML 在指定行上方插入整行，绕过 openpyxl load+save。

        新行样式继承自上一行同列（与慢路径 copy_row_style 对齐，无样式格不设 s）。
        行号经 _renumber_rows 重排。失败返回 None 回退 openpyxl 路径。
        """
        try:
            import engine.fast_apply as _fa
        except Exception:
            return None
        import zipfile
        import tempfile
        import os
        _xet = _fa._xet
        path = Path(path)
        try:
            with zipfile.ZipFile(path) as zf:
                sheet_map = _fa._sheet_map(zf)
                xml_path = sheet_map.get(sheet)
                if not xml_path:
                    return None
                xml_bytes = zf.read(xml_path)
            root = _xet.fromstring(xml_bytes)
            sheet_data = root.find(f"{{{_fa._NS}}}sheetData")
            if sheet_data is None:
                return None
            rows = list(sheet_data)
            idx = row - 1
            if idx < 0 or idx > len(rows):
                return None
            ref_row = rows[idx - 1] if idx > 0 else None
            new_row = _xet.Element(f"{{{_fa._NS}}}row", {"r": str(row)})
            if values:
                for col_idx in sorted(values):
                    col0 = col_idx - 1
                    style = _fa._cell_style(ref_row, col0) if ref_row is not None else None
                    new_row.append(_fa._make_cell(_fa._col_letter(col0) + str(row),
                                                  values[col_idx], style=style))
            rows.insert(idx, new_row)
            _fa._renumber_rows(rows, root)
            sheet_data[:] = rows
            new_xml = _xet.tostring(root, encoding="utf-8", xml_declaration=True)
            fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(path.parent))
            os.close(fd)
            try:
                with zipfile.ZipFile(path) as zf, \
                        zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
                    for item in zf.infolist():
                        data = new_xml if item.filename == xml_path else zf.read(item.filename)
                        out.writestr(item, data)
                os.replace(tmp, str(path))
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            return {
                "needs_manual_fix": False,
                "cache_message": "",
                "comment_replay": {"replayed": False, "still_lost": 0},
                "hold_events": [],
            }
        except Exception:
            return None

    def write_cell(self, path: Path, sheet: str, row: int, col: int, value: Any,
                   number_format: str | None = None) -> CLICallResult:
        """写入单个单元格并保留样式。

        写值会清掉 openpyxl 单元格样式，故写值前快照原样式，写值后写回；
        若原格无样式 → 从同列上一数据行继承 alignment + number_format，
        避免对齐丢失、日期列写值变序列号。

        Args:
            path: Excel 文件路径。
            sheet: Sheet 名称。
            row: 行号（1-based）。
            col: 列号（1-based）。
            value: 要写入的值。
            number_format: 可选，显式指定数字格式（如 'yyyy-mm-dd hh:mm:ss'），
                覆盖继承值。agent 写 date/datetime 列时从 value_constraints 传入。

        Returns:
            CLICallResult，ok=True 时 data 包含 {row, col, value}。
        """
        from copy import copy
        from ..core.style_utils import inherit_column_style, get_column_number_format_majority
        try:
            # 纯数据大表 zip+XML 快路径（绕过 openpyxl load+save）；number_format 需走慢路径处理格式
            if number_format is None and self._can_fast_write(path):
                fast_info = self._fast_write_cell(path, sheet, row, col, _serialize_cell_value(value))
                if fast_info is not None:
                    self._invalidate(path, sheet)
                    return CLICallResult(
                        ok=True, data={"row": row, "col": col, "value": value},
                        needs_manual_fix=fast_info["needs_manual_fix"],
                        cache_message=fast_info["cache_message"],
                        comment_replay=fast_info.get("comment_replay", {}),
                        hold_events=fast_info.get("hold_events", []),
                    )
            wb = self._load(path)
            ws = wb[sheet]
            src_cell = ws.cell(row, col)
            # 写值前快照原样式（has_style=False 时 snap=None）
            if src_cell.has_style:
                snap = (copy(src_cell.font), copy(src_cell.fill), copy(src_cell.border),
                        copy(src_cell.alignment), src_cell.number_format,
                        copy(src_cell.protection))
            else:
                snap = None
            ws.cell(row, col, _serialize_cell_value(value))
            dsr = self._resolve_data_start(path, sheet)
            if snap is not None:
                tgt = ws.cell(row, col)
                tgt.font, tgt.fill, tgt.border, tgt.alignment, \
                    tgt.number_format, tgt.protection = snap
            else:
                inherit_column_style(ws, row, col, dsr)
            # number_format 优先级：相近行众数 > 显式参数 > 继承值
            # date/datetime 列写值时查相近行实际格式自动对齐，yaml format 作兜底
            if number_format:
                majority = get_column_number_format_majority(ws, col, row, dsr)
                ws.cell(row, col).number_format = majority or number_format
            cache_info = self._save_with_cache_check(wb, path, sheet)
            return CLICallResult(
                ok=True, data={"row": row, "col": col, "value": value},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                comment_replay=cache_info.get("comment_replay", {}),
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def read_cell(self, path: Path, sheet: str, row: int, col: int) -> CLICallResult:
        """读取单个单元格值。

        内部调用 _read_cell_value，自动处理公式回退。

        Returns:
            CLICallResult，ok=True 时 data 为单元格值（公式单元格返回计算结果）。
        """
        try:
            # 无公式表：calamine 单行读，免 openpyxl load（大表 load ~22s）
            if not self._sheet_has_formula(path):
                try:
                    from python_calamine import CalamineWorkbook
                    cw = CalamineWorkbook.from_path(str(path))
                    raw = cw.get_sheet_by_name(sheet).to_python()
                    if 1 <= row <= len(raw) and 1 <= col <= len(raw[row - 1]):
                        return CLICallResult(ok=True,
                                             data=self._normalize_calamine(raw[row - 1][col - 1]))
                except BaseException:
                    # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退
                    pass
            return CLICallResult(ok=True, data=self._read_cell_value(path, sheet, row, col))
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def append_row(self, path: Path, sheet: str, values: dict[int, Any]) -> CLICallResult:
        """在数据末尾追加新行并复制基准行样式。

        新行各列写值后，从基准最后一数据行整行复制 font/fill/border/alignment/
        number_format/protection 到对应列，避免新行样式断裂（如日期列变序列号）。
        基准行无样式时跳过，新行保持默认。
        """
        from ..core.style_utils import copy_row_style
        try:
            # 纯数据大表 zip+XML 快路径：尾部追加整行，绕过 openpyxl load+save（各 ~10s+）
            if self._can_fast_write(path):
                fast_info = self._fast_append_row(path, sheet, values)
                if fast_info is not None:
                    self._invalidate(path, sheet)
                    return CLICallResult(
                        ok=True, data=fast_info["data"],
                        needs_manual_fix=fast_info["needs_manual_fix"],
                        cache_message=fast_info["cache_message"],
                        comment_replay=fast_info.get("comment_replay", {}),
                        hold_events=fast_info.get("hold_events", []),
                    )
            wb = self._load(path)
            ws = wb[sheet]
            dsr = self._resolve_data_start(path, sheet)
            last_row = self._last_data_row(ws, dsr)
            new_row = max(last_row + 1, dsr)
            max_col = ws.max_column
            for col_idx, val in values.items():
                ws.cell(new_row, col_idx, _serialize_cell_value(val))
            # 从基准最后一数据行整行复制样式到新行
            if last_row >= dsr:
                copy_row_style(ws, last_row, new_row, max_col)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            return CLICallResult(
                ok=True, data={"row": new_row, "values": values},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                comment_replay=cache_info.get("comment_replay", {}),
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def _get_table_index(self):
        """T1: 懒加载并缓存 _table_index.json（含 search_blob），供搜索预检用。

        mtime 校验：索引文件被 refresh_if_changed 重写后自动重载，避免写后脏读。
        """
        try:
            idx_path = Path(__file__).resolve().parent.parent / "_table_index.json"
            mtime = idx_path.stat().st_mtime if idx_path.exists() else 0.0
        except Exception:
            mtime = 0.0
        if self._table_idx_cache is not None and self._table_idx_mtime == mtime and mtime > 0:
            return self._table_idx_cache
        try:
            from ..locator.table_index import load_index
            self._table_idx_cache = load_index()
        except Exception:
            self._table_idx_cache = []
        self._table_idx_mtime = mtime
        return self._table_idx_cache

    def _table_likely_contains_any(self, path: Path, keywords: list[str]) -> bool:
        """T1 搜索预检：任一 keyword 可能出现在该表任意 sheet 的 search_blob 中 → True。

        全部 keyword 都不可能命中 → False（跳过 workbook 加载，避免 65+ 表全量扫描）。
        索引缺失/旧索引无 search_blob/keyword 太短 → 放行 True，走原扫描流程。
        """
        kws = [k for k in keywords if k and len(k) >= 2]
        if not kws:
            return True
        idx = self._get_table_index()
        if not idx:
            return True
        for t in idx:
            if t.stem != path.stem:
                continue
            for s in t.sheets:
                blob = s.search_blob or ""
                if not blob:
                    continue
                for kw in kws:
                    if kw.lower() in blob:
                        return True
            return False
        return True  # 索引中无该表，放行

    def search_rows(self, path: Path, keyword: str, sheet: str = "",
                    col: str = "") -> list[SearchResult]:
        """自然语言智能搜索，支持中英混合关键词。

        分层策略（从精确到模糊，逐层降级）：
        1. 完整关键词搜索 —— 直接匹配用户输入的整体字符串。
        2. 分词降级 —— 将关键词按中英混合规则切分为 token，按长度降序逐个搜索，首个命中即返回。
        3. N-gram 组合 —— 对 token 列表做 2/3/4-gram 拼接搜索，去重合并结果。

        Args:
            path: Excel 文件路径。
            keyword: 搜索关键词（支持中文、英文、数字混合）。
            sheet: 限定搜索的 sheet 名，空字符串表示搜索所有 sheet。
            col: 限定搜索的列名或列号，空字符串表示搜索所有列。

        Returns:
            SearchResult 列表，匹配同一行多列时每行仅返回首个匹配列。
        """
        # T1: search_blob 预检 — 完整词 + 分词 token 任一可能命中才加载 workbook，
        # 跳过 65+ 表中不可能命中的表，避免全量加载扫描。
        if not sheet and not col:
            pre_cands = [keyword] + [t for t in self._tokenize(keyword) if len(t) >= 2]
            if not self._table_likely_contains_any(path, pre_cands):
                return []

        # 第一层：完整关键词（可能会带自然语言描述词）
        results = self._search_single(path, keyword, sheet, col)
        if results:
            return results

        # 第二层：按 token 降级搜索，哪个 token 有匹配就用哪个
        tokens = self._tokenize(keyword)
        # 按长度降序：长的 token 优先（如 "TEST" 优先于 "神通" 的单个字）
        tokens.sort(key=lambda x: -len(x))
        for t in tokens:
            if len(t) <= 1:
                continue
            results = self._search_single(path, t, sheet, col)
            if results:
                return results

        # 第三层：n-gram 组合（如 "神通TEST" → ["神通","TEST"] → ["神通TEST", "TEST"]都试过了，再试 "神通"+后一个字 等）
        seen: set[tuple[str, int, int]] = set()
        merged: list[SearchResult] = []
        for n in (2, 3, 4):
            if n > len(tokens):
                break
            for i in range(len(tokens) - n + 1):
                gram = "".join(tokens[i:i + n])
                if len(gram) <= 1:
                    continue
                for r in self._search_single(path, gram, sheet, col):
                    key = (r.sheet, r.row, r.col)
                    if key not in seen:
                        seen.add(key)
                        merged.append(r)
            if merged:
                return merged

        return []

    @staticmethod
    def _tokenize(text: str) -> list[str]:
        """中英混合分词。中文逐字切，英文/数字按连续片段整体保留。"""
        import re
        if not text:
            return []
        tokens = []
        for chunk in re.findall(r"[\u4e00-\u9fff]+|[A-Za-z0-9_]+", text):
            if re.fullmatch(r"[A-Za-z0-9_]+", chunk):
                tokens.append(chunk.lower())
            else:
                tokens.extend(chunk)
        return tokens

    def _search_via_calamine(self, path: Path, keyword: str, sheet: str = "",
                             col: str = "") -> Optional[list[SearchResult]]:
        """无公式表 calamine 内存扫描搜索，替代 openpyxl 逐格 O(N²) 全表扫描。

        返回 None 表示不可用（含公式表/读失败），调用方回退原 openpyxl 路径。
        行号对齐 openpyxl（1-based 绝对行号 + data_row 相对行号）。
        """
        if self._sheet_has_formula(path):
            return None
        try:
            from python_calamine import CalamineWorkbook
            cw = CalamineWorkbook.from_path(str(path))
        except BaseException:
            # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退
            return None
        kw_lower = (keyword or "").lower()
        # 索引中取该表各 sheet 的表头（col 名匹配用）
        headers_by_sheet: dict[str, list[str]] = {}
        for t in self._get_table_index():
            if t.stem == path.stem:
                for s in t.sheets:
                    headers_by_sheet[s.name] = s.headers or s.header_names or []
                break
        results: list[SearchResult] = []
        targets = [sheet] if sheet else list(cw.sheet_names)
        for sn in targets:
            if sn not in cw.sheet_names:
                continue
            dsr = self._resolve_data_start(path, sn)
            headers = headers_by_sheet.get(sn, [])
            col_idx = None
            if col:
                if str(col).isdigit():
                    col_idx = int(col)
                else:
                    for ci, h in enumerate(headers, start=1):
                        if h and str(col) in str(h):
                            col_idx = ci
                            break
            raw = cw.get_sheet_by_name(sn).to_python()
            for i in range(max(dsr - 1, 0), len(raw)):
                r = i + 1
                data_row = r - dsr + 1 if r >= dsr else 0
                nrow = [self._normalize_calamine(v) for v in raw[i]]
                if col_idx is not None:
                    if col_idx <= len(nrow):
                        cv = nrow[col_idx - 1]
                        if cv is not None and kw_lower in str(cv).lower():
                            hname = headers[col_idx - 1] if col_idx <= len(headers) else ""
                            results.append(SearchResult(
                                sheet=sn, row=r, col=col_idx, col_name=str(hname or ""),
                                cell_value=cv, row_data=nrow, data_row=data_row))
                else:
                    for c in range(len(nrow)):
                        cv = nrow[c]
                        if cv is not None and kw_lower in str(cv).lower():
                            hname = headers[c] if c < len(headers) else ""
                            results.append(SearchResult(
                                sheet=sn, row=r, col=c + 1, col_name=str(hname or ""),
                                cell_value=cv, row_data=nrow, data_row=data_row))
                            break
        return results

    def _search_single(self, path: Path, keyword: str, sheet: str = "",
                       col: str = "") -> list[SearchResult]:
        """单关键词搜索：逐 sheet、逐行做子串匹配（大小写不敏感）。

        优化：先查 _table_index.json 的 row_index（名称/id 列倒排索引），
        命中直接返回对应行；未命中走 calamine 内存扫（无公式表）；
        含公式表回退 openpyxl 逐格扫描。
        """
        results: list[SearchResult] = []
        # 无公式表：calamine 内存快扫（10w 行 ~1.4s），优先于 row_index/openpyxl。
        # row_index 命中的行仍需逐格读整行（openpyxl 慢），大表命中多行时更慢。
        if not self._sheet_has_formula(path):
            fast = self._search_via_calamine(path, keyword, sheet, col)
            if fast is not None:
                return fast
        # 含公式表：row_index 倒排优先，未命中回退逐格扫描
        if not col and sheet == "":
            idx_hits = self._search_via_index(path, keyword)
            if idx_hits:
                return idx_hits
        wb = self._load(path)
        sheets = [sheet] if sheet else [ws.title for ws in wb.worksheets]
        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            dsr = self._browse_data_start(ws, path, sn)
            last_row = self._last_data_row(ws, dsr)
            header = self._header_of(ws)
            max_col = ws.max_column
            kw_lower = keyword.lower()
            for r in range(dsr, last_row + 1):
                row_data = [self._read_cell_value(path, sn, r, c)
                            for c in range(1, max_col + 1)]
                data_row = r - dsr + 1
                if col:
                    col_idx = None
                    if col.isdigit():
                        col_idx = int(col)
                    else:
                        for ci, h in enumerate(header, start=1):
                            if h and col in str(h):
                                col_idx = ci
                                break
                    if col_idx:
                        cell_val = row_data[col_idx - 1] if col_idx <= len(row_data) else None
                        if cell_val and kw_lower in str(cell_val).lower():
                            hname = header[col_idx - 1] if col_idx <= len(header) else ""
                            results.append(SearchResult(sheet=sn, row=r, col=col_idx,
                                                        col_name=str(hname or ""),
                                                        cell_value=cell_val, row_data=row_data,
                                                        data_row=data_row))
                else:
                    for c in range(1, max_col + 1):
                        cell_val = row_data[c - 1] if c <= len(row_data) else None
                        if cell_val and kw_lower in str(cell_val).lower():
                            hname = header[c - 1] if c <= len(header) else ""
                            results.append(SearchResult(sheet=sn, row=r, col=c,
                                                        col_name=str(hname or ""),
                                                        cell_value=cell_val, row_data=row_data,
                                                        data_row=data_row))
                            break
        return results

    def _search_via_index(self, path: Path, keyword: str) -> list[SearchResult]:
        """用 _table_index.json 的 row_index 倒排索引快速定位。

        row_index 结构：{col_name: {value_str: [row_nums]}}，只对名称/id 列建索引。
        keyword 子串匹配 col_name 下的 value，命中即返回该行。
        未加载索引或无命中返回空 list。
        """
        data = self._get_table_index()
        if not data:
            return []
        kw_lower = keyword.lower()
        results: list[SearchResult] = []
        wb = None
        for t in data:
            if t.stem != path.stem:
                continue
            for s in t.sheets:
                sn = s.name
                row_index = s.row_index or {}
                headers = s.headers or s.header_names or []
                for col_name, val_map in row_index.items():
                    for val, row_nums in val_map.items():
                        if val and kw_lower in str(val).lower():
                            if wb is None:
                                wb = self._load(path)
                            ws = wb[sn]
                            # 浏览页数据起始行，用于把绝对行号换算成 browse 相对行
                            dsr = self._browse_data_start(ws, path, sn)
                            for r in row_nums:
                                # 读该行整行数据（限 headers 长度）
                                max_c = len(headers) or ws.max_column
                                row_data = [self._read_cell_value(path, sn, r, c)
                                            for c in range(1, max_c + 1)]
                                # 找命中列号
                                col_idx = 0
                                for ci, h in enumerate(headers, start=1):
                                    if h and col_name in str(h):
                                        col_idx = ci
                                        break
                                if col_idx == 0:
                                    col_idx = 1
                                data_row = r - dsr + 1 if r >= dsr else 0
                                results.append(SearchResult(
                                    sheet=sn, row=r, col=col_idx,
                                    col_name=str(col_name),
                                    cell_value=val, row_data=row_data,
                                    data_row=data_row))
        return results

    def locate_row_via_index(self, path: Path, sheet: str, col_idx: int,
                             value: str, match_mode: str = "contains"
                             ) -> Optional[tuple[list[int], str, str]]:
        """用 row_index 倒排索引快速定位行（处理模式加速）。

        与 _search_via_index 同源，但面向 _locate_row 的"已知 sheet+col_idx"场景：
        直接按 col_idx 映射到 row_index 的 col_name，无需遍历全部索引列。

        Args:
            path: 表格文件路径
            sheet: sheet 名
            col_idx: 定位列序号（1-based，对齐 header_names）
            value: 定位值
            match_mode: exact/startswith/contains

        Returns:
            (row_nums, col_name, matched_value) 或 None（无索引/列非索引列/无命中）。
            col_name 为索引中实际命中的列名；matched_value 为命中的 key。
        """
        idx = self._get_table_index()
        if not idx:
            return None
        # search_blob 预检：value 子串不在该 sheet 任一可搜索列去重值集合 → 跳过
        # value 可能为 int/float（如 prefab_id=8005），统一 str() 转换避免 .lower() 崩溃
        val_lower = str(value).lower() if value is not None else ""
        blob_hit = False
        sheet_meta = None
        for t in idx:
            if t.stem != path.stem:
                continue
            for s in t.sheets:
                if s.name != sheet:
                    continue
                sheet_meta = s
                blob = s.search_blob or ""
                if val_lower in blob:
                    blob_hit = True
                break
            break
        if sheet_meta is None:
            return None
        if not blob_hit:
            return None

        # col_idx → header_name → row_index 的 col_name
        header_names = sheet_meta.header_names or []
        if col_idx - 1 >= len(header_names):
            return None
        target_hn = header_names[col_idx - 1]
        if not target_hn:
            return None
        row_index = sheet_meta.row_index or {}
        val_map: dict[str, list[int]] = {}
        for cn, vm in row_index.items():
            # row_index 的 col_name 与 header_names 经 _clean_header 对齐
            if cn == target_hn or target_hn in cn or cn in target_hn:
                val_map = vm
                break
        if not val_map:
            return None

        hits: list[tuple[int, str]] = []
        for k, row_nums in val_map.items():
            ks = str(k).strip()
            if match_mode == "exact":
                if ks.lower() == val_lower:
                    hits.extend((r, ks) for r in row_nums)
            elif match_mode == "startswith":
                if ks.lower().startswith(val_lower):
                    hits.extend((r, ks) for r in row_nums)
            else:  # contains
                if val_lower and val_lower in ks.lower():
                    hits.extend((r, ks) for r in row_nums)
        if not hits:
            return None
        # 去重保序（同一行可能在多 key 命中）
        seen: set[int] = set()
        row_nums: list[int] = []
        first_val = hits[0][1]
        for r, _ in hits:
            if r not in seen:
                seen.add(r)
                row_nums.append(r)
        return row_nums, target_hn, first_val

    def locate_row(self, path: Path, sheet: str, col: int, value: str,
                   mode: str = "contains") -> int | None:
        """按列值定位第一个匹配行号。

        无公式表走 calamine 内存扫（10w 行 ~1.4s），替代 openpyxl 逐格读取
        （每格 _read_cell_value 两次字典查找，10w 行秒级~分钟级）。
        含公式表保留原 openpyxl 路径（公式格需 data_only 回退）。
        """
        if not self._sheet_has_formula(path):
            try:
                from python_calamine import CalamineWorkbook
                cw = CalamineWorkbook.from_path(str(path))
                raw = cw.get_sheet_by_name(sheet).to_python()
                dsr = self._resolve_data_start(path, sheet)
                v_lower = str(value).lower()
                for i in range(max(dsr - 1, 0), len(raw)):
                    row = raw[i]
                    if col > len(row):
                        continue
                    cell_val = row[col - 1]
                    if cell_val is None:
                        continue
                    cs = str(cell_val).strip().lower()
                    if mode == "exact":
                        if cs == v_lower:
                            return i + 1
                    elif mode == "startswith":
                        if cs.startswith(v_lower):
                            return i + 1
                    else:
                        if v_lower in cs:
                            return i + 1
                return None
            except BaseException:
                # calamine 内部 Rust panic（PanicException 继承 BaseException）→ 回退 openpyxl
                pass
        ws = self._load(path)[sheet]
        dsr = self._resolve_data_start(path, sheet)
        last_row = self._last_data_row(ws, dsr)
        for r in range(dsr, last_row + 1):
            cell_val = self._read_cell_value(path, sheet, r, col)
            if cell_val is None:
                continue
            cs = str(cell_val).strip()
            if mode == "exact":
                if cs.lower() == value.lower():
                    return r
            elif mode == "startswith":
                if cs.lower().startswith(value.lower()):
                    return r
            else:
                if value.lower() in cs.lower():
                    return r
        return None

    def delete_row(self, path: Path, sheet: str, row: int) -> CLICallResult:
        """删除指定行并保存，操作后刷新缓存。

        Args:
            path: Excel 文件路径。
            sheet: Sheet 名称。
            row: 要删除的行号（1-based）。

        Returns:
            CLICallResult，ok=True 时 data 包含 {row}。
        """
        try:
            # 纯数据大表 zip+XML 快路径（无公式/批注/合并单元格，无需公式引用位移）
            if self._can_fast_write(path):
                fast_info = self._fast_delete_row(path, sheet, row)
                if fast_info is not None:
                    self._invalidate(path, sheet)
                    return CLICallResult(
                        ok=True, data={"row": row},
                        needs_manual_fix=fast_info["needs_manual_fix"],
                        cache_message=fast_info["cache_message"],
                        comment_replay=fast_info.get("comment_replay", {}),
                        hold_events=fast_info.get("hold_events", []),
                    )
            wb = self._load(path)
            ws = wb[sheet]
            ws.delete_rows(row)
            # 删行后重写公式内的行引用（openpyxl 不自动调整引用）
            shift_report = shift_workbook_formulas(wb, sheet, "row", row, -1)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            cache_info = shift_report.merge_into(cache_info)
            return CLICallResult(
                ok=True, data={"row": row},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def insert_row(self, path: Path, sheet: str, row: int,
                   values: dict[int, Any] | None = None) -> CLICallResult:
        """在指定行上方插入新行并复制上一行样式。

        与 delete_row 对称：物理 insert_rows(row) → shift_workbook_formulas(row,+1)
        重写公式引用 → 缓存校验。表头/类型行不动；values 按列索引写入新行数据区。
        插入后从上一行（row-1）整行复制样式到新行，保证对齐/数字格式一致。

        Args:
            row: 在此行上方插入（1-based）。应 >= 数据起始行。
            values: {col_idx: value}，可空（插空行后由 AI 再 write_cell 填值）。

        Returns:
            CLICallResult，ok=True 时 data 包含 {row}。
        """
        from ..core.style_utils import copy_row_style
        try:
            # 纯数据大表 zip+XML 快路径（无公式/批注/合并单元格，无需公式引用位移）
            if self._can_fast_write(path):
                fast_info = self._fast_insert_row(path, sheet, row, values)
                if fast_info is not None:
                    self._invalidate(path, sheet)
                    return CLICallResult(
                        ok=True, data={"row": row, "values": values or {}},
                        needs_manual_fix=fast_info["needs_manual_fix"],
                        cache_message=fast_info["cache_message"],
                        comment_replay=fast_info.get("comment_replay", {}),
                        hold_events=fast_info.get("hold_events", []),
                    )
            wb = self._load(path)
            ws = wb[sheet]
            max_col = ws.max_column
            ws.insert_rows(row)
            # 插行后重写公式内的行引用（openpyxl 不自动调整引用）
            shift_report = shift_workbook_formulas(wb, sheet, "row", row, 1)
            if values:
                for col_idx, val in values.items():
                    ws.cell(row, col_idx, val)
            # 从上一行整行复制样式到新行（row-1 插入后即为新行的上一行）
            if row > 1:
                copy_row_style(ws, row - 1, row, max_col)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            cache_info = shift_report.merge_into(cache_info)
            return CLICallResult(
                ok=True, data={"row": row, "values": values or {}},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def sort_sheet(self, path: Path, sheet: str, key_col: int = 1,
                   ascending: bool = True) -> CLICallResult:
        """按 key_col 对数据区行重排序。

        整行搬移值+样式；数据区内公式按行置换映射重写行引用（行内计算随行
        搬移，如 F3=SUM(B3:E3) 移到 row7 后改写为 SUM(B7:E7)）。区外公式
        （汇总行）不参与置换——其引用物理行号，纯排序不改变物理行号。
        聚合范围是否需扩展（插入新行后）由上层语义检测，本方法不处理。
        """
        from copy import copy
        try:
            wb = self._load(path)
            ws = wb[sheet]
            dsr = self._resolve_data_start(path, sheet)
            last_row = self._last_data_row(ws, dsr)
            if last_row < dsr:
                return CLICallResult(ok=True, data={"sorted": 0, "key_col": key_col,
                                                   "ascending": ascending})
            max_col = ws.max_column
            # 收集数据行：(排序键, 原始行号, [(value, 样式快照), ...])
            # 样式必须收集时深拷贝：写回是原地覆盖，若存单元格引用，
            # 交换行时后写行会读到被前一步污染的源 → 样式错位（如居中变左对齐）
            records: list[tuple[Any, int, list[tuple]]] = []
            for r in range(dsr, last_row + 1):
                key = ws.cell(r, key_col).value
                cells = []
                for c in range(1, max_col + 1):
                    src = ws.cell(r, c)
                    if src.has_style:
                        snap = (copy(src.font), copy(src.fill), copy(src.border),
                                copy(src.alignment), src.number_format,
                                copy(src.protection))
                    else:
                        snap = None
                    cells.append((src.value, snap))
                records.append((key, r, cells))

            def sort_key(item):
                v = item[0]
                if v is None or (isinstance(v, str) and not v.strip()):
                    return (2, 0.0, "")
                if isinstance(v, bool):
                    return (1, 0.0, str(v))
                if isinstance(v, (int, float)):
                    return (0, float(v), "")
                s = str(v).strip()
                try:
                    return (0, float(s), "")
                except ValueError:
                    return (1, 0.0, s)

            records.sort(key=sort_key, reverse=not ascending)

            # 写回：值 + 样式整行搬移，并构建 old→new 行置换映射
            row_map: dict[int, int] = {}
            for i, (_, old_row, cells) in enumerate(records):
                new_row = dsr + i
                row_map[old_row] = new_row
                for c in range(1, max_col + 1):
                    val, snap = cells[c - 1]
                    tgt = ws.cell(new_row, c)
                    tgt.value = val
                    if snap is not None:
                        tgt.font, tgt.fill, tgt.border, tgt.alignment, \
                            tgt.number_format, tgt.protection = snap

            # 重写数据区内公式的行引用（行内计算随行搬移）
            rewritten = 0
            skipped_dynamic = 0
            for r in range(dsr, dsr + len(records)):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        new_v, _, dynamic = permute_formula_rows(
                            v, row_map, sheet, current_sheet=sheet)
                        if dynamic:
                            skipped_dynamic += 1
                        elif new_v != v:
                            cell.value = new_v
                            rewritten += 1

            cache_info = self._save_with_cache_check(wb, path, sheet)
            msg = cache_info["cache_message"]
            if rewritten or skipped_dynamic:
                extra = f"公式行引用重写{rewritten}个"
                if skipped_dynamic:
                    extra += f"；{skipped_dynamic}个含动态函数需人工核对"
                msg = f"{msg} | {extra}" if msg else extra
            return CLICallResult(
                ok=True,
                data={"sorted": len(records), "key_col": key_col,
                      "ascending": ascending, "formula_rewritten": rewritten,
                      "formula_skipped_dynamic": skipped_dynamic},
                needs_manual_fix=cache_info["needs_manual_fix"] or skipped_dynamic > 0,
                cache_message=msg,
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    # ---- L2 公式语义原语实现（委托 formula_semantics 模块）----
    def interpret_formula(self, path: Path, sheet: str, cell: str):
        """解析单格公式语义。纯读，供 AI 理解公式意图。"""
        from ..formula.formula_semantics import interpret_formula as _interp
        dsr = self._resolve_data_start(path, sheet)
        return _interp(path, sheet, cell, data_start_row=dsr)

    def scan_sheet_formulas(self, path: Path, sheet: str):
        """扫描全 sheet 公式语义。纯读，供 AI 拿全貌。"""
        from ..formula.formula_semantics import scan_sheet_formulas as _scan
        dsr = self._resolve_data_start(path, sheet)
        return _scan(path, sheet, data_start_row=dsr)

    def preview_formula_impact(self, path: Path, sheet: str, op: dict):
        """dry-run 模拟增删改，返回机械位移结果 + 语义缺口标注。纯读不落盘。"""
        from ..formula.formula_semantics import preview_formula_impact as _preview
        dsr = self._resolve_data_start(path, sheet)
        return _preview(path, sheet, op, data_start_row=dsr)

    def rewrite_formula(self, path: Path, sheet: str, cell: str,
                        new_formula: str) -> CLICallResult:
        """按 AI 决策写入目标公式 + 缓存校验。判断权在 AI，不做语义校验。"""
        from ..formula.formula_semantics import rewrite_formula as _rewrite
        return _rewrite(self, path, sheet, cell, new_formula)

    # ---- 列级操作实现 ----
    def _resolve_col_index(self, ws, col: int | str) -> int | None:
        """把列名或列索引统一解析为 1-based 列索引。

        col 为数字字符串或 int → 直接取整。
        col 为列名 → 在表头中精确匹配（取冒号前部分），找不到再模糊子串匹配。
        """
        if isinstance(col, int):
            return col
        s = str(col).strip()
        if s.isdigit():
            return int(s)
        header = self._header_of(ws)
        # 精确匹配（冒号前部分，兼容 "技能id:int" 这类带类型标注的表头）
        for ci, h in enumerate(header, start=1):
            if h and str(h).split(":")[0].strip() == s:
                return ci
        # 模糊子串匹配
        for ci, h in enumerate(header, start=1):
            if h and s in str(h):
                return ci
        return None

    def list_columns(self, path: Path, sheet: str) -> list[tuple[int, str]]:
        """列出 sheet 所有列，返回 [(col_idx, col_name), ...]。"""
        try:
            ws = self._load(path)[sheet]
            return [(ci, str(h)) for ci, h in enumerate(self._header_of(ws), start=1)]
        except Exception:
            return []

    def insert_column(self, path: Path, sheet: str, name: str,
                      after: int | None = None, type_str: str | None = None,
                      default: Any = None) -> CLICallResult:
        """新增列并保存。

        after=None 追加到末尾（不移动现有列，最安全）；after=列索引 在其右侧插入空列。
        写入表头行；type_str 非空时写入表头下一行（类型行约定）。
        default 非空时填充数据区所有现有行。
        """
        try:
            wb = self._load(path)
            ws = wb[sheet]
            max_col = ws.max_column or len(self._header_of(ws))
            if after is None:
                new_idx = max_col + 1  # 追加末尾，不调用 insert_cols 避免移动
                shift_report = None
            else:
                new_idx = int(after) + 1
                ws.insert_cols(new_idx)
                # 中间插列后重写公式内的列引用（openpyxl 不自动调整引用）
                shift_report = shift_workbook_formulas(wb, sheet, "col", new_idx, 1)
            # 表头
            ws.cell(self.header_row, new_idx, name)
            # 类型行（表头下一行）
            if type_str:
                ws.cell(self.header_row + 1, new_idx, type_str)
            # 默认值填充数据区
            if default is not None:
                dsr = self._resolve_data_start(path, sheet)
                last_row = self._last_data_row(ws, dsr)
                for r in range(dsr, last_row + 1):
                    ws.cell(r, new_idx, default)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            if shift_report is not None:
                cache_info = shift_report.merge_into(cache_info)
            return CLICallResult(
                ok=True, data={"col": new_idx, "name": name},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def delete_column(self, path: Path, sheet: str, col: int | str) -> CLICallResult:
        """删除列并保存。col 可为列索引或列名。"""
        try:
            wb = self._load(path)
            ws = wb[sheet]
            col_idx = self._resolve_col_index(ws, col)
            if col_idx is None:
                return CLICallResult(ok=False, error=f"未找到列: {col}")
            ws.delete_cols(col_idx)
            # 删列后重写公式内的列引用（openpyxl 不自动调整引用）
            shift_report = shift_workbook_formulas(wb, sheet, "col", col_idx, -1)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            cache_info = shift_report.merge_into(cache_info)
            return CLICallResult(
                ok=True, data={"col": col_idx},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))

    def rename_column(self, path: Path, sheet: str, col: int | str,
                      new_name: str) -> CLICallResult:
        """重命名列表头并保存。col 可为列索引或列名。"""
        try:
            wb = self._load(path)
            ws = wb[sheet]
            col_idx = self._resolve_col_index(ws, col)
            if col_idx is None:
                return CLICallResult(ok=False, error=f"未找到列: {col}")
            old_name = ws.cell(self.header_row, col_idx).value
            ws.cell(self.header_row, col_idx, new_name)
            cache_info = self._save_with_cache_check(wb, path, sheet)
            return CLICallResult(
                ok=True, data={"col": col_idx, "old": old_name, "new": new_name},
                needs_manual_fix=cache_info["needs_manual_fix"],
                cache_message=cache_info["cache_message"],
                hold_events=cache_info.get("hold_events", []),
            )
        except Exception as e:
            return CLICallResult(ok=False, error=str(e), stderr=str(e))
