"""O16 方法 F 编号账本跨分支单测（validate_multi_branch + claim_id + agent 接线）。

覆盖：
- F1 validate_multi_branch(branch_roots) → CrossBranchReport
- F2 多分支冲突判定（同 branch 同 stem 白名单 / 跨 branch 冲突）
- F4 claim_id 查询：冲突 + 建议下一空闲号 + 单分支白名单
- F4 agent _validate_id_scope id-claim 注入：冲突 → 返 False + pre_commit_hold 事件

无 LLM/真实 SVN 分支，用 tmp_path 构造多分支目录结构 + openpyxl 写 ID 列。
id_mgr.xlsx SETTING sheet 需构造（load_id_mgr）。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from engine.id_scope import IdScopeValidator, CrossBranchReport, CrossBranchConflict
import engine.id_scope as _ids_module


@pytest.fixture(autouse=True)
def _reset_id_scope_singleton():
    """autouse：每测前后重置 id_scope 全局单例，避免 O16 测污染其他测（load_id_mgr
    留 _id_mgr_loaded=True 影响依赖 _id_mgr_loaded=False 的断言）。"""
    _saved = _ids_module._validator
    _ids_module._validator = None
    yield
    _ids_module._validator = _saved


def _write_id_table(path: Path, sheet: str, ids: list[int], id_col_name: str = "编号",
                    data_start_row: int = 5):
    """构造含 ID 列的表（header 行 1，数据行从 data_start_row 起，默认 5 匹配 CLI 默认）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.cell(1, 1, id_col_name)
    for i, v in enumerate(ids, start=data_start_row):
        ws.cell(i, 1, v)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_id_mgr(path: Path, segs: list[dict]):
    """构造 id_mgr.xlsx SETTING sheet。segs=[{module,id_min,id_max,used_min,used_max,status}]。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SETTING"
    ws.cell(1, 1, "模块"); ws.cell(1, 2, "编号下限"); ws.cell(1, 3, "编号上限")
    ws.cell(1, 4, "已用下限"); ws.cell(1, 5, "已用上限"); ws.cell(1, 6, "状态")
    for i, s in enumerate(segs, start=2):
        ws.cell(i, 1, s["module"]); ws.cell(i, 2, s["id_min"]); ws.cell(i, 3, s["id_max"])
        ws.cell(i, 4, s.get("used_min")); ws.cell(i, 5, s.get("used_max"))
        ws.cell(i, 6, s.get("status", ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class TestF1ValidateMultiBranch:
    def test_empty_branches_returns_empty_report(self):
        v = IdScopeValidator()
        r = v.validate_multi_branch([])
        assert isinstance(r, CrossBranchReport)
        assert r.cross_branch_conflicts == []
        assert r.scanned_tables == 0

    def test_single_branch_no_conflict(self, tmp_path):
        """单分支：同 stem 跨 sheet 白名单不算冲突。"""
        root = tmp_path / "trunk"
        _write_id_table(root / "pet.xlsx", "Pet", [1, 2, 3])
        v = IdScopeValidator()
        r = v.validate_multi_branch([root])
        assert r.scanned_tables >= 1
        assert r.cross_branch_conflicts == []
        assert r.branches_scanned == ["trunk"]

    def test_multi_branch_conflict(self, tmp_path):
        """两分支同 id → 跨分支冲突。"""
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_id_table(trunk / "pet.xlsx", "Pet", [100, 101])
        _write_id_table(dev / "pet.xlsx", "Pet", [100, 200])  # 100 跨分支冲突
        v = IdScopeValidator()
        r = v.validate_multi_branch([trunk, dev])
        assert len(r.cross_branch_conflicts) == 1, f"scanned={r.scanned_tables} branches={r.branches_scanned}"
        c = r.cross_branch_conflicts[0]
        assert c.conflict_value == 100
        branches = {loc["branch"] for loc in c.locations}
        assert branches == {"trunk", "dev"}
        assert set(r.branches_scanned) == {"trunk", "dev"}

    def test_same_branch_same_stem_whitelist(self, tmp_path):
        """同分支同 stem 跨 sheet 白名单（不算跨分支冲突）。"""
        trunk = tmp_path / "trunk"
        _write_id_table(trunk / "pet.xlsx", "Pet", [50])
        # 同文件不同 sheet 同 id（stem 相同）→ 白名单
        wb = Workbook(); ws = wb.active; ws.title = "Pet"; ws.cell(1, 1, "编号"); ws.cell(5, 1, 50)
        ws2 = wb.create_sheet("Pet2"); ws2.cell(1, 1, "编号"); ws2.cell(5, 1, 50)
        (trunk).mkdir(parents=True, exist_ok=True)
        wb.save(trunk / "pet.xlsx")
        v = IdScopeValidator()
        r = v.validate_multi_branch([trunk])
        assert r.cross_branch_conflicts == [], f"同 stem 应白名单: {r.cross_branch_conflicts}"


class TestF4ClaimId:
    def test_claim_no_conflict(self, tmp_path):
        """单分支 id=999 不存在 → claimed=False。"""
        root = tmp_path / "trunk"
        _write_id_table(root / "pet.xlsx", "Pet", [1, 2, 3])
        v = IdScopeValidator()
        r = v.claim_id(999, root)
        assert r["claimed"] is False
        assert r["conflict_locations"] == []
        assert r["suggested_next"] is None

    def test_claim_multi_branch_conflict_with_suggestion(self, tmp_path):
        """多分支同 id 冲突 → claimed=True + suggested_next（已用最大+1）。"""
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_id_table(trunk / "pet.xlsx", "Pet", [100, 101, 102])
        _write_id_table(dev / "pet.xlsx", "Pet", [100, 200])
        v = IdScopeValidator()
        r = v.claim_id(100, trunk, branches=[trunk, dev])
        assert r["claimed"] is True
        assert len(r["conflict_locations"]) >= 2
        assert r["suggested_next"] is not None
        assert r["suggested_next"] > 200  # 应大于已用最大

    def test_claim_single_branch_whitelist(self, tmp_path):
        """单分支同 stem 跨 sheet 同 id → 不算冲突（白名单）。"""
        root = tmp_path / "trunk"
        wb = Workbook(); ws = wb.active; ws.title = "Pet"; ws.cell(1, 1, "编号"); ws.cell(5, 1, 50)
        ws2 = wb.create_sheet("Pet2"); ws2.cell(1, 1, "编号"); ws2.cell(5, 1, 50)
        root.mkdir(parents=True, exist_ok=True)
        wb.save(root / "pet.xlsx")
        v = IdScopeValidator()
        r = v.claim_id(50, root)
        assert r["claimed"] is False  # 同 stem 白名单


class TestF4AgentInjection:
    def _make_agent(self):
        from agent.excel.core.agent import TableAgent
        ag = object.__new__(TableAgent)
        ag._agent_subtask_sink = None
        ag._resources_dir = None
        return ag

    def test_agent_validate_id_scope_no_mgr_returns_ok(self):
        """id_mgr 未加载 → ok=True 不阻断。"""
        ag = self._make_agent()
        ok, reason = ag._validate_id_scope("pet", "Pet", "编号", 999)
        assert ok is True
        assert reason == ""

    def test_agent_validate_id_scope_id_conflict_hold_event(self, tmp_path, monkeypatch):
        """F4：id-claim 跨分支冲突 → 返 False + pre_commit_hold 事件 SSE 推送。"""
        from engine.id_scope import get_id_scope_validator
        import engine.id_scope as _ids
        _ids._validator = None
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_id_mgr(trunk / "id_mgr.xlsx", [
            {"module": "pet.Pet", "id_min": 1, "id_max": 1000,
             "used_min": 1, "used_max": 100, "status": "active"}])
        _write_id_table(trunk / "pet.xlsx", "Pet", [100, 101])
        _write_id_table(dev / "pet.xlsx", "Pet", [100, 200])  # 100 跨分支冲突
        v = get_id_scope_validator()
        v.load_id_mgr(trunk / "id_mgr.xlsx")
        ag = self._make_agent()
        ag._resources_dir = str(trunk)
        # 多分支模式：注入 branches 让 agent 扫多根（复用 _validate_id_scope 的 claim_id）
        # 注：agent 默认单根；这里通过环境变量 CODEMAKER_ID_SCOPE_BRANCHES 传多根
        monkeypatch.setenv("CODEMAKER_ID_SCOPE_BRANCHES", str(dev))
        emitted = []
        ag._agent_subtask_sink = lambda event, data: emitted.append((event, data))
        ok, reason = ag._validate_id_scope("pet", "Pet", "编号", 100)
        assert ok is False
        assert "冲突" in reason or "conflict" in reason.lower()
        assert len(emitted) == 1
        assert emitted[0][0] == "pre_commit_hold"
        assert emitted[0][1]["kind"] == "id_conflict"

    def test_agent_validate_id_scope_no_conflict_ok(self, tmp_path, monkeypatch):
        """F4：id-claim 无冲突 → ok=True。"""
        import engine.id_scope as _ids
        _ids._validator = None
        trunk = tmp_path / "trunk"
        _write_id_mgr(trunk / "id_mgr.xlsx", [
            {"module": "pet.Pet", "id_min": 1, "id_max": 1000,
             "used_min": 1, "used_max": 100, "status": "active"}])
        _write_id_table(trunk / "pet.xlsx", "Pet", [1, 2])
        from engine.id_scope import get_id_scope_validator
        v = get_id_scope_validator()
        v.load_id_mgr(trunk / "id_mgr.xlsx")
        ag = self._make_agent()
        ag._resources_dir = str(trunk)
        ok, reason = ag._validate_id_scope("pet", "Pet", "编号", 999)
        assert ok is True
