"""O17 方法 E Schema 血缘联动单测（compute_column_lineage + sync_preview + column_added）。

覆盖：
- E1 compute_column_lineage(branch_roots) → ColumnLineageGraph
- E1 sync_preview(table, sheet) → trunk 有列其他分支缺
- E2 compute_column_changes(src_cols, tgt_cols) → column_added kind
- E1 _capped 后缀统一 table key（pet vs pet_capped 同 table）

无真实 SVN 分支，用 tmp_path 构造多分支目录 + openpyxl 写列定义。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from engine.column_lineage import (
    compute_column_lineage, ColumnLineageGraph, ColumnLineageEntry,
    _scan_sheet_columns,
)
from routers.structural import compute_column_changes


def _write_table_with_columns(path: Path, sheet: str,
                              headers: list[str],
                              types: list[str] = None,
                              constraints: list = None,
                              data_start_row: int = 5):
    """构造含列定义的表（row1=header, row2=type, row3=constraints）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)
    if types:
        for i, t in enumerate(types, start=1):
            ws.cell(2, i, t)
    if constraints:
        for i, c in enumerate(constraints, start=1):
            ws.cell(3, i, c)
    # 一行数据占位（data_start_row）
    ws.cell(data_start_row, 1, 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


class TestE1ComputeColumnLineage:
    def test_empty_branches_returns_empty_graph(self):
        g = compute_column_lineage([])
        assert isinstance(g, ColumnLineageGraph)
        assert g.tables == {}
        assert g.scanned_files == 0

    def test_single_branch_scans_columns(self, tmp_path):
        trunk = tmp_path / "trunk"
        _write_table_with_columns(
            trunk / "pet.xlsx", "Pet",
            headers=["编号", "名称", "类型"],
            types=["int", "str", "str"],
            constraints=[1, 1, 0],
        )
        g = compute_column_lineage([trunk])
        assert "pet" in g.tables
        assert "Pet" in g.tables["pet"]
        cols = g.tables["pet"]["Pet"]
        assert "编号" in cols
        assert "名称" in cols
        assert cols["编号"].col_type == "int"
        assert cols["编号"].not_empty is True
        assert cols["类型"].not_empty is False
        assert "trunk" in cols["编号"].present_in_branches

    def test_multi_branch_lineage_present_in(self, tmp_path):
        """两分支同表同列 → present_in_branches 含两分支。"""
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_table_with_columns(trunk / "pet.xlsx", "Pet", ["编号", "名称"])
        _write_table_with_columns(dev / "pet.xlsx", "Pet", ["编号", "名称"])
        g = compute_column_lineage([trunk, dev])
        cols = g.tables["pet"]["Pet"]
        assert set(cols["编号"].present_in_branches) == {"trunk", "dev"}

    def test_capped_suffix_unified_table_key(self, tmp_path):
        """pet_capped.xlsx 与 trunk/pet.xlsx 统一 table_key="pet"。"""
        trunk = tmp_path / "trunk"
        capped = tmp_path / "cappedbranch"
        _write_table_with_columns(trunk / "pet.xlsx", "Pet", ["编号", "名称"])
        _write_table_with_columns(capped / "pet_capped.xlsx", "Pet", ["编号"])
        g = compute_column_lineage([trunk, capped])
        # 两文件 stem 不同（pet vs pet_capped）但 _capped 后缀统一
        assert "pet" in g.tables
        cols = g.tables["pet"]["Pet"]
        assert "名称" in cols  # trunk 有
        assert "trunk" in cols["名称"].present_in_branches
        assert "cappedbranch" not in cols["名称"].present_in_branches


class TestE1SyncPreview:
    def test_trunk_add_column_missing_in_dev(self, tmp_path):
        """trunk 加列 → sync_preview 报 dev 缺该列。"""
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_table_with_columns(trunk / "pet.xlsx", "Pet", ["编号", "名称", "新列"])
        _write_table_with_columns(dev / "pet.xlsx", "Pet", ["编号", "名称"])  # 缺"新列"
        g = compute_column_lineage([trunk, dev])
        r = g.sync_preview("pet", "Pet")
        assert "新列" in r["trunk_columns"]
        missing_dev = [m for m in r["missing_in_branches"] if m["branch"] == "dev"]
        assert len(missing_dev) == 1
        assert "新列" in missing_dev[0]["missing_columns"]

    def test_no_missing_returns_empty(self, tmp_path):
        """两分支列集一致 → missing_in_branches 空。"""
        trunk = tmp_path / "trunk"
        dev = tmp_path / "dev"
        _write_table_with_columns(trunk / "pet.xlsx", "Pet", ["编号", "名称"])
        _write_table_with_columns(dev / "pet.xlsx", "Pet", ["编号", "名称"])
        g = compute_column_lineage([trunk, dev])
        r = g.sync_preview("pet", "Pet")
        assert r["missing_in_branches"] == []

    def test_table_not_in_graph_returns_empty(self, tmp_path):
        g = compute_column_lineage([tmp_path])
        r = g.sync_preview("nonexistent", "S1")
        assert r["trunk_columns"] == []
        assert r["missing_in_branches"] == []


class TestE2ComputeColumnChanges:
    def test_column_added_detected(self):
        """src(trunk) 有列 tgt(capped) 无 → column_added。"""
        src = {"pet": {"Pet": {"编号", "名称", "新列"}}}
        tgt = {"pet": {"Pet": {"编号", "名称"}}}
        out = compute_column_changes(src, tgt)
        added = [c for c in out if c["kind"] == "column_added"]
        assert len(added) == 1
        assert added[0]["column"] == "新列"
        assert added[0]["table"] == "pet"
        assert added[0]["sheet"] == "Pet"

    def test_no_added_when_identical(self):
        src = {"pet": {"Pet": {"编号", "名称"}}}
        tgt = {"pet": {"Pet": {"编号", "名称"}}}
        out = compute_column_changes(src, tgt)
        assert out == []

    def test_column_added_multi_sheet(self):
        src = {"pet": {"Pet": {"a", "b"}, "Pet2": {"c", "d", "e"}}}
        tgt = {"pet": {"Pet": {"a"}, "Pet2": {"c", "d"}}}
        out = compute_column_changes(src, tgt)
        cols_added = {(c["sheet"], c["column"]) for c in out}
        assert ("Pet", "b") in cols_added
        assert ("Pet2", "e") in cols_added


class TestScanSheetColumnsHelper:
    def test_empty_sheet_returns_none(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "S1"
        assert _scan_sheet_columns(ws) is None

    def test_reads_header_type_constraints(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pet"
        ws.cell(1, 1, "编号"); ws.cell(1, 2, "名称")
        ws.cell(2, 1, "int"); ws.cell(2, 2, "str")
        ws.cell(3, 1, 1); ws.cell(3, 2, 0)
        cols = _scan_sheet_columns(ws)
        assert "编号" in cols
        assert cols["编号"]["type"] == "int"
        assert cols["编号"]["not_empty"] is True
        assert cols["名称"]["not_empty"] is False
