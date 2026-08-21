"""D1 写后读回验证单测（capability: write-verification）。

验证 _verify_write_back + _values_equal：
1. 写盘成功值正确 → ok=True
2. 落盘值不符 → ok=False + mismatched_fields
3. 首列空行 → 按非首列字段比对（不依赖 pk 定位）
4. 读回失败（IO 错误）→ ok=False + error="read_back_failed"
5. _values_equal 容差：None/空等价、数值容差、str strip、list 递归
"""
from __future__ import annotations

import os
import sys
import types
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.agent import TableAgent, AgentResult, _values_equal
from agent.cli_interface import StubCodeMakerCLI


def _make_agent(cli, **extra) -> types.SimpleNamespace:
    agent = types.SimpleNamespace(cli=cli, live_index=False, _index_cache=None, **extra)
    for name in ("_verify_write_back", "_write_cell_and_verify",
                 "_refresh_index_after_write", "_do_append",
                 "_auto_sort_after_write", "_allocate_pk"):
        if hasattr(TableAgent, name):
            setattr(agent, name, getattr(TableAgent, name).__get__(agent))
    return agent


@pytest.fixture()
def sheet_xlsx(tmp_path):
    """建表：表头行1，数据行5起，预置一行。"""
    p = tmp_path / "t.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "Sheet1"
    s.cell(1, 1, "id")
    s.cell(1, 2, "名称")
    s.cell(1, 3, "类型")
    s.cell(5, 1, 5)
    s.cell(5, 2, "a")
    s.cell(5, 3, 1)
    wb.save(p)
    return p


def test_values_equal_semantics():
    """_values_equal 容差语义。"""
    assert _values_equal(None, None)
    assert _values_equal(None, "")  # None/空等价
    assert _values_equal("", None)
    assert _values_equal(1, 1.0)  # 数值容差
    assert _values_equal(1.0000001, 1.0)
    assert not _values_equal(1.1, 1.0)
    assert _values_equal("abc", "abc ")  # str strip
    assert _values_equal([1, 2], [1, 2])  # list 递归
    assert not _values_equal([1, 2], [1, 3])
    assert not _values_equal("a", "b")


def test_verify_write_back_success(sheet_xlsx):
    """写盘成功值正确 → ok=True。"""
    cli = StubCodeMakerCLI(workspace=sheet_xlsx.parent, header_row=1, data_start_row=5)
    agent = _make_agent(cli)
    # 写入行6：id=6, 名称=b, 类型=2
    r = cli.append_row(sheet_xlsx, "Sheet1", {1: 6, 2: "b", 3: 2})
    assert r.ok
    new_row = r.data["row"]
    verify = agent._verify_write_back(sheet_xlsx, "Sheet1", new_row, {1: 6, 2: "b", 3: 2})
    assert verify["ok"] is True, verify


def test_verify_write_back_mismatch(sheet_xlsx):
    """落盘值不符 → ok=False + mismatched_fields。"""
    cli = StubCodeMakerCLI(workspace=sheet_xlsx.parent, header_row=1, data_start_row=5)
    agent = _make_agent(cli)
    cli.append_row(sheet_xlsx, "Sheet1", {1: 6, 2: "b", 3: 2})
    # 期望值故意写错（类型期望 99，实际 2）
    verify = agent._verify_write_back(sheet_xlsx, "Sheet1", 6, {1: 6, 2: "b", 3: 99})
    assert verify["ok"] is False
    assert 3 in verify["mismatched_fields"]
    assert verify["mismatched_fields"][3]["expected"] == 99
    assert verify["mismatched_fields"][3]["actual"] == 2


def test_verify_write_back_first_col_empty(sheet_xlsx):
    """首列空行 → 按非首列字段比对（不依赖 pk 定位）。"""
    cli = StubCodeMakerCLI(workspace=sheet_xlsx.parent, header_row=1, data_start_row=5)
    agent = _make_agent(cli)
    # 写入行6：首列空，仅写名称+类型
    r = cli.append_row(sheet_xlsx, "Sheet1", {2: "c", 3: 3})
    assert r.ok
    new_row = r.data["row"]
    # expected_fields 不含首列（1），仅比对非首列
    verify = agent._verify_write_back(sheet_xlsx, "Sheet1", new_row, {2: "c", 3: 3})
    assert verify["ok"] is True, verify


def test_verify_write_back_read_failure(sheet_xlsx):
    """读回失败（IO 错误）→ ok=False + error="read_back_failed"。"""
    cli = StubCodeMakerCLI(workspace=sheet_xlsx.parent, header_row=1, data_start_row=5)
    agent = _make_agent(cli)
    # monkeypatch read_cell 抛异常
    orig_read_cell = cli.read_cell
    cli.read_cell = lambda *a, **kw: (_ for _ in ()).throw(IOError("锁"))
    try:
        verify = agent._verify_write_back(sheet_xlsx, "Sheet1", 6, {1: 6})
        assert verify["ok"] is False
        assert verify["error"] == "read_back_failed"
    finally:
        cli.read_cell = orig_read_cell
