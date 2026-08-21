"""方法 A 阶段一公式守门 audit 留痕单测。

验证 _save_with_cache_check 的 needs_manual_fix 分支：CODEMAKER_FORMULA_GATE 开关
控制 warning+audit 行为。on/hold=warning+audit_log 留痕；off=静默不记。
真实公式重算流程由 test_formula_cache.py 覆盖，此处聚焦 audit 留痕逻辑（mock snapshot_before
绕过真实公式检测 + mock validate_and_fix 返回 needs_manual_fix=True）。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from agent.excel.cli.cli_interface import StubCodeMakerCLI
from agent.excel.core.backup_audit import BackupAuditor


class _FakeResult:
    """模拟 validate_and_fix 的返回。"""
    def __init__(self, needs: bool, msg: str = "公式缓存丢失"):
        self.needs_manual_fix = needs
        self.message = msg


def _make_formula_workbook(p: Path) -> None:
    """构造含公式的表（实际 mock snapshot_before 绕过检测，公式仅占位）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(1, 1, "a")
    ws.cell(5, 1, 1)
    ws.cell(5, 2, "=A5+1")
    wb.save(p)


def _make_auditor(tmp_path: Path) -> BackupAuditor:
    """构造隔离的 BackupAuditor（backups+audit_log 全在 tmp_path，不污染 server/backups）。"""
    return BackupAuditor(
        workspace=tmp_path,
        backups_dir=tmp_path / "backups",
        audit_log_path=tmp_path / "audit_log.jsonl",
    )


def _read_audit_ops(audit_path: Path) -> list[str]:
    """读 audit_log.jsonl 所有 operation 字段。"""
    if not audit_path.exists():
        return []
    import json
    ops: list[str] = []
    with open(audit_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                d = json.loads(line)
                ops.append(d.get("operation", ""))
            except Exception:
                continue
    return ops


def _mock_formula_loss(cli, monkeypatch, needs: bool = True):
    """mock snapshot_before 返回非空快照 + validate_and_fix 返回 needs_manual_fix。"""
    monkeypatch.setattr(cli._formula_validator, "snapshot_before",
                        lambda path: {"_mock": 1})
    monkeypatch.setattr(cli._formula_validator, "validate_and_fix",
                        lambda path, before: _FakeResult(needs))


def test_no_formula_no_audit(tmp_path, monkeypatch):
    """无公式表：fast-path，needs_manual_fix=False，audit 无 formula_loss_detected。"""
    audit_path = tmp_path / "audit_log.jsonl"
    p = tmp_path / "plain.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "id"); ws.cell(5, 1, 1)
    wb.save(p)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    cli._auditor = _make_auditor(tmp_path)
    # 不 mock → 真实 snapshot_before 检测无公式 → fast-path
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok
    assert res.needs_manual_fix is False
    ops = _read_audit_ops(audit_path)
    assert "formula_loss_detected" not in ops


def test_formula_loss_detected_audit_on(tmp_path, monkeypatch):
    """GATE=on(默认) + mock needs_manual_fix=True → audit 有 formula_loss_detected。"""
    audit_path = tmp_path / "audit_log.jsonl"
    p = tmp_path / "fml.xlsx"
    _make_formula_workbook(p)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    cli._auditor = _make_auditor(tmp_path)
    _mock_formula_loss(cli, monkeypatch, needs=True)
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok
    assert res.needs_manual_fix is True
    ops = _read_audit_ops(audit_path)
    assert "formula_loss_detected" in ops, f"audit 未记录: {ops}"


def test_gate_off_silent(tmp_path, monkeypatch):
    """GATE=off + mock needs_manual_fix=True → 静默不记 audit。"""
    monkeypatch.setenv("CODEMAKER_FORMULA_GATE", "off")
    audit_path = tmp_path / "audit_log.jsonl"
    p = tmp_path / "fml.xlsx"
    _make_formula_workbook(p)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    cli._auditor = _make_auditor(tmp_path)
    _mock_formula_loss(cli, monkeypatch, needs=True)
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok
    assert res.needs_manual_fix is True
    ops = _read_audit_ops(audit_path)
    assert "formula_loss_detected" not in ops, f"off 模式不应记 audit: {ops}"


def test_gate_hold_audit(tmp_path, monkeypatch):
    """GATE=hold + mock needs_manual_fix=True → audit 留痕（阶段一等同 on，阻断留第二波）。"""
    monkeypatch.setenv("CODEMAKER_FORMULA_GATE", "hold")
    audit_path = tmp_path / "audit_log.jsonl"
    p = tmp_path / "fml.xlsx"
    _make_formula_workbook(p)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    cli._auditor = _make_auditor(tmp_path)
    _mock_formula_loss(cli, monkeypatch, needs=True)
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok
    assert res.needs_manual_fix is True
    ops = _read_audit_ops(audit_path)
    assert "formula_loss_detected" in ops


def test_formula_no_loss_no_audit(tmp_path, monkeypatch):
    """GATE=on + mock needs_manual_fix=False → audit 无 formula_loss_detected（重算成功）。"""
    audit_path = tmp_path / "audit_log.jsonl"
    p = tmp_path / "fml.xlsx"
    _make_formula_workbook(p)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    cli._auditor = _make_auditor(tmp_path)
    _mock_formula_loss(cli, monkeypatch, needs=False)
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok
    assert res.needs_manual_fix is False
    ops = _read_audit_ops(audit_path)
    assert "formula_loss_detected" not in ops
