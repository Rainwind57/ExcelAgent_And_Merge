"""方法 D 批注守门单测。

验证 _save_with_cache_check 的批注保护层：save 前 snapshot → save 后 reload 做差
→ 丢失则原 wb 回写 Comment → 二次 save → 二次做差记数。
核心指标：写表后批注文本保留率 100%（openpyxl save 偶发丢批注的预防层）。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment

from agent.excel.cli.cli_interface import StubCodeMakerCLI


def _make_workbook(p: Path, with_comment: bool = False) -> None:
    """构造最小表行5起（对齐 CLI 默认 data_start_row=5）。
    with_comment 时 A5 带批注。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(1, 1, "id")
    ws.cell(1, 2, "name")
    ws.cell(5, 1, 1)
    ws.cell(5, 2, "alice")
    if with_comment:
        ws.cell(5, 1).comment = Comment("主键不可改", "tester")
    wb.save(p)


def _comment_text(p: Path, sheet: str, coord: str) -> str:
    """独立 openpyxl 读单元格批注文本，无批注返回空串。"""
    import openpyxl
    wb = openpyxl.load_workbook(p, data_only=False)
    try:
        c = wb[sheet][coord].comment
        return c.text if c else ""
    finally:
        wb.close()


def test_no_comment_write_fast_path(tmp_path):
    """无批注表写一次：fast-path，comment_replay.replayed=False，不触发回写。"""
    p = tmp_path / "plain.xlsx"
    _make_workbook(p, with_comment=False)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    res = cli.write_cell(p, "S1", 5, 2, "bob")
    assert res.ok, f"write_cell failed: {res.error}"
    assert res.comment_replay["replayed"] is False
    assert res.comment_replay["still_lost"] == 0


def test_comment_write_other_column(tmp_path):
    """批注表改非批注列：A5 有批注，改 B5 值，写后 A5 批注文本一致。"""
    p = tmp_path / "cmt.xlsx"
    _make_workbook(p, with_comment=True)
    before = _comment_text(p, "S1", "A5")
    assert before == "主键不可改"
    cli = StubCodeMakerCLI(workspace=tmp_path)
    res = cli.write_cell(p, "S1", 5, 2, "bob")
    assert res.ok, f"write_cell failed: {res.error}"
    after = _comment_text(p, "S1", "A5")
    assert after == before, f"批注丢失: before={before!r} after={after!r}"


def test_comment_write_comment_column(tmp_path):
    """批注表改批注所在列：A5 有批注，改 A5 值（不改批注），写后批注文本一致。"""
    p = tmp_path / "cmt2.xlsx"
    _make_workbook(p, with_comment=True)
    before = _comment_text(p, "S1", "A5")
    cli = StubCodeMakerCLI(workspace=tmp_path)
    res = cli.write_cell(p, "S1", 5, 1, 999)
    assert res.ok, f"write_cell failed: {res.error}"
    after = _comment_text(p, "S1", "A5")
    assert after == before, f"批注丢失: before={before!r} after={after!r}"


def test_append_row_preserves_comment(tmp_path):
    """append_row 追加行：原 A5 批注不丢。"""
    p = tmp_path / "cmt3.xlsx"
    _make_workbook(p, with_comment=True)
    before = _comment_text(p, "S1", "A5")
    cli = StubCodeMakerCLI(workspace=tmp_path)
    res = cli.append_row(p, "S1", {1: 2, 2: "carol"})
    assert res.ok, f"append_row failed: {res.error}"
    after = _comment_text(p, "S1", "A5")
    assert after == before, f"批注丢失: before={before!r} after={after!r}"


def test_guard_off_env(tmp_path, monkeypatch):
    """CODEMAKER_COMMENT_GUARD=off 时跳过守门，comment_replay 返回默认空结构。"""
    monkeypatch.setenv("CODEMAKER_COMMENT_GUARD", "off")
    p = tmp_path / "off.xlsx"
    _make_workbook(p, with_comment=True)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    res = cli.write_cell(p, "S1", 5, 2, "bob")
    assert res.ok, f"write_cell failed: {res.error}"
    # off 时 before_comments 为空，replayed=False
    assert res.comment_replay["replayed"] is False
    assert res.comment_replay["still_lost"] == 0
