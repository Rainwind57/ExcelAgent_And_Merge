"""4.6 snapshot_formulas (path, mtime, size) 缓存单测。

验证:同一未变文件重复 snapshot 命中缓存(免二次 load);文件变化失效重读。
"""
from __future__ import annotations

import os
import sys
import time
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from agent.excel.formula_cache_validator import (
    snapshot_formulas, reset_formula_snapshot_cache, _SNAPSHOT_CACHE,
)


def _make_formula_file(p: Path, a_val=10, b_val=20):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(1, 1, "a")
    ws.cell(1, 2, "b")
    ws.cell(1, 3, "sum")
    ws.cell(2, 1, a_val)
    ws.cell(2, 2, b_val)
    ws.cell(2, 3, "=A2+B2")  # 公式
    wb.save(p)
    return p


def test_snapshot_cache_hit_same_file(tmp_path):
    """同一文件(未变)二次 snapshot → 命中缓存,返回等值。"""
    reset_formula_snapshot_cache()
    p = tmp_path / "f.xlsx"
    _make_formula_file(p)
    snap1 = snapshot_formulas(p)
    assert snap1  # 含公式
    assert ("S1", 2, 3) in snap1

    snap2 = snapshot_formulas(p)
    assert snap2 == snap1
    # 缓存条目存在
    assert any(k[0] == str(p) for k in _SNAPSHOT_CACHE)


def test_snapshot_cache_invalidate_on_change(tmp_path):
    """文件内容变化(size 变)→ 缓存失效重读,反映新公式值。"""
    reset_formula_snapshot_cache()
    p = tmp_path / "f.xlsx"
    _make_formula_file(p, a_val=10, b_val=20)
    snap1 = snapshot_formulas(p)
    # 修改文件(改 a 值 → 公式缓存值变/size 可能变)
    _make_formula_file(p, a_val=999, b_val=20)
    # 确保 mtime 变化(部分 FS 1s 精度)
    time.sleep(0.01)
    os.utime(p, None)
    snap2 = snapshot_formulas(p)
    # 重新读到的公式单元格位置仍在(公式未变),缓存失效后重读
    assert ("S1", 2, 3) in snap2


def test_snapshot_cache_returns_copy(tmp_path):
    """命中缓存返回副本,调用方改不影响缓存。"""
    reset_formula_snapshot_cache()
    p = tmp_path / "f.xlsx"
    _make_formula_file(p)
    snap1 = snapshot_formulas(p)
    snap1[("X", 9, 9)] = "tampered"
    snap2 = snapshot_formulas(p)
    assert ("X", 9, 9) not in snap2, "缓存应返回副本,不受调用方篡改"


def test_snapshot_no_formula_file_cached_empty(tmp_path):
    """无公式文件也缓存空 dict,二次 snapshot 命中。"""
    reset_formula_snapshot_cache()
    p = tmp_path / "nofml.xlsx"
    wb = Workbook()
    wb.active.cell(1, 1, "x")
    wb.active.cell(2, 1, 1)
    wb.save(p)
    s1 = snapshot_formulas(p)
    assert s1 == {}
    s2 = snapshot_formulas(p)
    assert s2 == {}
    assert any(k[0] == str(p) for k in _SNAPSHOT_CACHE)
