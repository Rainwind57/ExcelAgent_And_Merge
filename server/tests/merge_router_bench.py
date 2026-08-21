"""路由层性能基准：branch/subdir compare 端到端 优化前/后对比。

monkeypatch 模拟优化前：禁用 _prefill_rev_cache（每文件 svn log）+ 强制 openpyxl
读 sheet 名（模拟无 calamine，_dir_sheet_sets 每文件 0.3-0.5s）。对比优化后（一次
svn info 批量预填 + calamine 读 sheet）。

直接调 branch_compare / subdir_compare 函数（不走 HTTP），测端到端 compare 耗时。

指标：
  - branch_compare_baseline_ms / optimized_ms / speedup
  - subdir_compare_baseline_ms / optimized_ms / speedup
  - dirs_load_cold_ms / warm_ms（冷/热缓存，30s TTL）
  - preview_base_ms（LCA 反查 + HEAD rev）

用法（在 server/ 目录下，需 SVN demo 已初始化 + python-calamine 已装）:
    python -m tests.merge_router_bench
    python -m tests.merge_router_bench --runs 3   # 多轮取中位数
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
import time
from pathlib import Path

TESTS_DIR = Path(__file__).resolve().parent
SERVER_DIR = TESTS_DIR.parent
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))

from routers import merge_branch  # noqa: E402
from routers.merge_branch import branch_compare, BranchCompareRequest, list_branch_dirs  # noqa: E402
from routers import merge_subdir  # noqa: E402
from routers.merge_subdir import subdir_compare, SubdirCompareRequest  # noqa: E402

DEMO_SVN_WC = SERVER_DIR.parent / "merge" / "svn" / "demo_svn" / "wc"


# 保存原始实现，供 _patch_baseline 恢复
_orig_prefill = merge_branch._prefill_rev_cache
_orig_sheet_names_mb = merge_branch._sheet_names
_orig_sheet_names_ms = merge_subdir._sheet_names


def _sheet_names_openpyxl(fp):
    """强制 openpyxl 回退版（模拟优化前无 calamine）：每文件 0.3-0.5s。"""
    from openpyxl import load_workbook
    wb = None
    try:
        wb = load_workbook(fp, read_only=True, data_only=True)
        return set(wb.sheetnames)
    except Exception:
        return set()
    finally:
        if wb is not None:
            wb.close()


def _patch_baseline(on: bool):
    """on=True 模拟优化前：禁用 _prefill_rev_cache（逐文件 svn log）+ 强制 openpyxl 读 sheet。"""
    if on:
        merge_branch._prefill_rev_cache = lambda *a, **k: None
        merge_subdir._prefill_rev_cache = lambda *a, **k: None
        merge_branch._sheet_names = _sheet_names_openpyxl
        merge_subdir._sheet_names = _sheet_names_openpyxl
    else:
        merge_branch._prefill_rev_cache = _orig_prefill
        merge_subdir._prefill_rev_cache = _orig_prefill
        merge_branch._sheet_names = _orig_sheet_names_mb
        merge_subdir._sheet_names = _orig_sheet_names_ms


def _invalidate_dirs_cache():
    """清 /dirs 缓存，测冷加载。"""
    merge_branch._DIRS_CACHE["data"] = None
    merge_branch._DIRS_CACHE["ts"] = 0.0


def _median(samples_ms):
    return round(statistics.median(samples_ms), 1) if samples_ms else 0.0


def bench_dirs_load(runs: int):
    """/dirs 冷加载（清缓存）vs 热加载（命中缓存）。"""
    cold, warm = [], []
    for _ in range(runs):
        _invalidate_dirs_cache()
        t0 = time.perf_counter()
        list_branch_dirs()
        cold.append((time.perf_counter() - t0) * 1000)
        t0 = time.perf_counter()
        list_branch_dirs()
        warm.append((time.perf_counter() - t0) * 1000)
    return {
        "tables_listed": len((list_branch_dirs() or {}).get("dirs", [])),
        "cold_ms_median": _median(cold),
        "warm_ms_median": _median(warm),
        "cache_speedup": round(_median(cold) / _median(warm), 2) if _median(warm) else 0.0,
    }


def bench_preview_base(runs: int):
    """preview-base：LCA 反查 + HEAD rev（svn log 并行）。"""
    src = str(DEMO_SVN_WC / "branches" / "dev1")
    tgt = str(DEMO_SVN_WC / "trunk")
    samples = []
    for _ in range(runs):
        t0 = time.perf_counter()
        merge_branch.preview_base(merge_branch.PreviewBaseRequest(
            source_branch=src, target_branch=tgt))
        samples.append((time.perf_counter() - t0) * 1000)
    return {"src": "dev1", "tgt": "trunk", "ms_median": _median(samples)}


def bench_branch_compare(runs: int):
    """branch_compare（dev1→trunk 全表）：优化前(baseline) vs 优化后对比。

    baseline：禁用 _prefill_rev_cache（每文件 svn log）+ 强制 openpyxl 读 sheet 名
              （模拟优化前无 calamine，_dir_sheet_sets 每文件 0.3-0.5s）。
    optimized：当前优化版（_prefill_rev_cache 一次 svn info + calamine 读 sheet）。
    """
    src = str(DEMO_SVN_WC / "branches" / "dev1")
    tgt = str(DEMO_SVN_WC / "trunk")
    req = BranchCompareRequest(direction="absorb", source_branch=src, target_branch=tgt)

    baseline_ms, optimized_ms = [], []
    baseline_groups = optimized_groups = None
    for i in range(runs):
        _patch_baseline(True)
        t0 = time.perf_counter()
        resp = branch_compare(req)
        baseline_ms.append((time.perf_counter() - t0) * 1000)
        if baseline_groups is None:
            baseline_groups = {gn: len(fg.files) for gn, fg in resp.groups.items()}
        _patch_baseline(False)
        t0 = time.perf_counter()
        resp = branch_compare(req)
        optimized_ms.append((time.perf_counter() - t0) * 1000)
        if optimized_groups is None:
            optimized_groups = {gn: len(fg.files) for gn, fg in resp.groups.items()}

    b, o = _median(baseline_ms), _median(optimized_ms)
    return {
        "tables": len(optimized_groups or {}),
        "baseline_ms_median": b,
        "optimized_ms_median": o,
        "speedup": round(b / o, 2) if o else 0.0,
        "result_match": baseline_groups == optimized_groups,
    }


def bench_subdir_compare(runs: int):
    """subdir_compare（subdev_1→trunk 全表）：优化前(baseline) vs 优化后对比。"""
    src = str(DEMO_SVN_WC / "branches" / "subdev_1")
    tgt = str(DEMO_SVN_WC / "trunk")
    req = SubdirCompareRequest(source_dir=src, target_dir=tgt)

    baseline_ms, optimized_ms = [], []
    baseline_struct = optimized_struct = None
    for i in range(runs):
        _patch_baseline(True)
        t0 = time.perf_counter()
        resp = subdir_compare(req)
        baseline_ms.append((time.perf_counter() - t0) * 1000)
        if baseline_struct is None:
            baseline_struct = sorted((c.table, c.status) for c in resp.structural_changes)
        _patch_baseline(False)
        t0 = time.perf_counter()
        resp = subdir_compare(req)
        optimized_ms.append((time.perf_counter() - t0) * 1000)
        if optimized_struct is None:
            optimized_struct = sorted((c.table, c.status) for c in resp.structural_changes)

    b, o = _median(baseline_ms), _median(optimized_ms)
    return {
        "tables": len(optimized_struct or {}),
        "baseline_ms_median": b,
        "optimized_ms_median": o,
        "speedup": round(b / o, 2) if o else 0.0,
        "result_match": baseline_struct == optimized_struct,
        "false_source_deleted": sum(1 for _, st in (optimized_struct or []) if st == "source_deleted"),
    }


def run_all(runs: int):
    if not (DEMO_SVN_WC / "branches" / "dev1").is_dir():
        return {"skipped": f"demo_svn/wc 不存在（{DEMO_SVN_WC}）"}

    print(f"=== merge 路由层性能基准（runs={runs}，取中位数）===")
    print("[1/4] /dirs 加载（冷/热）...")
    dirs = bench_dirs_load(runs)
    print(f"  冷={dirs['cold_ms_median']}ms 热={dirs['warm_ms_median']}ms "
          f"speedup={dirs['cache_speedup']}x tables={dirs['tables_listed']}")

    print("[2/4] preview-base（LCA 反查）...")
    pb = bench_preview_base(runs)
    print(f"  {pb['src']}→{pb['tgt']} {pb['ms_median']}ms")

    print("[3/4] branch_compare 优化前/后（dev1→trunk 全表）...")
    bc = bench_branch_compare(runs)
    print(f"  优化前={bc['baseline_ms_median']}ms 优化后={bc['optimized_ms_median']}ms "
          f"speedup={bc['speedup']}x tables={bc['tables']} match={bc['result_match']}")

    print("[4/4] subdir_compare 优化前/后（subdev_1→trunk 全表）...")
    sc = bench_subdir_compare(runs)
    print(f"  优化前={sc['baseline_ms_median']}ms 优化后={sc['optimized_ms_median']}ms "
          f"speedup={sc['speedup']}x tables={sc['tables']} match={sc['result_match']} "
          f"false_src_deleted={sc['false_source_deleted']}")

    _patch_baseline(False)  # 恢复
    return {
        "dirs_load": dirs,
        "preview_base": pb,
        "branch_compare_ab": bc,
        "subdir_compare_ab": sc,
    }


def print_report(result: dict):
    print("\n=== 性能对比汇总（优化前 → 优化后）===")
    print(f"{'环节':<24} {'优化前(ms)':<14} {'优化后(ms)':<14} {'加速比':<10} {'说明'}")
    d = result["dirs_load"]
    print(f"{'/dirs 加载':<24} {d['cold_ms_median']:<14} {d['warm_ms_median']:<14} "
          f"{d['cache_speedup']}x 冷扫→热缓存命中")
    bc = result["branch_compare_ab"]
    print(f"{'branch_compare':<24} {bc['baseline_ms_median']:<14} {bc['optimized_ms_median']:<14} "
          f"{bc['speedup']}x svn info批量+calamine（dev1→trunk 74表）")
    sc = result["subdir_compare_ab"]
    print(f"{'subdir_compare':<24} {sc['baseline_ms_median']:<14} {sc['optimized_ms_median']:<14} "
          f"{sc['speedup']}x svn info批量+calamine（subdev_1→trunk）")
    pb = result["preview_base"]
    print(f"{'preview-base':<24} {'-':<14} {pb['ms_median']:<14} {'-':<10} LCA 反查（无优化项）")
    print(f"\n结果一致性：branch={'OK' if bc['result_match'] else 'MISMATCH'}  "
          f"subdir={'OK' if sc['result_match'] else 'MISMATCH'}  "
          f"false_source_deleted={sc['false_source_deleted']}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--runs", type=int, default=1, help="重复轮数（取中位数，默认 1）")
    ap.add_argument("--json-out", type=str, default=None, help="JSON 报告输出路径")
    args = ap.parse_args()

    result = run_all(args.runs)
    if isinstance(result, dict) and result.get("skipped"):
        print(result["skipped"])
        return

    print_report(result)

    report = {
        "summary": {
            "branch_speedup": result["branch_compare_ab"]["speedup"],
            "subdir_speedup": result["subdir_compare_ab"]["speedup"],
            "dirs_cache_speedup": result["dirs_load"]["cache_speedup"],
            "results_match": result["branch_compare_ab"]["result_match"]
                             and result["subdir_compare_ab"]["result_match"],
            "false_source_deleted": result["subdir_compare_ab"]["false_source_deleted"],
        },
        "details": result,
    }
    REPORT_DIR = TESTS_DIR / "reports"
    REPORT_DIR.mkdir(exist_ok=True)
    out = REPORT_DIR / "merge_router_bench_latest.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n报告已写: {out}")
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
