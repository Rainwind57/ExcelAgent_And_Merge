"""表格操作测试用例 —— 真实后端 Agent skill on/off A/B 评测脚本。

用途:
    针对 tests/cases/table_operation_test_cases.json 中的每条自然语言指令，
    分别用 TableAgent(enable_skill=False) 和 TableAgent(enable_skill=True)
    通过 **真实后端 AgentService**（走 codemaker serve 真实 LLM 解析）跑一遍，
    在 resources/ 的临时沙箱副本上真实执行增删改，然后用"跑前/跑后 Excel
    文件差异"作为 ground truth，与 expected_answer 逐条比对，产出：

    - 精准程度（字段值级别的正确率）
    - 时间性能（每条指令耗时）
    - 定位功能（是否命中正确的 表/sheet/操作类型）
    - 覆盖度（多表级联指令中，应该产生的行操作有多少真正产生了）

    并输出每个样例的详细运行记录、skill on/off 对比表格、总体指标与结论报告。

设计要点:
    - 每个样例独立：从 resources/ 拷贝一份全新临时副本作为沙箱，Agent 在沙箱内
      真实读写（不影响真实配表），跑完即删除沙箱。
    - 判分不依赖 Agent 返回的结构化字段（不同 intent/多表级联时格式差异较大），
      而是直接对比沙箱内 xlsx 文件"跑前 vs 跑后"的行级差异（增/删/改），
      这是脚本能拿到的最贴近事实的 ground truth。
    - expected_answer 里的 <new_id> 之类占位符字段视为"只要求出现该字段被写入
      /变更"，不做具体值比对；其余具体值字段要求内容相等（数值容差、list/tuple
      递归比较）。

用法（在 server/ 目录下）:
    python -m tests.table_case_eval --quick 8      # 冒烟：只跑前 8 条
    python -m tests.table_case_eval                # 跑全部（可能很久，见控制台耗时统计）
    python -m tests.table_case_eval --quick 8 --cases-file tests/cases/table_operation_test_cases.json

前提: codemaker serve 已启动（.env 配置好 CODEMAKER_SERVER_URL 等）。
"""
from __future__ import annotations

import argparse
import ast
import json
import os
import shutil
import statistics
import sys
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl

# ── 路径 & 环境变量 ──────────────────────────────────────
TESTS_DIR = Path(__file__).resolve().parent
SERVER_DIR = TESTS_DIR.parent
ROOT = SERVER_DIR.parent
RES = ROOT / "resources"
REPORT_DIR = TESTS_DIR / "reports"
DEFAULT_CASES_FILE = TESTS_DIR / "cases" / "table_operation_test_cases.json"


def _load_dotenv(env_path: Path) -> None:
    """极简 .env 加载：不覆盖已存在的环境变量。codemaker_client 等模块在 import
    时读取 os.environ 构造模块级常量，因此必须在 import agent/services 之前调用。"""
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k, v = k.strip(), v.strip()
        if k and k not in os.environ:
            os.environ[k] = v


_load_dotenv(ROOT / ".env")

if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))

from services.agent_service import AgentService  # noqa: E402


# ── Excel 读取 & 解析 ────────────────────────────────────

_SKIP_SHEETS = {"config"}


def _is_business_sheet(name: str) -> bool:
    if not name:
        return False
    if name.strip().lower() in _SKIP_SHEETS:
        return False
    if "说明" in name:
        return False
    return True


def _field_key(raw: Any) -> str:
    """把类型标注行的单元格（如 'prefab_id:int' / 'effect.data.3001.combat_id: int'）
    转为规整字段名 'prefab_id' / 'effect.data.3001.combat_id'。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    if ":" in s:
        s = s.rsplit(":", 1)[0]
    return s.strip()


def _parse_cell(v: Any) -> Any:
    """把单元格原始值规整为可比较的 Python 结构（list/tuple 字面量尝试解析）。"""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).strip()
    if not s:
        return None
    if (s[0] in "[(") and (s[-1] in "])"):
        try:
            return ast.literal_eval(s)
        except Exception:
            try:
                return json.loads(s)
            except Exception:
                return s
    return s


def read_table_rows(path: Path, sheet: str) -> tuple[list[str], list[dict]]:
    """读取一个 sheet 的业务数据行。

    约定（见 schema_infer.py）：Row1=中文表头 Row2=英文字段名:类型 Row3=约束
    Row4=默认值 Row5+=数据。返回 (field_keys, rows)，rows 为
    [{field_key: parsed_value}, ...]（跳过全空行）。
    """
    if not path.exists():
        return [], []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return [], []
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if len(rows) < 2:
        return [], []
    field_keys = [_field_key(c) for c in rows[1]]
    data: list[dict] = []
    for r in rows[4:]:
        if all(v is None for v in r):
            continue
        row_dict: dict = {}
        for i, key in enumerate(field_keys):
            if not key:
                continue
            row_dict[key] = _parse_cell(r[i] if i < len(r) else None)
        if row_dict:
            data.append(row_dict)
    return field_keys, data


def _values_equal(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return str(a if a is not None else "").strip() == str(b if b is not None else "").strip()
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) < 1e-6
    except (TypeError, ValueError):
        pass
    if isinstance(a, (list, tuple)) and isinstance(b, (list, tuple)):
        if len(a) != len(b):
            return False
        return all(_values_equal(x, y) for x, y in zip(a, b))
    return str(a).strip() == str(b).strip()


def _is_placeholder(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("<") and v.endswith(">")


# ── 沙箱 vs 原始资源：文件级 & 行级差异 ──────────────────

@dataclass
class ActualOp:
    table: str          # 相对 resources 的路径（正斜杠）
    sheet: str
    operation: str       # add | modify | delete
    row_before: dict = field(default_factory=dict)
    row_after: dict = field(default_factory=dict)
    changed_fields: dict = field(default_factory=dict)   # modify 专用：仅变更字段
    consumed: bool = False


def _changed_files(sandbox: Path, pristine: Path) -> list[Path]:
    changed = []
    for p in sandbox.rglob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        rel = p.relative_to(sandbox)
        orig = pristine / rel
        try:
            if not orig.exists() or p.stat().st_size != orig.stat().st_size \
                    or p.read_bytes() != orig.read_bytes():
                changed.append(rel)
        except Exception:
            changed.append(rel)
    return changed


def diff_sandbox(sandbox: Path, pristine: Path) -> list[ActualOp]:
    """对比沙箱执行前后（pristine=原始 resources，sandbox=跑完后的临时副本），
    返回所有检测到的行级增/删/改操作（跨全部 xlsx 文件，含未在 expected_answer
    中提及的表——用于识别"多写"的假阳性）。"""
    ops: list[ActualOp] = []
    for rel in _changed_files(sandbox, pristine):
        table_key = str(rel).replace("\\", "/")
        s_path, p_path = sandbox / rel, pristine / rel
        try:
            wb = openpyxl.load_workbook(s_path, read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)
            wb.close()
        except Exception:
            continue
        for sn in sheet_names:
            if not _is_business_sheet(sn):
                continue
            fk_after, after_rows = read_table_rows(s_path, sn)
            fk_before, before_rows = read_table_rows(p_path, sn)
            if not fk_after and not fk_before:
                continue
            pk = (fk_after or fk_before)[0]
            if not pk:
                continue
            before_idx = {r.get(pk): r for r in before_rows if r.get(pk) is not None}
            after_idx = {r.get(pk): r for r in after_rows if r.get(pk) is not None}
            # D5: 首列空行（pk=None/空）按全字段签名判断新增/删除，修复首列盲区
            before_sigs = set()
            for r in before_rows:
                before_sigs.add(tuple(sorted((k, _to_hashable(v)) for k, v in r.items())))
            after_sigs = set()
            for r in after_rows:
                after_sigs.add(tuple(sorted((k, _to_hashable(v)) for k, v in r.items())))
            for k, v in after_idx.items():
                if k not in before_idx:
                    ops.append(ActualOp(table=table_key, sheet=sn, operation="add",
                                        row_after=v))
                else:
                    bv = before_idx[k]
                    changed = {fk: av for fk, av in v.items()
                              if not _values_equal(av, bv.get(fk))}
                    if changed:
                        ops.append(ActualOp(table=table_key, sheet=sn, operation="modify",
                                            row_before=bv, row_after=v, changed_fields=changed))
            # D5: pk 空行（首列盲区）纳入 add/delete 候选
            for v in after_rows:
                pk_val = v.get(pk)
                if pk_val is not None and pk_val != "":
                    continue
                sig = tuple(sorted((k, _to_hashable(x)) for k, x in v.items()))
                if sig not in before_sigs:
                    ops.append(ActualOp(table=table_key, sheet=sn, operation="add",
                                        row_after=v))
            for v in before_rows:
                pk_val = v.get(pk)
                if pk_val is not None and pk_val != "":
                    continue
                sig = tuple(sorted((k, _to_hashable(x)) for k, x in v.items()))
                if sig not in after_sigs:
                    ops.append(ActualOp(table=table_key, sheet=sn, operation="delete",
                                        row_before=v))
            for k, v in before_idx.items():
                if k not in after_idx:
                    ops.append(ActualOp(table=table_key, sheet=sn, operation="delete",
                                        row_before=v))
    return ops


# ── expected_answer 逐条比对 ─────────────────────────────

@dataclass
class EntryResult:
    index: int
    table: str
    sheet: str
    operation: str
    status: str            # matched | partial | located_only | missing | precondition_missing
    table_sheet_hit: bool = False
    row_located: bool = False
    field_score: float = 0.0
    concrete_total: int = 0
    concrete_matched: int = 0
    note: str = ""


def _norm_table(t: str) -> str:
    return t.replace("\\", "/").strip()


def _stem_from_table(table: str) -> str:
    """从 table 路径提取 stem（文件名去扩展名）。如 'pet/pet.xlsx' → 'pet'。"""
    if not table:
        return ""
    name = table.rsplit("/", 1)[-1]
    if "." in name:
        name = name.rsplit(".", 1)[0]
    return name


def _build_eval_sheet_aliases() -> dict[tuple[str, str], set[str]]:
    """D6: 构建评估用 sheet 别名正向映射 {(stem, real_sheet): {alias, ...}}。

    用于 match_case：expected.sheet 写别名时，actual sheet 为真实名 → 视为命中。
    反转 SheetAliasConfig.mapping（alias→sheet）为 (stem,sheet)→set(aliases)，
    含通配 '*' 表别名合并到所有表。

    返回空 dict（yaml 缺失/异常）时 match_case 退化为精确匹配。
    """
    out: dict[tuple[str, str], set[str]] = {}
    try:
        from agent.excel.skill_loader import SheetAliasConfig
        cfg = SheetAliasConfig.load()
        for stem, aliases in (cfg.mapping or {}).items():
            if not isinstance(aliases, dict):
                continue
            for alias, real_sheet in aliases.items():
                if not alias or not real_sheet:
                    continue
                key = (stem, str(real_sheet))
                out.setdefault(key, set()).add(alias)
                # 真实 sheet 名本身也视为"别名"（精确匹配）
                out[key].add(str(real_sheet))
    except Exception:
        pass
    return out


def _sheet_matches(expected_sheet: str, actual_sheet: str,
                   stem: str, alias_map: dict) -> bool:
    """D6: 判断 expected sheet 是否匹配 actual sheet（含别名）。

    alias_map 为 _build_eval_sheet_aliases 产物 {(stem, real_sheet): set(aliases)}。
    策略：
      1. 精确相等 → 命中
      2. expected_sheet 是 actual_sheet 的别名（(stem, actual_sheet) 别名集含 expected）→ 命中
      3. actual_sheet 是 expected_sheet 的别名（(stem, expected_sheet) 别名集含 actual）→ 命中
    """
    if not expected_sheet or not actual_sheet:
        return expected_sheet == actual_sheet
    if expected_sheet == actual_sheet:
        return True
    if not alias_map:
        return False
    # expected 是 actual 的别名？
    if expected_sheet in alias_map.get((stem, actual_sheet), set()):
        return True
    # actual 是 expected 的别名？
    if actual_sheet in alias_map.get((stem, expected_sheet), set()):
        return True
    return False


def match_case(expected_answer: list[dict], actual_ops: list[ActualOp],
              pristine_index: dict,
              sheet_alias_map: dict | None = None) -> tuple[list[EntryResult], list[ActualOp]]:
    """把 expected_answer 与 actual_ops 逐条匹配。

    pristine_index: {(table,sheet): {pk_field_first_col: set(存在的第一列pk值)}}
                     用于判断 modify/delete 的目标行在原始数据里是否真实存在
                     （precondition_missing = 测试夹具本身没有这行数据，
                     不应算作 Agent 的定位失败）。

    sheet_alias_map: D6 sheet 别名映射，expected.sheet 写别名时匹配 actual 真实名。
                     None → 精确匹配。

    返回 (EntryResult 列表, 未被 expected 消费的 ActualOp 列表)。
    """
    results: list[EntryResult] = []
    for i, e in enumerate(expected_answer):
        table = _norm_table(e.get("table", ""))
        sheet = e.get("sheet", "")
        op = e.get("operation", "")
        row_key = e.get("row_key") or {}
        row_content = e.get("row_content") or {}

        same_ts_ops = [o for o in actual_ops
                      if o.table == table
                      and _sheet_matches(sheet, o.sheet, _stem_from_table(table), sheet_alias_map or {})
                      and o.operation == op]
        table_sheet_hit = len(same_ts_ops) > 0

        if op in ("modify", "delete"):
            precondition_ok = True
            if row_key:
                key_field, key_val = next(iter(row_key.items()))
                existing = pristine_index.get((table, sheet), {}).get(key_field, set())
                precondition_ok = any(_values_equal(x, key_val) for x in existing)
            if not precondition_ok:
                results.append(EntryResult(i, table, sheet, op, "precondition_missing",
                                           table_sheet_hit=table_sheet_hit,
                                           note="原始数据中未找到 row_key 对应的行，测试夹具与当前配表不一致"))
                continue

            cand = None
            for o in same_ts_ops:
                if o.consumed:
                    continue
                base = o.row_before if op == "modify" else o.row_before
                if all(_values_equal(base.get(k), v) for k, v in row_key.items()):
                    cand = o
                    break
            if cand is None:
                results.append(EntryResult(i, table, sheet, op, "missing",
                                           table_sheet_hit=table_sheet_hit,
                                           note="未定位到 row_key 匹配的行操作"))
                continue
            cand.consumed = True
            if op == "delete":
                results.append(EntryResult(i, table, sheet, op, "matched",
                                           table_sheet_hit=True, row_located=True,
                                           field_score=1.0, concrete_total=1, concrete_matched=1))
                continue
            # modify: 比对 row_content 每个字段
            total = matched = 0
            for k, v in row_content.items():
                total += 1
                if _is_placeholder(v):
                    if k in cand.changed_fields and cand.changed_fields.get(k) not in (None, ""):
                        matched += 1
                else:
                    if k in cand.changed_fields and _values_equal(cand.changed_fields.get(k), v):
                        matched += 1
            score = (matched / total) if total else 1.0
            status = "matched" if score >= 0.999 else ("partial" if score > 0 else "located_only")
            results.append(EntryResult(i, table, sheet, op, status,
                                       table_sheet_hit=True, row_located=True,
                                       field_score=score, concrete_total=total, concrete_matched=matched))
            continue

        # ── add ──
        concrete = {k: v for k, v in row_content.items() if not _is_placeholder(v)}
        best, best_score, best_matched, best_total = None, -1.0, 0, len(row_content)
        for o in same_ts_ops:
            if o.consumed:
                continue
            matched = sum(1 for k, v in concrete.items()
                          if _values_equal(o.row_after.get(k), v))
            score = (matched / len(concrete)) if concrete else (1.0 if o.row_after else 0.0)
            if score > best_score:
                best, best_score, best_matched = o, score, matched
        if best is None:
            results.append(EntryResult(i, table, sheet, op, "missing",
                                       table_sheet_hit=table_sheet_hit,
                                       concrete_total=len(concrete),
                                       note="未找到新增行（该表/sheet 无新增记录，或全部已被其它 expected 项占用）"))
            continue
        if concrete and best_score <= 0:
            # 该 sheet 有新增行，但字段内容完全不匹配——仍占用一个候选，判 located_only
            best.consumed = True
            results.append(EntryResult(i, table, sheet, op, "located_only",
                                       table_sheet_hit=True, row_located=True,
                                       field_score=0.0, concrete_total=len(concrete), concrete_matched=0))
            continue
        best.consumed = True
        status = "matched" if (not concrete or best_score >= 0.999) else "partial"
        results.append(EntryResult(i, table, sheet, op, status,
                                   table_sheet_hit=True, row_located=True,
                                   field_score=best_score if concrete else 1.0,
                                   concrete_total=len(concrete), concrete_matched=best_matched))

    extra_ops = [o for o in actual_ops if not o.consumed]
    return results, extra_ops


def _to_hashable(v: Any) -> Any:
    """把 _parse_cell 产出的值转为可哈希形式（list→tuple，dict→排序 items tuple）。
    pristine_index 用 set 存值做 precondition 判断，list/dict 不可哈希会抛 TypeError。"""
    if isinstance(v, list):
        return tuple(_to_hashable(x) for x in v)
    if isinstance(v, dict):
        return tuple(sorted((k, _to_hashable(x)) for k, x in v.items()))
    return v


def build_pristine_index(expected_answer: list[dict]) -> dict:
    """为本 case 涉及到的 (table, sheet) 建 {field: set(该表所有列中出现过的值)}
    索引（仅第一列 pk 列 + row_key 涉及的字段），供 precondition 判断。

    D6: sheet 别名 resolve 为真实 sheet 名读行，key 仍用 expected sheet（保持
    match_case 调用一致）。
    """
    idx: dict = {}
    seen_ts = set()
    try:
        from agent.excel.skill_loader import SheetAliasConfig
        sa_cfg = SheetAliasConfig.load()
    except Exception:
        sa_cfg = None
    for e in expected_answer:
        table = _norm_table(e.get("table", ""))
        sheet = e.get("sheet", "")
        if (table, sheet) in seen_ts:
            continue
        seen_ts.add((table, sheet))
        path = RES / table
        # D6: 别名 resolve
        real_sheet = sheet
        if sa_cfg is not None:
            stem = _stem_from_table(table)
            resolved = sa_cfg.resolve(stem, sheet)
            if resolved:
                real_sheet = resolved
        try:
            field_keys, rows = read_table_rows(path, real_sheet)
        except Exception:
            field_keys, rows = [], []
        field_vals: dict = {}
        for r in rows:
            for k, v in r.items():
                field_vals.setdefault(k, set()).add(_to_hashable(v))
        # 记录主键字段（Row2 首个非空字段名），供 _validate_fixture 判 add 冲突用
        field_vals["__pk__"] = next((fk for fk in field_keys if fk), "")
        idx[(table, sheet)] = field_vals
    return idx


def _validate_fixture(expected_answer: list[dict],
                      pristine_idx: dict) -> list[dict]:
    """D7: 校验测试夹具与 pristine 配表一致性，返回 fixture_error 列表。

    每条 error: {index, table, sheet, kind, detail}
      kind:
        - add_id_already_exists: add 用例 row_contentpristine 已存在该 ID
        - modify_delete_row_missing: modify/delete row_key 目标行 pristine 不存在

    返回空列表 = 夹具一致。
    """
    errors: list[dict] = []
    for i, e in enumerate(expected_answer):
        table = _norm_table(e.get("table", ""))
        sheet = e.get("sheet", "")
        op = e.get("operation", "")
        try:
            from agent.excel.skill_loader import SheetAliasConfig
            sa_cfg = SheetAliasConfig.load()
            stem = _stem_from_table(table)
            real_sheet = sa_cfg.resolve(stem, sheet) or sheet
        except Exception:
            real_sheet = sheet
        field_vals = pristine_idx.get((table, sheet), {})
        if op == "add":
            row_content = e.get("row_content") or {}
            # 只校验主键字段冲突。非主键字段（枚举值/外键/效果码等）在 pristine
            # 重复属正常业务数据，不应判夹具冲突——原逻辑遍历所有非 placeholder 字段，
            # 导致 entity_class / space_id / effect.key 这类共享值被误报，整条 case
            # 被踢出统计，虚假拉低 Agent 得分。
            pk_field = field_vals.get("__pk__", "")
            for k, v in row_content.items():
                if _is_placeholder(v):
                    continue
                if k != pk_field:
                    continue
                if k in field_vals and any(_values_equal(x, v) for x in field_vals[k]):
                    errors.append({
                        "index": i, "table": table, "sheet": sheet,
                        "kind": "add_id_already_exists",
                        "detail": f"add 用例 主键[{k}]={v!r} 在 pristine 已存在",
                    })
                    break
        elif op in ("modify", "delete"):
            row_key = e.get("row_key") or {}
            if row_key:
                key_field, key_val = next(iter(row_key.items()))
                existing = field_vals.get(key_field, set())
                if not any(_values_equal(x, key_val) for x in existing):
                    errors.append({
                        "index": i, "table": table, "sheet": sheet,
                        "kind": "modify_delete_row_missing",
                        "detail": f"{op} 用例 row_key[{key_field}]={key_val!r} 在 pristine 未找到目标行",
                    })
    return errors


# ── 单样例执行 ────────────────────────────────────────────

@dataclass
class CaseRunResult:
    cid: int
    input_text: str
    skill_enabled: bool
    ok: bool
    error: str
    elapsed_ms: float
    needs_confirm_used: bool
    entries: list[EntryResult]
    extra_ops: int
    n_expected: int
    n_effective: int          # 排除 precondition_missing
    locate_rate: float        # table_sheet_hit 均值
    coverage: float           # row_located 均值
    field_accuracy: float     # located 条目的 field_score 均值
    strict_pass: bool         # 所有 effective 条目 status==matched 且 extra_ops==0
    truth_ok: bool = False    # D8: 全 expected row_located + field_score=1 + 无异表 extra_ops
    extra_ops_off_table: int = 0  # D8: 异表/异语义 extra_ops 数（用于扣分）
    fixture_errors: list = field(default_factory=list)  # D7: 夹具错误列表
    fixture_error: bool = False    # D7: 有夹具错误 → 排除出统计
    # capability: error-type-distribution / llm-call-instrumentation
    error_type: str = "unknown"    # 失败时按 ErrorType 分类
    llm_stats: dict = field(default_factory=dict)  # per-case LLM 调用统计


def run_one_case(cid: int, case: dict, enable_skill: bool,
                 legacy: bool = False) -> CaseRunResult:
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"tblcase_{cid}_{'on' if enable_skill else 'off'}_"))
    sandbox = tmp_dir / "resources"
    shutil.copytree(RES, sandbox)
    service = None
    try:
        try:
            service = AgentService(resources_dir=sandbox, enable_skill=enable_skill)
            session_id = f"case{cid}_{'on' if enable_skill else 'off'}"
            t0 = time.perf_counter()
            resp = service.chat(text=case["input"], session_id=session_id, dry_run=False)
            needs_confirm_used = False
            if getattr(resp, "needs_confirm", False) and getattr(resp, "confirm_token", None):
                needs_confirm_used = True
                resp = service.chat(text=case["input"], session_id=session_id, dry_run=False,
                                    confirm_token=resp.confirm_token, confirm_cascade=True)
            elapsed_ms = (time.perf_counter() - t0) * 1000

            actual_ops = diff_sandbox(sandbox, RES)
            pristine_idx = build_pristine_index(case["expected_answer"])
            if legacy:
                # D9 legacy 模式：旧算法（无 sheet 别名/fixture 校验/truth_ok/扣分）
                entries, extra_ops_list = match_case(case["expected_answer"], actual_ops,
                                                     pristine_idx, sheet_alias_map=None)
                extra_ops = len(extra_ops_list)
                fixture_errors = []
            else:
                # D7: 跑前校验夹具一致性
                fixture_errors = _validate_fixture(case["expected_answer"], pristine_idx)
                sheet_alias_map = _build_eval_sheet_aliases()
                entries, extra_ops_list = match_case(case["expected_answer"], actual_ops,
                                                     pristine_idx, sheet_alias_map=sheet_alias_map)
                extra_ops = len(extra_ops_list)

            n_expected = len(entries)
            effective = [r for r in entries if r.status != "precondition_missing"]
            n_effective = len(effective)
            locate_rate = (sum(r.table_sheet_hit for r in effective) / n_effective) if n_effective else 0.0
            coverage = (sum(r.row_located for r in effective) / n_effective) if n_effective else 0.0
            located = [r for r in effective if r.row_located]
            field_accuracy = (sum(r.field_score for r in located) / len(located)) if located else 0.0
            if not legacy:
                # D8: 异表 extra_ops（op.table/sheet 不在 expected 集合）
                expected_ts = {(_norm_table(e.get("table", "")), e.get("sheet", ""))
                               for e in case.get("expected_answer", [])}
                extra_ops_off_table = sum(1 for o in extra_ops_list
                                          if (o.table, o.sheet) not in expected_ts)
                # D8: 异表 extra_ops 扣分（每条扣 1/n_effective）
                penalty = (extra_ops_off_table / n_effective) if n_effective else 0.0
                coverage = max(0.0, coverage - penalty)
                field_accuracy = max(0.0, field_accuracy - penalty)
                strict_pass = bool(effective) and all(r.status == "matched" for r in effective) and extra_ops == 0
                # D8 truth_ok: 所有 effective row_located + field_score=1 + 无异表 extra_ops
                truth_ok = (n_effective > 0
                            and all(r.row_located and r.field_score >= 0.999 for r in effective)
                            and extra_ops_off_table == 0)
            else:
                extra_ops_off_table = 0
                strict_pass = bool(effective) and all(r.status == "matched" for r in effective) and extra_ops == 0
                truth_ok = False

            return CaseRunResult(
                cid=cid, input_text=case["input"], skill_enabled=enable_skill,
                ok=bool(getattr(resp, "ok", False)), error=str(getattr(resp, "error", "") or ""),
                elapsed_ms=round(elapsed_ms, 1), needs_confirm_used=needs_confirm_used,
                entries=entries, extra_ops=extra_ops,
                n_expected=n_expected, n_effective=n_effective,
                locate_rate=round(locate_rate, 4), coverage=round(coverage, 4),
                field_accuracy=round(field_accuracy, 4), strict_pass=strict_pass,
                truth_ok=truth_ok, extra_ops_off_table=extra_ops_off_table,
                fixture_errors=fixture_errors, fixture_error=bool(fixture_errors),
                error_type=_classify_error(service, resp),
                llm_stats=_collect_llm_stats(service),
            )
        except Exception as e:
            # 单 case 崩（AgentService 构造失败/索引损坏/超时等）不中断整轮，
            # 记为失败结果，已跑结果不丢。
            import traceback
            tb = traceback.format_exc()
            print(f"        [error] case {cid} ({'on' if enable_skill else 'off'}) 异常: {e}")
            print(tb)
            return CaseRunResult(
                cid=cid, input_text=case["input"], skill_enabled=enable_skill,
                ok=False, error=f"{type(e).__name__}: {e}",
                elapsed_ms=0.0, needs_confirm_used=False,
                entries=[], extra_ops=0,
                n_expected=len(case.get("expected_answer", [])), n_effective=0,
                locate_rate=0.0, coverage=0.0, field_accuracy=0.0, strict_pass=False,
            )
    finally:
        try:
            if service is not None and getattr(service, "_file_watcher", None) is not None:
                service._file_watcher.stop()
        except Exception:
            pass
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _classify_error(service, resp) -> str:
    """capability: error-type-distribution —— 失败 case 按 ErrorType 分类。"""
    if resp is None or getattr(resp, "ok", False):
        return "unknown"
    try:
        from agent.excel.repair.error_classifier import classify, VerifyResult
        # service.chat 返回 AgentChatResponse，无 steps；构造 res-like 取 message
        res_like = type("R", (), {"steps": [], "message": getattr(resp, "message", "") or ""})()
        classified = classify(None, res_like, VerifyResult(), context={})
        return classified.error_type.value
    except Exception:
        return "unknown"


def _collect_llm_stats(service) -> dict:
    """capability: llm-call-instrumentation —— 读 agent._llm_counter 快照。"""
    try:
        agent = getattr(service, "agent", None)
        if agent is not None and getattr(agent, "_llm_counter", None) is not None:
            return agent._llm_counter.as_dict()
    except Exception:
        pass
    return {}


# ── 聚合 & 报告 ──────────────────────────────────────────

def aggregate(results: list[CaseRunResult]) -> dict:
    if not results:
        return {}
    # D7: 排除 fixture_error 用例（夹具与配表不一致，非 Agent 责任）
    valid = [r for r in results if not r.fixture_error]
    n = len(valid)
    if n == 0:
        # 补全所有 diff_metrics/render_report 需要的 key（全 0），避免下游 KeyError
        return {
            "n": 0, "n_total": len(results), "n_excluded": len(results),
            "fixture_error_rate": 1.0 if results else 0.0,
            "ok_rate": 0.0, "truth_ok_rate": 0.0, "error_rate": 0.0,
            "confirm_rate": 0.0, "locate_rate": 0.0, "coverage": 0.0,
            "field_accuracy": 0.0, "strict_pass_rate": 0.0,
            "avg_extra_ops": 0.0, "avg_elapsed_ms": 0.0,
            "p50_elapsed_ms": 0.0, "p95_elapsed_ms": 0.0, "total_elapsed_s": 0.0,
        }
    elapsed = [r.elapsed_ms for r in valid]
    elapsed_sorted = sorted(elapsed)

    def pct(p):
        if not elapsed_sorted:
            return 0.0
        idx = min(len(elapsed_sorted) - 1, int(len(elapsed_sorted) * p))
        return elapsed_sorted[idx]

    return {
        "n": n,
        "n_total": len(results),
        "n_excluded": len(results) - n,
        "fixture_error_rate": round(sum(r.fixture_error for r in results) / len(results), 4),
        "ok_rate": round(sum(r.ok for r in valid) / n, 4),
        "truth_ok_rate": round(sum(r.truth_ok for r in valid) / n, 4),
        "error_rate": round(sum(1 for r in valid if r.error) / n, 4),
        "confirm_rate": round(sum(r.needs_confirm_used for r in valid) / n, 4),
        "locate_rate": round(mean_or0([r.locate_rate for r in valid]), 4),
        "coverage": round(mean_or0([r.coverage for r in valid]), 4),
        "field_accuracy": round(mean_or0([r.field_accuracy for r in valid]), 4),
        "strict_pass_rate": round(sum(r.strict_pass for r in valid) / n, 4),
        "avg_extra_ops": round(mean_or0([r.extra_ops for r in valid]), 4),
        "avg_elapsed_ms": round(statistics.mean(elapsed), 1) if elapsed else 0.0,
        "p50_elapsed_ms": round(pct(0.50), 1),
        "p95_elapsed_ms": round(pct(0.95), 1),
        "total_elapsed_s": round(sum(elapsed) / 1000, 1),
        # capability: error-type-distribution
        "error_type_distribution": _aggregate_error_types(valid),
        # capability: llm-call-instrumentation
        "total_llm_calls": sum(r.llm_stats.get("total_calls", 0) for r in valid),
        "total_tokens": sum(r.llm_stats.get("total_tokens", 0) for r in valid),
        "avg_llm_calls": round(mean_or0([r.llm_stats.get("total_calls", 0) for r in valid]), 1),
        "success_path_calls": sum(r.llm_stats.get("success_path_calls", 0) for r in valid),
        "failure_path_calls": sum(r.llm_stats.get("failure_path_calls", 0) for r in valid),
    }


def _aggregate_error_types(results) -> dict:
    """聚合失败 case 的 ErrorType 分布。"""
    dist: dict[str, int] = {}
    for r in results:
        if not r.ok:
            et = r.error_type or "unknown"
            dist[et] = dist.get(et, 0) + 1
    return dist


def mean_or0(xs: list) -> float:
    return statistics.mean(xs) if xs else 0.0


def diff_metrics(off: dict, on: dict) -> dict:
    def d(key, lower_better=False):
        o, n = off.get(key, 0), on.get(key, 0)
        dv = round(n - o, 4)
        return -dv if lower_better else dv
    return {
        "locate_rate_delta": d("locate_rate"),
        "coverage_delta": d("coverage"),
        "field_accuracy_delta": d("field_accuracy"),
        "strict_pass_delta": d("strict_pass_rate"),
        "ok_rate_delta": d("ok_rate"),
        "truth_ok_delta": d("truth_ok_rate"),
        "extra_ops_delta": d("avg_extra_ops", lower_better=True),
        "elapsed_delta_ms": d("avg_elapsed_ms", lower_better=True),
    }


_STATUS_LABEL = {
    "matched": "✅ 完全匹配", "partial": "🟡 部分匹配", "located_only": "🟠 定位到但字段不符",
    "missing": "❌ 未产生", "precondition_missing": "⚪ 夹具缺失(跳过)",
}


def render_case_detail(off: CaseRunResult, on: CaseRunResult, case_text: str) -> str:
    lines = [f"### 样例 {off.cid}: {case_text}", ""]
    lines.append("| 指标 | skill=off | skill=on |")
    lines.append("|---|---|---|")
    lines.append(f"| 响应ok | {off.ok} | {on.ok} |")
    lines.append(f"| 定位功能(table/sheet命中率) | {off.locate_rate:.2f} | {on.locate_rate:.2f} |")
    lines.append(f"| 覆盖度(行操作产出率) | {off.coverage:.2f} | {on.coverage:.2f} |")
    lines.append(f"| 精准程度(字段值正确率) | {off.field_accuracy:.2f} | {on.field_accuracy:.2f} |")
    lines.append(f"| 严格通过 | {off.strict_pass} | {on.strict_pass} |")
    lines.append(f"| 多余写入(误写行数) | {off.extra_ops} | {on.extra_ops} |")
    lines.append(f"| 耗时(ms) | {off.elapsed_ms:.0f} | {on.elapsed_ms:.0f} |")
    if off.error or on.error:
        lines.append(f"| 错误 | {off.error or '-'} | {on.error or '-'} |")
    lines.append("")
    lines.append("expected_answer 逐条判定（skill=off → skill=on）：")
    lines.append("")
    lines.append("| # | table.sheet | op | off | on |")
    lines.append("|---|---|---|---|---|")
    for i in range(off.n_expected):
        eo = off.entries[i] if i < len(off.entries) else None
        en = on.entries[i] if i < len(on.entries) else None
        ts = f"{eo.table}.{eo.sheet}" if eo else ""
        opv = eo.operation if eo else ""
        lo = _STATUS_LABEL.get(eo.status, eo.status) if eo else "-"
        ln = _STATUS_LABEL.get(en.status, en.status) if en else "-"
        eo_score = f"{eo.field_score:.2f}" if eo else "-"
        en_score = f"{en.field_score:.2f}" if en else "-"
        lines.append(f"| {i+1} | {ts} | {opv} | {lo} ({eo_score}) | {ln} ({en_score}) |")
    lines.append("")
    return "\n".join(lines)


def render_report(cases: list[dict], off_results: list[CaseRunResult],
                  on_results: list[CaseRunResult]) -> str:
    off = aggregate(off_results)
    on = aggregate(on_results)
    diff = diff_metrics(off, on)
    lines = [
        "# 后端 Agent skill on/off A/B 测试报告",
        "",
        f"- 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 样例来源: table_operation_test_cases.json（{off.get('n', 0)} 条）",
        "- 执行方式: 进程内直接调用 AgentService（真实 codemaker serve LLM），"
        "每条样例在 resources/ 临时沙箱副本内真实执行增删改，跑前/跑后 xlsx 差异作为 ground truth",
        "- skill=off: TableAgent(enable_skill=False)，列/行定位仅靠原始表头，不挂列别名/行规则/反模式/短形式",
        "- skill=on : TableAgent(enable_skill=True)，正常加载 server/agent/excel/skills/ 全部 skill",
        "",
        "## 一、总体指标对比",
        "",
        "| 指标 | 说明 | skill=off | skill=on | 变化 |",
        "|---|---|---|---|---|",
        f"| 定位功能 | 命中正确 table+sheet+操作类型的比例 | {off['locate_rate']:.4f} | {on['locate_rate']:.4f} | {diff['locate_rate_delta']:+.4f} |",
        f"| 覆盖度 | expected 行操作中真正被产出的比例（含级联多表） | {off['coverage']:.4f} | {on['coverage']:.4f} | {diff['coverage_delta']:+.4f} |",
        f"| 精准程度 | 被定位到的行里，字段值完全正确的比例 | {off['field_accuracy']:.4f} | {on['field_accuracy']:.4f} | {diff['field_accuracy_delta']:+.4f} |",
        f"| 严格通过率 | 整条样例 expected_answer 100%命中且无多余写入 | {off['strict_pass_rate']:.4f} | {on['strict_pass_rate']:.4f} | {diff['strict_pass_delta']:+.4f} |",
        f"| 响应ok率 | Agent 自报告执行成功比例 | {off['ok_rate']:.4f} | {on['ok_rate']:.4f} | {diff['ok_rate_delta']:+.4f} |",
        f"| 平均多余写入 | 未被 expected 认领的行改动数（越低越好） | {off['avg_extra_ops']:.4f} | {on['avg_extra_ops']:.4f} | {diff['extra_ops_delta']:+.4f} |",
        f"| 平均耗时(ms) | 单条指令端到端耗时（含二次确认续传） | {off['avg_elapsed_ms']:.1f} | {on['avg_elapsed_ms']:.1f} | {diff['elapsed_delta_ms']:+.1f} |",
        f"| P50耗时(ms) | | {off['p50_elapsed_ms']:.1f} | {on['p50_elapsed_ms']:.1f} | |",
        f"| P95耗时(ms) | | {off['p95_elapsed_ms']:.1f} | {on['p95_elapsed_ms']:.1f} | |",
        f"| 需二次确认比例 | 触发级联删除等确认流程的比例 | {off['confirm_rate']:.4f} | {on['confirm_rate']:.4f} | |",
        f"| 总耗时(s) | 本轮全部样例累计耗时 | {off['total_elapsed_s']:.1f} | {on['total_elapsed_s']:.1f} | |",
        "",
        "## 二、按操作类型细分（skill=on）",
        "",
    ]

    def by_op_metrics(results: list[CaseRunResult], op: str) -> dict:
        entries = [en for r in results for en in r.entries
                  if en.operation == op and en.status != "precondition_missing"]
        n = len(entries)
        if not n:
            return {"n": 0, "locate": 0.0, "coverage": 0.0, "field_acc": 0.0}
        return {
            "n": n,
            "locate": round(sum(e.table_sheet_hit for e in entries) / n, 4),
            "coverage": round(sum(e.row_located for e in entries) / n, 4),
            "field_acc": round(mean_or0([e.field_score for e in entries if e.row_located]), 4),
        }

    lines += ["| 操作类型 | n | 定位率 | 覆盖率 | 字段精准度 | (对比 off) |",
             "|---|---|---|---|---|---|"]
    for op in ("add", "modify", "delete"):
        m_on = by_op_metrics(on_results, op)
        m_off = by_op_metrics(off_results, op)
        lines.append(f"| {op} | {m_on['n']} | {m_on['locate']:.4f} | {m_on['coverage']:.4f} | "
                     f"{m_on['field_acc']:.4f} | off: locate={m_off['locate']:.4f} "
                     f"cov={m_off['coverage']:.4f} acc={m_off['field_acc']:.4f} |")

    lines += ["", "## 三、每个样例详细运行情况", ""]
    for i, case in enumerate(cases):
        off_r = off_results[i]
        on_r = on_results[i]
        lines.append(render_case_detail(off_r, on_r, case["input"]))

    worst_on = sorted(on_results, key=lambda r: (r.coverage + r.field_accuracy))[:10]
    lines += ["## 四、skill=on 表现最差的样例（覆盖度+精准度最低 Top10）", "",
             "| cid | input | 覆盖度 | 精准度 | 定位率 | 错误 |",
             "|---|---|---|---|---|---|"]
    for r in worst_on:
        txt = r.input_text[:40].replace("|", "/")
        lines.append(f"| {r.cid} | {txt} | {r.coverage:.2f} | {r.field_accuracy:.2f} | "
                     f"{r.locate_rate:.2f} | {r.error[:60] if r.error else '-'} |")

    # D7: 夹具错误清单
    fixture_errors = [r for r in on_results if r.fixture_error]
    lines += ["", f"## 五、夹具错误清单（D7，共 {len(fixture_errors)} 条，已排除出统计）", "",
             "| cid | input | 错误类型 | 详情 |",
             "|---|---|---|---|"]
    for r in fixture_errors:
        txt = r.input_text[:40].replace("|", "/")
        for fe in r.fixture_errors:
            lines.append(f"| {r.cid} | {txt} | {fe.get('kind', '')} | {fe.get('detail', '')[:80]} |")
    # 重编号后续段（原五→六）
    lines += [
        "",
        "## 六、结论总结",
        "",
        f"- skill 挂载后定位功能变化: {diff['locate_rate_delta']:+.4f}"
        f"（{'提升' if diff['locate_rate_delta'] > 0 else ('下降' if diff['locate_rate_delta'] < 0 else '持平')}）",
        f"- skill 挂载后覆盖度变化: {diff['coverage_delta']:+.4f}"
        f"（{'提升' if diff['coverage_delta'] > 0 else ('下降' if diff['coverage_delta'] < 0 else '持平')}）",
        f"- skill 挂载后精准程度变化: {diff['field_accuracy_delta']:+.4f}"
        f"（{'提升' if diff['field_accuracy_delta'] > 0 else ('下降' if diff['field_accuracy_delta'] < 0 else '持平')}）",
        f"- skill 挂载后严格通过率变化: {diff['strict_pass_delta']:+.4f}",
        f"- skill 挂载后耗时变化: {diff['elapsed_delta_ms']:+.1f}ms"
        f"（{'变快' if diff['elapsed_delta_ms'] > 0 else ('变慢' if diff['elapsed_delta_ms'] < 0 else '持平')}，"
        "正数=on比off快，因为定义为 off-on）",
        f"- 综合判定: skill 增强 "
        f"{'有效' if (diff['coverage_delta'] > 0 or diff['field_accuracy_delta'] > 0 or diff['strict_pass_delta'] > 0) else '效果不明显'}",
        "",
        "注意事项：",
        "- ⚪ 夹具缺失(跳过) 表示 expected_answer 引用的 row_key 在当前 resources/ 真实数据中不存在"
        "（测试夹具与配表现状不一致），该条不计入定位/覆盖/精准分母，不代表 Agent 缺陷。",
        "- 每条样例都在独立临时沙箱执行，互不影响；测试完成后沙箱已删除，不会污染真实 resources/。",
        "- 「多余写入」統计的是全量 resources/ 目录里未被 expected_answer 认领的行改动"
        "（含 expected 未提及的表），用于发现 Agent 误改/过度级联的副作用。",
    ]
    return "\n".join(lines)


# ── 主流程 ──────────────────────────────────────────────

def load_cases(path: Path) -> list[dict]:
    return json.loads(path.read_text(encoding="utf-8"))


def _ser(r: CaseRunResult) -> dict:
    d = r.__dict__.copy()
    d["entries"] = [e.__dict__ for e in r.entries]
    return d


def _save_results(out_dir: Path, run_cases: list[dict],
                  off_results: list[CaseRunResult],
                  on_results: list[CaseRunResult],
                  meta: dict) -> None:
    """增量/最终保存：JSON（含已跑结果 + 进度 meta）+ MD（基于已跑子集渲染）。
    每跑完一个 case 调用一次，中断后落盘文件即为最新部分结果。"""
    try:
        payload = {
            "meta": meta,
            "off": aggregate(off_results), "on": aggregate(on_results),
            "off_results": [_ser(r) for r in off_results],
            "on_results": [_ser(r) for r in on_results],
        }
        (out_dir / "table_case_eval_latest.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8")
        # MD 用已跑子集渲染（cases 与 off_results 同序对齐）
        report_md = render_report(run_cases, off_results, on_results)
        (out_dir / "table_case_eval_latest.md").write_text(report_md, encoding="utf-8")
    except Exception as e:
        print(f"  [warn] 增量保存失败（不影响继续跑）: {e}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--quick", type=int, default=0, help="只跑前 N 条冒烟（与 --start 同时用时仅当 start=1 生效）")
    ap.add_argument("--start", type=int, default=1, help="从第 N 条开始（含，1-based）")
    ap.add_argument("--end", type=int, default=0, help="到第 N 条结束（含），0=跑到最后")
    ap.add_argument("--cases-file", type=str, default=str(DEFAULT_CASES_FILE))
    ap.add_argument("--out", type=str, default=str(REPORT_DIR))
    ap.add_argument("--legacy-eval", action="store_true",
                    help="D9: 用旧评估算法（无 sheet 别名/fixture 校验/truth_ok/extra_ops 扣分），供对照")
    ap.add_argument("--induce", action="store_true",
                    help="#6: skill=on 轮跑完后用 LLM 从失败 trace 归纳反模式候选（也可用 EVAL_INDUCE_AP=1）")
    args = ap.parse_args()

    # D11: 标记 eval 运行中，阻止 skill_updater._run_mini_regression 嵌套跑 case
    # （否则 promote_with_guard 同步触发 mini 回归 → run_one_case → AgentService + LLM
    #  → 30 case × 双轮 × LLM 递归爆炸卡死）
    os.environ["TABLE_CASE_EVAL_RUNNING"] = "1"

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    legacy = bool(args.legacy_eval)
    if legacy:
        print("⚠ D9 legacy 模式：使用旧评估算法（无 sheet 别名/fixture 校验/truth_ok/扣分），仅供对照")

    cases_all = load_cases(Path(args.cases_file))
    total_all = len(cases_all)
    start = max(1, min(args.start, total_all))
    end = args.end if args.end > 0 else total_all
    end = min(end, total_all)
    # --quick 兼容：仅当未显式跳过起点时才裁到前 N
    if args.quick and start == 1:
        end = min(end, args.quick)

    run_cases: list[dict] = []
    off_results: list[CaseRunResult] = []
    on_results: list[CaseRunResult] = []
    print(f"共 {total_all} 条样例，本轮跑 [{start}-{end}]，"
          f"开始 skill=off / skill=on 双轮真实 Agent 执行...")

    for i, case in enumerate(cases_all, start=1):
        if i < start or i > end:
            continue
        print(f"[{i}/{total_all}] (off) {case['input']}")
        t0 = time.time()
        r_off = run_one_case(i, case, enable_skill=False, legacy=legacy)
        print(f"        ok={r_off.ok} cov={r_off.coverage:.2f} acc={r_off.field_accuracy:.2f} "
             f"{r_off.elapsed_ms:.0f}ms (总耗时 {time.time()-t0:.1f}s)")
        off_results.append(r_off)

        print(f"[{i}/{total_all}] (on)  {case['input'][:50]}")
        t0 = time.time()
        r_on = run_one_case(i, case, enable_skill=True, legacy=legacy)
        print(f"        ok={r_on.ok} cov={r_on.coverage:.2f} acc={r_on.field_accuracy:.2f} "
             f"{r_on.elapsed_ms:.0f}ms (总耗时 {time.time()-t0:.1f}s)")
        on_results.append(r_on)
        run_cases.append(case)

        # 增量保存：每跑完一个 case 立即落盘，中断不丢
        done = i - start + 1
        span = end - start + 1
        _save_results(out_dir, run_cases, off_results, on_results, {
            "start": start, "end": end, "done": done, "span": span,
            "total_all": total_all, "last_cid": i,
        })
        print(f"        [已保存 {done}/{span} → table_case_eval_latest.json/.md]")

    print("\n" + "=" * 60)
    print(f"报告已写: {out_dir / 'table_case_eval_latest.md'}")
    print(f"原始数据: {out_dir / 'table_case_eval_latest.json'}")

    # 最终控制台汇总（跑完直接看，不必打开文件）
    if on_results and off_results:
        off_agg = aggregate(off_results)
        on_agg = aggregate(on_results)
        print(f"\n本轮汇总（cases {start}-{end}，共跑 {len(off_results)} 条）：")
        print(f"  skill=off: 定位={off_agg['locate_rate']:.4f} 覆盖={off_agg['coverage']:.4f} "
              f"精准={off_agg['field_accuracy']:.4f} 严格通过={off_agg['strict_pass_rate']:.4f} "
              f"ok率={off_agg['ok_rate']:.4f}")
        print(f"  skill=on : 定位={on_agg['locate_rate']:.4f} 覆盖={on_agg['coverage']:.4f} "
              f"精准={on_agg['field_accuracy']:.4f} 严格通过={on_agg['strict_pass_rate']:.4f} "
              f"ok率={on_agg['ok_rate']:.4f}")
        print(f"  耗时: off avg={off_agg['avg_elapsed_ms']:.0f}ms  on avg={on_agg['avg_elapsed_ms']:.0f}ms  "
              f"on 总计={on_agg['total_elapsed_s']:.1f}s")
        d = diff_metrics(off_agg, on_agg)
        print(f"  变化(正=on更好): 定位{d['locate_rate_delta']:+.4f} 覆盖{d['coverage_delta']:+.4f} "
              f"精准{d['field_accuracy_delta']:+.4f} 严格通过{d['strict_pass_delta']:+.4f}")
        # capability: error-type-distribution / llm-call-instrumentation
        print(f"  ErrorType 分布(on): {on_agg.get('error_type_distribution', {})}")
        print(f"  LLM 调用(on): total={on_agg.get('total_llm_calls',0)} "
              f"avg={on_agg.get('avg_llm_calls',0)} tokens={on_agg.get('total_tokens',0)}")
    print("=" * 60)

    # capability: eval-baseline-management —— 归档本次 run
    try:
        from tests.eval_baseline import archive_run, make_run_id
        tag = os.environ.get("EVAL_BASELINE_TAG", "")
        rid = make_run_id(tag=tag or None)
        json_data = json.loads((out_dir / "table_case_eval_latest.json").read_text(encoding="utf-8"))
        if isinstance(json_data, dict) and "summary" not in json_data:
            json_data = {"summary": on_agg if on_results else {}, **json_data}
        else:
            json_data["summary"] = on_agg if on_results else {}
        md = (out_dir / "table_case_eval_latest.md").read_text(encoding="utf-8") if (out_dir / "table_case_eval_latest.md").exists() else ""
        archive_run("table_case_eval", json_data, md, run_id=rid)
        print(f"  归档: reports/archive/table_case_eval_{rid}.json")
    except Exception as e:
        print(f"  [warn] 归档失败（不阻断）: {e}")

    # #6: AI 反模式归纳（可选，--induce 或 EVAL_INDUCE_AP=1 开启）
    induce_on = args.induce or bool(os.environ.get("EVAL_INDUCE_AP"))
    if induce_on and on_results:
        print("-" * 60)
        print("#6 AI 反模式归纳...")
        try:
            from agent.codemaker_client import CodemakerClient
            from agent.excel.core.step_ai_enhancer import StepAIEnhancer
            from agent.excel.skill_updater import get_skill_updater
            failed_traces = []
            for r in on_results:
                if (not r.ok) or r.error_type:
                    ents = r.entries or []
                    failed_traces.append({
                        "input": r.input_text or "",
                        "error_type": r.error_type or "",
                        "error_detail": (r.error or "")[:120],
                        "entries_summary": "; ".join(
                            (getattr(e, "note", "") or "")[:60] for e in ents[:3]
                        ),
                    })
            print(f"  失败 trace: {len(failed_traces)} 条")
            if failed_traces:
                client = CodemakerClient()
                enhancer = StepAIEnhancer(client, directory=str(RES))
                produced = get_skill_updater().induce_anti_patterns(failed_traces, enhancer)
                print(f"  归纳产出: {len(produced)} 条反模式候选 (status=pending_review)")
                for ap in produced:
                    print(f"    - [{ap.type}] trigger={ap.trigger_pattern} → {ap.action} | {(ap.rationale or '')[:50]}")
            else:
                print("  无失败 trace，跳过归纳")
        except Exception as e:
            print(f"  [warn] #6 AI 反模式归纳失败（不阻断 eval）: {e}")
    print("=" * 60)


if __name__ == "__main__":
    main()
