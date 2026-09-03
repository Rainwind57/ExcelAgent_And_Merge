"""multi-table-orchestration 单测（capability: cross-table-transaction + relation-graph-consumption）。

验证：
- D2 跨表模式检测信号（detect_cross_table_action）
- D4 跨表事务：全成功提交；中间失败标记 dirty_data + failed_tables
- D5 RelationGraph.get_related_tables + 运行时 merge

§去硬模板：原 D1/D3 段测试 cross_table_splitter 的 11 个硬编码 _build_*_intents
模板函数已随生产代码整体移除，相关测试一并删除。detect_cross_table_action 本身
（跨表模式检测信号，非模板生成）保留，其测试继续保留于 D2 段。
"""
from __future__ import annotations

import os
import sys
import types
from pathlib import Path

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.excel.cross_table_splitter import detect_cross_table_action
from agent.excel.agent import AgentResult
from agent.nl_parser import NLIntent


# ── D2 跨表模式检测信号 ──────────────────────────────────

def test_detect_npc_teleport():
    """D2: 传送 NPC 命中 npc_teleport。"""
    assert detect_cross_table_action("新增NPC传送使者传送到10001") == "npc_teleport"


def test_detect_npc_combat():
    """D2: 战斗 NPC 命中 npc_combat。"""
    assert detect_cross_table_action("新增NPC擂台挑战者战斗ID为102") == "npc_combat"


def test_detect_npc_reward():
    """D2: 奖励 NPC 命中 npc_reward。"""
    assert detect_cross_table_action("新增NPC奖励使者获得reward_id 10006") == "npc_reward"


# ── D4 跨表事务 ──────────────────────────────────────────

def test_transaction_all_success():
    """D4: 全成功提交（所有 res.ok=True）。"""
    from agent.excel.operation_orchestrator import OperationOrchestrator

    def fake_run_single(intent, confirm_token, session_id):
        res = AgentResult(intent=intent)
        res.add("write", True, "成功")
        return res

    orch = OperationOrchestrator(fake_run_single)
    intents = [NLIntent(action="add", table_hint="t1", raw="t1"),
               NLIntent(action="add", table_hint="t2", raw="t2")]
    results = orch.run(intents)
    assert len(results) == 2
    assert all(r.ok is True for r in results)


def test_transaction_middle_failure_marks_dirty():
    """D4: 中间失败 → 标记 dirty_data + failed_tables + 中断后续。"""
    from agent.excel.operation_orchestrator import OperationOrchestrator

    call_count = [0]

    def fake_run_single(intent, confirm_token, session_id):
        call_count[0] += 1
        res = AgentResult(intent=intent)
        if call_count[0] == 2:  # 第2条失败
            res.add("write", False, "写后验证失败")
        else:
            res.add("write", True, "成功")
        return res

    orch = OperationOrchestrator(fake_run_single)
    intents = [NLIntent(action="add", table_hint="t1", raw="t1"),
               NLIntent(action="add", table_hint="t2", raw="t2"),
               NLIntent(action="add", table_hint="t3", raw="t3")]
    results = orch.run(intents)
    # 第3条被跳过（事务中断）
    assert call_count[0] == 2, "第3条应被跳过"
    # 失败结果有 dirty_data + failed_tables
    failed = [r for r in results if r.ok is not True]
    assert len(failed) >= 1
    assert any(r.dirty_data for r in failed)
    assert any(r.failed_tables for r in failed)


# ── D5 RelationGraph ─────────────────────────────────────

def test_relation_graph_npc_three_tables():
    """D5: NPC 三表关系存在（entity_prefab↔interaction↔spawn_world_entity）。"""
    from agent.excel.table_relations import RelationGraph
    g = RelationGraph.load()
    related = g.get_related_tables("entity_prefab")
    related_stems = [s for s, _ in related]
    assert "interaction" in related_stems
    assert "spawn_world_entity" in related_stems


def test_get_related_tables_returns_relation_type():
    """D5: get_related_tables 返回 (related_stem, relation_type)。"""
    from agent.excel.table_relations import RelationGraph
    g = RelationGraph.load()
    related = g.get_related_tables("entity_prefab")
    for stem, rtype in related:
        assert isinstance(stem, str)
        assert isinstance(rtype, str)
        assert rtype  # 非空


def test_relation_graph_load_merges_runtime():
    """D5: load() merge 运行时 table_relations.runtime.json（静态优先）。"""
    from agent.excel.table_relations import RelationGraph
    g = RelationGraph.load()
    # 静态 6 条 + 运行时 merge（若有 runtime 文件，静态优先去重）
    assert len(g.relations) >= 6  # 至少静态 6 条


def test_get_related_tables_empty_for_unknown_stem():
    """D5: 未知 stem 返回空列表。"""
    from agent.excel.table_relations import RelationGraph
    g = RelationGraph.load()
    assert g.get_related_tables("nonexistent_table") == []


# ── D5.2 schema_infer regex 扩展 + 落盘逻辑 ───────────────

def test_infer_ref_table_extended_patterns():
    """D5.2: _infer_ref_table 覆盖实际列名格式（含空格/大写 ID/数字冒号前缀）。"""
    from agent.excel.schema_infer import _infer_ref_table
    # 原格式兼容
    assert _infer_ref_table("pet_id") == "pet"
    assert _infer_ref_table("宠物id") == "宠物"
    assert _infer_ref_table("物品编号") == "物品"
    # D5 扩展：大写 ID + 空格
    assert _infer_ref_table("实体Prefab ID") == "实体Prefab"
    assert _infer_ref_table("对话ID") == "对话"
    assert _infer_ref_table("战斗ID") == "战斗"
    # D5 扩展：数字冒号前缀（raw 兜底）
    assert _infer_ref_table("3006", "3006:对话ID") == "对话"
    assert _infer_ref_table("3004", "3004: spawn ID") == "spawn"


def test_generate_table_relations_format():
    """D5.2: generate_table_relations 输出 RelationGraph 兼容格式。"""
    from agent.excel.schema_infer import generate_table_relations
    # 空表输入 → 空 relations
    result = generate_table_relations({})
    assert "relations" in result
    assert isinstance(result["relations"], list)


def test_regenerate_skills_merges_table_relations(tmp_path, monkeypatch):
    """D5.2: regenerate_skills 自动发现关系 merge 到 table_relations.json（不覆盖既有）。"""
    from agent.excel import schema_infer
    from agent.excel.table_relations import RelationGraph

    # 建最小 xlsx（pet + pet_evolve 外键关系）
    import openpyxl
    res = tmp_path / "res"
    pet_dir = res / "pet"; pet_dir.mkdir(parents=True)
    wb = openpyxl.Workbook()
    s = wb.active; s.title = "Pet"
    s.cell(1, 1, "灵兽id"); s.cell(2, 1, "pet_id:int")
    s.cell(5, 1, 1)
    wb.save(pet_dir / "pet.xlsx")

    ev_dir = res / "pet_evolve"; ev_dir.mkdir(parents=True)
    wb2 = openpyxl.Workbook()
    s2 = wb2.active; s2.title = "PetEvolveData"
    s2.cell(1, 1, "宠物id"); s2.cell(2, 1, "pet_id:int")
    s2.cell(5, 1, 1)
    wb2.save(ev_dir / "pet_evolve.xlsx")

    # 备份原 table_relations.json
    from agent.excel.table_relations import _relations_path
    orig_path = _relations_path()
    orig_content = orig_path.read_text(encoding="utf-8") if orig_path.exists() else None

    try:
        # skills_dir 隔离到 tmp：避免把 tmp 表内容写进生产 L1_derived yamls（污染真实 skills）
        schema_infer.regenerate_skills(res, skills_dir=tmp_path / "skills")
        # 重新加载，确认 pet_evolve→pet 关系被 merge
        g = RelationGraph.load()
        pet_rels = [r for r in g.relations if "pet.xlsx" in r.to_path]
        assert len(pet_rels) >= 1, "应发现 pet_evolve→pet 外键关系"
    finally:
        # 恢复原 table_relations.json
        if orig_content is not None:
            orig_path.write_text(orig_content, encoding="utf-8")
