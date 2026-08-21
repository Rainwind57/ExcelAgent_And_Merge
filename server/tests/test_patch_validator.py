"""方法 C PATCH_CONFIG 守门单测。

验证 validate_capped_workbook 5 坑硬规则（ca-overview §3.3）：
  ① sheet 名不在 trunk（坑1）② PATCH_CONFIG A 列重复登记（坑5）
  ③ PATCH_CONFIG 缺失（坑3）④ _capped 有 CONFIG sheet（坑4）
  ⑤ B 列融合方式非法（§3.2）
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from engine.patch_validator import validate_capped_workbook, Violation


def _make_capped(p: Path, patch_config=None, extra_sheets=None, has_config=False):
    """构造 _capped.xlsx 样例。

    patch_config: [(sheet_name, fusion_mode)] PATCH_CONFIG 行；None=无 PATCH_CONFIG
    extra_sheets: {sheet_name: [[row...],...]} 数据 sheet
    has_config: True 加 CONFIG sheet（坑4）
    """
    wb = Workbook()
    default = wb.active
    if has_config:
        default.title = "CONFIG"
        default.cell(1, 1, "表名")
        if patch_config is not None:
            ws_pc = wb.create_sheet("PATCH_CONFIG")
            for i, (s, m) in enumerate(patch_config, 1):
                ws_pc.cell(i, 1, s)
                ws_pc.cell(i, 2, m)
    elif patch_config is not None:
        default.title = "PATCH_CONFIG"
        for i, (s, m) in enumerate(patch_config, 1):
            default.cell(i, 1, s)
            default.cell(i, 2, m)
    else:
        default.title = "DATA"
    if extra_sheets:
        for sname, rows in extra_sheets.items():
            ws = wb.create_sheet(sname)
            for r, row in enumerate(rows, 1):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, val)
    wb.save(p)


def _rules(violations):
    return {v.rule for v in violations}


def test_compliant_capped(tmp_path):
    """合规 _capped：PATCH_CONFIG + 数据 sheet，无 CONFIG，sheet 名在 trunk。"""
    p = tmp_path / "ok.xlsx"
    _make_capped(p, patch_config=[("Item", "PATCH_GEN"), ("Npc", "SHEET_GEN")],
                 extra_sheets={"Item": [["id", "name"], [1, "a"]], "Npc": [["id"], [1]]})
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item", "Npc", "Other"])
    assert vs == [], f"合规文件不应有违规: {[v.to_dict() for v in vs]}"


def test_pit4_has_config_sheet(tmp_path):
    """坑4：_capped 有 CONFIG sheet → violation rule=4。"""
    p = tmp_path / "bad4.xlsx"
    _make_capped(p, patch_config=[("Item", "PATCH_GEN")], has_config=True)
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item"])
    assert 4 in _rules(vs), f"坑4 应命中: {[v.to_dict() for v in vs]}"
    assert any("CONFIG" in v.message for v in vs)


def test_pit3_no_patch_config(tmp_path):
    """坑3：_capped 无 PATCH_CONFIG sheet → violation rule=3。"""
    p = tmp_path / "bad3.xlsx"
    _make_capped(p, patch_config=None, extra_sheets={"Item": [[1]]})
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item"])
    assert 3 in _rules(vs), f"坑3 应命中: {[v.to_dict() for v in vs]}"
    assert any("PATCH_CONFIG" in v.message for v in vs)


def test_pit1_sheet_not_in_trunk(tmp_path):
    """坑1：PATCH_CONFIG A 列 sheet 不在 trunk → violation rule=1。"""
    p = tmp_path / "bad1.xlsx"
    _make_capped(p, patch_config=[("Item", "PATCH_GEN"), ("UnknownSheet", "PATCH_GEN")])
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item", "Npc"])
    assert 1 in _rules(vs), f"坑1 应命中: {[v.to_dict() for v in vs]}"
    assert any("UnknownSheet" in v.message for v in vs)


def test_pit5_duplicate_sheet_in_patch_config(tmp_path):
    """坑5：PATCH_CONFIG A 列重复登记 → violation rule=2（先写者优先）。"""
    p = tmp_path / "bad5.xlsx"
    _make_capped(p, patch_config=[("Item", "PATCH_GEN"), ("Item", "SHEET_GEN")])
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item"])
    assert 2 in _rules(vs), f"坑5 应命中: {[v.to_dict() for v in vs]}"
    assert any("重复" in v.message for v in vs)


def test_invalid_fusion_mode(tmp_path):
    """§3.2：B 列融合方式非法 → violation rule=5。"""
    p = tmp_path / "bad_mode.xlsx"
    _make_capped(p, patch_config=[("Item", "INVALID_MODE")])
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item"])
    assert 5 in _rules(vs), f"非法融合方式应命中 rule=5: {[v.to_dict() for v in vs]}"
    assert any("INVALID_MODE" in v.message for v in vs)


def test_no_trunk_sheets_skip_rule1(tmp_path):
    """trunk_sheets 空 → 跳过规则①（无 trunk 参照），不报 sheet 不在 trunk。"""
    p = tmp_path / "notrunk.xlsx"
    _make_capped(p, patch_config=[("WhateverSheet", "PATCH_GEN")])
    vs = validate_capped_workbook(p, trunk_sheet_names=None)
    assert 1 not in _rules(vs), "无 trunk 参照应跳过规则①"


def test_violation_to_dict(tmp_path):
    """Violation.to_dict 可 JSON 序列化（路由返回用）。"""
    import json
    p = tmp_path / "bad4.xlsx"
    _make_capped(p, patch_config=[("Item", "PATCH_GEN")], has_config=True)
    vs = validate_capped_workbook(p, trunk_sheet_names=["Item"])
    d = vs[0].to_dict()
    s = json.dumps(d, ensure_ascii=False)
    assert "patch_config" in s
    assert "hold" in s
