"""O15 方法 A/D 升级接 pre_commit_hold 通道（AD1/AD2）+ A2 agent 消费 + D4 audit 单测。

覆盖：
- AD1：CODEMAKER_FORMULA_GATE=hold + needs_manual_fix=True → CLICallResult.hold_events
  含 kind=formula_loss + record_hold_audit 留痕 pre_commit_hold。
- AD2：批注二次回写后仍丢（still_lost>0）→ CLICallResult.hold_events 含 kind=comment_loss
  + record_hold_audit 留痕 + D4 comment_replay_partial audit。
- A2：agent 层 _run_set/_run_add 消费 hold_events → res.failures 追加 #40 软失败 +
  经 _agent_subtask_sink 推 pre_commit_hold SSE 事件。

真实公式重算/批注丢失流程由 test_formula_cache.py / test_comment_guard.py 覆盖，
此处聚焦 hold 事件构造 + agent 消费逻辑（mock 绕过真实 IO）。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment

from agent.excel.cli.cli_interface import StubCodeMakerCLI, CLICallResult
from agent.excel.core.backup_audit import BackupAuditor
from agent.excel.core.agent import TableAgent
from agent.excel.parser.nl_parser import NLIntent


class _FakeResult:
    def __init__(self, needs: bool, msg: str = "公式缓存丢失"):
        self.needs_manual_fix = needs
        self.message = msg


def _make_auditor(tmp_path: Path) -> BackupAuditor:
    return BackupAuditor(
        workspace=tmp_path,
        backups_dir=tmp_path / "backups",
        audit_log_path=tmp_path / "audit_log.jsonl",
    )


def _read_audit_ops(audit_path: Path) -> list[dict]:
    if not audit_path.exists():
        return []
    import json
    ops: list[dict] = []
    with open(audit_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                ops.append(json.loads(line))
            except Exception:
                continue
    return ops


def _mock_formula_loss(cli, monkeypatch, needs: bool = True):
    monkeypatch.setattr(cli._formula_validator, "snapshot_before",
                        lambda path: {"_mock": 1})
    monkeypatch.setattr(cli._formula_validator, "validate_and_fix",
                        lambda path, before: _FakeResult(needs))


def _make_formula_workbook(p: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(1, 1, "a")
    ws.cell(5, 1, 1)
    ws.cell(5, 2, "=A5+1")
    wb.save(p)


def _make_agent():
    """轻量 TableAgent（绕过重 __init__）。"""
    ag = object.__new__(TableAgent)
    ag._agent_subtask_sink = None
    ag._auditor = None
    return ag


def _make_res():
    """轻量 AgentResult（只挂 failures + final + add）。"""
    from agent.excel.core.agent import AgentResult
    from agent.excel.parser.nl_parser import NLIntent
    r = AgentResult(intent=NLIntent(action="set", raw="test"))
    r.failures = []
    return r


class TestAD1FormulaLossHoldEvent:
    def test_hold_mode_produces_formula_loss_hold_event(self, tmp_path, monkeypatch):
        """AD1：GATE=hold + needs=True → CLICallResult.hold_events 含 kind=formula_loss。"""
        monkeypatch.setenv("CODEMAKER_FORMULA_GATE", "hold")
        audit_path = tmp_path / "audit_log.jsonl"
        p = tmp_path / "fml.xlsx"
        _make_formula_workbook(p)
        cli = StubCodeMakerCLI(workspace=tmp_path)
        cli._auditor = _make_auditor(tmp_path)
        _mock_formula_loss(cli, monkeypatch, needs=True)
        res = cli.write_cell(p, "S1", 5, 1, 999)
        assert res.ok
        assert res.needs_manual_fix is True
        assert res.hold_events, "hold_events 不应为空"
        kinds = [getattr(e, "kind", "") for e in res.hold_events]
        assert "formula_loss" in kinds, f"缺 formula_loss: {kinds}"
        ops = _read_audit_ops(audit_path)
        ops_types = [d.get("operation", "") for d in ops]
        assert "pre_commit_hold" in ops_types
        assert "formula_loss_detected" in ops_types

    def test_on_mode_no_formula_loss_hold_event(self, tmp_path, monkeypatch):
        """AD1：GATE=on（非 hold）+ needs=True → 无 formula_loss hold 事件（仅 audit warning）。"""
        monkeypatch.setenv("CODEMAKER_FORMULA_GATE", "on")
        p = tmp_path / "fml.xlsx"
        _make_formula_workbook(p)
        cli = StubCodeMakerCLI(workspace=tmp_path)
        cli._auditor = _make_auditor(tmp_path)
        _mock_formula_loss(cli, monkeypatch, needs=True)
        res = cli.write_cell(p, "S1", 5, 1, 999)
        assert res.ok
        kinds = [getattr(e, "kind", "") for e in (res.hold_events or [])]
        assert "formula_loss" not in kinds, f"on 模式不应产 hold 事件: {kinds}"

    def test_off_mode_no_hold_events(self, tmp_path, monkeypatch):
        """AD1：GATE=off → 无 hold_events + 无 audit。"""
        monkeypatch.setenv("CODEMAKER_FORMULA_GATE", "off")
        audit_path = tmp_path / "audit_log.jsonl"
        p = tmp_path / "fml.xlsx"
        _make_formula_workbook(p)
        cli = StubCodeMakerCLI(workspace=tmp_path)
        cli._auditor = _make_auditor(tmp_path)
        _mock_formula_loss(cli, monkeypatch, needs=True)
        res = cli.write_cell(p, "S1", 5, 1, 999)
        assert res.ok
        assert res.hold_events == []
        ops = _read_audit_ops(audit_path)
        assert all(d.get("operation") != "pre_commit_hold" for d in ops)


class TestAD2D4CommentLossHoldEvent:
    def _make_comment_workbook(self, p: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "S1"
        ws.cell(1, 1, "id")
        ws.cell(2, 1, 1)
        ws.cell(2, 2, "name")
        c = Comment("保留我", "tester")
        ws.cell(2, 2).comment = c
        wb.save(p)

    def test_comment_loss_produces_hold_event_and_partial_audit(self, tmp_path, monkeypatch):
        """AD2+D4：批注二次回写后仍丢 → kind=comment_loss hold 事件 + comment_replay_partial audit。

        构造批注丢失场景：mock _detect_comment_loss 永远返回非空（模拟二次 save 后仍丢）。
        """
        audit_path = tmp_path / "audit_log.jsonl"
        p = tmp_path / "cmt.xlsx"
        self._make_comment_workbook(p)
        cli = StubCodeMakerCLI(workspace=tmp_path)
        cli._auditor = _make_auditor(tmp_path)
        # mock 永远返回 1 个丢失（模拟回写后二次做差仍丢）
        lost_coord = ("S1", "B2")
        monkeypatch.setattr(cli, "_detect_comment_loss",
                            lambda path, before: {lost_coord: ("保留我", "tester")})
        res = cli.write_cell(p, "S1", 2, 1, 999)
        assert res.ok
        cr = res.comment_replay
        assert cr.get("replayed") is True
        assert cr.get("still_lost") == 1
        assert res.hold_events, "should have comment_loss hold event"
        kinds = [getattr(e, "kind", "") for e in res.hold_events]
        assert "comment_loss" in kinds, f"缺 comment_loss: {kinds}"
        ops = _read_audit_ops(audit_path)
        ops_types = [d.get("operation", "") for d in ops]
        assert "comment_replay_partial" in ops_types, f"缺 D4 audit: {ops_types}"
        assert "pre_commit_hold" in ops_types

    def test_no_comment_loss_no_hold_event(self, tmp_path, monkeypatch):
        """AD2：批注无丢失 → 无 comment_loss hold 事件。"""
        p = tmp_path / "cmt.xlsx"
        self._make_comment_workbook(p)
        cli = StubCodeMakerCLI(workspace=tmp_path)
        cli._auditor = _make_auditor(tmp_path)
        monkeypatch.setattr(cli, "_detect_comment_loss", lambda path, before: {})
        res = cli.write_cell(p, "S1", 2, 1, 999)
        assert res.ok
        assert res.comment_replay.get("replayed") is False
        kinds = [getattr(e, "kind", "") for e in (res.hold_events or [])]
        assert "comment_loss" not in kinds


class TestA2AgentConsumesHoldEvents:
    def test_agent_consumes_hold_events_to_failures(self, tmp_path, monkeypatch):
        """A2：agent._handle_cli_hold_events → res.failures 追加 #40 软失败 dict。"""
        ag = _make_agent()
        res = _make_res()
        # 构造含 2 hold_events 的 CLICallResult
        from routers.precommit_hold import PreCommitHoldEvent
        cli_result = CLICallResult(
            ok=True, data=None,
            hold_events=[
                PreCommitHoldEvent(kind="formula_loss", severity="hold", count=1,
                                   message="公式丢失", recommendation="manual_fix"),
                PreCommitHoldEvent(kind="comment_loss", severity="hold", count=2,
                                   message="批注丢失", recommendation="manual_fix"),
            ],
        )
        ag._handle_cli_hold_events(res, cli_result, sheet="S1")
        assert len(res.failures) == 2
        f0 = res.failures[0]
        assert f0["code"] == 40
        assert f0["kind"] == "formula_loss"
        assert f0["sheet"] == "S1"
        assert res.failures[1]["kind"] == "comment_loss"

    def test_agent_emits_sse_when_sink_set(self, tmp_path, monkeypatch):
        """A2：_agent_subtask_sink 注入 → 推 pre_commit_hold SSE 事件。"""
        ag = _make_agent()
        emitted = []

        def _sink(event, data):
            emitted.append((event, data))

        ag._agent_subtask_sink = _sink
        res = _make_res()
        from routers.precommit_hold import PreCommitHoldEvent
        cli_result = CLICallResult(
            ok=True, data=None,
            hold_events=[PreCommitHoldEvent(kind="formula_loss", severity="hold",
                                            count=1, message="m")])
        ag._handle_cli_hold_events(res, cli_result, sheet="S1")
        assert len(emitted) == 1
        assert emitted[0][0] == "pre_commit_hold"
        assert emitted[0][1]["kind"] == "formula_loss"

    def test_agent_no_hold_events_noop(self, tmp_path, monkeypatch):
        """A2：CLICallResult.hold_events 空 → res.failures 不变。"""
        ag = _make_agent()
        res = _make_res()
        cli_result = CLICallResult(ok=True, data=None, hold_events=[])
        ag._handle_cli_hold_events(res, cli_result, sheet="S1")
        assert res.failures == []

    def test_agent_sink_failure_swallowed(self, tmp_path, monkeypatch):
        """A2：sink 抛异常 → 静默吞（不阻断），res.failures 仍追加。"""
        ag = _make_agent()

        def _bad_sink(event, data):
            raise RuntimeError("sink broken")

        ag._agent_subtask_sink = _bad_sink
        res = _make_res()
        from routers.precommit_hold import PreCommitHoldEvent
        cli_result = CLICallResult(
            ok=True, data=None,
            hold_events=[PreCommitHoldEvent(kind="comment_loss", severity="hold",
                                            count=1, message="m")])
        ag._handle_cli_hold_events(res, cli_result, sheet="S1")
        assert len(res.failures) == 1  # 软失败仍入，sink 异常被吞
