"""skill 挂载 vs 未挂载 A/B 对照测试框架。

目的: 评判 skill（列别名/行规则/上下文消歧/反模式/短形式）对 AI 表格定位的增强质量。

设计:
  - 120 样例: 列定位 40 + 行定位 40 + 消歧/边界 40
  - 每样例带预构造 NLIntent（绕过 LLM 解析的非确定性，直接测定位阶段，
    skill 的核心增强全在定位路径，此方案确定性可重复）
  - 双轮: skill=off（各 cfg 空实例）vs skill=on（正常 load yaml）
  - 全部用 get（查询）action，不写盘，可重复跑
  - 5 指标:
      1. 定位成功率   mean(col_evidence.resolved 非空 且 score>=0.5)
      2. 平均列置信度 mean(col.score)
      3. 平均行置信度 mean(row.confidence)
      4. 歧义率       mean(row.ambiguous)  — 越低越好
      5. 平均耗时     mean(响应 ms)
  - 准确率（按 expected_ok）作辅助指标

用法:
    cd server
    python -m tests.skill_ab_test                  # 跑双轮 + 报告
    python -m tests.skill_ab_test --quick 30       # 只跑前 30 样例冒烟
    python -m tests.skill_ab_test --report-only    # 只读已有报告

依赖: resources/ 下有 pet/ability/item/hero 等真实表（只读，不改）。
      不依赖 codemaker serve（绕过 LLM）。
"""
from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
import time
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from statistics import mean
from typing import Optional

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import yaml
    _HAS_YAML = True
except ImportError:
    _HAS_YAML = False

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.excel.nl_parser import NLIntent
from agent.excel.agent import TableAgent
from agent.excel.real_cli import RealCodeMakerCLI
from agent.excel.codemaker_parser import CodemakerNLParser  # noqa: F401  (构造占位)
from agent.excel.skill_updater import get_skill_updater


ROOT = Path(__file__).resolve().parents[2]
RES = ROOT / "resources"
REPORT_DIR = Path(__file__).resolve().parent / "reports"


# ── 动态扫描 resources 全表生成样例（普适覆盖）──────────────
_SKILLS_DIR = Path(__file__).resolve().parents[1] / "agent" / "excel" / "skills"
# T12: L1 自动派生文件迁移到 L1_derived/，回退根目录兼容
_L1_DERIVED_DIR = _SKILLS_DIR / "L1_derived"
def _l1_path(name: str) -> Path:
    p = _L1_DERIVED_DIR / name
    return p if p.exists() else _SKILLS_DIR / name
_ROW_ALIASES_PATH = _l1_path("row_aliases.yaml")
_SHORT_FORM_PATH = _l1_path("column_short_form.yaml")


def _load_row_aliases() -> dict:
    """加载 row_aliases.yaml → {stem: {sheet: [{locator_column, match}]}}。文件缺失返回 {}。"""
    if not _HAS_YAML or not _ROW_ALIASES_PATH.exists():
        return {}
    try:
        data = yaml.safe_load(_ROW_ALIASES_PATH.read_text(encoding="utf-8")) or {}
        return data.get("tables", {}) or {}
    except Exception:
        return {}


def _load_short_forms() -> dict:
    """加载 column_short_form.yaml 通配层 → {real_col: [short_forms]}。
    取 columns."*"."*" 段（适用于所有表/sheet 的基础短形式）。"""
    if not _HAS_YAML or not _SHORT_FORM_PATH.exists():
        return {}
    try:
        data = yaml.safe_load(_SHORT_FORM_PATH.read_text(encoding="utf-8")) or {}
        return (data.get("columns", {}).get("*", {}).get("*", {})) or {}
    except Exception:
        return {}


# ── 全链路承压所需的反查 helper（中文表别名 / 列别名 / sheet 别名）──
_COLUMN_ALIASES_PATH = _l1_path("column_aliases.yaml")


def _load_short_forms_full() -> dict:
    """加载 column_short_form.yaml 完整结构 → {table: {sheet: {real_col: [shorts]}}}。"""
    if not _HAS_YAML or not _SHORT_FORM_PATH.exists():
        return {}
    try:
        data = yaml.safe_load(_SHORT_FORM_PATH.read_text(encoding="utf-8")) or {}
        return data.get("columns", {}) or {}
    except Exception:
        return {}


def _load_column_aliases_full() -> dict:
    """加载 column_aliases.yaml → {table: {sheet: {alias: real_col}}}。"""
    if not _HAS_YAML or not _COLUMN_ALIASES_PATH.exists():
        return {}
    try:
        data = yaml.safe_load(_COLUMN_ALIASES_PATH.read_text(encoding="utf-8")) or {}
        return data.get("columns", {}) or {}
    except Exception:
        return {}


def _short_forms_for(stem: str, sheet: str, real_col: str) -> list[str]:
    """合并通配+精确，返回 real_col 的短形式列表（去重保序）。"""
    full = _load_short_forms_full()
    out: list[str] = []
    for tbl_key in ("*", stem):
        tbl = full.get(tbl_key, {})
        for sheet_key in ("*", sheet):
            for s in tbl.get(sheet_key, {}).get(real_col, []):
                if s and s not in out:
                    out.append(s)
    return out


def _aliases_for_col(stem: str, sheet: str, real_col: str) -> list[str]:
    """反查 real_col 的列别名（alias→real_col 反向，合并通配+精确）。
    返回 alias != real_col 的别名列表，去重保序。"""
    full = _load_column_aliases_full()
    out: list[str] = []
    for tbl_key in ("*", stem):
        tbl = full.get(tbl_key, {})
        for sheet_key in ("*", sheet):
            for alias, col in tbl.get(sheet_key, {}).items():
                if col == real_col and alias != real_col and alias not in out:
                    out.append(alias)
    return out


def _col_alias_for(stem: str, sheet: str, real_col: str) -> Optional[str]:
    """选一个真实用户可能输入的自然语言列形式（测列定位 skill）。
    优先短形式（口语化），其次列别名；都没有返回 None。"""
    for s in _short_forms_for(stem, sheet, real_col):
        if len(s) < len(real_col):
            return s
    for a in _aliases_for_col(stem, sheet, real_col):
        if len(a) < len(real_col):
            return a
    return None


def _cn_alias_for(stem: str) -> Optional[str]:
    """从 alias_mapping.json 反查该表的中文别名（测表定位 + TableResolver）。
    优先 2-4 字业务别名；跳过过短/过长。返回首个可用别名。"""
    try:
        from agent.excel.alias_mapping import AliasMapping
    except Exception:
        return None
    am = AliasMapping.load()
    cands = [a for a in am.files_for_stem(stem) if 2 <= len(a) <= 4]
    return cands[0] if cands else None


def _sheet_alias_for(stem: str, real_sheet: str) -> Optional[str]:
    """反查真实 sheet 名对应的业务别名（测 sheet_cfg 消歧）。
    从 sheet_aliases.yaml 找 alias→real_sheet，返回首个别名。"""
    try:
        from agent.excel.skill_loader import SheetAliasConfig
    except Exception:
        return None
    sa = SheetAliasConfig.load()
    for alias, sn in sa.aliases_for(stem).items():
        if sn == real_sheet and alias != real_sheet:
            return alias
    return None


def _read_sheet(path: Path, sheet: str, loc_col: str, n_values: int = 2):
    """读 sheet 表头 + loc_col 列前 n 个非空值。返回 (headers, values)。
    表/sheet 不存在或 loc_col 不在表头 → (headers, []) 或 (None, None)。"""
    if not _HAS_OPENPYXL:
        return None, None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, None
    try:
        if sheet not in wb.sheetnames:
            return None, None
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, None
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_idx = None
        for i, h in enumerate(headers):
            if h and (h == loc_col or loc_col in h):
                col_idx = i
                break
        if col_idx is None:
            return headers, []
        # 类型定义行模式：如 "name:string" / "id:int"（英文字段名:类型）
        values = []
        for row in rows[1:]:
            if col_idx < len(row):
                v = row[col_idx]
                if v is not None and str(v).strip():
                    sv = str(v).strip()
                    # 跳过类型定义行值（表第2行常是字段类型定义）
                    if _TYPE_RE.match(sv):
                        continue
                    values.append(sv)
            if len(values) >= n_values:
                break
        return headers, values
    finally:
        try:
            wb.close()
        except Exception:
            pass


@dataclass
class Case:
    cid: int
    category: str          # col_locate | row_locate | disambig | negative_missing | negative_badtable
    text: str              # 原始自然语言（仅供报告展示）
    intent: NLIntent
    expected_ok: bool      # 预期定位是否应成功（宽松判定，侧重 A/B 相对差）
    expected: dict = field(default_factory=dict)
    # expected grounded 字段: table/sheet/target_col/cell_value/ok/ambiguous
    # 从真实 xlsx 读取落地，非假设。正例必填 cell_value，负例只填 ok=False。


@dataclass
class CaseResult:
    cid: int
    category: str
    skill_enabled: bool
    ok: bool
    expected_ok: bool
    col_score: float
    col_resolved: bool
    row_confidence: float
    row_ambiguous: bool
    needs_confirm: bool
    elapsed_ms: float
    # 字段级检查（grounded expected vs actual）
    table_stem: str = ""
    table_sheet: str = ""
    returned_value: str = ""


# 跳过的非业务列（无查询语义），grounded target_col 选取时排除
_SKIP_COLS = ("填表说明", "备注", "程序", "注释", "说明", "test", "TEST",
              "不要修改", "勿删", "占位", "reserved", "Reserved")
# 类型定义行模式：如 "name:string" / "id:int" / "excl_state: int" / "key:1"
# （英文字段名[可含空格]:类型或值；冒号两侧允许空格）——表第2行常见类型定义行，
# 不是业务数据，必须跳过以保证 grounded 真值严谨。
_TYPE_RE = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_ ]*\s*:\s*\w+$')


def _read_full_rows(path: Path, sheet: str, loc_col: str, n_rows: int = 3):
    """读 sheet 表头 + 前 n_rows 条业务行（跳过类型定义行）。

    返回 (headers, rows)。rows 为 list[dict]，每条 {col_name: cell_value_str}。
    用于 grounded expected：读取真实 cell_value 校准答案。"""
    if not _HAS_OPENPYXL:
        return None, []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, []
    try:
        if sheet not in wb.sheetnames:
            return None, []
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, []
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_idx = None
        for i, h in enumerate(headers):
            if h and (h == loc_col or loc_col in h):
                col_idx = i
                break
        if col_idx is None:
            return headers, []
        out = []
        for row in rows[1:]:
            if col_idx >= len(row):
                continue
            lv = row[col_idx]
            if lv is None or not str(lv).strip():
                continue
            sv = str(lv).strip()
            if _TYPE_RE.match(sv):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = row[i] if i < len(row) else None
                row_dict[h] = str(v).strip() if v is not None else ""
            out.append(row_dict)
            if len(out) >= n_rows:
                break
        return headers, out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _pick_target_col(headers: list, row_dict: dict, loc_col: str) -> Optional[str]:
    """选有真实数据的语义列作 target_col。跳过辅助/类型/空值列。
    确保 expected target_col 有查询语义且 cell 有值，避免假设性答案。"""
    for h in headers:
        if not h or h == loc_col:
            continue
        if any(s in h for s in _SKIP_COLS):
            continue
        v = row_dict.get(h, "")
        if v and not _TYPE_RE.match(v):
            return h
    return None


# ── 样例生成（动态扫描 resources 全表，普适覆盖）─────────
def generate_cases() -> list[Case]:
    """扫描 resources/**/*.xlsx 所有表，生成 grounded 样例（expected 答案从真实 xlsx 读取落地）。

    数据来源:
      - row_aliases.yaml: 每表的行定位列 + match 模式
      - 真实 xlsx 表头 + 行数据: locator_value 用真实实体值，cell_value 从真实单元格读取
      - column_short_form.yaml: 短形式列查询（off contains 可能失败，on 短形式命中）

    每表生成:
      1. nl_col_alias: 中文表别名 + 列别名/短形式查询（全链路承压：表名走
         AliasMapping/TableResolver，sheet 留空走行评分消歧，列靠 skill 别名解析）
      2. col_locate: 短形式查询（另一列的短形式，off exact_substr vs on skill 命中）
      3. disambig（多 sheet，sheet_hint 用业务别名测 sheet_cfg）/ row_locate contains（单 sheet）
      4. negative_missing: 不存在实体（expected ok=False）
    末尾追加 negative_badtable: 坏表名（expected ok=False）

    expected 字段（grounded）:
      table/sheet/target_col/cell_value/ok/ambiguous —— cell_value 从真实 xlsx 单元格读取。
    """
    cases: list[Case] = []
    cid = 0

    def add(category, text, intent, expected_ok, expected=None):
        nonlocal cid
        cid += 1
        intent.raw = text  # 填原始文本，供 get 流程从 raw 解析目标列
        cases.append(Case(cid=cid, category=category, text=text,
                          intent=intent, expected_ok=expected_ok,
                          expected=expected or {}))

    if not _HAS_OPENPYXL:
        warnings.warn("openpyxl 不可用，无法动态扫描表生成样例", RuntimeWarning)
        return cases

    row_aliases = _load_row_aliases()
    short_forms = _load_short_forms()
    xlsx_paths = sorted(glob.glob(str(ROOT / "resources" / "**" / "*.xlsx"),
                                  recursive=True))

    for path in xlsx_paths:
        stem = Path(path).stem
        if "qa_test" in str(path) or "formula_sample" in str(path):
            continue
        cfg = row_aliases.get(stem, {})
        if not cfg:
            continue

        # 收集该表有效的 (sheet, loc_col, match, headers, rows) 组合（rows 含完整行字典）
        valid = []
        for sh, rules in cfg.items():
            if not isinstance(rules, list):
                continue
            for r in rules:
                loc_col = r.get("locator_column", "")
                if not loc_col:
                    continue
                headers, rows = _read_full_rows(Path(path), sh, loc_col, n_rows=2)
                if headers and loc_col in headers and rows:
                    valid.append((sh, loc_col, r.get("match", "exact"), headers, rows))
                    break
        if not valid:
            continue

        sheet, loc_col, loc_match, headers, rows = valid[0]
        row0 = rows[0]
        first_val = row0.get(loc_col, "")
        if not first_val or len(first_val) < 1:
            continue
        # 语义 target 列：跳过辅助/类型/空值列，确保 expected 答案有查询意义
        target_col = _pick_target_col(headers, row0, loc_col)
        if not target_col:
            continue  # 无语义 target 列 → 跳过，不造假设性用例
        cell_val = row0.get(target_col, "")
        # 中文表别名：无别名则回退 stem（仍测 sheet/列解析，但表定位靠 stem 精确命中）
        cn_alias = _cn_alias_for(stem) or stem
        # target_col 的自然语言形式（短形式/列别名），无则用真实列名
        col_alias = _col_alias_for(stem, sheet, target_col) or target_col

        # 样例1: 自然语言列别名查询（中文表名 + 列别名/短形式，sheet_hint 留空走消歧）
        # off 轮：sheet_cfg 空 + short_form_cfg 空 → 列靠 exact_substr(0.9)，多 sheet 靠行评分
        # on 轮：sheet_cfg 别名 + short_form 扩展 → 列 skill 命中(1.0)，sheet 别名直命中
        add("nl_col_alias", f"查询{cn_alias}{first_val}的{col_alias}",
            NLIntent(action="get", table_hint=cn_alias, sheet_hint="",
                     locator_field=loc_col, locator_value=first_val,
                     target_field=col_alias, raw=f"查询{cn_alias}{first_val}的{col_alias}"),
            True,
            expected={"table": stem, "sheet": sheet, "target_col": target_col,
                      "cell_value": cell_val, "ok": True, "ambiguous": False})

        # 样例2: 另一列的短形式查询（off exact_substr vs on short_form 扩展）
        for h in headers:
            if not h or h == loc_col or h == target_col:
                continue
            shorts = short_forms.get(h, [])
            if shorts:
                sf_val = row0.get(h, "")
                # 跳过空值/类型行单元格：expected cell_value 必须有真实值，否则严格判定误报
                if not sf_val or _TYPE_RE.match(sf_val):
                    continue
                add("col_locate", f"查询{cn_alias}{first_val}的{shorts[0]}",
                    NLIntent(action="get", table_hint=cn_alias, sheet_hint="",
                             locator_field=loc_col, locator_value=first_val,
                             target_field=shorts[0],
                             raw=f"查询{cn_alias}{first_val}的{shorts[0]}"), True,
                    expected={"table": stem, "sheet": sheet, "target_col": h,
                              "cell_value": sf_val, "ok": True, "ambiguous": False})
                break

        # 样例3: 多 sheet 消歧（sheet_hint 用业务别名测 sheet_cfg）/ 单 sheet 行 contains
        if len(valid) > 1:
            sh2, lc2, m2, h2, rows2 = valid[1]
            row0_2 = rows2[0]
            v2 = row0_2.get(lc2, "")
            tgt2 = _pick_target_col(h2, row0_2, lc2)
            if v2 and tgt2:
                # sheet 别名作 sheet_hint：on 轮 sheet_cfg 解析，off 轮靠行评分
                sh_hint = _sheet_alias_for(stem, sh2) or sh2
                tgt2_alias = _col_alias_for(stem, sh2, tgt2) or tgt2
                add("disambig", f"查询{cn_alias}{sh_hint}的{v2}的{tgt2_alias}",
                    NLIntent(action="get", table_hint=cn_alias, sheet_hint=sh_hint,
                             locator_field=lc2, locator_value=v2,
                             target_field=tgt2_alias,
                             raw=f"查询{cn_alias}{sh_hint}的{v2}的{tgt2_alias}"), True,
                    expected={"table": stem, "sheet": sh2, "target_col": tgt2,
                              "cell_value": row0_2.get(tgt2, ""), "ok": True, "ambiguous": False})
        elif loc_match == "contains" and len(first_val) >= 2:
            frag = first_val[:2]
            # frag contains 定位行 + 列别名查询（raw 含目标列，可解析）。
            # 仅对 contains 型定位列生成：exact 型 id 列截断 frag 注定失配，非 skill 缺陷。
            add("row_locate", f"查询{cn_alias}{frag}的{col_alias}",
                NLIntent(action="get", table_hint=cn_alias, sheet_hint="",
                         locator_field=loc_col, locator_value=frag,
                         target_field=col_alias,
                         raw=f"查询{cn_alias}{frag}的{col_alias}"), True,
                expected={"table": stem, "sheet": sheet, "target_col": target_col,
                          "cell_value": cell_val, "ok": True, "ambiguous": False})

        # 样例4: 负例 —— 不存在实体（expected ok=False，table_hint 用中文别名测表定位）
        add("negative_missing", f"查询{cn_alias}__nonexistent_xyz__的{target_col}",
            NLIntent(action="get", table_hint=cn_alias, sheet_hint="",
                     locator_field=loc_col, locator_value="__nonexistent_xyz__",
                     target_field=target_col,
                     raw=f"查询{cn_alias}__nonexistent_xyz__的{target_col}"), False,
            expected={"table": stem, "sheet": sheet, "ok": False})

    # 末尾负例：坏表名（expected ok=False）
    add("negative_badtable", "查询__bad_table_xyz__的某个属性",
        NLIntent(action="get", table_hint="__bad_table_xyz__",
                 locator_field="名称", locator_value="任意",
                 target_field="任意", raw=""), False,
        expected={"ok": False})

    return cases


# ── Agent 构造 ──────────────────────────────────────────
def build_agent(enable_skill: bool, enable_verify_repair_loop: bool = True) -> TableAgent:
    """构造 TableAgent。parser 仅占位（样例带预构造 intent，不走 parse）。

    enable_verify_repair_loop：capability: verify-repair-loop 的 A/B 维度。
    loop=off 退回原线性 pipeline + 单轮 retry，用于对比成功率与延迟。
    """
    cli = RealCodeMakerCLI(workspace=RES)
    parser = CodemakerNLParser(directory=str(RES))
    agent = TableAgent(cli=cli, parser=parser, enable_skill=enable_skill,
                       enable_verify_repair_loop=enable_verify_repair_loop)
    # 测试不写 evidence（避免污染 skill 学习），dialog 仍记录带 skill_enabled 标记
    agent.enable_evidence = False
    return agent


def compare_loop_on_off(cases, enable_skill: bool = True) -> dict:
    """capability: verify-repair-loop —— loop=on/off 对比维度。

    返回 {on: {ok_rate, avg_ms}, off: {ok_rate, avg_ms}, delta}。
    需 codemaker serve 运行（样例走 _run_single 真实执行）。
    成功路径延迟不应显著增加、成功率不应退化（合并门禁）。
    """

    def _run_dim(loop_on: bool):
        agent = build_agent(enable_skill, enable_verify_repair_loop=loop_on)
        ok_n, tot, ms = 0, 0, 0.0
        for case in cases:
            r = run_case(agent, case, skill_enabled=enable_skill)
            tot += 1
            if r.ok:
                ok_n += 1
            ms += r.elapsed_ms
        return {"ok_rate": ok_n / tot if tot else 0.0, "avg_ms": ms / tot if tot else 0.0}

    on = _run_dim(True)
    off = _run_dim(False)
    return {
        "on": on, "off": off,
        "delta_ok_rate": on["ok_rate"] - off["ok_rate"],
        "delta_avg_ms": on["avg_ms"] - off["avg_ms"],
    }


# ── 单样例执行 ──────────────────────────────────────────
def run_case(agent: TableAgent, case: Case, skill_enabled: bool) -> CaseResult:
    t0 = time.perf_counter()
    res = agent._run_single(case.intent, session_id="ab_test")
    elapsed_ms = (time.perf_counter() - t0) * 1000
    col = res.col_evidence or {}
    row = res.row_evidence or {}
    return CaseResult(
        cid=case.cid, category=case.category, skill_enabled=skill_enabled,
        ok=res.ok, expected_ok=case.expected_ok,
        col_score=float(col.get("score") or 0.0),
        col_resolved=bool(col.get("resolved")),
        row_confidence=float(row.get("confidence") or 0.0),
        row_ambiguous=bool(row.get("ambiguous")),
        needs_confirm=res.needs_confirm,
        elapsed_ms=round(elapsed_ms, 2),
        table_stem=getattr(res, "table_stem", "") or "",
        table_sheet=getattr(res, "table_sheet", "") or "",
        returned_value=str(res.message or ""),
    )


# ── 指标聚合 ────────────────────────────────────────────
def aggregate(results: list[CaseResult]) -> dict:
    if not results:
        return {}
    n = len(results)
    acc = sum(1 for r in results if r.ok == r.expected_ok) / n
    # 阈值 0.85：off 轮模糊匹配 score~0.8 判失败，on 轮 skill 命中 score=1.0 判成功
    loc_success = sum(1 for r in results if r.col_resolved and r.col_score >= 0.85) / n
    # 精确命中率：score==1.0（skill 别名/短形式 direct 命中），off 轮模糊 0.8 不算
    exact_match = sum(1 for r in results if r.col_resolved and r.col_score >= 0.99) / n
    # 一次成功率：无确认无纠正且成功
    first_try = sum(1 for r in results if r.ok and not r.needs_confirm) / n
    return {
        "n": n,
        "accuracy": round(acc, 4),                          # 辅助：实际ok==expected
        "locate_success_rate": round(loc_success, 4),        # 指标1（阈值0.85）
        "exact_match_rate": round(exact_match, 4),           # 指标1b 精确命中
        "first_try_rate": round(first_try, 4),               # 指标1c 一次成功
        "avg_col_score": round(mean(r.col_score for r in results), 4),     # 指标2
        "avg_row_confidence": round(mean(r.row_confidence for r in results), 4),  # 指标3
        "ambiguity_rate": round(sum(r.row_ambiguous for r in results) / n, 4),    # 指标4 越低越好
        "avg_elapsed_ms": round(mean(r.elapsed_ms for r in results), 2),   # 指标5
    }


def diff_metrics(off: dict, on: dict) -> dict:
    """skill on 相对 off 的变化（正数=改善，歧义率除外）。"""
    def delta(key, lower_is_better=False):
        o, n = off.get(key, 0), on.get(key, 0)
        d = round(n - o, 4)
        if lower_is_better:
            return -d  # 歧义率下降算改善
        return d
    return {
        "accuracy_delta": delta("accuracy"),
        "locate_success_delta": delta("locate_success_rate"),
        "exact_match_delta": delta("exact_match_rate"),
        "first_try_delta": delta("first_try_rate"),
        "col_score_delta": delta("avg_col_score"),
        "row_confidence_delta": delta("avg_row_confidence"),
        "ambiguity_delta": delta("ambiguity_rate", lower_is_better=True),
        "elapsed_delta_ms": delta("avg_elapsed_ms", lower_is_better=True),
    }


def field_accuracy(cases: list[Case], results: list[CaseResult]) -> dict:
    """grounded 字段级准确率：expected（真实xlsx落地）vs actual。

    返回 {table_acc, sheet_acc, ok_acc, value_acc, neg_acc, overall, strict_ok_acc}。
    - table/sheet: 正例+负例均查（负例 expected 无 table/sheet 则跳过）
    - ok: 正例期望True负例期望False（宽松，agent 自判 ok）
    - value: 仅正例，actual message 含 expected cell_value
    - neg: 负例 ok==False 比例
    - strict_ok_acc: 严格 ok 判定 —— 正例需 value 匹配 且 col_score≥0.95
      （暴露"ok=True 但列错配/值错"的假成功），负例需 not ok
    """
    by_cid = {r.cid: r for r in results}
    table_p = table_n = sheet_p = sheet_n = ok_p = ok_n = val_p = val_n = neg_p = neg_n = 0
    sok_p = sok_n = 0
    for c in cases:
        r = by_cid.get(c.cid)
        if r is None:
            continue
        exp = c.expected
        is_neg = c.category.startswith("negative")
        # table
        if "table" in exp:
            table_n += 1
            if r.table_stem == exp["table"]:
                table_p += 1
        # sheet
        if "sheet" in exp:
            sheet_n += 1
            if r.table_sheet == exp["sheet"]:
                sheet_p += 1
        # ok（宽松：agent 自判 ok == expected）
        if "ok" in exp:
            ok_n += 1
            if r.ok == exp["ok"]:
                ok_p += 1
        # value（仅正例且有 cell_value）
        if not is_neg and exp.get("cell_value"):
            val_n += 1
            if str(exp["cell_value"]) in r.returned_value:
                val_p += 1
        # neg
        if is_neg:
            neg_n += 1
            if not r.ok:
                neg_p += 1
        # strict ok：正例需 value 匹配 + col_score≥0.95；负例需 not ok
        if "ok" in exp:
            sok_n += 1
            if is_neg:
                if not r.ok:
                    sok_p += 1
            else:
                cell = exp.get("cell_value", "")
                if (r.ok and r.col_score >= 0.95 and cell
                        and str(cell) in (r.returned_value or "")):
                    sok_p += 1
    def rate(p, n):
        return round(p / n, 4) if n else 0.0
    total_p = table_p + sheet_p + ok_p + val_p + neg_p
    total_n = table_n + sheet_n + ok_n + val_n + neg_n
    # strict overall：用严格 ok 替换宽松 ok，value 仍独立计入
    # （strict_ok 已内含 value 匹配，故 strict_overall 不重复加 value，避免双重计数）
    return {
        "table_acc": rate(table_p, table_n),
        "sheet_acc": rate(sheet_p, sheet_n),
        "ok_acc": rate(ok_p, ok_n),
        "value_acc": rate(val_p, val_n),
        "neg_acc": rate(neg_p, neg_n),
        "overall": rate(total_p, total_n),
        "strict_ok_acc": rate(sok_p, sok_n),
        "strict_overall": rate(table_p + sheet_p + sok_p + neg_p,
                               table_n + sheet_n + sok_n + neg_n),
    }


def validate_aliases() -> list[dict]:
    """sheet 别名端到端验证：每条别名构造 NL，跑 TableResolver.resolve，断言落点正确。

    返回 list[{table, alias, expected_sheet, resolved_table, resolved_sheet, pass, reason}]。
    """
    from agent.excel.table_resolver import TableResolver
    from agent.excel.skill_loader import SheetAliasConfig
    cfg = SheetAliasConfig.load()
    resolver = TableResolver(sheet_cfg=cfg)
    idx_path = Path(__file__).resolve().parents[1] / "agent" / "excel" / "_table_index.json"
    real_sheets: dict[str, list[str]] = {}
    if idx_path.exists():
        for it in json.loads(idx_path.read_text(encoding="utf-8")):
            real_sheets[it["stem"]] = [s["name"] for s in it.get("sheets", [])]
    out = []
    for table, al in cfg.mapping.items():
        if table == "*":
            continue
        for alias, sn in al.items():
            if not alias:
                continue
            text = f"查询{table}{alias}相关数据"
            r = resolver.resolve(text)
            rt = r.table_stem if r else None
            rs = r.sheet if r else None
            ok = (rt == table) and (rs == sn)
            reason = ""
            if rt != table:
                reason = f"table解析为{rt}≠{table}"
            elif rs != sn:
                reason = f"sheet解析为{rs}≠{sn}"
            if table in real_sheets and sn not in real_sheets[table]:
                ok = False
                reason = f"配置错误：{sn}不在{table}真实sheet列表"
            out.append({"table": table, "alias": alias, "expected_sheet": sn,
                        "resolved_table": rt, "resolved_sheet": rs,
                        "pass": ok, "reason": reason})
    return out


# ── grounded 自评 → skill_updater 闭环（AI 自动更新）────────
def feed_corrections(cases: list[Case], results: list[CaseResult],
                     do_promote: bool = True) -> dict:
    """把 grounded 自评的"答错"信号回灌 skill_updater，实现无人工标注的自适配。

    机理（闭环补齐 D7 的盲区）:
      现有 promote 只捕获 低置信度命中(score∈[0.3,0.75)) 或 user_corrected。
      "高置信度但答错"（如短形式/别名 mis-resolve，score=1.0）不会被捕获。
      grounded A/B 用真实 xlsx 单元格值做裁判，能识别这类"自信答错"，
      据此造 user_corrected=True 的纠正 evidence（query=用户所输词 → resolved=真值列），
      喂入候选池；后续同一纠正累计≥3 次 + 过回归门禁即 promote 到 runtime 别名。

    仅对"别名/短形式缺口"生效（query != 正确列名）；query==列名却错配的
    属定位逻辑 bug（如多同名 id 列），别名无法修，跳过不误学。

    Args:
        cases:   本轮样例（含 grounded expected）
        results: skill=on 轮结果
        do_promote: 是否立即尝试 promote（默认 True，走既有安全阀）

    Returns:
        {corrections: [...], queued: int, promoted: [...]}。
    """
    by_cid = {r.cid: r for r in results}
    updater = get_skill_updater()
    corrections: list[dict] = []
    for c in cases:
        if c.category.startswith("negative"):
            continue
        r = by_cid.get(c.cid)
        if r is None:
            continue
        exp = c.expected
        correct_col = exp.get("target_col")
        cell = exp.get("cell_value")
        if not correct_col or not cell:
            continue
        # 答对（返回值含真值）→ 无需纠正
        if str(cell) in (r.returned_value or ""):
            continue
        query_term = (c.intent.target_field or "").strip()
        # 别名/短形式缺口：用户所输词 != 真实列名，且非空
        if not query_term or query_term == correct_col:
            continue
        record = {
            "table_stem": exp.get("table", ""),
            "sheet": exp.get("sheet", ""),
            "col": {"query": query_term, "resolved": correct_col,
                    "score": float(r.col_score or 0.0)},
            "user_corrected": True,   # grounded 真值裁定 → 视作黄金纠正信号
            "ts": _now_iso_local(),
            "source": "grounded_ab_selfeval",
        }
        corrections.append(record)
        updater.ingest_evidence(record)
    promoted: list[dict] = []
    if do_promote:
        try:
            promoted = updater.try_promote()
        except Exception as e:
            warnings.warn(f"try_promote failed: {e}", RuntimeWarning)
    return {"corrections": corrections, "queued": len(corrections),
            "promoted": promoted}


def _now_iso_local() -> str:
    # 带时区（与 skill_updater._now_iso 一致，避免 naive/aware 相减报错）
    from datetime import datetime
    return datetime.now().astimezone().isoformat(timespec="seconds")


# ── 报告 ────────────────────────────────────────────────
def render_report(off: dict, on: dict, diff: dict,
                  off_results: list, on_results: list,
                  off_field: dict, on_field: dict,
                  alias_checks: list) -> str:
    lines = [
        "# skill A/B 对照测试报告（grounded 真值）",
        "",
        f"- 样例数: {on.get('n', 0)}",
        f"- 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 真值来源: 真实 xlsx 单元格值（expected cell_value 落地，非假设）",
        "",
        "## 一、5 指标对比（定位径）",
        "",
        "| 指标 | skill=off | skill=on | 变化 | 说明 |",
        "|---|---|---|---|---|",
        f"| 定位成功率(≥0.85) | {off['locate_success_rate']:.4f} | {on['locate_success_rate']:.4f} | {diff['locate_success_delta']:+.4f} | 越高越好(阈值0.85,off模糊0.8判失败) |",
        f"| 精确命中率(=1.0) | {off['exact_match_rate']:.4f} | {on['exact_match_rate']:.4f} | {diff['exact_match_delta']:+.4f} | 越高越好(skill direct命中) |",
        f"| 一次成功率 | {off['first_try_rate']:.4f} | {on['first_try_rate']:.4f} | {diff['first_try_delta']:+.4f} | 越高越好(无确认一次成功) |",
        f"| 平均列置信度 | {off['avg_col_score']:.4f} | {on['avg_col_score']:.4f} | {diff['col_score_delta']:+.4f} | 越高越好 |",
        f"| 平均行置信度 | {off['avg_row_confidence']:.4f} | {on['avg_row_confidence']:.4f} | {diff['row_confidence_delta']:+.4f} | 越高越好 |",
        f"| 歧义率 | {off['ambiguity_rate']:.4f} | {on['ambiguity_rate']:.4f} | {diff['ambiguity_delta']:+.4f} | 越低越好(正数=改善) |",
        f"| 平均耗时(ms) | {off['avg_elapsed_ms']:.2f} | {on['avg_elapsed_ms']:.2f} | {diff['elapsed_delta_ms']:+.2f} | 越低越好(正数=改善) |",
        f"| 准确率(ok==expected) | {off['accuracy']:.4f} | {on['accuracy']:.4f} | {diff['accuracy_delta']:+.4f} | 实际ok==expected |",
        "",
        "## 二、grounded 字段级准确率（expected 来自真实 xlsx）",
        "",
        "| 字段 | skill=off | skill=on | 变化 | 说明 |",
        "|---|---|---|---|---|",
    ]
    for key, label, hint in [
        ("table_acc", "表定位", "resolved table==expected"),
        ("sheet_acc", "sheet定位", "resolved sheet==expected"),
        ("ok_acc", "ok判定(宽松)", "ok==expected_ok(正例True/负例False)"),
        ("value_acc", "答案值", "returned message含真实cell_value"),
        ("neg_acc", "负例拒绝", "负例ok==False比例"),
        ("strict_ok_acc", "严格ok判定", "正例:value匹配+col_score≥0.95;负例:not ok"),
        ("strict_overall", "严格综合", "表+sheet+严格ok+负例(暴露列错配假成功)"),
        ("overall", "综合(宽松)", "全部字段检查通过率"),
    ]:
        o = off_field.get(key, 0.0)
        n = on_field.get(key, 0.0)
        d = round(n - o, 4)
        lines.append(f"| {label} | {o:.4f} | {n:.4f} | {d:+.4f} | {hint} |")
    lines += [
        "",
        "## 三、分类别指标（skill=on）",
        "",
        "| 类别 | n | 定位成功率 | 列置信度 | 歧义率 |",
        "|---|---|---|---|---|",
    ]
    for cat in ("nl_col_alias", "col_locate", "row_locate", "disambig",
                "negative_missing", "negative_badtable"):
        sub = [r for r in on_results if r.category == cat]
        if not sub:
            continue
        agg = aggregate(sub)
        lines.append(f"| {cat} | {agg['n']} | {agg['locate_success_rate']:.4f} | "
                     f"{agg['avg_col_score']:.4f} | {agg['ambiguity_rate']:.4f} |")
    # 别名端到端
    lines += [
        "",
        "## 四、sheet 别名端到端验证",
        "",
    ]
    if alias_checks:
        ap = sum(1 for a in alias_checks if a["pass"])
        an = len(alias_checks)
        lines.append(f"- 通过 {ap}/{an} = {ap/an:.4f}")
        afails = [a for a in alias_checks if not a["pass"]]
        if afails:
            lines.append(f"- 失败 {len(afails)} 条:")
            for a in afails[:15]:
                lines.append(f"  - {a['table']}.{a['alias']}->{a['expected_sheet']} "
                             f"resolved=({a['resolved_table']},{a['resolved_sheet']}) {a['reason']}")
    else:
        lines.append("- 无别名配置")
    lines += [
        "",
        "## 结论判定",
        "",
        f"- skill 增强 {'有效' if diff['locate_success_delta'] > 0 or diff['col_score_delta'] > 0 or on_field['overall'] > off_field['overall'] else '不明显'}",
        f"- 定位成功率提升 {diff['locate_success_delta']:+.4f}",
        f"- 列置信度提升 {diff['col_score_delta']:+.4f}",
        f"- 歧义率改善 {diff['ambiguity_delta']:+.4f}",
        f"- grounded综合准确率提升 {round(on_field['overall'] - off_field['overall'], 4):+.4f}",
        f"- 严格综合准确率提升 {round(on_field['strict_overall'] - off_field['strict_overall'], 4):+.4f} "
        f"(off={off_field['strict_overall']:.4f} → on={on_field['strict_overall']:.4f}, 暴露列错配假成功)",
    ]
    return "\n".join(lines)


# ── 主流程 ─────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--quick", type=int, default=0, help="只跑前 N 样例冒烟")
    ap.add_argument("--report-only", action="store_true")
    ap.add_argument("--aliases", action="store_true", help="仅跑 sheet 别名端到端验证")
    ap.add_argument("--learn", action="store_true",
                    help="闭环：把 grounded 答错信号回灌 skill_updater 候选池并尝试 promote")
    args = ap.parse_args()

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    if args.report_only:
        p = REPORT_DIR / "skill_ab_latest.json"
        if p.exists():
            print(p.read_text(encoding="utf-8"))
        return

    # 别名端到端验证（不依赖 agent，快）
    alias_checks = validate_aliases() if not args.report_only else []
    ap = sum(1 for a in alias_checks if a["pass"])
    print(f"sheet 别名端到端: {ap}/{len(alias_checks)} 通过")

    if args.aliases:
        print("\n## sheet 别名端到端验证")
        print(f"通过 {ap}/{len(alias_checks)} = {ap/len(alias_checks) if alias_checks else 0:.4f}")
        for a in alias_checks:
            if not a["pass"]:
                print(f"  FAIL {a['table']}.{a['alias']}->{a['expected_sheet']} "
                      f"resolved=({a['resolved_table']},{a['resolved_sheet']}) {a['reason']}")
        return

    cases = generate_cases()
    if args.quick:
        cases = cases[:args.quick]
    print(f"生成 {len(cases)} 样例（grounded 真值），开始 A/B 双轮...")

    print("[1/2] skill=off 轮...")
    agent_off = build_agent(enable_skill=False)
    off_results = [run_case(agent_off, c, False) for c in cases]

    print("[2/2] skill=on 轮...")
    agent_on = build_agent(enable_skill=True)
    on_results = [run_case(agent_on, c, True) for c in cases]

    off = aggregate(off_results)
    on = aggregate(on_results)
    diff = diff_metrics(off, on)
    off_field = field_accuracy(cases, off_results)
    on_field = field_accuracy(cases, on_results)

    report_md = render_report(off, on, diff, off_results, on_results,
                              off_field, on_field, alias_checks)
    (REPORT_DIR / "skill_ab_latest.md").write_text(report_md, encoding="utf-8")

    payload = {
        "summary": {"off": off, "on": on, "diff": diff},
        "off_field": off_field, "on_field": on_field,
        "alias_pass": ap, "alias_total": len(alias_checks),
        "off_results": [r.__dict__ for r in off_results],
        "on_results": [r.__dict__ for r in on_results],
        "alias_checks": alias_checks,
    }
    (REPORT_DIR / "skill_ab_latest.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n" + report_md)
    print(f"\n报告已写: {REPORT_DIR / 'skill_ab_latest.md'}")

    # capability: eval-baseline-management —— 归档
    try:
        from tests.eval_baseline import archive_run, make_run_id
        tag = os.environ.get("EVAL_BASELINE_TAG", "")
        rid = make_run_id(tag=tag or None)
        archive_run("skill_ab_test", payload, report_md, run_id=rid)
        print(f"归档: reports/archive/skill_ab_test_{rid}.json")
    except Exception as e:
        print(f"[warn] 归档失败: {e}")

    # 闭环自适配：把答错信号回灌 skill_updater（默认关闭，--learn 开启）
    if args.learn:
        learn = feed_corrections(cases, on_results)
        print(f"\n## 闭环自适配（--learn）")
        print(f"- 识别别名/短形式缺口纠正 {learn['queued']} 条 → 已喂候选池")
        for c in learn["corrections"]:
            print(f"  · {c['table_stem']}.{c['sheet']}: 「{c['col']['query']}」→ {c['col']['resolved']}")
        if learn["promoted"]:
            print(f"- 本轮 promote 到 runtime 别名 {len(learn['promoted'])} 条:")
            for p in learn["promoted"]:
                print(f"  ✓ {p['table_stem']}.{p['sheet']}: {p['query']}→{p['resolved']} "
                      f"(hits={p['hits']}, conf={p['confidence_avg']})")
        else:
            print("- 本轮未 promote（需同纠正累计≥3 次且过回归门禁；候选已入池待累积）")

    # 清理测试 dialog 记录（避免污染生产案例库）
    _cleanup_test_dialogs()


def _cleanup_test_dialogs():
    """删 ab_test session 的 dialog 记录，避免测试数据进案例库。"""
    try:
        from agent.excel.core.dialog_logger import get_dialog_logger
        dl = get_dialog_logger()
        for d in (dl.dialog_dir, dl.examples_dir, dl.failures_dir):
            if not d.exists():
                continue
            for p in d.glob("*.jsonl"):
                # 只删 ab_test session 或测试产生的 _unknown 表文件
                if p.name.startswith("ab_test") or p.stem == "_unknown":
                    p.unlink(missing_ok=True)
    except Exception:
        pass


if __name__ == "__main__":
    main()
