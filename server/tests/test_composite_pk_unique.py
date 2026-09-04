"""复合主键唯一性单测（rules primary_key + Step2 field_layer + Step3 _do_append）。

覆盖缺口：FabaoLevel 这类"(法宝id, 法宝等级)"联合键被单列唯一性误判冲突，
导致合法多行被拦。验证三层一致：
  - rules_loader.get_primary_key_overlay 解析 primary_key 声明
  - schema_bundle._composite_existing_from_rows 建组合值集合
  - validator._check_composite_unique (5,1)冲突检出 / (5,2)(5,3)放行
  - agent._check_composite_pk_conflict 写盘组合冲突检测
不依赖 serve LLM，用临时 xlsx + StubCodeMakerCLI。
"""
from pathlib import Path

from agent.excel.cli.cli_interface import StubCodeMakerCLI


def _make_test_xlsx(tmp: Path) -> Path:
    """建临时 FabaoLevel sheet：复合主键 (法宝id, 法宝等级)。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FabaoLevel"
    # row1 中文表头, row2 规范名, row3-4 占位, row5 起数据
    ws.append(["法宝id", "法宝等级", "技能id", "技能等级"])
    ws.append(["fabao_id:int", "level:int", "spell_id:int", "spell_level:int"])
    ws.append(["", "", "", ""])
    ws.append(["", "", "", ""])
    ws.append([1, 1, 100001, 1])
    ws.append([1, 2, 100001, 20])
    ws.append([2, 1, 100008, 1])
    ws.append([2, 2, 100008, 20])
    p = tmp / "_test_composite_pk_unique.xlsx"
    wb.save(p)
    # 同时建 fabao 规则文件（rules/validate/fabao.md）让 primary_key overlay 生效。
    rules_dir = tmp / "rules" / "validate"
    rules_dir.mkdir(parents=True, exist_ok=True)
    (rules_dir / "fabao.md").write_text(
        "# fabao 表校验约束\n\n```yaml\ntables:\n  fabao:\n    FabaoLevel:\n"
        "      primary_key: [法宝id, 法宝等级]\n      columns: {}\n```\n",
        encoding="utf-8")
    return p


def test_get_primary_key_overlay(tmp_path, monkeypatch):
    """rules_loader 解析 sheet 级 primary_key 复合键声明。"""
    _make_test_xlsx(tmp_path)
    rules_dir = tmp_path / "rules"
    from agent.excel.core import rules_loader
    monkeypatch.setattr(rules_loader, "_RULES_DIR", rules_dir)
    monkeypatch.setattr(rules_loader, "_VALIDATE_DIR", rules_dir / "validate")
    monkeypatch.setattr(rules_loader, "_FILL_DIR", rules_dir / "fill")
    rules_loader.reset_cache()
    overlay = rules_loader.get_primary_key_overlay()
    assert "fabao" in overlay, f"应解析到 fabao，实际 {list(overlay.keys())}"
    assert overlay["fabao"].get("FabaoLevel") == ["法宝id", "法宝等级"], \
        f"复合键声明应正确，实际 {overlay['fabao']}"
    print("PASS get_primary_key_overlay: 解析 primary_key 复合键声明")


def test_composite_existing_from_rows():
    """schema_bundle 组合值集合构造。"""
    from agent.excel.core.schema_bundle import _composite_existing_from_rows
    headers = ["法宝id", "法宝等级", "技能id"]
    rows = [[1, 1, 100001], [1, 2, 100008], [2, 1, 100009]]
    s = _composite_existing_from_rows(headers, rows, ["法宝id".lower(), "法宝等级".lower()])
    assert ("1", "1") in s and ("2", "1") in s, f"应含组合，实际 {s}"
    assert len(s) == 3, f"应 3 个组合，实际 {len(s)}"
    # 任一 PK 列不在表头 → 空
    s2 = _composite_existing_from_rows(headers, rows, ["法宝id", "不存在的列"])
    assert s2 == set(), "缺 PK 列应返空放行"
    # 单列 PK → 返空（不适用复合检测）
    assert _composite_existing_from_rows(headers, rows, ["法宝id"]) == set()
    print("PASS _composite_existing_from_rows: 组合值集合构造")


def test_check_composite_unique_conflict_pass(tmp_path, monkeypatch):
    """(1,1) 与现有冲突 → 检出；三条新组合 → 全放行（纯逻辑，不依赖 Stub read_sheet）。"""
    from agent.excel.subagent.validator_agent import ValidatorAgent
    v = ValidatorAgent(); v._ask_callback = None; v._cli = None

    # 手工构造 existing + composite_existing（模拟表里已有行 (1,1)(1,2)(2,1)）
    existing_values = {
        "法宝id": {"1", "2"},
        "法宝等级": {"1", "2"},
    }
    composite_existing = {("1", "1"), ("1", "2"), ("2", "1")}
    pk_cols = ["法宝id", "法宝等级"]

    # (1,1) 与现有冲突
    iss = v._check_composite_unique(
        {"法宝id": "1", "法宝等级": "1"}, pk_cols, existing_values,
        composite_existing=composite_existing)
    assert any(getattr(i, "issue_type", "") == "unique_violation" for i in iss), \
        f"(1,1) 重复应检冲突，实际 {iss}"
    print("PASS _check_composite_unique (conflict): (1,1) 与现有冲突检出")

    # 三条新组合 → 全放行
    for combo in [("5", "1"), ("5", "2"), ("5", "3")]:
        iss2 = v._check_composite_unique(
            {"法宝id": combo[0], "法宝等级": combo[1]}, pk_cols, existing_values,
            composite_existing=composite_existing)
        assert not iss2, f"新组合 {combo} 不应报冲突，实际 {iss2}"
    print("PASS _check_composite_unique (pass): 三条新组合全放行")


def test_check_composite_pk_conflict_write(tmp_path, monkeypatch):
    """agent._check_composite_pk_conflict 写盘组合检测：(1,1) 冲突 / (1,3)(5,1) 放行。

    用最小 MockCLI 替代 Stub（避免 Stub read_sheet data_start 依赖），只实现
    _check_composite_pk_conflict 需要的 read_header + read_sheet。
    """
    p = _make_test_xlsx(tmp_path)

    class _MockParser:
        def __init__(self):
            self.client = None; self.directory = ""; self.model = ""

    # 最小 mock cli：read_header / read_sheet 读临时 xlsx 用 openpyxl
    import openpyxl
    class _MockCLI:
        name = "mock"
        def read_header(self, _p, sheet):
            wb = openpyxl.load_workbook(_p, read_only=True)
            ws = wb[sheet]
            hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            wb.close()
            return hdr
        def read_sheet(self, _p, sheet):
            wb = openpyxl.load_workbook(_p, read_only=True)
            ws = wb[sheet]
            out = []
            for r in ws.iter_rows(min_row=5, values_only=True):
                if any(v is not None for v in r):
                    out.append(list(r))
            wb.close()
            return out

    from agent.excel.core.agent import TableAgent
    cli = _MockCLI()
    agent = TableAgent(cli=cli, parser=_MockParser())

    # PK 列名按表头对齐 → 写盘 values {1-based col: val}
    # (1,1) 与现有 (1,1) 冲突
    desc = agent._check_composite_pk_conflict(
        p, "FabaoLevel", {1: '1', 2: '1'}, ["法宝id", "法宝等级"])
    assert desc is not None, "(1,1) 应判冲突"
    assert "1" in desc and "法宝id" in desc, f"冲突描述应含列名+值，实际 {desc}"
    # (1,3) 新组合放行
    desc2 = agent._check_composite_pk_conflict(
        p, "FabaoLevel", {1: '1', 2: '3'}, ["法宝id", "法宝等级"])
    assert desc2 is None, f"(1,3) 不应冲突，实际 {desc2}"
    # (5,1) 新组合放行
    desc3 = agent._check_composite_pk_conflict(
        p, "FabaoLevel", {1: '5', 2: '1'}, ["法宝id", "法宝等级"])
    assert desc3 is None, f"(5,1) 不应冲突，实际 {desc3}"
    print("PASS _check_composite_pk_conflict: (1,1)冲突 / (1,3)(5,1)放行")


if __name__ == "__main__":
    import tempfile
    class _M:
        def __init__(self): self._mp = {}
        def setenv(self, k, v): self._mp[k] = v
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        m = _M()
        test_composite_existing_from_rows()
        test_get_primary_key_overlay(td, m)
        test_check_composite_unique_conflict_pass(td, m)
        test_check_composite_pk_conflict_write(td, m)
    print("ALL PASS")
