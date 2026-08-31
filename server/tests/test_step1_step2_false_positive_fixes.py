"""Step1/Step2 误报修复回归测试（太虚剑宗复合指令实证用例）。

覆盖 8 个缺陷，判据全部来自**真实 resources 数据**或**纯文本规则**，不绑业务词：

  F1 业务必填列「历史全空列」豁免（SchoolAbility 功效描述，全表 0 行填过）
  F1b 业务必填列「同族列已写值」豁免（用户只给一个「描述」，不该要求填满表里
      所有描述列；反向用例断言同族列全空时仍要报缺，不误伤原 Pack3 契约）
  F2 「不带奖励」否定式 → 识别为显式置空而非 LLM 漏产
  F3 可共用引用列不是唯一键/身份列（大世界model_id=1027 被 7 个门派共用）
  F4 显式 FK 引用值不被转占位符（buff_id=600003 被同一天赋 1/2 级共用）
  F5 子表主键即外键时绑定上游产出（school_spirit.神通id → 新神通 id）
  F6 ask 卡片从错误输入派生形式合规的建议值（「200,0,150」→ 200）
  F7 点分嵌套列不再查到兄弟列的类型（to_pos 取到 tuple 而非 to_space_id:int）
"""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from agent.excel.parse_agent import ParseAgent  # noqa: E402
from agent.excel.schema_bundle import (  # noqa: E402
    _existing_values_from_rows, _rows_to_dicts,
)
from agent.excel.subagent.validator_agent import (  # noqa: E402
    Issue, IssueType, ValidatorAgent, _derive_suggestion_from_value,
    _duplicate_value_cols,
)

RES = Path(__file__).resolve().parents[2] / "resources"


class _ResCli:
    """最小 CLI：直读 resources 下的真实 xlsx（row1 表头 / row2 类型 / row3+ 数据）。"""

    def __init__(self):
        self._wb = {}

    def _load(self, path):
        key = str(path)
        if key not in self._wb:
            self._wb[key] = openpyxl.load_workbook(key, read_only=True)
        return self._wb[key]

    def list_tables(self):
        return sorted(RES.rglob("*.xlsx"))

    def get_sheets(self, path):
        return self._load(path).sheetnames

    def read_header(self, path, sheet):
        for row in self._load(path)[sheet].iter_rows(min_row=1, max_row=1,
                                                      values_only=True):
            return list(row)
        return []

    def read_type_row(self, path, sheet):
        for row in self._load(path)[sheet].iter_rows(min_row=2, max_row=2,
                                                      values_only=True):
            return list(row)
        return []

    def read_sheet(self, path, sheet):
        return [list(r) for r in self._load(path)[sheet].iter_rows(
            min_row=3, values_only=True)]


def _path_of(stem: str):
    cli = _ResCli()
    return next(p for p in cli.list_tables() if p.stem == stem)


def _schema(stem: str, sheet: str):
    cli = _ResCli()
    path = _path_of(stem)
    return cli.read_header(path, sheet), cli.read_type_row(path, sheet)


def _rows(stem: str, sheet: str):
    cli = _ResCli()
    return cli.read_sheet(_path_of(stem), sheet)


# ───────────────────────── F3：可共用引用列 ─────────────────────────


def test_duplicate_value_cols_marks_shared_reference_columns():
    """school/School 里被多行共用的 model_id 列应被识别为重复值列。"""
    hdr = _schema("school", "School")[0]
    dup = _duplicate_value_cols(
        _rows_to_dicts(hdr, _rows("school", "School")), hdr)
    assert "大世界model_id" in dup
    assert "战斗model_id" in dup


def test_pk_like_col_excludes_shared_reference_columns():
    """大世界model_id 不算唯一键；表首列「门派」仍是 PK。"""
    hdr = _schema("school", "School")[0]
    dup = _duplicate_value_cols(
        _rows_to_dicts(hdr, _rows("school", "School")), hdr)
    v = ValidatorAgent()
    assert v._is_pk_like_col("大世界model_id", stem="school", headers=hdr,
                             sheet="School", non_unique_cols=dup) is False
    assert v._is_pk_like_col("门派", stem="school", headers=hdr,
                             sheet="School", non_unique_cols=dup) is True


def test_pick_single_pk_column_skips_shared_reference_columns():
    """PK 扫描列不能选中被多行共用的 model_id（否则 1027 被误改 1029）。"""
    hdr = _schema("school", "School")[0]
    dup = _duplicate_value_cols(
        _rows_to_dicts(hdr, _rows("school", "School")), hdr)
    assert ValidatorAgent._pick_single_pk_column(hdr, dup) != "大世界model_id"


# ─────────────────────── F1/F1b/F2：业务必填列 ───────────────────────

_RAW_ABILITY = ("新增一个门派叫'太虚剑宗'，先配四个神通：第一个神通'太虚剑意'，"
                "描述'以虚御实，剑出无痕'，配两套心法，心法一叫'剑意·蓄'")


class _AddIntent:
    action = "add"
    raw = _RAW_ABILITY


def _ability_fields():
    return {
        "神通id": "<new_ability1_id>", "名称": "太虚剑意",
        "神通描述": "以虚御实，剑出无痕，每回合自动凝成太虚剑气",
        "心法1名称": "剑意·蓄", "心法1描述": "开局即得剑气，快速进入节奏",
        "心法2名称": "剑意·绝", "心法2描述": "击杀敌方时额外获得剑气",
    }


def test_never_filled_col_exempt():
    """功效描述全表从没填过 → 不算业务必填列。"""
    hdr = _schema("school_ability", "SchoolAbility")[0]
    ev = _existing_values_from_rows(hdr, _rows("school_ability", "SchoolAbility"))
    assert "功效描述" in ValidatorAgent._never_filled_cols(hdr, ev)


def test_same_family_exempt_and_still_reports_when_family_empty():
    """同族（描述族）已有列写入 → 同族其余列不报；整族全空时仍要报（不误伤）。"""
    hdr = _schema("school_ability", "SchoolAbility")[0]
    ev = _existing_values_from_rows(hdr, _rows("school_ability", "SchoolAbility"))
    v = ValidatorAgent()
    v._pk_cols_cache = {}
    cols = [getattr(i, "col", "") for i in v._check_business_required_pre_add(
        _AddIntent(), hdr, _ability_fields(), _RAW_ABILITY, existing_values=ev)]
    assert "功效描述" not in cols and "升级描述" not in cols
    # 反向：描述族一列都没写 → 仍要报神通描述缺失
    cols2 = [getattr(i, "col", "") for i in v._check_business_required_pre_add(
        _AddIntent(), hdr, {"神通id": 9999, "名称": "某某"}, _RAW_ABILITY,
        existing_values=ev)]
    assert "神通描述" in cols2


def test_negation_exempt_fills_zero_for_int_column():
    """「不带奖励」= 显式置空，不报缺；数值列补 0。"""
    hdr, trow = _schema("mail", "GlobalMail")
    raw = ("最后给新门派发一封开宗立派的全服邮件：邮件模板标题'太虚剑宗开宗立派'，"
           "全服邮件 global_id 20，邮件类型 1，发送人'系统'，"
           "发送时间 2026-09-01 00:00:00，不带奖励。")
    assert "奖励" in ValidatorAgent._explicitly_empty_cols(raw, hdr)
    fields = {"全服邮件ID": 20, "发送人": "系统"}
    v = ValidatorAgent()
    v._pk_cols_cache = {}
    issues = v._check_business_required_pre_add(
        _AddIntent(), hdr, fields, raw, existing_values={}, type_row=trow)
    assert all(getattr(i, "col", "") != "奖励" for i in issues)
    assert fields.get("奖励") == 0


# ───────────────────────── F4：显式 FK 引用值 ─────────────────────────


def test_shared_value_columns_marks_buff_id():
    """SchoolTalentLevel 的 buff_id 被同一天赋 1/2 级共用 → 可共用引用列。"""
    pa = ParseAgent(cli=_ResCli())
    shared = pa._shared_value_columns("school_talent", "SchoolTalentLevel")
    assert "buffid" in shared       # row2 规范键
    assert "被动效果buffid" in shared  # row1 中文键
    assert "id" not in shared       # 主键不是可共用列


# ───────────────── F5：子表主键即外键 → 绑定上游 ─────────────────


class _It:
    def __init__(self, stem, sheet, fields, label):
        self.action = "add"
        self.table_hint = stem
        self.sheet_hint = sheet
        self.produces_label = label
        self.consumes_labels = []
        self.extras = {"fields": fields, "produces": label, "consumes": {}}


class _Edge:
    def __init__(self, fs, fsh, fc, ts, tsh, tc):
        self.from_stem, self.from_sheet, self.from_column = fs, fsh, fc
        self.to_stem, self.to_sheet, self.to_column = ts, tsh, tc


def test_bind_self_primary_to_upstream():
    """school_spirit.神通id（既是主键又是指向神通的外键）绑到 4 个新神通。"""
    pa = ParseAgent(cli=_ResCli())
    abilities = [
        _It("school_ability", "SchoolAbility",
            {"school_ability_id": f"<new_ability{i}_id>", "name": n},
            f"new_ability{i}_id")
        for i, n in enumerate(["太虚剑意", "裂空斩", "万剑归宗", "无相护体"], 1)
    ]
    spirits = [
        _It("school_spirit", "SchoolSpirit",
            {"school_ability_id": f"<new_school_spirit_id{'' if i == 1 else '_' + str(i)}>",
             "school_id": 9, "spirit_id": i, "spirit_buffs[0]": 600001 + i % 2},
            f"new_school_spirit_id{'' if i == 1 else '_' + str(i)}")
        for i in range(1, 5)
    ]
    edges = [_Edge("school_spirit", "SchoolSpirit", "school_ability_id",
                   "school_ability", "SchoolAbility", "school_ability_id")]
    intents = abilities + spirits
    assert pa._bind_self_primary_to_upstream(intents, edges) == 4
    assert [s.extras["fields"]["school_ability_id"] for s in spirits] == \
        [f"<new_ability{i}_id>" for i in range(1, 5)]
    # _drop_foreign_placeholder_primary_fields 不得把绑定打回自引用
    pa._drop_foreign_placeholder_primary_fields(intents, edges)
    assert spirits[0].extras["fields"]["school_ability_id"] == "<new_ability1_id>"


# ───────────────── F6：ask 必须给出建议值 ─────────────────


def test_derive_suggestion_from_value():
    """形式问题（多值串/百分号/单位/全角/布尔词）都能整理出建议值。"""
    assert _derive_suggestion_from_value("200,0,150", "int") == 200
    assert _derive_suggestion_from_value("20%", "int") == 20
    assert _derive_suggestion_from_value("3万", "int") == 30000
    assert _derive_suggestion_from_value("是", "bool") == 1
    assert _derive_suggestion_from_value("否", "bool") == 0
    assert _derive_suggestion_from_value("１，２", "float") == 1.0
    assert _derive_suggestion_from_value("第3档", "int") == 3
    # 纯中文塞 int 列：猜不出就不猜（宁可不给建议，也不写脏数据）
    assert _derive_suggestion_from_value("二百五", "int") is None


def test_ask_always_carries_a_suggestion():
    """TYPE_MISMATCH 无枚举/无现有值时，ask 也要带从错误输入派生的建议值。"""
    v = ValidatorAgent()
    captured = {}

    def _cb(question):
        captured.update(question)
        return {"mode": "field", "value": question.get("suggested_id")}

    v._ask_callback = _cb

    class _ItV:
        action = "add"
        table_hint = "interaction"
        sheet_hint = "Interaction"
        raw = "BOSS 战斗也一起配上：在战场 10050 的坐标 (200,0,150)"

    v._ask_hard_issue(
        _ItV(),
        Issue(col="effect.data.3005.to_space_id",
              issue_type=IssueType.TYPE_MISMATCH.value,
              expected="effect.data.3005.to_space_id: int",
              suggestion="", value="200,0,150"),
        data_getter=lambda *_: {})
    assert str(captured.get("suggested_id")) == "200"
    assert captured.get("mode_hint") == "value_input"
    assert captured.get("suggestion")


# ───────────────── F7：点分嵌套列类型查错 ─────────────────


def test_dotted_nested_col_type_not_from_sibling():
    """to_pos 必须取到自己的 tuple 类型，不再撞到 to_space_id:int。"""
    hdr, trow = _schema("interaction", "Interaction")
    v = ValidatorAgent()
    t_pos = v._lookup_col_type("3005", hdr, trow,
                               orig_col="effect.data.3005.to_pos")
    t_sid = v._lookup_col_type("3005", hdr, trow,
                               orig_col="effect.data.3005.to_space_id")
    assert "to_pos" in t_pos and "tuple" in t_pos
    assert "to_space_id" in t_sid and "int" in t_sid
    # 坐标「200,0,150」在 tuple 列上是合法的
    assert v._coerce_field_simple(t_pos, "200,0,150")[0] is True
    assert re.search(r"\d", t_pos)


# ───────────── A+：对话选项悬空占位符规则兜底补空壳 ─────────────
# 背景：InteractionConv.options[N] 引用了 <opt_xxx_id>，但 DecomposeAgent
# 漏拆对应的 InteractionConvOption producer（「下次再来/结束」类选项常被当
# 作结束节点省略）。且 options[N]→InteractionConvOption 的 FK 边在 fk_edges
# 里往往缺失，原 _backfill_dangling_placeholder_producers 查不到 target 就
# 静默跳过 → Step1 unresolved_placeholder 硬失败。
# 修复：无 FK 边时按「InteractionConv.options[N] 引用 opt_* 标签 → 目标表
# InteractionConvOption」的规则确定性补空壳（零 LLM、零失败可能）。


def test_aplus_rule_shell_backfills_dangling_option_producer():
    """InteractionConv.options[N] 引用悬空 opt_* 标签 → 规则兜底补 InteractionConvOption。"""
    pa = ParseAgent(cli=_ResCli())
    pa._decompose_agent = None  # 模拟降级环境（纯规则兜底，不走 LLM 重拆）
    raw = ("新增对话：点击弹出'欢迎光临，要不要看看我的货物？'，"
           "选项'好的，看看'和'下次再来'。")
    conv = _It("interaction", "InteractionConv",
               {"prompt_text": "欢迎光临，要不要看看我的货物？",
                "options[0]": "<opt_yes_id>", "options[1]": "<opt_no_id>"},
               "conv_root_id")
    # 有 opt_yes_id 的 producer，但 opt_no_id 悬空（无 producer、无 FK 边）
    opt_yes = _It("interaction", "InteractionConvOption",
                  {"option_text": "好的，看看"}, "opt_yes_id")
    intents = [conv, opt_yes]
    n = pa._backfill_dangling_placeholder_producers(intents, raw, [])
    assert n == 1
    shells = [it for it in intents
              if (it.table_hint == "interaction"
                  and it.sheet_hint == "InteractionConvOption"
                  and it.produces_label == "opt_no_id")]
    assert len(shells) == 1
    # 空壳保证引用闭环：option_id 填上标签自身作为自引用 PK，可被 produced 集合接收
    assert shells[0].extras["fields"]["option_id"] == "<opt_no_id>"
    assert shells[0].extras.get("source") == "dangling_option_rule_shell"
    # 补全后所有被引用占位符都应有 producer（不再 unresolved）
    produced = {it.produces_label for it in intents}
    assert "opt_no_id" in produced and "opt_yes_id" in produced


def test_aplus_rule_shell_extracts_option_text_from_raw():
    """空壳能从原文抽出 option_text（「下次再来」→ 不落空）。"""
    pa = ParseAgent(cli=_ResCli())
    pa._decompose_agent = None
    raw = ("对话'请选择'，选项'确认前往'和'下次再来'。")
    assert pa._extract_option_text_for_label(raw, "opt_no_id") == "下次再来"
    # 抽不到的标签（如奖励对话的领取选项）返回空，不强填脏数据
    assert pa._extract_option_text_for_label("随便一段没有选项的话", "opt_done_id") == ""


def test_aplus_rule_shell_non_option_dangling_still_skipped():
    """非对话选项的悬空占位符（无 FK 边）仍按原逻辑跳过，不臆测目标表。"""
    pa = ParseAgent(cli=_ResCli())
    pa._decompose_agent = None
    raw = "新增物品并用到一个奖励包"
    # item 的 reward_id 引用悬空，但消费者不是 InteractionConv.options → 不兜底
    item = _It("item", "Item", {"reward_id": "<some_reward_id>"}, "new_item_id")
    intents = [item]
    n = pa._backfill_dangling_placeholder_producers(intents, raw, [])
    assert n == 0
    assert len(intents) == 1
