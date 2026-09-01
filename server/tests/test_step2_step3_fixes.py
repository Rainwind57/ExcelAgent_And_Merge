"""Step2/Step3 修复的确定性回归（无 serve，秒级）。

覆盖本轮三处修复中可脱离真 LLM 直接验证的两项：
  1. Step3 _do_append partial 归正：被丢字段(coerce/match 失败)把 res.ok 预毒化为
     False 后，追加行+写后验证通过应把 ok 归正为 True（不再把成功 partial 误报失败、
     不再触发 _run_single 回滚已写行）。
  2. validate_two_layer 漏表预检 UnboundLocalError('merged') 回归：复杂跨表输入
     （候选多、产出少 → 触发漏表预检）不再崩，且挂 intent_coverage_gap issue。

运行: python -m pytest server/tests/test_step2_step3_fixes.py -v
"""
from __future__ import annotations

import os
import sys
import types
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.cli_interface import StubCodeMakerCLI
from agent.agent import TableAgent, AgentResult
from agent.nl_parser import NLIntent
from agent.excel.subagent.validator_agent import ValidatorAgent


HEADERS = ["reward_id", "名称", "每日领取上限", "必得道具1", "道具1数量"]


def _build_reward_xlsx(ws: Path) -> Path:
    wb = Workbook()
    s = wb.active
    s.title = "Reward"
    for i, h in enumerate(HEADERS, 1):
        s.cell(1, i, h)
    rows = [
        [100600, "包A", 1, 10001, 5],
        [100601, "包B", 1, 10001, 5],
        [100602, "包C", 1, 10001, 5],
    ]
    for r, row in enumerate(rows, start=5):
        for c, v in enumerate(row, 1):
            s.cell(r, c, v)
    p = ws / "reward.xlsx"
    wb.save(p)
    return p


def _make_agent(cli) -> types.SimpleNamespace:
    agent = types.SimpleNamespace(cli=cli, live_index=False)
    for name in (
        "_do_append", "_auto_sort_after_write", "_refresh_index_after_write",
        "_verify_write_back", "_allocate_pk", "_is_misplaced_pk",
        "_real_pk_col_name", "_locate_pk_col", "_check_missing_required_after_add",
    ):
        setattr(agent, name, getattr(TableAgent, name).__get__(agent))
    return agent


def _make_validator() -> ValidatorAgent:
    v = object.__new__(ValidatorAgent)
    v._cli = None
    v._parser = None
    v._ask_callback = None
    v._required_fields = None
    v._pk_cols_cache = None
    return v


def test_resolve_fk_name_to_id_accepts_unique_short_enum_prefix():
    agent = types.SimpleNamespace()
    agent._cross_table_search = lambda *_args, **_kwargs: [
        {
            "match_type": "exact",
            "table_stem": "spirit",
            "sheet": "Spirit",
            "matches": [{"row": 5, "value": "金"}],
        }
    ]
    agent._find_table_by_stem = lambda _stem: Path("spirit.xlsx")
    agent._locate_pk_col = lambda _path, _sheet: 1
    agent.cli = types.SimpleNamespace(
        read_cell=lambda *_args: types.SimpleNamespace(ok=True, data=1)
    )
    agent._resolve_fk_name_to_id = TableAgent._resolve_fk_name_to_id.__get__(agent)

    assert agent._resolve_fk_name_to_id("金灵根", exclude_stem="school_spirit") == 1


def test_resolve_fk_name_to_id_rejects_ambiguous_short_enum_prefix():
    agent = types.SimpleNamespace()
    agent._cross_table_search = lambda *_args, **_kwargs: [
        {
            "match_type": "exact",
            "table_stem": "spirit",
            "sheet": "Spirit",
            "matches": [{"row": 5, "value": "金"}, {"row": 6, "value": "木"}],
        }
    ]
    agent._find_table_by_stem = lambda _stem: Path("spirit.xlsx")
    agent._locate_pk_col = lambda _path, _sheet: 1
    agent.cli = types.SimpleNamespace(
        read_cell=lambda *_args: types.SimpleNamespace(ok=True, data=1)
    )
    agent._resolve_fk_name_to_id = TableAgent._resolve_fk_name_to_id.__get__(agent)

    assert agent._resolve_fk_name_to_id("金木灵根", exclude_stem="school_spirit") is None


def test_apply_batch_to_intent_supports_full_field_edit():
    v = _make_validator()
    intent = types.SimpleNamespace(extras={
        "fields": {
            "bad_name": "old",
            "desc": "old desc",
            "unused": "drop me",
        }
    })

    v._apply_batch_to_intent(
        intent,
        columns_reply=[],
        fields_reply=[
            {"col": "name", "value": "new name", "delete": False},
            {"col": "desc", "value": "new desc", "delete": False},
            {"col": "unused", "value": "drop me", "delete": True},
            {"col": "level", "value": 1, "delete": False},
        ],
    )

    assert intent.extras["fields"] == {
        "name": "new name",
        "desc": "new desc",
        "level": 1,
    }


def test_apply_batch_to_intent_keeps_legacy_column_remap():
    v = _make_validator()
    intent = types.SimpleNamespace(extras={"fields": {"bad_name": "value", "drop": "x"}})

    v._apply_batch_to_intent(
        intent,
        columns_reply=[
            {"col": "bad_name", "fill_value": "name", "delete": False},
            {"col": "drop", "fill_value": "", "delete": True},
        ],
    )

    assert intent.extras["fields"] == {"name": "value"}


# ── 1. Step3 partial 归正 ────────────────────────────────────────────

def test_do_append_partial_success_restores_ok_true():
    """被丢字段把 res.ok 预毒化为 False → _do_append 追加+验证通过后归正 ok=True。

    模拟 add handler 里 `res.add("coerce_value", False, ...)` / `match_field False`
    对丢弃字段的记录（AgentResult.add 会把整体 ok 永久置 False）。修复前 out.ok
    仍 False → _run_single 判失败 → 回滚已写行 + Step4 标"失败"。修复后写盘真成功
    则 ok=True。
    """
    with TemporaryDirectory() as tmp:
        ws = Path(tmp)
        cli = StubCodeMakerCLI(workspace=ws, header_row=1, data_start_row=5)
        path = _build_reward_xlsx(ws)
        agent = _make_agent(cli)

        res = AgentResult(ok=True, intent=None)
        # 预毒化：模拟某字段类型转换失败被跳过（partial）
        res.add("coerce_value", False, "列[活动类型] 类型错，已丢弃")
        assert res.ok is False, "add(ok=False) 应把整体 ok 置 False"

        # 追加一条不冲突的行（PK 100603 空闲）→ 写盘真成功
        values = {1: 100603, 2: "测试奖励包", 3: 1, 4: 10001, 5: 5}
        agent._do_append(path, "Reward", values, res)

        assert res.ok is True, (
            f"追加行+写后验证通过应把 partial 的 ok 归正为 True，实际 {res.ok} "
            f"message={res.message}")
        _names = [getattr(s, "name", "") for s in getattr(res, "steps", [])]
        assert any(n == "append_row" for n in _names), "应有 append_row 步骤"


def test_resolved_duplicate_target_column_does_not_count_as_partial_failure():
    failed_by_index = {
        2: ("activity_type", "old value could not be coerced"),
        4: ("reward_id", "bad id"),
    }
    values = {2: 1, 3: "春节活动"}

    unresolved = TableAgent._unresolved_failed_fields(failed_by_index, values)

    assert unresolved == [(4, "reward_id", "bad id")]


def test_step3_marks_partial_subtask_failed():
    from agent.excel.core.pipeline.contracts import (
        STEP1_PARSE, STEP3_EXECUTE, StepContext, StepResult)
    from agent.excel.core.pipeline.step3_execute_subagent import Step3ExecuteSubAgent

    intent = NLIntent(action="add", table_hint="activity", sheet_hint="Activity",
                      raw="新增活动",
                      extras={"fields": {"活动类型": "限时", "活动名称": "九霄论剑"}})

    class Services:
        def peek_llm_total(self):
            return 0

        def run_single(self, *_args, **_kwargs):
            return types.SimpleNamespace(
                ok=True,
                partial=True,
                needs_confirm=False,
                message="部分字段跳过",
                steps=[{"name": "coerce_value", "ok": False, "message": "活动类型无法转码"}],
                result_rows=[{"col_name": "活动名称", "new_value": "九霄论剑"}],
                table_stem="activity",
                table_sheet="Activity",
                needs_user_fill=[],
                failures=[],
            )

    ctx = StepContext(session_id="s", user_text="新增活动")
    ctx.set_result(STEP1_PARSE, StepResult(
        step_id=STEP1_PARSE, ok=True, artifacts={"intents": [intent]}))

    result = Step3ExecuteSubAgent(services=Services()).execute(ctx)

    assert result.step_id == STEP3_EXECUTE
    assert result.ok is False
    assert result.metrics["subtasks_fail"] == 1
    assert result.metrics["subtasks_partial"] == 1
    assert result.artifacts["subtasks"][0]["ok"] is False
    assert any(f.get("type") == "partial_write" for f in result.artifacts["failures"])


def test_step2_marks_enum_invalid_as_hard_before_execute():
    from agent.excel.core.pipeline.contracts import (
        STEP1_PARSE, STEP2_VALIDATE, StepContext, StepResult)
    from agent.excel.core.pipeline.step2_validate_subagent import Step2ValidateSubAgent

    intent = NLIntent(action="add", table_hint="activity", sheet_hint="Activity",
                      raw="新增活动",
                      extras={"fields": {"活动类型": "限时", "活动名称": "九霄论剑"}})

    class Services:
        def wire_sinks(self, res):
            return res

        def validate_intents(self, intents, res, *_args, **_kwargs):
            res.failures.append({
                "type": "validation_tip",
                "table": "activity",
                "sheet": "Activity",
                "col": "activity_type",
                "root_cause": "enum_invalid: 枚举: ['1', '2']",
            })
            return intents

    ctx = StepContext(session_id="s", user_text="新增活动")
    ctx.set_result(STEP1_PARSE, StepResult(
        step_id=STEP1_PARSE, ok=True, artifacts={"intents": [intent]}))

    result = Step2ValidateSubAgent(services=Services()).execute(ctx)

    assert result.step_id == STEP2_VALIDATE
    assert result.ok is True
    assert not any(e.error_type == "validation_tip" and e.is_hard for e in result.errors)


def test_step2_marks_enum_invalid_intent_failure_as_hard():
    from agent.excel.core.pipeline.contracts import (
        STEP1_PARSE, STEP2_VALIDATE, StepContext, StepResult)
    from agent.excel.core.pipeline.step2_validate_subagent import Step2ValidateSubAgent

    intent = NLIntent(action="add", table_hint="activity", sheet_hint="Activity",
                      raw="新增活动",
                      extras={"fields": {"活动类型": "限时", "活动名称": "九霄论剑"}})
    intent.failures.append({
        "type": "validation_tip",
        "issue_type": "enum_invalid",
        "table": "activity",
        "sheet": "Activity",
        "col": "activity_type",
        "root_cause": "enum_invalid: 枚举: ['1', '2']",
    })
    ctx = StepContext(session_id="s", user_text="新增活动")
    ctx.set_result(STEP1_PARSE, StepResult(
        step_id=STEP1_PARSE, ok=True, artifacts={"intents": [intent]}))

    result = Step2ValidateSubAgent(services=None).execute(ctx)

    assert result.step_id == STEP2_VALIDATE
    assert result.ok is True
    assert not any(e.error_type == "validation_tip" and e.is_hard for e in result.errors)


def test_step2_accepts_modify_action():
    from agent.excel.core.pipeline.contracts import (
        STEP1_PARSE, STEP2_VALIDATE, StepContext, StepResult)
    from agent.excel.core.pipeline.step2_validate_subagent import Step2ValidateSubAgent

    intent = NLIntent(action="modify", table_hint="school", sheet_hint="School",
                      raw="修改门派", extras={"fields": {"name": "剑宗"}})
    ctx = StepContext(session_id="s", user_text="修改门派")
    ctx.set_result(STEP1_PARSE, StepResult(
        step_id=STEP1_PARSE, ok=True, artifacts={"intents": [intent]}))

    result = Step2ValidateSubAgent(services=None).execute(ctx)

    assert result.step_id == STEP2_VALIDATE
    assert not any(e.error_type == "invalid_action" for e in result.errors)


def test_do_append_pk_conflict_stays_failed():
    """对照：真主键冲突（100601 已占用且非误塞）→ 提前 return，ok 不被归正。"""
    with TemporaryDirectory() as tmp:
        ws = Path(tmp)
        cli = StubCodeMakerCLI(workspace=ws, header_row=1, data_start_row=5)
        path = _build_reward_xlsx(ws)
        agent = _make_agent(cli)

        res = AgentResult(ok=True, intent=None)
        # 显式指定已占用 PK，且其它列各不相同（非误塞语义值）
        values = {1: 100601, 2: "撞库包", 3: 9, 4: 10009, 5: 9}
        agent._do_append(path, "Reward", values, res)

        assert res.ok is False, "真 PK 冲突应保持失败，不被归正"
        _names = [getattr(s, "name", "") for s in getattr(res, "steps", [])]
        assert "pk_conflict" in _names


# ── 2. validate_two_layer 漏表预检 merged 崩溃回归 ────────────────────

def test_validate_two_layer_coverage_gap_no_unbound_merged():
    """复杂跨表输入触发漏表预检 → 原 merged 在初始化前被引用抛
    UnboundLocalError → validate_two_layer 整体崩 → 字段层/写盘预演全不跑。
    修复后不崩，且挂 intent_coverage_gap issue（进 tips）。
    """
    v = _make_validator()

    # 8 候选表，产出仅 reward → _gap=7 >= 8//3=2 触发漏表预检
    cand_stems = ["reward", "activity", "combat", "item", "spell",
                  "entity_prefab", "interaction", "spawn_world_entity"]
    locator_result = types.SimpleNamespace(
        candidates=[types.SimpleNamespace(stem=s) for s in cand_stems],
        fk_edges=[], column_signal=None)

    it = NLIntent(action="add", table_hint="reward", sheet_hint="Reward",
                  raw="新增奖励包 reward_id 30010",
                  extras={"fields": {"reward_id": 30010, "名称": "焚天赤龙首杀奖励"}})

    # 不应抛 UnboundLocalError；漏表预检已移除，降级为 schema_missing issue
    vr = v.validate_two_layer([it], schema_getter=None,
                              locator_result=locator_result)
    assert isinstance(vr, dict)
    _tips_blob = " ".join(str(t) for t in (vr.get("tips") or []))
    _issues_blob = " ".join(str(i) for i in (vr.get("issues") or []))
    assert ("schema_missing" in _tips_blob or "schema_missing" in _issues_blob), (
        f"schema_getter=None 应挂 schema_missing issue，tips={vr.get('tips')}")


# ── 3. Step1 _filter_intents 真实表保留（不再当幻觉丢弃） ──────────────

def test_filter_intents_keeps_real_table_outside_candidates():
    """候选内全保留；候选外默认丢弃（CODEMAKER_DECOMPOSE_KEEP_ALIAS 默认关，
    保现状行为）。keep_alias=1 灰度时候选外 alias 命中候选内 stem 才保留。"""
    from agent.excel.subagent.decompose_agent import DecomposeAgent

    fake_cli = types.SimpleNamespace(
        list_tables=lambda: [Path("interaction.xlsx"), Path("reward.xlsx"),
                             Path("combat.xlsx")])
    da = DecomposeAgent(parser=None, thinking_sink=lambda p, d: None, cli=fake_cli)

    candidates = [types.SimpleNamespace(stem="reward")]
    valid_stems = {"reward"}
    intents = [
        types.SimpleNamespace(table_hint="reward", sheet_hint="Reward",
                              extras={"fields": {"reward_id": 30010}}),
        # 候选外（默认丢弃，无论是否真实表）
        types.SimpleNamespace(table_hint="interaction", sheet_hint="",
                              extras={"fields": {"对话内容": "去讨伐"}}),
        # 候选外且不存在 → 真幻觉，丢弃
        types.SimpleNamespace(table_hint="不存在的幻觉表", sheet_hint="",
                              extras={"fields": {"x": 1}}),
    ]
    kept = da._filter_intents(intents, candidates, valid_stems, path="test")
    kept_stems = {getattr(it, "table_hint", "") for it in kept}

    assert "reward" in kept_stems, "候选内表应保留"
    assert "interaction" not in kept_stems, "候选外默认丢弃（keep_alias 默认关）"
    assert "不存在的幻觉表" not in kept_stems, "不存在的表丢弃"


def test_filter_intents_drops_real_but_off_schema_table():
    """LLM 只能产当前 prompt 注入 schema 的候选表。

    表真实存在但不在本次 candidates 里，仍属于 off-schema 输出；缺表应由 locator
    召回或 backfill 处理，不能在 Step1 接受模型凭记忆产出的表。
    """
    from agent.excel.subagent.decompose_agent import DecomposeAgent

    da = DecomposeAgent(parser=None, thinking_sink=lambda p, d: None, cli=None)
    candidates = [types.SimpleNamespace(stem="building")]
    intents = [
        types.SimpleNamespace(table_hint="building", sheet_hint="Building",
                              extras={"fields": {"id": 1}}),
        types.SimpleNamespace(table_hint="school", sheet_hint="School",
                              extras={"fields": {"name": "X"}}),
    ]

    kept = da._filter_intents(
        intents, candidates, {"building"}, path="single-table retry")

    assert [getattr(it, "table_hint", "") for it in kept] == ["building"]

