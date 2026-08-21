"""公式语义原语（L2）+ insert_row 端到端测试。

验证 P0-1 两个缺口的闭环：
  §1.2 insert_row（中间插入行 + 公式位移）
  §1.3 L2 语义原语（interpret/scan/preview/rewrite）+ agent 流程

自构 xlsx（不依赖 resources 样本，泛化证明）。文本断言为基线；
重算值断言需 libreoffice，不可用则跳过。
"""
import tempfile
from pathlib import Path

import openpyxl

from agent.excel.cli_interface import StubCodeMakerCLI


def _build_sum_sample(d: Path) -> Path:
    """数据行 F3:F12 各 =SUM(B:E)，汇总 F13=SUM(F3:F12)。data_start=3。"""
    p = d / "sum.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "a", "b", "c", "d", "sum"])  # row1 表头
    ws.append(["int"] * 6)                         # row2 类型
    for i in range(1, 11):                         # row3-12 数据
        r = 2 + i
        ws.append([i, i, i, i, i, f"=SUM(B{r}:E{r})"])
    ws.append([99, "", "", "", "", "=SUM(F3:F12)"])  # row13 汇总
    wb.save(p)
    wb.close()
    return p


def _build_summary_sep_sample(d: Path) -> Path:
    """汇总放 G 列（与被汇总 F 列分离），避免 append+rewrite 循环引用。

    数据行 F3:F12 各 =SUM(B:E)，汇总 G13=SUM(F3:F12)。data_start=3。
    """
    p = d / "sum_sep.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["id", "a", "b", "c", "d", "sum", "total"])  # row1 表头
    ws.append(["int"] * 7)                                  # row2 类型
    for i in range(1, 11):                                  # row3-12 数据
        r = 2 + i
        ws.append([i, i, i, i, i, f"=SUM(B{r}:E{r})", ""])
    ws.append([99, "", "", "", "", "", "=SUM(F3:F12)"])     # row13 汇总(G13)
    wb.save(p)
    wb.close()
    return p


def _load_formulas(path: Path, sheet: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                out[c.coordinate] = c.value
    wb.close()
    return out


# ============ §1.2 insert_row ============
def test_insert_row_shift():
    """中间插行 → 下方公式行号位移 + 汇总范围扩展含新行。"""
    print("\n=== E2E insert_row 位移 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        # 在第5行上方插一行 → 汇总 F13→F14，范围 F3:F12→F3:F13（扩展含新行）
        r = cli.insert_row(p, "S", 5, values={1: 999})
        print(f"  ok={r.ok} needs_manual_fix={r.needs_manual_fix}")
        assert r.ok
        formulas = _load_formulas(p, "S")
        # 汇总原 F13 下移到 F14，范围扩展 F3:F13
        assert formulas.get("F14", "").lower() == "=sum(f3:f13)", f"F14={formulas.get('F14')!r}"
        # 原 F6(=SUM(B6:E6)) 下移到 F7，行号 6→7
        assert formulas.get("F7", "").lower() == "=sum(b7:e7)", f"F7={formulas.get('F7')!r}"
        # 原 F4 行号<5 不受影响
        assert formulas.get("F4", "").lower() == "=sum(b4:e4)", f"F4={formulas.get('F4')!r}"
        print("  PASS：插行后汇总范围扩展 + 下方公式行号位移")


# ============ §1.3 L2 原语 ============
def test_interpret_formula():
    """interpret_formula 解析汇总公式语义。"""
    print("\n=== E2E interpret_formula ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        sem = cli.interpret_formula(p, "S", "F13")
        print(f"  formula={sem.formula}")
        print(f"  funcs={sem.funcs} is_aggregate={sem.is_aggregate}")
        print(f"  covers_data_area={sem.covers_data_area} is_last_row={sem.is_last_row}")
        print(f"  notes={sem.notes}")
        assert sem.formula == "=SUM(F3:F12)"
        assert "SUM" in sem.funcs
        assert sem.is_aggregate
        assert sem.covers_data_area
        assert sem.is_last_row
        # 行内计算公式
        sem2 = cli.interpret_formula(p, "S", "F3")
        assert sem2.formula == "=SUM(B3:E3)"
        assert sem2.is_aggregate
        assert not sem2.is_last_row
        print("  PASS：汇总与行内公式语义解析正确")


def test_scan_sheet_formulas():
    """scan_sheet_formulas 扫描全 sheet 公式。"""
    print("\n=== E2E scan_sheet_formulas ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        sems = cli.scan_sheet_formulas(p, "S")
        print(f"  共 {len(sems)} 个公式")
        aggregates = [s for s in sems if s.is_aggregate and s.covers_data_area and s.is_last_row]
        print(f"  末行汇总(覆盖数据区): {[s.cell for s in aggregates]}")
        assert len(sems) == 11  # 10 行内 + 1 汇总
        assert len(aggregates) == 1
        assert aggregates[0].cell == "F13"
        print("  PASS：扫描识别出末行汇总公式")


def test_preview_append_gap():
    """preview_formula_impact(append_rows) 标注汇总范围未含新行的语义缺口。"""
    print("\n=== E2E preview_formula_impact append ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        report = cli.preview_formula_impact(p, "S", {"kind": "append_rows", "count": 3})
        print(f"  op={report.op}")
        print(f"  needs_agent_decision={report.needs_agent_decision}")
        for g in report.semantic_gaps:
            print(f"  缺口: {g}")
        assert report.needs_agent_decision
        assert any("F3:F12" in g and "未纳入" in g for g in report.semantic_gaps)
        print("  PASS：append 预演标注汇总范围缺口")


def test_preview_insert_no_gap():
    """preview_formula_impact(insert_row) 机械位移正确，无语义缺口。"""
    print("\n=== E2E preview_formula_impact insert ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        report = cli.preview_formula_impact(p, "S", {"kind": "insert_row", "row": 5})
        changed = [i for i in report.impacts if i.changed]
        gaps = [i for i in report.impacts if i.semantic_gap]
        print(f"  机械改写单元格: {len(changed)}, 语义缺口: {len(gaps)}")
        # 汇总 F13 应被机械改写为 F14（行号位移），范围扩展 F3:F13
        f13 = next(i for i in report.impacts if i.cell == "F13")
        print(f"  F13 before={f13.formula_before} after={f13.formula_after}")
        assert f13.changed
        assert f13.formula_after.lower() == "=sum(f3:f13)"
        assert not gaps  # insert 机械位移正确，无缺口
        print("  PASS：insert 预演机械位移正确无缺口")


# ============ 完整 agent 流程（L3 模拟）============
def test_agent_flow_append_extend_summary():
    """完整 agent 流程：scan→preview→append→rewrite→校验。

    模拟 AI：append 3 行数据后，末行汇总需含新行 → rewrite 扩展范围。
    汇总放 G 列（与被汇总 F 列分离），rewrite G13=SUM(F3:F16) 不含自身，无循环引用。
    """
    print("\n=== E2E agent 流程：append 扩展汇总 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_summary_sep_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)

        # 步骤1：scan → AI 理解 G13 是"统计全部数据行"汇总
        sems = cli.scan_sheet_formulas(p, "S")
        summary = next(s for s in sems if s.is_aggregate and s.is_last_row and s.covers_data_area)
        print(f"  1.scan: 汇总在 {summary.cell} = {summary.formula}")
        assert summary.cell == "G13"

        # 步骤2：preview(append 3) → 机械结果：汇总不变；语义标注：新行未纳入
        report = cli.preview_formula_impact(p, "S", {"kind": "append_rows", "count": 3})
        assert report.needs_agent_decision
        print(f"  2.preview: {report.semantic_gaps[0]}")

        # 步骤3：AI 推理 → 汇总语义=全部数据行，新3行是数据行 → 目标 =SUM(F3:F15)
        # 步骤4：append 3 行数据（append 到机械末行13之后 → 14,15,16）
        for i in range(3):
            cli.append_row(p, "S", {1: 100 + i, 2: 1, 3: 1, 4: 1, 5: 1, 6: f"=SUM(B{14+i}:E{14+i})"})
        formulas = _load_formulas(p, "S")
        # append 后汇总仍在 G13（append 不位移），范围未变 F3:F12（缺新行）
        assert formulas["G13"] == "=SUM(F3:F12)"
        print("  4.append 3 行后汇总仍 =SUM(F3:F12)（未含新行）")

        # 步骤5：rewrite_formula 执行 AI 决策 → 扩展含新行 F3:F16
        r = cli.rewrite_formula(p, "S", "G13", "=SUM(F3:F16)")
        print(f"  5.rewrite: ok={r.ok} cache={r.cache_message}")
        assert r.ok
        formulas = _load_formulas(p, "S")
        assert formulas["G13"] == "=SUM(F3:F16)"
        print("  6.汇总已扩展为 =SUM(F3:F16)，含全部数据行")
        print("  PASS：agent 流程闭环（scan→preview→append→rewrite）")


def test_agent_flow_insert_no_rewrite_needed():
    """agent 流程变体：insert_row 机械位移已正确，AI 无需 rewrite。"""
    print("\n=== E2E agent 流程：insert 无需 rewrite ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        # preview insert → 无缺口 → AI 判定无需 rewrite
        report = cli.preview_formula_impact(p, "S", {"kind": "insert_row", "row": 5})
        assert not report.needs_agent_decision
        print("  preview insert 无语义缺口 → AI 直接 insert_row，无需 rewrite")
        cli.insert_row(p, "S", 5, values={1: 999})
        formulas = _load_formulas(p, "S")
        assert formulas["F14"] == "=SUM(F3:F13)"  # 汇总位移+范围扩展
        print("  PASS：insert 后机械位移已正确处理汇总")


# ============ 边界场景 ============
def _build_cross_sample(d: Path) -> Path:
    """主表 B列=VLOOKUP(A,Ref!A:B,2)，Ref 表 A:key B:val。data_start=3。"""
    p = d / "cross.xlsx"
    wb = openpyxl.Workbook()
    main = wb.active
    main.title = "Main"
    ref = wb.create_sheet("Ref")
    ref.append(["key", "val"])
    ref.append(["int", "int"])
    for i in range(1, 6):
        ref.append([i, i * 100])
    main.append(["id", "lookup"])
    main.append(["int", "int"])
    for i in range(1, 6):
        r = 2 + i
        main.append([i, f"=VLOOKUP(A{r},Ref!A:B,2,FALSE)"])
    wb.save(p)
    wb.close()
    return p


def _build_dynamic_sample(d: Path) -> Path:
    """含 OFFSET 动态函数。4列，dyn 在 D 列。data_start=3。"""
    p = d / "dyn.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dyn"
    ws.append(["id", "v", "x", "dyn"])
    ws.append(["int", "int", "int", "int"])
    for i in range(1, 4):
        r = 2 + i
        ws.append([i, i * 10, 0, f"=OFFSET(A1,{i},1)"])
    wb.save(p)
    wb.close()
    return p


def test_interpret_cross_sheet():
    """跨表引用：interpret 标 is_cross_sheet + 目标 sheet。"""
    print("\n=== E2E interpret 跨表引用 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_cross_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        sem = cli.interpret_formula(p, "Main", "B3")
        print(f"  formula={sem.formula}")
        print(f"  is_lookup={sem.is_lookup} is_cross_sheet={sem.is_cross_sheet}")
        print(f"  refs={[(r.raw, r.sheet) for r in sem.refs]}")
        assert sem.is_lookup
        assert sem.is_cross_sheet
        cross_refs = [r for r in sem.refs if r.sheet == "Ref"]
        assert cross_refs, "应有指向 Ref 的跨表引用"
        print("  PASS：跨表 VLOOKUP 语义解析正确")


def test_interpret_lookup():
    """查表函数 VLOOKUP → is_lookup=True，is_aggregate=False。"""
    print("\n=== E2E interpret 查表函数 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_cross_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        sem = cli.interpret_formula(p, "Main", "B3")
        assert "VLOOKUP" in sem.funcs
        assert sem.is_lookup
        assert not sem.is_aggregate
        print(f"  funcs={sem.funcs} is_lookup={sem.is_lookup} is_aggregate={sem.is_aggregate}")
        print("  PASS：查表函数识别正确")


def test_interpret_dynamic_func():
    """OFFSET 动态函数 → has_dynamic_func=True，preview 标 skipped_dynamic。"""
    print("\n=== E2E interpret 动态函数 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_dynamic_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        sem = cli.interpret_formula(p, "Dyn", "D3")
        print(f"  formula={sem.formula} has_dynamic_func={sem.has_dynamic_func}")
        assert sem.has_dynamic_func
        # preview delete_row → 动态函数跳过，不位移
        report = cli.preview_formula_impact(p, "Dyn", {"kind": "delete_row", "row": 5})
        dyn = [i for i in report.impacts if i.skipped_dynamic]
        print(f"  preview delete_row skipped_dynamic={len(dyn)}")
        assert dyn, "动态函数应被跳过标记"
        assert all(not i.changed for i in dyn), "动态函数不应被机械改写"
        print("  PASS：动态函数识别 + preview 跳过")


def test_preview_delete_row_ref_error():
    """删行 → 引用被删行的单点公式产生 #REF!，preview 标 has_ref_error。"""
    print("\n=== E2E preview 删行 #REF! ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        # 在 H1 放 =A5（引用数据行5）
        wb = openpyxl.load_workbook(p)
        wb["S"]["H1"] = "=A5"
        wb.save(p)
        wb.close()
        cli = StubCodeMakerCLI(d, data_start_row=3)
        report = cli.preview_formula_impact(p, "S", {"kind": "delete_row", "row": 5})
        h1 = next(i for i in report.impacts if i.cell == "H1")
        print(f"  H1: {h1.formula_before} -> {h1.formula_after} has_ref_error={h1.has_ref_error}")
        assert h1.has_ref_error
        assert "#REF!" in h1.formula_after
        # 汇总 F13=SUM(F3:F12) 删第5行 → 范围收缩 F3:F11，无 #REF!
        f13 = next(i for i in report.impacts if i.cell == "F13")
        print(f"  F13: {f13.formula_before} -> {f13.formula_after} has_ref_error={f13.has_ref_error}")
        assert not f13.has_ref_error
        assert f13.formula_after.lower() == "=sum(f3:f11)"
        print("  PASS：删行单点引用 #REF! + 汇总范围收缩")


def test_preview_delete_col():
    """删列 preview：被删列引用 → #REF!，右侧列位移。"""
    print("\n=== E2E preview 删列 ===")
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "mul.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "M"
        ws.append(["id", "a", "c", "d", "prod"])
        ws.append(["int"] * 5)
        ws.append([1, 1, 2, 3, "=C3*D3"])  # E3=C3*D3
        wb.save(p)
        wb.close()
        cli = StubCodeMakerCLI(d, data_start_row=3)
        report = cli.preview_formula_impact(p, "M", {"kind": "delete_col", "col": 3})
        e3 = next(i for i in report.impacts if i.cell == "E3")
        print(f"  E3: {e3.formula_before} -> {e3.formula_after} ref_error={e3.has_ref_error}")
        # 删 C 列(第3)：C3 被删 #REF!，D3(col4>=3)→C3 → =#REF!*C3
        assert e3.has_ref_error
        assert e3.formula_after.lower() == "=#ref!*c3"
        print("  PASS：删列 #REF! + 右侧列位移")


def test_rewrite_invalid_cell():
    """rewrite 不存在的 sheet/cell → ok=False，不崩溃。"""
    print("\n=== E2E rewrite 失败路径 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        # 不存在的 sheet
        r = cli.rewrite_formula(p, "NoSuchSheet", "F13", "=SUM(F3:F12)")
        print(f"  无效sheet: ok={r.ok} error={r.error}")
        assert not r.ok
        # 不存在的 cell（越界坐标也能写，openpyxl 会扩展，故测无效 sheet 即可）
        print("  PASS：rewrite 失败路径不崩溃")


def test_scan_multi_sheet():
    """scan 只扫指定 sheet，不串表。"""
    print("\n=== E2E scan 多 sheet 隔离 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_cross_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        main_sems = cli.scan_sheet_formulas(p, "Main")
        ref_sems = cli.scan_sheet_formulas(p, "Ref")
        print(f"  Main 公式数={len(main_sems)} Ref 公式数={len(ref_sems)}")
        assert len(main_sems) == 5  # B3-B7
        assert len(ref_sems) == 0   # Ref 无公式
        assert all(s.sheet == "Main" for s in main_sems)
        print("  PASS：scan 按 sheet 隔离，不串表")


def test_insert_row_at_boundary():
    """insert_row 在数据起始行边界：插在范围首行(idx==rlo)→范围下移跟数据，新空行排除。

    符合 Excel 行为：插在 SUM 范围首行位置，范围跟随原数据下移(F3:F12→F4:F13)，
    新空行(row3)不在范围内。若需含新行由 AI rewrite（语义决策，非机械）。
    """
    print("\n=== E2E insert_row 边界（范围首行）===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        # 在数据起始行(3)插行 → 范围首行==idx，整体下移
        r = cli.insert_row(p, "S", 3, values={1: 777})
        assert r.ok
        formulas = _load_formulas(p, "S")
        # 汇总原 F13→F14，范围 F3:F12→F4:F13（下移跟数据，新行3排除）
        assert formulas["F14"].lower() == "=sum(f4:f13)", f"F14={formulas.get('F14')}"
        # 原数据行 F3(=SUM(B3:E3)) 下移 F4，行号 3→4
        assert formulas["F4"].lower() == "=sum(b4:e4)", f"F4={formulas.get('F4')}"
        print("  PASS：范围首行插行→范围下移跟数据（Excel 兼容）")


def test_insert_row_inside_range_expands():
    """insert_row 在范围内部(idx>rlo)→范围扩展含新行（对照边界用例）。"""
    print("\n=== E2E insert_row 范围内部扩展 ===")
    with tempfile.TemporaryDirectory() as d:
        p = _build_sum_sample(Path(d))
        cli = StubCodeMakerCLI(d, data_start_row=3)
        # 在第5行插（范围 F3:F12 内部）→ 范围扩展 F3:F13 含新行
        r = cli.insert_row(p, "S", 5, values={1: 888})
        assert r.ok
        formulas = _load_formulas(p, "S")
        assert formulas["F14"].lower() == "=sum(f3:f13)", f"F14={formulas.get('F14')}"
        print("  PASS：范围内部插行→扩展含新行")


if __name__ == "__main__":
    test_insert_row_shift()
    test_insert_row_at_boundary()
    test_insert_row_inside_range_expands()
    test_interpret_formula()
    test_interpret_cross_sheet()
    test_interpret_lookup()
    test_interpret_dynamic_func()
    test_scan_sheet_formulas()
    test_scan_multi_sheet()
    test_preview_append_gap()
    test_preview_insert_no_gap()
    test_preview_delete_row_ref_error()
    test_preview_delete_col()
    test_rewrite_invalid_cell()
    test_agent_flow_append_extend_summary()
    test_agent_flow_insert_no_rewrite_needed()
    print("\n=== 全部 L2 + insert_row E2E 测试完成 ===")
