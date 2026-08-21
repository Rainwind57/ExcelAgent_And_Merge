"""3.4 _save_with_cache_check / _invalidate 缓存清除范围收敛单测。

验证:无公式 fast-path 写操作 → row 缓存仅清受影响 sheet,他表保留命中。
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import pytest
from openpyxl import Workbook

from agent.cli_interface import StubCodeMakerCLI


def test_invalidate_sheet_scoped_clears_only_target(tmp_path):
    """_invalidate(path, sheet) 仅清该 sheet row 缓存,他表保留。"""
    cli = StubCodeMakerCLI(workspace=tmp_path, header_row=1, data_start_row=5)
    p = str(tmp_path / "t.xlsx")
    cli._row_cache[(p, "Sheet1")] = [["a"]]
    cli._row_cache[(p, "Sheet2")] = [["b"]]
    cli._cache[("wb", p, False)] = "wb_obj_1"

    cli._invalidate(Path(p), "Sheet1")

    assert (p, "Sheet1") not in cli._row_cache
    assert (p, "Sheet2") in cli._row_cache  # 他表保留
    assert ("wb", p, False) not in cli._cache  # wb 缓存按文件全清


def test_invalidate_no_sheet_clears_all(tmp_path):
    """_invalidate(path) 缺 sheet → 清该文件全部 sheet(保守,公式重算场景)。"""
    cli = StubCodeMakerCLI(workspace=tmp_path, header_row=1, data_start_row=5)
    p = str(tmp_path / "t.xlsx")
    cli._row_cache[(p, "Sheet1")] = [["a"]]
    cli._row_cache[(p, "Sheet2")] = [["b"]]

    cli._invalidate(Path(p))

    assert (p, "Sheet1") not in cli._row_cache
    assert (p, "Sheet2") not in cli._row_cache


def test_append_row_scoped_cache_survival(tmp_path):
    """无公式 fast-path:写 Sheet1 → Sheet1 row 缓存清,Sheet2 保留。"""
    p = tmp_path / "multi.xlsx"
    wb = Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s1.cell(1, 1, "id")
    s1.cell(1, 2, "名称")
    s1.cell(5, 1, 1)
    s1.cell(5, 2, "a")
    s2 = wb.create_sheet("Sheet2")
    s2.cell(1, 1, "id")
    s2.cell(1, 2, "名称")
    s2.cell(5, 1, 2)
    s2.cell(5, 2, "b")
    wb.save(p)

    cli = StubCodeMakerCLI(workspace=tmp_path, header_row=1, data_start_row=5)
    # 读两 sheet 填充 row cache
    cli.read_sheet(p, "Sheet1")
    cli.read_sheet(p, "Sheet2")
    assert (str(p), "Sheet1") in cli._row_cache
    assert (str(p), "Sheet2") in cli._row_cache

    # 写 Sheet1(无公式 → fast-path → scoped 失效)
    r = cli.append_row(p, "Sheet1", {1: 99, 2: "z"})
    assert r.ok

    assert (str(p), "Sheet1") not in cli._row_cache  # 受影响 sheet 清除
    assert (str(p), "Sheet2") in cli._row_cache      # 他表保留命中
