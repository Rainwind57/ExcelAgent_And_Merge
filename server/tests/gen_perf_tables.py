"""生成大数据量 perf 测试表（10k / 50k / 100k 行 × pet/ability/item 三表）。

克隆真实表 schema（表头 + 类型行），合成游戏化数据，写 manifest 供压测脚本读取
期望值。所有数据确定性生成（固定种子），保证压测可复现。

产物:
  resources/perf/perf_{table}_{tier}.xlsx   (9 个文件)
  resources/perf/_manifest.json             (压测脚本读期望值用)

用法:
  uv run python -m server.tests.gen_perf_tables                  # 全量
  uv run python -m server.tests.gen_perf_tables --tiers 10k      # 仅 10k
  uv run python -m server.tests.gen_perf_tables --tables pet,ability
  uv run python -m server.tests.gen_perf_tables --regen-manifest # 只重写 manifest
"""
from __future__ import annotations

import argparse
import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet

WORKSPACE = Path(__file__).resolve().parent.parent.parent
RESOURCES = WORKSPACE / "resources"
PERF_DIR = RESOURCES / "perf"
MANIFEST_PATH = PERF_DIR / "_manifest.json"

SEED = 42

TIERS = {"10k": 10_000, "50k": 50_000, "100k": 100_000}
TABLES = ("pet", "ability", "item")


# ------------------------------ schema 定义 ------------------------------
# 每表: (sheet 名, [(列名, 类型), ...], id 列名, 数据起始行)
# 布局: 行1=表头, 行2=类型行("列名:type"), 行3+=数据（统一 data_start=3）

PET_SHEET = "Pet"
PET_COLS = [
    ("灵兽id", "int"), ("名称", "string"), ("灵兽model_id", "int"),
    ("灵兽类型", "int"), ("灵兽品质", "int"), ("灵兽元素类型", "string"),
    ("体力资质", "int"), ("物攻资质", "int"), ("法攻资质", "int"),
    ("物防资质", "int"), ("法防资质", "int"), ("速度资质", "int"),
    ("成长率", "float"), ("天赋技能", "int"),
    ("面板气血上限", "int"), ("面板物理攻击", "int"), ("面板法术攻击", "int"),
    ("面板物理防御", "int"), ("面板速度", "int"), ("面板物理暴击", "float"),
]
PET_ID_COL = "灵兽id"

ABILITY_SHEET = "Ability"
ABILITY_COLS = [
    ("神通id", "int"), ("名称", "string"), ("神通描述", "string"),
    ("图标", "string"), ("技能id", "int"), ("技能等级", "int"),
    ("被动id", "int"), ("属性", "string"),
]
ABILITY_ID_COL = "神通id"

ITEM_SHEET = "ItemBase"
ITEM_COLS = [
    ("物品编号", "int"), ("名称", "string"), ("备注", "string"),
    ("品质", "int"), ("道具类型", "int"), ("程序处理type", "string"),
    ("道具描述", "string"), ("图标", "string"), ("最大堆叠数量", "int"),
    ("可丢弃", "int"), ("可使用的场景", "int"), ("是否自动使用", "int"),
    ("境界", "int"),
]
ITEM_ID_COL = "物品编号"

SCHEMAS = {
    "pet": (PET_SHEET, PET_COLS, PET_ID_COL),
    "ability": (ABILITY_SHEET, ABILITY_COLS, ABILITY_ID_COL),
    "item": (ITEM_SHEET, ITEM_COLS, ITEM_ID_COL),
}

# ------------------------------ 游戏化数据池 ------------------------------

PET_BASE_NAMES = [
    "火焰犬", "烈焰犬", "炽焰獒", "小白鲸", "木结花", "小顽蝠", "慢慢龟", "岩岩石",
    "朱雀", "玄武", "青龙", "白虎", "麒麟", "凤凰", "饕餮", "穷奇", "梼杌", "混沌",
    "金翎雕", "墨鳞蛇", "碧水蛟", "赤焰马", "青霜鹤", "紫电豹", "黄沙蝎", "玄冰龟",
]
PET_ELEMENTS = ["火", "水", "木", "金", "土", "风", "雷", "光", "暗"]
PET_TYPES = [1, 2, 3]  # 1=普通, 2=稀有, 3=神兽
PET_QUALITIES = [1, 2, 3, 4, 5]  # 凡/良/上/珍/绝

ABILITY_BASE_NAMES = [
    "三味真火", "蛮牛狂击", "冰封千里", "雷霆万钧", "枯木逢春", "金钟罩", "厚土壁垒",
    "风卷残云", "光明普照", "暗影袭杀", "烈焰风暴", "寒冰刺骨", "万剑归宗", "天雷诀",
    "回春术", "金刚不坏", "流沙陷阱", "九天雷劫", "圣光治愈", "幽冥毒雾",
]
ABILITY_ICONS = [f"icon_spell_{i:03d}" for i in range(1, 60)]
ABILITY_DESC_TEMPLATES = [
    "对{target}造成{dmg}点{elem}系伤害",
    "复活死亡{target}，恢复{dmg}点气血",
    "对{target}施加{elem}系减速效果，持续{dmg}回合",
    "提升己方{target}的{elem}系抗性{dmg}点",
    "对全体敌方造成{dmg}点{elem}系法术伤害",
]

ITEM_BASE_NAMES = {
    1: ["回血丹", "回蓝丹", "解毒散", "复活丹", "怒气丸", "经验丹"],  # 药品
    2: ["玄铁剑", "青锋剑", "紫电枪", "烈焰刀", "寒冰杖", "金钟盾"],  # 装备
    3: ["翡翠石", "光芒石", "太阳石", "月亮石", "神秘石", "黄宝石"],  # 宝石
    4: ["铁矿", "木材", "兽皮", "灵草", "灵石", "妖丹"],  # 材料
    5: ["测试道具", "礼包", "宝箱", "钥匙", "信物", "卷轴"],  # 杂物
}
ITEM_QUALITIES = [1, 2, 3, 4, 5]  # 凡/良/上/珍/绝
ITEM_ICONS = [f"Icon_256_{9000000 + i}" for i in range(0, 200, 7)]
ITEM_DESCS = ["道具", "药品", "装备", "宝石", "材料", "礼包", "卷轴", "信物"]


# ------------------------------ 行生成器 ------------------------------

def gen_pet_row(i: int, rng: random.Random) -> dict:
    """i 从 0 起，灵兽id = 100001 + i。"""
    pet_id = 100001 + i
    base = PET_BASE_NAMES[i % len(PET_BASE_NAMES)]
    variant = i // len(PET_BASE_NAMES) + 1
    name = f"{base}·{variant}" if variant > 1 else base
    quality = PET_QUALITIES[(i * 3) % len(PET_QUALITIES)]
    pet_type = PET_TYPES[(i * 5) % len(PET_TYPES)]
    element = PET_ELEMENTS[(i * 7) % len(PET_ELEMENTS)]
    # 资质: 品质越高基数越大，加确定性扰动
    base_apt = 800 + quality * 200
    phys_atk = base_apt + (i * 7) % 1500          # 物攻资质 (压测 get 的目标列)
    hp_apt = base_apt + (i * 11) % 2000           # 体力资质
    mag_atk = base_apt + (i * 13) % 1400
    phys_def = base_apt + (i * 17) % 1200
    mag_def = base_apt + (i * 19) % 1200
    speed = base_apt + (i * 23) % 1000
    growth = round(1.0 + quality * 0.15 + (i % 50) * 0.01, 2)  # 成长率
    talent = 5001 + (i % 50)
    # 面板 (资质 × 品质系数)
    panel_hp = hp_apt * 10 + quality * 500
    panel_patk = phys_atk * 2 + quality * 100
    panel_matk = mag_atk * 2 + quality * 100
    panel_pdef = phys_def * 2
    panel_speed = speed // 10
    panel_crit = round(0.05 + (i % 20) * 0.01, 2)
    return {
        "灵兽id": pet_id, "名称": name, "灵兽model_id": pet_id,
        "灵兽类型": pet_type, "灵兽品质": quality, "灵兽元素类型": element,
        "体力资质": hp_apt, "物攻资质": phys_atk, "法攻资质": mag_atk,
        "物防资质": phys_def, "法防资质": mag_def, "速度资质": speed,
        "成长率": growth, "天赋技能": talent,
        "面板气血上限": panel_hp, "面板物理攻击": panel_patk,
        "面板法术攻击": panel_matk, "面板物理防御": panel_pdef,
        "面板速度": panel_speed, "面板物理暴击": panel_crit,
    }


def gen_ability_row(i: int, rng: random.Random) -> dict:
    """神通id = 5001 + i。"""
    ab_id = 5001 + i
    base = ABILITY_BASE_NAMES[i % len(ABILITY_BASE_NAMES)]
    variant = i // len(ABILITY_BASE_NAMES) + 1
    name = f"{base}·{variant}" if variant > 1 else base
    elem = PET_ELEMENTS[(i * 7) % len(PET_ELEMENTS)]
    dmg = 100 + (i * 13) % 900
    desc = rng.choice(ABILITY_DESC_TEMPLATES).format(
        target="敌方目标", dmg=dmg, elem=elem)
    icon = ABILITY_ICONS[(i * 3) % len(ABILITY_ICONS)]
    spell_id = 100001 + (i % 5000)
    spell_level = 1 + (i % 10)
    passive_id = (5001 + (i * 17) % 3000) if i % 4 == 0 else None
    attrs = f"{{\"{elem}Atk\":{dmg}}}" if i % 3 == 0 else None
    return {
        "神通id": ab_id, "名称": name, "神通描述": desc,
        "图标": icon, "技能id": spell_id, "技能等级": spell_level,
        "被动id": passive_id, "属性": attrs,
    }


def gen_item_row(i: int, rng: random.Random) -> dict:
    """物品编号 = 20001 + i。"""
    item_id = 20001 + i
    item_type = (i % 5) + 1  # 1..5
    base_pool = ITEM_BASE_NAMES[item_type]
    base = base_pool[i % len(base_pool)]
    variant = i // len(base_pool) + 1
    name = f"{base}·{variant}" if variant > 1 else base
    quality = ITEM_QUALITIES[(i * 3) % len(ITEM_QUALITIES)]
    icon = ITEM_ICONS[(i * 5) % len(ITEM_ICONS)]
    desc = ITEM_DESCS[(i * 7) % len(ITEM_DESCS)]
    max_stack = 99 if item_type == 1 else 1
    droppable = 1
    use_scene = 0
    auto_use = 1 if item_type == 1 else 0
    realm = 1 + (i % 10)
    note = None
    handler = None
    return {
        "物品编号": item_id, "名称": name, "备注": note,
        "品质": quality, "道具类型": item_type, "程序处理type": handler,
        "道具描述": desc, "图标": icon, "最大堆叠数量": max_stack,
        "可丢弃": droppable, "可使用的场景": use_scene,
        "是否自动使用": auto_use, "境界": realm,
    }


ROW_GENS = {"pet": gen_pet_row, "ability": gen_ability_row, "item": gen_item_row}


# ------------------------------ 写表 ------------------------------

def _cell_value(v):
    """openpyxl write_only 不接受 None 之外的某些类型，统一规整。"""
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def write_table(table: str, tier: str, rows: int) -> dict:
    sheet_name, cols, id_col = SCHEMAS[table]
    out_path = PERF_DIR / f"perf_{table}_{tier}.xlsx"
    rng = random.Random(SEED + hash(table + tier) % 100000)

    wb = Workbook(write_only=True)
    ws: WriteOnlyWorksheet = wb.create_sheet(sheet_name)

    # 行1: 表头
    ws.append([name for name, _ in cols])
    # 行2: 类型行 "列名:type"
    ws.append([f"{name}:{typ}" for name, typ in cols])

    gen = ROW_GENS[table]
    sample_id = None
    sample_values = None
    # 取第 5 行（i=4）作为压测样本，记录期望值
    sample_i = 4
    for i in range(rows):
        row_dict = gen(i, rng)
        ws.append([_cell_value(row_dict.get(name)) for name, _ in cols])
        if i == sample_i:
            sample_id = row_dict[id_col]
            sample_values = dict(row_dict)  # 记录全部字段供压测读期望值

    wb.save(out_path)
    wb.close()

    # write_only 模式不写 <dimension> 元数据，read_only 模式读回 max_row=None，
    # 会导致 table_index.build_index 崩溃。用 normal 模式 re-save 嵌入 dimension。
    from openpyxl import load_workbook as _load
    wb2 = _load(out_path)  # normal 模式，计算并写入 dimension
    wb2.save(out_path)
    wb2.close()

    return {
        "table": table,
        "tier": tier,
        "file": f"perf/perf_{table}_{tier}.xlsx",
        "absolute_path": str(out_path),
        "sheet": sheet_name,
        "rows": rows,
        "cols": len(cols),
        "id_col": id_col,
        "header_row": 1,
        "type_row": 2,
        "data_start_row": 3,
        "columns": [{"name": n, "type": t} for n, t in cols],
        "sample_row_index": sample_i,
        "sample_id": sample_id,
        "sample_values": sample_values,
    }


def build_manifest(entries: list[dict]) -> dict:
    by_key = {}
    for e in entries:
        by_key[f"perf_{e['table']}_{e['tier']}"] = e
    return {
        "seed": SEED,
        "tables": TABLES,
        "tiers": list(TIERS.keys()),
        "entries": by_key,
    }


def main():
    ap = argparse.ArgumentParser(description="生成大数据量 perf 测试表")
    ap.add_argument("--tiers", default="10k,50k,100k",
                    help="档位，逗号分隔 (默认 10k,50k,100k)")
    ap.add_argument("--tables", default="pet,ability,item",
                    help="表，逗号分隔 (默认 pet,ability,item)")
    ap.add_argument("--regen-manifest", action="store_true",
                    help="只重写 manifest（不重写 xlsx）")
    args = ap.parse_args()

    tiers = [t.strip() for t in args.tiers.split(",") if t.strip()]
    tables = [t.strip() for t in args.tables.split(",") if t.strip()]
    for t in tiers:
        if t not in TIERS:
            raise SystemExit(f"未知档位: {t} (可选: {list(TIERS)})")
    for t in tables:
        if t not in TABLES:
            raise SystemExit(f"未知表: {t} (可选: {list(TABLES)})")

    PERF_DIR.mkdir(parents=True, exist_ok=True)

    if args.regen_manifest:
        print(f"[manifest] 仅重写 manifest → {MANIFEST_PATH}")
        entries = []
        for table in tables:
            for tier in tiers:
                # 从现有 manifest 读取条目（若存在）
                if MANIFEST_PATH.exists():
                    old = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
                    key = f"perf_{table}_{tier}"
                    if key in old.get("entries", {}):
                        entries.append(old["entries"][key])
        manifest = build_manifest(entries)
        MANIFEST_PATH.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[manifest] 完成，{len(entries)} 条")
        return

    entries = []
    for table in tables:
        for tier in tiers:
            rows = TIERS[tier]
            print(f"[gen] perf_{table}_{tier} ({rows} 行) ...", flush=True)
            t0_path = PERF_DIR / f"perf_{table}_{tier}.xlsx"
            entry = write_table(table, tier, rows)
            size_mb = t0_path.stat().st_size / 1024 / 1024
            entry["file_size_mb"] = round(size_mb, 2)
            entries.append(entry)
            print(f"       → {t0_path.name}  {size_mb:.1f} MB  "
                  f"sample_id={entry['sample_id']}  "
                  f"sample_values={entry['sample_values']}", flush=True)

    manifest = build_manifest(entries)
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[manifest] 写入 {MANIFEST_PATH} ({len(entries)} 条)")
    print("[done] 全部生成完毕")


if __name__ == "__main__":
    main()
