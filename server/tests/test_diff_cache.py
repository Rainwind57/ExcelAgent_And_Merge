"""4.8 diff 结果缓存单测:同文件命中、文件变化失效、指纹。"""
from __future__ import annotations

import os
import sys
import time
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook

from routers import diff


def _make(p: Path, v=1):
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.cell(1, 1, "id")
    ws.cell(2, 1, v)
    wb.save(p)
    return p


def test_file_fingerprint_sorted_and_missing(tmp_path):
    p1 = _make(tmp_path / "a.xlsx")
    p2 = _make(tmp_path / "b.xlsx")
    fp = diff._file_fingerprint([str(p2), str(p1)])
    assert fp is not None
    # 排序后 a 在前
    assert fp[0][0] == "a.xlsx" and fp[1][0] == "b.xlsx"
    # 不存在文件 → None(不缓存)
    assert diff._file_fingerprint([str(tmp_path / "nope.xlsx")]) is None


def test_diff_cache_hit_same_files(tmp_path, monkeypatch):
    """同一文件集二次 compare → 第二次命中缓存(compare_sheet 仅调一次)。"""
    diff.reset_diff_cache()
    p1 = _make(tmp_path / "base.xlsx")
    p2 = _make(tmp_path / "dev.xlsx", v=2)
    calls = {"n": 0}

    def fake_compare(file_sheets, base_name, sheet_name, *a, **kw):
        calls["n"] += 1
        return {"rows": [], "headers": ["id"], "stats": {"total_rows": 0},
                "missing_rows": [], "structure_diff": None}

    monkeypatch.setattr(diff, "compare_sheet", fake_compare)
    fl = [str(p1), str(p2)]
    r1 = diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", False)
    r2 = diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", False)
    assert calls["n"] == 1, "第二次应命中缓存不调 compare_sheet"
    assert r1 is r2, "命中应返回同一缓存对象"


def test_diff_cache_invalidate_on_change(tmp_path, monkeypatch):
    """文件内容变化 → 指纹变 → 缓存失效重算。"""
    diff.reset_diff_cache()
    p1 = _make(tmp_path / "base.xlsx")
    p2 = _make(tmp_path / "dev.xlsx", v=2)
    calls = {"n": 0}

    def fake_compare(*a, **kw):
        calls["n"] += 1
        return {"rows": [], "headers": ["id"], "stats": {"total_rows": 0},
                "missing_rows": [], "structure_diff": None}

    monkeypatch.setattr(diff, "compare_sheet", fake_compare)
    fl = [str(p1), str(p2)]
    diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", False)
    # 改文件内容(size 变)+ 强制 mtime 变
    _make(p2, v=999)
    time.sleep(0.01)
    os.utime(p2, None)
    diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", False)
    assert calls["n"] == 2, "文件变化后应失效重算"


def test_diff_cache_detect_missing_distinct_key(tmp_path, monkeypatch):
    """detect_missing 不同 → 不同缓存条目(都需重算)。"""
    diff.reset_diff_cache()
    p1 = _make(tmp_path / "base.xlsx")
    p2 = _make(tmp_path / "dev.xlsx", v=2)
    calls = {"n": 0}

    def fake_compare(*a, **kw):
        calls["n"] += 1
        return {"rows": [], "headers": ["id"], "stats": {"total_rows": 0},
                "missing_rows": [], "structure_diff": None}

    monkeypatch.setattr(diff, "compare_sheet", fake_compare)
    fl = [str(p1), str(p2)]
    diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", False)
    diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", True)  # detect_missing 变
    assert calls["n"] == 2, "detect_missing 不同应分别重算"
    # 再调同 detect_missing=True → 命中
    diff._cached_compare_sheet(fl, {}, None, None, "base.xlsx", "S1", True)
    assert calls["n"] == 2
