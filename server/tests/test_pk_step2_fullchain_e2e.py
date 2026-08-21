"""核心4 PK 冲突前移 Step2 全链路 e2e（输入→输出）。

复现用户场景：新增 reward_id=99001（已占用）奖励包。
改动前：Step2 漏检 → 落 Step3 _do_append 才抓 pk_conflict → id_reallocate
  自撞（轮1 写 100603、轮2 又算 100603）→ post-exhaustion accept 救不回 →
  汇总报 verify_repair_exhausted 但表里躺着 100603 半成品。
改动后：Step2 validate_two_layer 检测 + ask + 改写 intent（BEFORE 写盘）→
  Step3 _do_append 用改写后的 100603 干净写入，零 pk_conflict。

覆盖：
  A. 交互接受建议：intent PK 改写 → _do_append 成功 → 无 pk_conflict
  B. 无 cb：intent 标 skipped=True（不落 Step3 半成品）
  C. id_reallocate 自撞防护：_allocate_pk(exclude=) 跳过已试 ID
  D. _apply_repair_fix 用真实 PK 列名（reward_id 而非泛 id）

运行: python -m pytest server/tests/test_pk_step2_fullchain_e2e.py -v
"""
from __future__ import annotations

import os
import sys
import types
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.cli_interface import StubCodeMakerCLI
from agent.agent import TableAgent, AgentResult
from agent.nl_parser import NLIntent
from agent.excel.schema_bundle import build_data_getter
from agent.excel.subagent.validator_agent import ValidatorAgent


# reward 表头（1-based 列号 → 列名）
HEADERS = ["reward_id", "名称", "每日领取上限", "必得道具1", "道具1数量"]


def _build_reward_xlsx(ws: Path) -> Path:
    """建 reward.xlsx：表头行1，数据行5起，预置 99001 + 100600..100602。"""
    wb = Workbook()
    s = wb.active
    s.title = "Reward"
    for i, h in enumerate(HEADERS, 1):
        s.cell(1, i, h)
    # 数据行5起：99001（用户要新增的，已占用）+ 100600/100601/100602
    rows = [
        [99001, "已存在的包", 1, 10001, 5],
        [100600, "包A", 1, 10001, 5],
        [100601, "包B", 1, 10001, 5],
        [100602, "包C", 1, 10001, 5],
    ]
    for r, row in enumerate(rows, start=5):
        for c, v in enumerate(row, 1):
            s.cell(r, c, v)
    p = ws / "reward.xlsx"
    wb.save(p)
    return p


def _make_agent(cli) -> types.SimpleNamespace:
    """绑定 _do_append 及其依赖 + _allocate_pk/_real_pk_col_name/_locate_pk_col。"""
    agent = types.SimpleNamespace(cli=cli, live_index=False)
    for name in (
        "_do_append", "_auto_sort_after_write", "_refresh_index_after_write",
        "_verify_write_back", "_allocate_pk", "_is_misplaced_pk",
        "_real_pk_col_name", "_locate_pk_col",
    ):
        setattr(agent, name, getattr(TableAgent, name).__get__(agent))
    return agent


def _make_validator() -> ValidatorAgent:
    v = object.__new__(ValidatorAgent)
    v._cli = None
    v._parser = None
    v._ask_callback = None
    v._required_fields = None
    return v


def _schema_getter(agent):
    from agent.excel.schema_bundle import _stem_to_path, _resolve_sheet, _resolve_path
    def _sg(intent):
        stem = getattr(intent, "table_hint", "") or ""
        sheet = getattr(intent, "sheet_hint", "") or ""
        # 镜像 agent.py _4step_sg：sheet/path 空时经 resolver + cli 回退
        if not sheet and stem:
            sheet = _resolve_sheet(agent, stem)
        path = _stem_to_path(agent, stem)
        if path is None and stem:
            path = _resolve_path(agent, stem)
        if path is None or not sheet:
            return [], []
        try:
            return (list(agent.cli.read_header(path, sheet)),
                    list(agent.cli.read_type_row(path, sheet)))
        except Exception:
            return [], []
    return _sg


def _intent(fields=None):
    return NLIntent(
        action="add", table_hint="reward", sheet_hint="Reward",
        raw="新增一个奖励包叫测试奖励包,reward_id 99001,每日限领 1 次,"
            "必给道具 10001 共 5 个",
        extras={"fields": fields or {}})


@pytest.fixture
def env():
    with TemporaryDirectory() as tmp:
        ws = Path(tmp)
        cli = StubCodeMakerCLI(workspace=ws, header_row=1, data_start_row=5)
        path = _build_reward_xlsx(ws)
        agent = _make_agent(cli)
        v = _make_validator()
        sg = _schema_getter(agent)
        dg = build_data_getter(agent)
        yield agent, cli, path, v, sg, dg


# ───────────────────── A. 交互接受建议：干净写入 ─────────────────────

def test_A_accept_suggest_rewrites_then_clean_write(env):
    """Step2 检测 99001 占用 → ask accept → intent PK 改 100603 →
    Step3 _do_append 用 100603 写盘成功，零 pk_conflict。"""
    agent, cli, path, v, sg, dg = env
    v.set_ask_callback(lambda q: {"accept_suggest": True})

    it = _intent({"reward_id": 99001, "名称": "测试奖励包",
                  "每日领取上限": 1, "必得道具1": 10001, "道具1数量": 5})
    vr = v.validate_two_layer([it], schema_getter=sg, data_getter=dg)

    # Step2 改写 intent PK 到建议值 max(100602)+1=100603
    assert it.extras["fields"]["reward_id"] == 100603, \
        f"accept 后 PK 应改写为 100603，实际 {it.extras['fields'].get('reward_id')}"
    # 非阻断 ok=True，但 issue 已从 tips 移除（冲突消除）
    assert vr["ok"] is True

    # Step3：用改写后的 intent 跑 _do_append（模拟写盘）
    new_pk = it.extras["fields"]["reward_id"]
    values = {1: new_pk, 2: "测试奖励包", 3: 1, 4: 10001, 5: 5}
    res = AgentResult(ok=True, intent=it)
    agent._do_append(path, "Reward", values, res)

    # 干净写入：ok=True，无 pk_conflict，行追加成功
    assert res.ok, f"应写入成功，实际 message={res.message}"
    _names = _step_names(res)
    assert "pk_conflict" not in _names, \
        "Step2 已改写 PK，Step3 不应再出现 pk_conflict"
    assert any(getattr(s, "name", "") == "append_row" and getattr(s, "ok", False)
               for s in _steps(res)), "应有 append_row 成功步骤"


def test_A_suggested_id_is_max_plus_one(env):
    """建议 ID = 现有最大值(100602)+1 = 100603。"""
    agent, cli, path, v, sg, dg = env
    asked = []
    v.set_ask_callback(lambda q: (asked.append(q), {"accept_suggest": True})[1])
    it = _intent({"reward_id": 99001, "名称": "测试奖励包"})
    v.validate_two_layer([it], schema_getter=sg, data_getter=dg)
    assert asked, "应触发 ask"
    assert asked[0]["suggested_id"] == 100603
    assert asked[0]["mode_hint"] == "pk_conflict"
    assert "99001" in asked[0]["suggestion"]


# ───────────────────── B. 无 cb：标 skipped 不落 Step3 ─────────────────────

def test_B_no_callback_marks_skipped(env):
    """无 cb（非交互）→ Step2 检测到冲突标 validation.skipped=True，
    intent PK 不改（避免静默改用户显式 ID），不落 Step3 半成品路径。"""
    agent, cli, path, v, sg, dg = env
    # v._ask_callback 已为 None（_make_validator 设的）
    it = _intent({"reward_id": 99001, "名称": "测试奖励包"})
    vr = v.validate_two_layer([it], schema_getter=sg, data_getter=dg)

    assert vr["ok"] is False  # 要求 A：无 cb 标 skipped → 真阻断（非 ok=True 恒）
    _v = getattr(it, "validation", None)
    assert _v is not None and getattr(_v, "skipped", False) is True, \
        "无 cb 时应标 skipped=True，让 _phase_execute 跳写盘"
    # PK 未被静默改写
    assert it.extras["fields"]["reward_id"] == 99001
    # issue 留在 tips → 走软失败上报（Step4 显式列出未解决）
    assert vr.get("tips"), "skipped 的 PK 冲突应留在 tips 走软失败上报"


def test_B_user_skip_marks_skipped(env):
    """用户主动 skip → 同样标 skipped=True。"""
    agent, cli, path, v, sg, dg = env
    v.set_ask_callback(lambda q: {"mode": "skip"})
    it = _intent({"reward_id": 99001, "名称": "测试奖励包"})
    v.validate_two_layer([it], schema_getter=sg, data_getter=dg)
    _v = getattr(it, "validation", None)
    assert _v is not None and getattr(_v, "skipped", False) is True
    assert it.extras["fields"]["reward_id"] == 99001  # 不改写


# ───────────────────── C. id_reallocate 自撞防护 ─────────────────────

def test_C_allocate_pk_exclude_avoids_self_collision(env):
    """轮1 写 100603 后（未回滚/读取时序问题），轮2 _allocate_pk
    带 exclude={100603} 应返 100604，不再自撞。"""
    agent, cli, path, v, sg, dg = env
    # 表里 max=100602 → 不带 exclude 返 100603
    n1 = agent._allocate_pk(path, "Reward", 1)
    assert n1 == 100603
    # 带 exclude={100603}（模拟本轮已试）→ 跳过返 100604
    n2 = agent._allocate_pk(path, "Reward", 1, exclude={100603})
    assert n2 == 100604, f"exclude 100603 后应返 100604，实际 {n2}"


def test_C_apply_repair_fix_real_pk_col_name(env):
    """_apply_repair_fix(allocate_new_id) 用真实 PK 列名 reward_id 写入 fields，
    非泛 'id' 键（matcher 别名回退脆弱）。"""
    from agent.nl_parser import NLIntent
    agent, cli, path, v, sg, dg = env
    # 绑定 _apply_repair_fix（依赖 _allocate_pk + _real_pk_col_name 已绑）
    agent._apply_repair_fix = TableAgent._apply_repair_fix.__get__(agent)
    it = NLIntent(action="add", table_hint="reward", sheet_hint="Reward",
                 raw="test", extras={"fields": {"reward_id": 99001, "名称": "x"}})
    applied = agent._apply_repair_fix(it, path, "Reward", {"allocate_new_id": True})
    assert applied
    fields = it.extras["fields"]
    # 真实 PK 列名 reward_id 被填入新 ID（非 "id" 键）
    assert fields.get("reward_id") == 100603, \
        f"应用新 PK 后 reward_id 应为 100603，实际 {fields.get('reward_id')}"
    assert "id" not in fields or fields.get("id") in (None,), \
        "不应再产生泛 'id' 键"
    assert it.extras["pk_value"] == 100603


def test_C_apply_repair_fix_excludes_attempted_ids(env):
    """连续两次 _apply_repair_fix(allocate_new_id) 带 exclude 集 →
    第二次不再返同值（防 id_reallocate 多轮自撞）。"""
    from agent.nl_parser import NLIntent
    agent, cli, path, v, sg, dg = env
    agent._apply_repair_fix = TableAgent._apply_repair_fix.__get__(agent)
    it = NLIntent(action="add", table_hint="reward", sheet_hint="Reward",
                 raw="test", extras={"fields": {"reward_id": 99001, "名称": "x"}})
    attempted: set = set()
    # 轮1
    agent._apply_repair_fix(it, path, "Reward", {"allocate_new_id": True},
                            exclude=attempted)
    id1 = it.extras["fields"]["reward_id"]
    attempted.add(id1)
    # 轮2：带 exclude（模拟 loop 维护的 _attempted_ids）
    agent._apply_repair_fix(it, path, "Reward", {"allocate_new_id": True},
                            exclude=attempted)
    id2 = it.extras["fields"]["reward_id"]
    assert id1 != id2, f"两轮 allocate_new_id 不应同值（{id1}），否则 id_reallocate 自撞"


# ───────────────────── D. sheet 回退（sheet_hint 空也能检测） ───────────

def test_D_sheet_fallback_when_sheet_hint_empty(env):
    """sheet_hint 空时 data_getter 经 TableResolver 回退解析 sheet，
    existing_values 非空 → 核心4 仍能检测 PK 占用。"""
    agent, cli, path, v, sg, dg = env
    v.set_ask_callback(lambda q: {"accept_suggest": True})
    # intent 不带 sheet_hint，仅 table_hint="reward"
    it = NLIntent(action="add", table_hint="reward", sheet_hint="",
                 raw="test", extras={"fields": {"reward_id": 99001, "名称": "x"}})
    data = dg(it)
    # data_getter 经回退解析到 sheet → existing_values 含 reward_id
    ev = data.get("existing_values") or {}
    assert ev.get("reward_id"), \
        f"sheet 回退后 existing_values 应含 reward_id 列，实际 {list(ev.keys())}"
    # 核心4 能检测 99001 占用 → ask
    asked = []
    v.set_ask_callback(lambda q: (asked.append(q), {"accept_suggest": True})[1])
    v.validate_two_layer([it], schema_getter=sg, data_getter=dg)
    assert asked, "sheet 回退后核心4 应检测到 99001 占用并 ask"


# ───────────────────── E. 前后端契约：ask payload ↔ reply 回传 ───────────

def test_E_ask_payload_matches_frontend_reply_contract(env):
    """前后端统一契约（AgentChatView.replyAskPk ↔ _ask_pk_conflict）：
    1. 后端 ask payload 含 mode_hint=pk_conflict + suggested_id + suggestion(含原值)
    2. 前端「接受并续跑」回传 {mode:field, accept_suggest:true} → 后端
       _ask_pk_conflict 命中 accept_suggest 分支 → _apply_pk_to_intent 改写 PK
    3. 前端「自定义」回传 {mode:field, custom_id:N} → 命中 custom_id 分支
    4. 改写后 intent PK = 建议/自定义值，Step3 可干净写入（无 pk_conflict）
    """
    agent, cli, path, v, sg, dg = env

    # ── 契约 1+2：accept_suggest ──
    asked = []
    v.set_ask_callback(lambda q: (asked.append(q), {"accept_suggest": True})[1])
    it1 = _intent({"reward_id": 99001, "名称": "测试奖励包"})
    v.validate_two_layer([it1], schema_getter=sg, data_getter=dg)
    assert asked, "应触发 ask"
    q = asked[0]
    # 后端 ask payload 字段（前端 AgentChatView.vue:960 依此渲染）
    assert q["mode_hint"] == "pk_conflict"
    assert q["suggested_id"] == 100603
    assert "99001" in q["suggestion"] and "100603" in q["suggestion"]
    # 前端 replyAskPk 回传 accept_suggest → 后端改写 intent PK
    assert it1.extras["fields"]["reward_id"] == 100603

    # ── 契约 3：custom_id ──
    v.set_ask_callback(lambda q: {"custom_id": 99999})
    it2 = _intent({"reward_id": 99001, "名称": "测试奖励包2"})
    v.validate_two_layer([it2], schema_getter=sg, data_getter=dg)
    assert it2.extras["fields"]["reward_id"] == 99999

    # ── 契约 4：改写后 Step3 _do_append 干净写入，无 pk_conflict ──
    values = {1: it1.extras["fields"]["reward_id"], 2: "测试奖励包",
              3: 1, 4: 10001, 5: 5}
    res = AgentResult(ok=True, intent=it1)
    agent._do_append(path, "Reward", values, res)
    assert res.ok, f"改写后应干净写入，message={res.message}"
    assert "pk_conflict" not in _step_names(res)


# ───────────────────── F. 真根因回归：list 型字段值不崩 validate ──────

def test_F_list_typed_field_value_does_not_crash_validate(env):
    """真根因回归：字段值为 list（LLM 产 [10001,10002] 等）时，
    validate_field_layer 旧代码 `val in ev`(set) 抛 TypeError 中断
    validate_two_layer → Core4 PK 前移检查跑不到 → 冲突落 Step3。
    修后走 str 比较，不崩，Core4 仍能检测 PK 冲突。
    """
    agent, cli, path, v, sg, dg = env
    asked = []
    v.set_ask_callback(lambda q: (asked.append(q), {"accept_suggest": True})[1])
    # 必得道具1 给 list 型值（模拟 LLM 过产/数组字段）
    it = _intent({"reward_id": 99001, "名称": "测试奖励包",
                  "每日领取上限": 1, "必得道具1": [10001, 10002],
                  "道具1数量": 5})
    # 不应抛 TypeError
    vr = v.validate_two_layer([it], schema_getter=sg, data_getter=dg)
    # Core4 仍检测到 PK 冲突并 ask
    assert asked, "list 型字段值不应阻断 Core4 PK 检测"
    assert asked[0]["suggested_id"] == 100603
    assert it.extras["fields"]["reward_id"] == 100603  # accept 后改写


def _steps(res) -> list:
    """从 AgentResult 取 step 列表（AgentStep dataclass，属性访问）。"""
    s = getattr(res, "steps", None) or getattr(res, "result_steps", None)
    return list(s or [])


def _step_names(res) -> list:
    return [getattr(s, "name", "") for s in _steps(res)]
