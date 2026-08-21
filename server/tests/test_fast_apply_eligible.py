"""4.7 fast_apply _eligible 资格扩展单测。

含公式/批注且仅更新 → 放行;含合并单元格或结构变更(插入/删除)与公式/批注共存 → 回退。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment

import engine.fast_apply as fa
from engine.fast_apply import _eligible


def _save(p: Path) -> Path:
    # 多填一些行保证 openpyxl 输出非空(eligible 的 size 检查测试里 monkeypatch 为 0)
    return p


def _updates(sheet="S1"):
    return {sheet: {"updates": {"1": {2: "v"}}, "deleted": [], "inserts": []}}


def _structural(sheet="S1"):
    return {sheet: {"updates": {}, "deleted": ["1"], "inserts": []}}


def test_eligible_plain_file(tmp_path, monkeypatch):
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "plain.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "id"); ws.cell(2, 1, 1)
    wb.save(p)
    assert _eligible(p, _updates()) is None


def test_eligible_formula_update_only_passes(tmp_path, monkeypatch):
    """含公式 + 仅更新 → 放行(4.7 扩展)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "fml.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "a"); ws.cell(1, 2, "b"); ws.cell(1, 3, "sum")
    ws.cell(2, 1, 1); ws.cell(2, 2, 2); ws.cell(2, 3, "=A2+B2")
    wb.save(p)
    assert _eligible(p, _updates()) is None, "含公式仅更新应放行快路径"


def test_eligible_formula_structural_rejects(tmp_path, monkeypatch):
    """含公式 + 结构变更(删除)→ 回退(行号位移致公式引用漂移)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "fml.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(2, 3, "=A2+B2"); ws.cell(2, 1, 1); ws.cell(2, 2, 2)
    wb.save(p)
    reason = _eligible(p, _structural())
    assert reason is not None and "公式" in reason


def test_eligible_comment_update_only_passes(tmp_path, monkeypatch):
    """含批注 + 仅更新 → 放行(批注独立 xml,纯更新不改行号)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "cmt.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "id"); ws.cell(2, 1, 1)
    ws.cell(2, 1).comment = Comment("hello", "tester")
    wb.save(p)
    assert _eligible(p, _updates()) is None, "含批注仅更新应放行快路径"


def test_eligible_comment_structural_rejects(tmp_path, monkeypatch):
    """含批注 + 结构变更 → 回退(行号位移致批注引用漂移)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "cmt.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "id"); ws.cell(2, 1, 1)
    ws.cell(2, 1).comment = Comment("hello", "tester")
    wb.save(p)
    reason = _eligible(p, _structural())
    assert reason is not None and "批注" in reason


def test_eligible_merged_always_rejects(tmp_path, monkeypatch):
    """含合并单元格 → 始终回退(非 anchor 写入静默失效)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "merge.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(1, 1, "id"); ws.cell(2, 1, 1)
    ws.merge_cells("B2:C2")
    wb.save(p)
    reason = _eligible(p, _updates())
    assert reason is not None and "合并" in reason


def test_eligible_no_edits_passes_formula(tmp_path, monkeypatch):
    """无编辑(纯复制)+ 含公式 → 放行(shutil.copy2,不触碰公式)。"""
    monkeypatch.setattr(fa, "_FAST_MIN_SIZE", 0)
    p = tmp_path / "fml.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    ws.cell(2, 3, "=A2+B2"); ws.cell(2, 1, 1); ws.cell(2, 2, 2)
    wb.save(p)
    assert _eligible(p, {}) is None
    assert _eligible(p, None) is None
