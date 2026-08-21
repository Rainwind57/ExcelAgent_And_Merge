"""缺口5 delete 反向引用清理单测（case6 删 quest 连带清 spawn/reward）。

不依赖 serve LLM。mock cli + cascade_resolver，验证 _do_cascade_delete
主动调 get_referencing_tables 追溯子表并删行。
"""
from pathlib import Path
from unittest.mock import MagicMock, patch

from agent.excel.cli_interface import StubCodeMakerCLI
from agent.excel.core.agent import TableAgent
from agent.excel.parser.nl_parser import NLIntent


class _MockParser:
    def __init__(self):
        self.client = None
        self.directory = ""
        self.model = ""


def _make_xlsx(tmp: Path) -> tuple[Path, Path, Path]:
    """建 quest / spawn_quest_entity / reward 三 xlsx（跨目录）。"""
    import openpyxl
    quest_p = tmp / "quest.xlsx"
    spawn_p = tmp / "spawn" / "spawn_quest_entity.xlsx"
    reward_p = tmp / "reward" / "reward.xlsx"
    spawn_p.parent.mkdir(parents=True, exist_ok=True)
    reward_p.parent.mkdir(parents=True, exist_ok=True)

    for p, sheet, hdrs, row in [
        (quest_p, "Quest", ["任务id", "名称"], [250003, "废弃任务"]),
        (spawn_p, "Spawn", ["entity_prefab_id", "pos"], [10045, "x,y"]),
        # reward 表 reward_id=250003（与 quest 同 id 值，反向追溯场景）
        (reward_p, "Reward", ["reward_id", "items"], [250003, "item1"]),
    ]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(hdrs)
        ws.append(hdrs)
        ws.append(["", ""])
        ws.append(["", ""])
        ws.append(row)
        wb.save(p)
    return quest_p, spawn_p, reward_p


def test_reverse_cascade_delete(tmp_path, monkeypatch):
    quest_p, spawn_p, reward_p = _make_xlsx(tmp_path)
    cli = StubCodeMakerCLI(workspace=tmp_path)
    agent = TableAgent(cli=cli, parser=_MockParser())
    # row_data 快照：quest 250003 行的 id 值
    row_data = {1: 250003}

    # mock cascade_resolver.get_referencing_tables 返 spawn_quest_entity + reward
    import agent.excel.core.cascade_resolver as cr_mod
    def fake_ref(stem):
        if stem == "quest":
            return [
                {"target_stem": "spawn_quest_entity", "source_col": "entity_prefab_id"},
                {"target_stem": "reward", "source_col": "reward_id"},
            ]
        return []
    monkeypatch.setattr(cr_mod, "get_referencing_tables", fake_ref)

    # mock list_tables 返三 path（让 stem→path 映射命中）
    cli.list_tables = MagicMock(return_value=[quest_p, spawn_p, reward_p])

    # 执行 _do_cascade_delete
    agent._do_cascade_delete(quest_p, "Quest", ["任务id", "名称"], row_data, "quest")

    # 验证 spawn_quest_entity 和 reward 的行 5（含对应 id 值）被删
    # 注意：spawn 的 entity_prefab_id=10045 ≠ quest 250003，靠主行快照扫不到
    # （需 DecomposeAgent 产独立 delete intent，属 prompt 改造范围，非 _do_cascade_delete）
    # reward 的 reward_id=250003 == quest id，反向追溯应删
    import openpyxl
    reward_wb = openpyxl.load_workbook(reward_p)
    reward_val = reward_wb["Reward"].cell(5, 1).value
    assert reward_val is None or str(reward_val).strip() == "", \
        f"reward 行5 reward_id=250003 应被反向追溯清空, 实际={reward_val}"
    print(f"PASS reverse_cascade: reward 行5={reward_val} (spawn 需独立 intent, 不在快照追溯范围)")


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v", "-s"])
