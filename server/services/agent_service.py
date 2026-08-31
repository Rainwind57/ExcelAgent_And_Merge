"""Agent 服务：包装 TableAgent 为 Web API 可用的服务层。

提供会话管理、操作历史、预览（dry-run）、撤销等功能。
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import os
import re as _re
import shutil
import sys
import tempfile
import time
import uuid
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# agent 包使用相对导入，需要通过 sys.path 确保能找到
_AGENT_DIR = Path(__file__).resolve().parent.parent / "agent"
if str(_AGENT_DIR.parent) not in sys.path:
    sys.path.insert(0, str(_AGENT_DIR.parent))

# 短句确认意图词表（从 parser_config.yaml 加载，减硬编码）
_CONFIRM_KW_PATH = _AGENT_DIR / "excel" / "skills" / "parser_config.yaml"
_CONFIRM_AFFIRM_DEFAULT = ("无误", "确认", "确定", "是的", "是", "好的", "好", "继续",
                           "执行", "没问题", "可以", "ok", "yes", "y", "嗯", "对", "删除")
_CONFIRM_DENY_DEFAULT = ("取消", "不要", "不删", "否", "不", "停止", "no", "n", "cancel")


def _load_confirm_keywords() -> tuple[set, set]:
    """从 parser_config.yaml 加载 confirm_affirm/confirm_deny；失败回退默认。

    返回 (affirm_set, deny_set)，元素均已 lower 供与 lower 后的用户输入比对。
    """
    try:
        import yaml
        if _CONFIRM_KW_PATH.exists():
            data = yaml.safe_load(_CONFIRM_KW_PATH.read_text(encoding="utf-8")) or {}
            aff = data.get("confirm_affirm") or []
            den = data.get("confirm_deny") or []
            if aff and den:
                return {str(k).lower() for k in aff}, {str(k).lower() for k in den}
    except Exception:
        pass
    return {k.lower() for k in _CONFIRM_AFFIRM_DEFAULT}, {k.lower() for k in _CONFIRM_DENY_DEFAULT}

from agent.cli_interface import StubCodeMakerCLI
from agent.real_cli import RealCodeMakerCLI
from agent.agent import TableAgent, AgentResult, AgentStep
from agent.nl_parser import NLIntent
from agent.codemaker_parser import CodemakerNLParser
from agent.codemaker_client import CodemakerClient, set_token_sink
from agent.table_index import load_index, TableMeta, refresh_if_changed
from agent.file_watcher import TableFileWatcher, has_watchdog
from agent.orchestrator import OrchestratorAgent, RouteResult
from agent.qa_handler import QAHandler
from agent.prompts import MERGE_SUGGEST_SYSTEM_PROMPT, MERGE_SUGGEST_BATCH_SYSTEM_PROMPT

from routers.workflow import create_snapshot_sync, restore_snapshot_sync

from config import RESOURCES_DIR
from models.agent_models import (
    AgentChatRequest, AgentChatResponse, AgentStepInfo,
    CellChange, DiffPreview, ResultColumn, ResultTable,
    SubTaskInfo,
    TableInfo, SheetSummary, SheetDataPage,
    SearchResult, SearchResponse,
    PreviewRequest, PreviewResponse,
    ValidateRequest, ValidateResponse, ValidationIssue,
    BatchRequest, BatchResponse, BatchItemResult,
    FormColumn, AddFormBuildRequest, AddFormResponse,
    FormValidateRequest, FieldIssue, AddFormValidateResponse,
    FormCommitRequest, AddFormCommitResponse,
)


def _json_hash(obj) -> str:
    """对 dict 做稳定哈希，用于缓存 key。"""
    raw = json.dumps(obj, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


# 置信度融合权重：final = w_llm·llm_conf + w_rev·rev_conf + w_val·value_conf
# 三者总和 = 1.0。LLM 自评权重最高（语义判断主信号），rev/值域为客观校正因子。
# 规则因子无信号时返回 None，融合时用 llm_conf 兜底，保证"无规则信息不稀释 LLM 自评"。
_CONF_W_LLM = 0.60
_CONF_W_REV = 0.25
_CONF_W_VALUE = 0.15


def _rev_confidence(sv_rev: Any, other_revs: list) -> Optional[float]:
    """修订号因子：被采纳版本相对其他候选的 rev 领先程度，映射到 [0,1]。

    领先 1 档 → 0.6，领先 2 档 → 0.8，领先 ≥3 档 → 0.95；
    落后 → 0.4（采纳"较旧"版本，负信号拉低）。无 rev 信息 → None（不校正）。
    """
    try:
        sv = float(sv_rev)
    except (TypeError, ValueError):
        return None
    nums = []
    for r in other_revs:
        try:
            nums.append(float(r))
        except (TypeError, ValueError):
            continue
    if not nums:
        return None
    lead = sv - max(nums)
    if lead >= 3:
        return 0.95
    if lead == 2:
        return 0.8
    if lead == 1:
        return 0.6
    return 0.4


def _value_confidence(sv_value: Any, col_name: str) -> Optional[float]:
    """值域因子：被采纳值在该列语义下的合理性，映射到 [0,1]。

    ID/编号列：正整数（>0）合理 → 0.9，否则 0.5。
    数值列：有限数值合理 → 0.8，否则 0.4。
    文本列：非空且较长（≥3 字符）合理 → 0.7，否则 0.5。
    未知列 / 无法判断 → None（不校正）。
    """
    col_lower = (col_name or "").lower()
    if any(kw in col_lower for kw in ("id", "编号", "等级", "level")):
        try:
            return 0.9 if float(sv_value) > 0 else 0.5
        except (TypeError, ValueError):
            return 0.5
    try:
        fv = float(sv_value)
        if fv != fv or fv in (float("inf"), float("-inf")):
            return 0.4
        return 0.8
    except (TypeError, ValueError):
        pass
    if sv_value is not None and len(str(sv_value).strip()) >= 3:
        return 0.7
    if sv_value is not None and str(sv_value).strip():
        return 0.5
    return None


def _fuse_confidence(llm_conf: float, rev_conf: Optional[float],
                     value_conf: Optional[float]) -> float:
    """融合 LLM 自评 + 修订号因子 + 值域因子为最终置信度，截断到 [0,1]。

    final = w_llm·llm_conf + w_rev·rev_conf + w_val·value_conf
    规则因子为 None（无信号）时以 llm_conf 替代，确保无规则信息时不改变 LLM 自评。
    """
    r = llm_conf if rev_conf is None else rev_conf
    v = llm_conf if value_conf is None else value_conf
    fused = (_CONF_W_LLM * llm_conf
             + _CONF_W_REV * r
             + _CONF_W_VALUE * v)
    return round(min(1.0, max(0.0, fused)), 4)


class AgentService:
    """无状态的 Agent 服务包装器。

    职责：
    - 初始化 TableAgent，指向 resources/ 目录
    - 管理每个 session 的操作历史（用于撤销）
    - 提供 dry-run 预览能力
    - 生成 agent 修改的 diff 表示
    - 构建验证报告
    """

    def __init__(self, resources_dir: Path = None, enable_skill: bool = True,
                 enable_verify_repair_loop: bool | None = None,
                 enable_skill_tools_recovery: bool | None = None,
                 verify_repair_max_rounds: int | None = None,
                 skill_tool_call_limit: int | None = None):
        self.resources_dir = Path(resources_dir or RESOURCES_DIR)
        self.enable_skill = enable_skill
        # verify-repair 迭代环开关：None 时从 env 读默认（与 Configuration 默认一致）
        self.enable_verify_repair_loop = (
            enable_verify_repair_loop if enable_verify_repair_loop is not None
            else os.environ.get("ENABLE_VERIFY_REPAIR_LOOP", "1") != "0"
        )
        self.enable_skill_tools_recovery = (
            enable_skill_tools_recovery if enable_skill_tools_recovery is not None
            else os.environ.get("ENABLE_SKILL_TOOLS_RECOVERY", "1") != "0"
        )
        self.verify_repair_max_rounds = (
            verify_repair_max_rounds if verify_repair_max_rounds is not None
            else int(os.environ.get("VERIFY_REPAIR_MAX_ROUNDS", "3"))
        )
        self.skill_tool_call_limit = (
            skill_tool_call_limit if skill_tool_call_limit is not None
            else int(os.environ.get("SKILL_TOOL_CALL_LIMIT", "4"))
        )
        os.makedirs(self.resources_dir, exist_ok=True)

        # 是否操作主 resources/ 目录。沙箱副本（测试/预览）不应刷新全局索引——
        # _table_index.json 位于 agent/excel/（全局共享），对沙箱调 refresh_if_changed
        # 会把临时数据写进全局索引污染主进程，且与 TableFileWatcher 后台线程并发写
        # 触发竞态。沙箱 Agent 只读全局索引（沙箱是 resources 的副本，内容一致），
        # 不写；live_index=False 让 TableAgent 跳过写后刷新。
        from config import RESOURCES_DIR as _RES
        self._is_main = self.resources_dir.resolve() == Path(_RES).resolve()

        # 全局索引自愈：若 _table_index.json 缺失/损坏/为空（如被并发竞态污染），
        # 从真实 resources 重建一次。沙箱 Agent 也依赖全局索引做路由，必须保证有效。
        try:
            from agent.table_index import load_index as _load_strict
            _need_rebuild = False
            try:
                _need_rebuild = len(_load_strict()) == 0
            except Exception:
                _need_rebuild = True
            if _need_rebuild:
                from agent.table_index import build_index
                build_index(Path(_RES))
        except Exception as e:
            logger.warning("表索引刷新失败，可能影响表定位: %s", e, exc_info=True)

        # 启动时增量刷新索引：仅主 resources 才刷新全局索引
        # refresh_if_changed 内部在索引不存在时自动全量 build_index
        if self._is_main:
            try:
                refresh_if_changed(self.resources_dir)
            except Exception:
                pass

        # 文件监听：仅主 resources 启用后台 watcher（沙箱副本是临时的，监听无意义且
        # 后台线程会并发写全局索引引发竞态）
        self._file_watcher: Optional[TableFileWatcher] = None
        if self._is_main and has_watchdog():
            try:
                self._file_watcher = TableFileWatcher(
                    workspace=self.resources_dir,
                    on_refresh=self._on_file_changed,
                )
                self._file_watcher.start()
            except Exception as e:
                self._file_watcher = None
                logger.warning("文件监听启动失败，合并后缓存可能不自动刷新: %s", e, exc_info=True)

        # Excel CLI 层：优先 RealCodeMakerCLI（openpyxl 基座 + codemaker 增强）
        try:
            self.cli = RealCodeMakerCLI(workspace=self.resources_dir)
        except Exception as e:
            logger.warning("RealCodeMakerCLI 初始化失败，降级到 StubCodeMakerCLI，写操作可能不生效: %s", e, exc_info=True)
            self.cli = StubCodeMakerCLI(workspace=self.resources_dir)

        # NL 解析器：codemaker LLM 解析（HTTP API 调用 codemaker serve）
        # directory 必须传，否则 /api/session 建会话会 400（见 codemaker_parser.py 注释）
        parser = CodemakerNLParser(directory=str(self.resources_dir),
                                   enable_skill=enable_skill)
        # 沙箱副本 live_index=False：不写全局索引，避免污染主进程 + 并发竞态崩溃
        self.agent = TableAgent(cli=self.cli, parser=parser, enable_skill=enable_skill,
                                live_index=self._is_main,
                                enable_verify_repair_loop=self.enable_verify_repair_loop,
                                enable_skill_tools_recovery=self.enable_skill_tools_recovery,
                                verify_repair_max_rounds=self.verify_repair_max_rounds,
                                skill_tool_call_limit=self.skill_tool_call_limit)

        # OrchestratorAgent：主智能体（聊天 + 调度 QA/CRUD 子智能体）
        cm_client = CodemakerClient()
        self.router = OrchestratorAgent(
            table_agent=self.agent,
            resources_dir=self.resources_dir,
            client=cm_client,
        )

        # 合并建议 LLM（懒加载，复用 router 的 codemaker client；serve 不可用时为 None，回退规则）
        self._llm: Any = None

        # 合并建议缓存：{(table_stem, sheet, col_name, hash(versions_json)): (result_dict, ts)}
        # LRU + TTL：上限 _SUGGEST_CACHE_MAX 条，单条存活 _SUGGEST_CACHE_TTL 秒。
        # 跨多次 compare 缓存 key 含 hash(versions) 永不自然淘汰，无界增长会内存膨胀，
        # 故加 LRU（写时 move_to_end + 超限 popitem）与 TTL（读时校验过期）。
        self._suggest_cache: "OrderedDict[tuple, tuple]" = OrderedDict()
        self._suggest_cache_max = 2000
        self._suggest_cache_ttl = 3600.0  # 1 小时

        # 会话历史 {session_id: [operation_record]}
        self._sessions: Dict[str, list] = {}

        # 会话 checkpoint 列表 {session_id: [{checkpoint_id, snap_id, timestamp, text}]}
        # 每次执行写动作（set/add/delete/col）前拍一个 checkpoint（resources/ 全量快照）。
        # rollback_to_checkpoint() 据此把表格还原到"用户某次输入之前"的状态。
        # 一个 checkpoint 对应一次自然语言输入（可能含多个原子写操作），非原子粒度。
        self._session_checkpoints: Dict[str, list] = {}
        # P27：4-step NL 路径中间态 checkpoint {session_id: {stage: [NLIntent dict]}}。
        # parse/validate 后拍 NLIntent 序列化，stall 可从中间态续跑免 Step1 重 LLM
        # decompose。opt-in CODEMAKER_4STEP_CHECKPOINT=1 时写。接线（save/resume）留 follow-up。
        self._nl_checkpoints: Dict[str, dict] = {}

        # 预览暂存 {operation_id: preview_data}
        self._previews: Dict[str, dict] = {}

        # 待确认危险操作 {session_id: {"token": str, "text": str, "expires_at": float}}
        # 级联删除/列删除返回 needs_confirm 时暂存；前端点确认回传 token → 校验未超时后执行。
        self._pending_confirms: Dict[str, dict] = {}

        # D6.4: 上一轮定位证据缓存 {session_id: {table_stem, sheet, col_resolved, row_resolved}}
        # 下一轮 result 返回后比对同表同 sheet 但 resolved 不同 → 标 user_corrected=True + corrected_to。
        # 纠正关键词触发："不是/搞错了/应该是/改回/不对" → 强制纠正标记。
        self._session_last_evidence: Dict[str, dict] = {}

        # T1: 搜索结果缓存 {(keyword, table): (SearchResponse, ts)}，30s 内同关键词不重复查询
        self._search_cache: Dict[tuple, tuple] = {}
        self._search_cache_ttl = 30.0

        # T9 D9.2: 进程启动跑一次全量衰减扫描，清过期 runtime 条目（30 天降权 / 60 天移除）
        # 失败只 warn，不阻断启动
        try:
            from agent.excel.skill_updater import get_skill_updater
            get_skill_updater().decay_scan()
        except Exception:
            pass

    # 确认令牌有效期（秒）
    _CONFIRM_TTL_SECONDS = 300

    # ── 上下文记忆 ──

    _MAX_CONTEXT_ENTRIES = 10

    @classmethod
    def _build_context(cls, session_history: list) -> str:
        """从会话历史构建上下文文本，注入 LLM prompt。"""
        if not session_history:
            return ""

        recent = session_history[-cls._MAX_CONTEXT_ENTRIES:]
        lines: list[str] = []
        for rec in recent:
            op_id = rec.get('id', '?')
            text = rec.get('text', '')
            intent = rec.get('intent', '')
            table_hint = rec.get('table_hint', '')
            sheet_hint = rec.get('sheet_hint', '')
            target_field = rec.get('target_field', '')
            value = rec.get('value', '')
            steps = rec.get('steps', [])
            step_summary = "; ".join(
                f"{s.get('name','?')}={'OK' if s.get('ok') else 'FAIL'}"
                for s in steps
            ) if steps else ""

            location = ""
            if table_hint:
                location += f"表={table_hint}"
            if sheet_hint:
                location += f" sheet={sheet_hint}"
            if target_field:
                location += f" 列={target_field}"
            if value:
                location += f" 值={value}"

            lines.append(f"- [{op_id}] {intent}: \"{text}\" {location} → {step_summary}")

        return "\n".join(lines)

    # ── 表格浏览 ──

    def get_tables(self) -> List[TableInfo]:
        """列出所有表格及其 sheet 摘要。"""
        try:
            index = load_index()
        except Exception:
            return []

        result = []
        for t in index:
            sheets = []
            for s in t.sheets:
                sheets.append(SheetSummary(
                    name=s.name,
                    row_count=s.row_count,
                    headers=s.headers[:20],  # 只返回前20列避免过大
                ))
            result.append(TableInfo(stem=t.stem, path=t.path, sheets=sheets))
        return result

    def get_table_detail(self, stem: str) -> Optional[TableInfo]:
        """获取单个表格详情。"""
        tables = self.get_tables()
        for t in tables:
            if t.stem == stem:
                return t
        return None

    def get_sheet_data(self, stem: str, sheet: str,
                       page: int = 1, page_size: int = 100,
                       include_columns: bool = False) -> Optional[SheetDataPage]:
        """读取指定 sheet 的分页数据。

        走 read_browse（自动探测表头行），避免索引 data_start 错位导致空数据。
        表头仍为空时从 _table_index.json 回退取索引中的 headers。

        R20: 始终填充 columns_meta（列名+列号），供同名列消歧。
        R21: include_columns=True 时同时填充 columns（列约束），减少配表模式往返。
        """
        file_path = self.resources_dir / f"{stem}.xlsx"
        # 模糊查找
        if not file_path.exists():
            for p in self.resources_dir.rglob(f"{stem}.xlsx"):
                file_path = p
                break
        if not file_path.exists():
            return None

        try:
            headers, page_rows, total = self.cli.read_browse(
                file_path, sheet, page=page, page_size=page_size)

            # 表头全空 → 从索引回退（索引构建时已清洗过表头）
            if not headers or all(h is None or str(h).strip() == "" for h in headers):
                idx_headers = self._headers_from_index(stem, sheet)
                if idx_headers:
                    headers = idx_headers

            # 规范化表头：None/非字符串 → ""，避免 SheetDataPage(List[str]) 校验失败
            # （部分 sheet 表头行存在空单元格，如 ability 的 A1、exclusive_state 第3列）
            headers = ["" if h is None else str(h) for h in headers]

            # R20: 构建 columns_meta（列名清理 + 1-based 列号）
            from models.agent_models import ColumnMeta
            columns_meta = [
                ColumnMeta(col_index=i + 1,
                           name=str(h).split(":")[0].strip() if h else "")
                for i, h in enumerate(headers)
            ]

            # R8: 算数据起始行，供前端原地编辑计算 Excel 绝对行号
            try:
                ws = self.cli._load(file_path)[sheet]
                data_start = self.cli._browse_data_start(ws, file_path, sheet)
            except Exception:
                data_start = 1

            # R21: 可选填充列约束
            columns = []
            if include_columns:
                columns = self.get_sheet_columns(stem, sheet) or []

            return SheetDataPage(
                sheet=sheet,
                headers=headers,
                rows=page_rows,
                total_rows=total,
                page=page,
                page_size=page_size,
                data_start=data_start,
                columns_meta=columns_meta,
                columns=columns,
            )
        except Exception:
            return None

    def _headers_from_index(self, stem: str, sheet: str) -> list:
        """从 _table_index.json 取指定 sheet 的表头（浏览表头探测失败时回退）。"""
        try:
            for t in load_index():
                if t.stem == stem:
                    for s in t.sheets:
                        if s.name == sheet:
                            return list(s.headers)
        except Exception:
            pass
        return []

    def get_sheet_columns(self, stem: str, sheet: str) -> Optional[List[FormColumn]]:
        """R8: 读取指定 sheet 的列约束（类型/必填/唯一/外键），供前端 hover 列头 tooltip。

        复用 scan_workbook 的列元数据扫描，按列号顺序返回 FormColumn 列表。
        scan_workbook 跳过 CONFIG/说明 sheet 时回退用 read_header 构建最小结构（仅列名）。
        """
        path = self._resolve_table_path(stem)
        if path is None:
            return None
        from agent.excel.schema_infer import scan_workbook
        try:
            tm = scan_workbook(path)
            sm = tm.sheets.get(sheet)
        except Exception:
            sm = None

        if sm is None:
            # 回退：仅列名（无约束信息）
            try:
                headers = self.cli.read_header(path, sheet)
            except Exception:
                headers = []
            return [
                FormColumn(col=i + 1, col_name=str(h or "").split(":")[0].strip())
                for i, h in enumerate(headers)
            ]

        columns: List[FormColumn] = []
        for col in sm.columns:
            ci = col.index + 1  # ColumnMeta.index 0-based → CLI 1-based
            desc_parts = []
            if col.is_required:
                desc_parts.append("必填")
            if col.col_type:
                desc_parts.append(f"类型:{col.col_type}")
            if col.is_id_column:
                desc_parts.append("唯一")
            if col.ref_table:
                desc_parts.append(f"引用:{col.ref_table}")
            columns.append(FormColumn(
                col=ci,
                col_name=col.clean_name or col.header,
                col_type=col.col_type,
                required=col.is_required,
                unique=col.is_id_column,
                default=col.default_value,
                ref_table=col.ref_table,
                is_id=col.is_id_column,
                is_name=col.is_name_column,
                description="；".join(desc_parts),
            ))
        return columns

    def search(self, keyword: str, table: str = "") -> SearchResponse:
        """全文搜索。

        T1: 30s 结果缓存——同 (keyword, table) 在 TTL 内直接返回缓存，避免重复扫描。
        """
        cache_key = (keyword or "", table or "")
        now = time.time()
        cached = self._search_cache.get(cache_key)
        if cached is not None:
            resp, ts = cached
            if now - ts < self._search_cache_ttl:
                return resp

        tables = self.cli.list_tables()
        results = []

        for path in tables:
            if table and table not in path.stem:
                continue
            try:
                hits = self.cli.search_rows(path, keyword)
                for h in hits:
                    results.append(SearchResult(
                        table_stem=path.stem,
                        table_path=str(path.relative_to(self.resources_dir)),
                        sheet=h.sheet,
                        row=h.row,
                        col=h.col,
                        col_name=h.col_name,
                        cell_value=h.cell_value,
                        row_data=h.row_data,
                        data_row=h.data_row,
                    ))
            except Exception:
                continue

        resp = SearchResponse(keyword=keyword, results=results, total=len(results))
        self._search_cache[cache_key] = (resp, now)
        # 缓存条目超过 200 个时清理最旧的，避免无界增长
        if len(self._search_cache) > 200:
            for k in sorted(self._search_cache, key=lambda k: self._search_cache[k][1])[:50]:
                del self._search_cache[k]
        return resp

    # ── Add Form（表单式新增）──

    _alias_expanded = False

    @staticmethod
    def expand_alias_mapping(dry_run: bool = False) -> dict:
        """从所有表列名自动推断中文别名，扩充 alias_mapping.json。

        规则：扫描每表每 sheet 的列名，提取形如「中文前缀 + id/编号/名称/类型...」
        的前缀作为该表的中文别名（如「建筑名称」→「建筑」→building）。
        不覆盖已注册别名。返回 {新增别名: stem} 统计。

        Args:
            dry_run: True 时只返回候选不写盘。
        """
        import re
        from agent.excel.alias_mapping import AliasMapping
        from agent.table_index import load_index

        am = AliasMapping.load()
        index = load_index()
        # 后缀按长度降序排，避免短后缀（id）先吃掉长后缀（model_id）
        suffixes = [
            "等级模板id", "model_id", "配置id", "编号", "名称", "名字",
            "类型", "描述", "图标", "大图", "model", "id",
        ]
        suffix_re = re.compile(r'^(.+?)(' + '|'.join(suffixes) + r')$')

        def _extract_prefix(hn: str) -> str:
            if not hn:
                return ""
            m = suffix_re.match(hn)
            if not m:
                return ""
            p = m.group(1).strip()
            if not (2 <= len(p) <= 5):
                return ""
            if not re.fullmatch(r'[\u4e00-\u9fff]+', p):
                return ""
            return p

        def _register(prefix: str, stem: str) -> None:
            if prefix and not am.lookup(prefix):
                am.set(prefix, f"{stem}.xlsx")
                added[prefix] = stem

        added: Dict[str, str] = {}
        # 第一轮：每表第一列（主键列）前缀，优先级最高（避免被其他表的同名列抢占）
        for tm in index:
            for s in tm.sheets:
                hns = s.header_names or []
                if hns:
                    _register(_extract_prefix(hns[0]), tm.stem)
        # 第二轮：其他列前缀，补注册（不覆盖第一轮与已注册项）
        for tm in index:
            for s in tm.sheets:
                for hn in (s.header_names or [])[1:]:
                    _register(_extract_prefix(hn), tm.stem)
        if added and not dry_run:
            am.save()
        return added

    def _ensure_alias_expanded(self) -> None:
        """首次调用时扩充 alias_mapping（类标志防重复）。"""
        if AgentService._alias_expanded:
            return
        try:
            AgentService.expand_alias_mapping()
        except Exception:
            pass
        AgentService._alias_expanded = True

    def _resolve_table_path(self, stem: str) -> Optional[Path]:
        """按 stem 定位 Excel 文件（优先根目录，回退递归模糊查找）。"""
        p = self.resources_dir / f"{stem}.xlsx"
        if p.exists():
            return p
        for q in self.resources_dir.rglob(f"{stem}.xlsx"):
            if not q.name.startswith("~$"):
                return q
        return None

    def build_add_form(self, text: str) -> AddFormResponse:
        """自然语言描述 → 匹配表 + sheet → 返回表头 + 约束 + 空行。

        流程：
            1. 用 alias_mapping 在文本中反查别名（如「灵兽」→pet），命中则设 table_hint
            2. 复用 TableAgent._resolve_table / _resolve_sheet 定位表与 sheet
            3. scan_workbook 读取列约束（类型/必填/ID/外键）
            4. 构建 FormColumn 列表 + 空行（默认值或空串）
        """
        self._ensure_alias_expanded()
        from agent.nl_parser import NLIntent
        from agent.excel.alias_mapping import AliasMapping
        from agent.excel.schema_infer import scan_workbook

        # 1. 别名反查设 table_hint（最长别名优先）
        am = AliasMapping.load()
        hits = am.lookup_in_text(text or "")
        table_hint = Path(hits[0][1]).stem if hits else None

        intent = NLIntent(raw=text or "", action="add", table_hint=table_hint)

        # 2. 定位表 + sheet（复用 agent 既有能力）
        path, sheet_hint = self.agent._resolve_table(intent)
        if path is None:
            return AddFormResponse(
                ok=False,
                message=f"未匹配到表格：「{text}」（hint={table_hint}）",
            )
        intent.sheet_hint = sheet_hint
        sheet = self.agent._resolve_sheet(path, intent)
        if not sheet:
            return AddFormResponse(
                ok=False, table_stem=path.stem,
                message=f"已匹配表 {path.stem}，但无法确定 sheet",
            )

        # 3. 读取列约束
        try:
            tm = scan_workbook(path)
            sm = tm.sheets.get(sheet)
        except Exception as e:
            return AddFormResponse(
                ok=False, table_stem=path.stem, sheet=sheet,
                message=f"读取表结构失败：{e}",
            )

        # scan_sheet 会跳过 CONFIG/说明 sheet；回退用 read_header 构建最小结构
        if sm is None:
            try:
                headers = self.cli.read_header(path, sheet)
            except Exception:
                headers = []
            columns = [
                FormColumn(col=i + 1, col_name=str(h or "").split(":")[0].strip())
                for i, h in enumerate(headers)
            ]
            empty_row = {str(c.col): "" for c in columns}
        else:
            columns = []
            empty_row = {}
            for col in sm.columns:
                ci = col.index + 1  # ColumnMeta.index 为 0-based，CLI 约定 1-based
                desc_parts = []
                if col.is_required:
                    desc_parts.append("必填")
                if col.col_type:
                    desc_parts.append(f"类型:{col.col_type}")
                if col.is_id_column:
                    desc_parts.append("唯一")
                if col.ref_table:
                    desc_parts.append(f"引用:{col.ref_table}")
                columns.append(FormColumn(
                    col=ci,
                    col_name=col.clean_name or col.header,
                    col_type=col.col_type,
                    required=col.is_required,
                    unique=col.is_id_column,
                    default=col.default_value,
                    ref_table=col.ref_table,
                    is_id=col.is_id_column,
                    is_name=col.is_name_column,
                    description="；".join(desc_parts),
                ))
                empty_row[str(ci)] = ""

        try:
            rel_path = str(path.relative_to(self.resources_dir))
        except ValueError:
            rel_path = path.name

        return AddFormResponse(
            ok=True,
            table_stem=path.stem,
            sheet=sheet,
            table_path=rel_path,
            columns=columns,
            empty_row=empty_row,
            message=f"已匹配：{path.stem} / {sheet}（{len(columns)} 列）",
        )

    def _try_intercept_add_form(self, text: str) -> Optional[AddFormResponse]:
        """纯类别新增拦截：输入如「增加灵兽」（无具体对象/字段）→ 返回表单，不写盘。

        判定：文本以新增动词开头 + 实体别名，且别名后无具体对象名/字段描述。
        命中返回 AddFormResponse；不命中返回 None，走正常 CRUD 流程。

        避免把「增加灵兽」误当对象名直接写入（如 名称=灵兽 污染主表）。
        有具体信息的（如「增加灵兽白虎」「增加灵兽 名称白虎」）不拦截，走直接新增。
        """
        self._ensure_alias_expanded()
        from agent.excel.alias_mapping import AliasMapping

        t = (text or "").strip()
        add_verbs = ("增加", "新增", "添加", "新建", "加一个")
        if not any(t.startswith(v) for v in add_verbs):
            return None
        body = t
        for v in add_verbs:
            if body.startswith(v):
                body = body[len(v):].strip()
                break
        if not body:
            return None

        am = AliasMapping.load()
        for alias, _fp in am.lookup_in_text(body):
            if body.startswith(alias):
                rest = body[len(alias):].strip().strip("，,。.！!？?、 ")
                if not rest:
                    return self.build_add_form(text)
                break
        return None

    def _resolve_form_values(self, path: Path, sheet: str,
                             values: Dict[str, Any]) -> Dict[int, Any]:
        """把表单提交的 {列号或列名: 值} 归一为 {1-based 列号: 值}。"""
        headers = self.cli.read_header(path, sheet)
        matcher = self.agent._make_matcher(headers, path.stem, sheet)
        out: Dict[int, Any] = {}
        for k, v in (values or {}).items():
            if isinstance(k, int):
                out[k] = v
                continue
            ks = str(k).strip()
            if ks.isdigit():
                out[int(ks)] = v
                continue
            m = matcher.match(ks) or matcher.match_best(ks)
            if m is not None:
                out[m.index] = v
                continue
            # 表头原文子串兜底
            for i, h in enumerate(headers, start=1):
                if h and ks in str(h):
                    out[i] = v
                    break
        return out

    def validate_add_row(self, req: FormValidateRequest) -> AddFormValidateResponse:
        """逐列校验：必填 / 类型 / 唯一 / 范围。返回 errors（阻断）+ warnings（提示）。

        前端据 errors 将对应单元格标红。
        """
        from agent.excel.schema_infer import scan_workbook

        path = self._resolve_table_path(req.table_stem)
        if path is None:
            return AddFormValidateResponse(
                ok=False, message=f"表格不存在：{req.table_stem}",
            )

        try:
            resolved = self._resolve_form_values(path, req.sheet, req.values)
        except Exception as e:
            return AddFormValidateResponse(
                ok=False, message=f"解析列值失败：{e}",
            )

        # 读列约束
        try:
            tm = scan_workbook(path)
            sm = tm.sheets.get(req.sheet)
        except Exception:
            sm = None
        if sm is None:
            # 无约束元信息 → 仅做宽松校验（不报错，原样返回）
            return AddFormValidateResponse(
                ok=True,
                coerced_values={str(k): v for k, v in resolved.items()},
                message="未读到列约束，跳过强校验",
            )

        # 现有数据：用于唯一性校验 {1-based col: set(str 值)}
        try:
            existing_rows = self.cli.read_sheet(path, req.sheet)
        except Exception:
            existing_rows = []
        existing_by_col: Dict[int, set] = {}
        for col in sm.columns:
            ci = col.index + 1
            vals = set()
            for r in existing_rows:
                if r and len(r) >= ci and r[ci - 1] is not None:
                    s = str(r[ci - 1]).strip()
                    if s:
                        vals.add(s)
            existing_by_col[ci] = vals

        errors: List[FieldIssue] = []
        warnings: List[FieldIssue] = []
        coerced: Dict[int, Any] = {}

        for col in sm.columns:
            ci = col.index + 1
            col_name = col.clean_name or col.header
            raw = resolved.get(ci)

            # 必填
            if col.is_required and (raw is None or str(raw).strip() == ""):
                errors.append(FieldIssue(
                    col=ci, col_name=col_name, severity="error",
                    message=f"必填项「{col_name}」不能为空",
                ))
                continue
            # 空且非必填 → 留空，不补默认值（用户填什么就提交什么）
            if raw is None or str(raw).strip() == "":
                continue

            # 类型转换（复用 TableAgent._coerce_value，含枚举映射）
            col_type = col.col_type or self.agent._get_col_type(
                req.table_stem, req.sheet, col_name)
            cv, warn, err = self.agent._coerce_value(
                col_type, raw, req.table_stem, req.sheet, col_name)
            if err:
                errors.append(FieldIssue(
                    col=ci, col_name=col_name, severity="error", message=err,
                ))
                continue
            if warn:
                warnings.append(FieldIssue(
                    col=ci, col_name=col_name, severity="warning", message=warn,
                ))

            # 唯一性（ID 列）
            if col.is_id_column:
                val_str = str(cv).strip()
                if val_str in existing_by_col.get(ci, set()):
                    errors.append(FieldIssue(
                        col=ci, col_name=col_name, severity="error",
                        message=f"「{col_name}」值「{val_str}」已存在，需唯一",
                    ))
                    continue

            # 范围（概率/比率类 0-100）
            cn = col_name.lower()
            if any(k in cn for k in ("概率", "rate", "ratio", "chance")):
                try:
                    fv = float(cv)
                    if fv < 0 or fv > 100:
                        errors.append(FieldIssue(
                            col=ci, col_name=col_name, severity="error",
                            message=f"「{col_name}」取值应在 0-100 之间（当前 {fv}）",
                        ))
                        continue
                except (ValueError, TypeError):
                    pass

            coerced[ci] = cv

        ok = not errors
        return AddFormValidateResponse(
            ok=ok,
            errors=errors,
            warnings=warnings,
            coerced_values={str(k): v for k, v in coerced.items()},
            message="校验通过" if ok else f"存在 {len(errors)} 个错误",
        )

    def commit_add_row(self, req: FormCommitRequest) -> AddFormCommitResponse:
        """校验通过后插入新行并按主键排序。

        confirm=False 时先校验，有 error 阻止提交；
        confirm=True 时跳过校验直接提交（调用方已自行校验过）。
        """
        from agent.agent import AgentResult
        from agent.nl_parser import NLIntent

        path = self._resolve_table_path(req.table_stem)
        if path is None:
            return AddFormCommitResponse(
                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                message=f"表格不存在：{req.table_stem}",
            )

        # 1. 校验（除非 confirm）
        if not req.confirm:
            v = self.validate_add_row(FormValidateRequest(
                table_stem=req.table_stem, sheet=req.sheet, values=req.values))
            if not v.ok:
                return AddFormCommitResponse(
                    ok=False, table_stem=req.table_stem, sheet=req.sheet,
                    dry_run=req.dry_run, message="校验未通过，已阻止提交", errors=v.errors,
                )
            values: Dict[int, Any] = {
                int(k): val for k, val in v.coerced_values.items()
            }
        else:
            try:
                values = self._resolve_form_values(path, req.sheet, req.values)
            except Exception as e:
                return AddFormCommitResponse(
                    ok=False, table_stem=req.table_stem, sheet=req.sheet,
                    message=f"解析列值失败：{e}",
                )

        if not values:
            return AddFormCommitResponse(
                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                message="无可写入的值",
            )

        # R23: dry_run 二段提交第一阶段 — 只校验+返回预览，不写盘
        if req.dry_run:
            preview = {str(k): v for k, v in values.items()}
            return AddFormCommitResponse(
                ok=True, table_stem=req.table_stem, sheet=req.sheet,
                dry_run=True, inserted_values=preview,
                message="预览模式：以下值校验通过，确认后用 dry_run=false 提交写盘",
            )

        # 2. 复用 _do_append：含主键唯一校验 + 排序 + 索引刷新
        res = AgentResult(
            ok=True,
            intent=NLIntent(raw="", action="add",
                            table_hint=req.table_stem, sheet_hint=req.sheet),
        )
        res.table_stem = req.table_stem
        res.table_sheet = req.sheet
        res = self.agent._do_append(path, req.sheet, values, res)

        if not res.ok or res.final is None or not res.final.ok:
            # 主键冲突等 → 从 steps/message 提取原因
            msg = res.message or "插入失败"
            pk_conflict = any(s.name == "pk_conflict" and not s.ok for s in res.steps)
            if pk_conflict:
                msg = "主键（第1列）值已存在，请更换后重试"
            return AddFormCommitResponse(
                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                message=msg,
            )

        new_row = res.final.data.get("row") if isinstance(res.final.data, dict) else None
        sorted_done = any(s.name == "auto_sort" and s.ok for s in res.steps)
        # R26: 记录新增行历史（供回滚删除）
        if new_row:
            try:
                from engine.table_history import record_change
                record_change(req.table_stem, req.sheet, "add_commit",
                              row=new_row, new_value={str(k): v for k, v in values.items()},
                              source="add_form_commit")
            except Exception:
                pass
        return AddFormCommitResponse(
            ok=True,
            table_stem=req.table_stem,
            sheet=req.sheet,
            new_row=new_row,
            inserted_values={str(k): v for k, v in values.items()},
            sorted=sorted_done,
            message=res.message or f"已新增到 {req.sheet} 行 {new_row}",
        )

    # ── R8: 单元格原地更新 ──

    def update_cell(self, req: "CellUpdateRequest") -> "CellUpdateResponse":
        """原地更新单个单元格：校验类型/唯一性 → 转换值 → 写值。

        跳过 NL 解析直接写值，供前端 TablesView 双击编辑。
        """
        from models.agent_models import CellUpdateResponse
        from agent.excel.schema_infer import scan_workbook

        path = self._resolve_table_path(req.table_stem)
        if path is None:
            return CellUpdateResponse(
                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                message=f"表格不存在：{req.table_stem}",
                error="table_not_found",
            )

        # 读旧值
        old_val = None
        try:
            rc = self.cli.read_cell(path, req.sheet, req.row, req.col)
            if rc.ok:
                old_val = rc.data
        except Exception:
            pass

        # 读列约束（类型/唯一）
        col_name = ""
        col_type = ""
        is_unique = False
        try:
            tm = scan_workbook(path)
            sm = tm.sheets.get(req.sheet)
            if sm:
                for col in sm.columns:
                    if col.index + 1 == req.col:
                        col_name = col.clean_name or col.header
                        col_type = col.col_type or self.agent._get_col_type(
                            req.table_stem, req.sheet, col_name)
                        is_unique = col.unique
                        break
        except Exception:
            pass

        # 读表头拿列名（兜底）
        if not col_name:
            try:
                headers = self.cli.read_header(path, req.sheet)
                if 1 <= req.col <= len(headers):
                    col_name = str(headers[req.col - 1] or "").split(":")[0].strip()
            except Exception:
                pass

        # 类型转换 + 枚举映射（复用 _coerce_value）
        new_val: Any = req.value
        if col_type:
            try:
                cv, warn, err = self.agent._coerce_value(
                    col_type, req.value, req.table_stem, req.sheet, col_name)
                if err:
                    return CellUpdateResponse(
                        ok=False, table_stem=req.table_stem, sheet=req.sheet,
                        row=req.row, col=req.col, col_name=col_name,
                        old_value=old_val, message=f"类型校验失败：{err}",
                        error="type_error",
                    )
                new_val = cv
            except Exception as e:
                return CellUpdateResponse(
                    ok=False, table_stem=req.table_stem, sheet=req.sheet,
                    row=req.row, col=req.col, col_name=col_name,
                    old_value=old_val, message=f"类型转换异常：{e}",
                    error="type_error",
                )

        # 唯一性校验
        if is_unique:
            try:
                existing_rows = self.cli.read_sheet(path, req.sheet)
                for r in existing_rows:
                    if r and len(r) >= req.col and r[req.col - 1] is not None:
                        if str(r[req.col - 1]).strip() == str(new_val).strip():
                            return CellUpdateResponse(
                                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                                row=req.row, col=req.col, col_name=col_name,
                                old_value=old_val,
                                message=f"唯一性校验失败：{col_name}={new_val} 已存在",
                                error="unique_conflict",
                            )
            except Exception:
                pass

        # 写值
        try:
            res = self.cli.write_cell(path, req.sheet, req.row, req.col, new_val)
            if not res.ok:
                return CellUpdateResponse(
                    ok=False, table_stem=req.table_stem, sheet=req.sheet,
                    row=req.row, col=req.col, col_name=col_name,
                    old_value=old_val, message=f"写值失败：{res.error}",
                    error="write_error",
                )
        except Exception as e:
            return CellUpdateResponse(
                ok=False, table_stem=req.table_stem, sheet=req.sheet,
                row=req.row, col=req.col, col_name=col_name,
                old_value=old_val, message=f"写值异常：{e}",
                error="write_error",
            )

        # 清缓存（写盘后让后续读取拿新值）
        try:
            self.cli._cache.clear()
        except Exception:
            pass

        # R26: 记录操作历史
        try:
            from engine.table_history import record_change
            record_change(req.table_stem, req.sheet, "cell_update",
                          row=req.row, col=req.col, col_name=col_name,
                          old_value=old_val, new_value=new_val, source="cell_update")
        except Exception:
            pass

        return CellUpdateResponse(
            ok=True, table_stem=req.table_stem, sheet=req.sheet,
            row=req.row, col=req.col, col_name=col_name,
            old_value=old_val, new_value=new_val,
            message=f"已更新 {req.sheet} 行{req.row} 列{col_name}={new_val}",
        )

    def update_cells_batch(self, req: "BatchCellUpdateRequest") -> "BatchCellUpdateResponse":
        """R22: 同行多列事务性批量改值。

        atomic=True（默认）：先逐列校验（类型/唯一），任一失败则全部不写，返回每列结果。
        atomic=False：逐列校验+写，失败的列跳过，成功的照写。
        """
        from models.agent_models import (BatchCellUpdateResponse, BatchCellUpdateResult)
        from agent.excel.schema_infer import scan_workbook

        resp = BatchCellUpdateResponse(
            ok=False, table_stem=req.table_stem, sheet=req.sheet, row=req.row)
        path = self._resolve_table_path(req.table_stem)
        if path is None:
            resp.message = f"表格不存在：{req.table_stem}"
            return resp
        if not req.updates:
            resp.message = "updates 为空"
            return resp

        # 读列约束（类型/唯一/列名）
        col_meta: dict[int, dict] = {}  # col(1-based) → {name, type, unique}
        try:
            tm = scan_workbook(path)
            sm = tm.sheets.get(req.sheet)
            if sm:
                for col in sm.columns:
                    ci = col.index + 1
                    col_meta[ci] = {
                        "name": col.clean_name or col.header,
                        "type": col.col_type or "",
                        "unique": col.is_id_column,
                    }
        except Exception:
            pass
        try:
            headers = self.cli.read_header(path, req.sheet)
        except Exception:
            headers = []

        def _col_name(ci: int) -> str:
            m = col_meta.get(ci)
            if m and m.get("name"):
                return m["name"]
            if 1 <= ci <= len(headers):
                return str(headers[ci - 1] or "").split(":")[0].strip()
            return f"列{ci}"

        try:
            all_rows = self.cli.read_sheet(path, req.sheet)
        except Exception:
            all_rows = []
        start = (self.cli._resolve_data_start(path, req.sheet)
                 if hasattr(self.cli, "_resolve_data_start")
                 else getattr(self.cli, "data_start_row", 5))
        row_idx = req.row - start
        cur_row = all_rows[row_idx] if 0 <= row_idx < len(all_rows) else []

        results: list[BatchCellUpdateResult] = []
        has_error = False
        for item in req.updates:
            ci = item.col
            cn = _col_name(ci)
            old_val = cur_row[ci - 1] if 0 <= ci - 1 < len(cur_row) else None
            res = BatchCellUpdateResult(col=ci, col_name=cn, old_value=old_val)
            m = col_meta.get(ci, {})
            col_type = m.get("type", "")
            new_val: Any = item.value
            if col_type:
                try:
                    cv, warn, err = self.agent._coerce_value(
                        col_type, item.value, req.table_stem, req.sheet, cn)
                    if err:
                        res.ok = False
                        res.error = f"type_error:{err}"
                        res.new_value = item.value
                        results.append(res)
                        has_error = True
                        continue
                    new_val = cv
                except Exception as e:
                    res.ok = False
                    res.error = f"type_error:{e}"
                    res.new_value = item.value
                    results.append(res)
                    has_error = True
                    continue
            res.new_value = new_val
            if m.get("unique"):
                dup = False
                for ri, r in enumerate(all_rows):
                    if ri == row_idx:
                        continue
                    if r and len(r) >= ci and r[ci - 1] is not None:
                        if str(r[ci - 1]).strip() == str(new_val).strip():
                            dup = True
                            break
                if dup:
                    res.ok = False
                    res.error = f"unique_conflict:{cn}={new_val} 已存在"
                    results.append(res)
                    has_error = True
                    continue
            results.append(res)

        if req.atomic and has_error:
            resp.ok = False
            resp.results = results
            failed = [r for r in results if not r.ok]
            resp.message = f"原子模式：{len(failed)} 列校验失败，全部未写"
            return resp

        written = 0
        for res in results:
            if not res.ok:
                continue
            try:
                w = self.cli.write_cell(path, req.sheet, req.row, res.col, res.new_value)
                if w.ok:
                    written += 1
                else:
                    res.ok = False
                    res.error = f"write_error:{w.error}"
                    has_error = True
            except Exception as e:
                res.ok = False
                res.error = f"write_error:{e}"
                has_error = True

        try:
            self.cli._cache.clear()
        except Exception:
            pass

        # R26: 记录每列操作历史
        try:
            from engine.table_history import record_change
            for res in results:
                if res.ok:
                    record_change(req.table_stem, req.sheet, "batch_update",
                                  row=req.row, col=res.col, col_name=res.col_name,
                                  old_value=res.old_value, new_value=res.new_value,
                                  source="batch_update")
        except Exception:
            pass

        resp.results = results
        resp.ok = not has_error
        if has_error:
            resp.message = f"已写 {written}/{len(results)} 列（部分失败）"
        else:
            resp.message = f"已批量更新 {req.sheet} 行{req.row} 共 {written} 列"
        return resp

    # ── R24 行/列增删 ──

    def delete_row(self, req: "RowDeleteRequest") -> "RowOpResponse":
        """R24: 删行 + 公式引用位移提示。"""
        from models.agent_models import RowOpResponse
        resp = RowOpResponse(ok=False, table_stem=req.table_stem, sheet=req.sheet, row=req.row)
        path = self._resolve_table_path(req.table_stem)
        if path is None:
            resp.message = f"表格不存在：{req.table_stem}"
            return resp
        # R26: 删行前记录旧行数据（供回滚重插）
        deleted_values: list = []
        try:
            all_rows = self.cli.read_sheet(path, req.sheet)
            start = (self.cli._resolve_data_start(path, req.sheet)
                     if hasattr(self.cli, "_resolve_data_start")
                     else getattr(self.cli, "data_start_row", 5))
            ri = req.row - start
            if 0 <= ri < len(all_rows):
                deleted_values = list(all_rows[ri])
        except Exception:
            pass
        try:
            r = self.cli.delete_row(path, req.sheet, req.row)
            if not r.ok:
                resp.message = f"删行失败：{r.error}"
                return resp
            try:
                self.cli._cache.clear()
            except Exception:
                pass
            resp.ok = True
            resp.message = f"已删除 {req.sheet} 行{req.row}"
            if r.needs_manual_fix:
                resp.warnings.append("删行后公式引用已机械位移，请检查跨行引用是否正确")
            if r.cache_message:
                resp.warnings.append(r.cache_message)
            # R26: 记录历史
            try:
                from engine.table_history import record_change
                record_change(req.table_stem, req.sheet, "row_delete",
                              row=req.row, old_value=None, new_value=None,
                              source="row_delete",
                              extra={"deleted_values": [str(v) for v in deleted_values]})
            except Exception:
                pass
        except Exception as e:
            resp.message = f"删行异常：{e}"
        return resp

    def insert_row(self, req: "RowInsertRequest") -> "RowOpResponse":
        """R24: 插行 + 样式继承 + 公式位移提示。"""
        from models.agent_models import RowOpResponse
        resp = RowOpResponse(ok=False, table_stem=req.table_stem, sheet=req.sheet, row=req.row)
        path = self._resolve_table_path(req.table_stem)
        if path is None:
            resp.message = f"表格不存在：{req.table_stem}"
            return resp
        values = {int(k): v for k, v in req.values.items()} if req.values else None
        try:
            r = self.cli.insert_row(path, req.sheet, req.row, values)
            if not r.ok:
                resp.message = f"插行失败：{r.error}"
                return resp
            try:
                self.cli._cache.clear()
            except Exception:
                pass
            resp.ok = True
            resp.message = f"已在 {req.sheet} 行{req.row} 上方插入新行"
            if r.needs_manual_fix:
                resp.warnings.append("插行后公式引用已机械位移，请检查跨行引用是否正确")
            if r.cache_message:
                resp.warnings.append(r.cache_message)
            # R26: 记录历史
            try:
                from engine.table_history import record_change
                record_change(req.table_stem, req.sheet, "row_insert",
                              row=req.row, source="row_insert")
            except Exception:
                pass
        except Exception as e:
            resp.message = f"插行异常：{e}"
        return resp

    def add_column(self, req: "ColumnOpRequest") -> "RowOpResponse":
        """R24: 新增列（高风险，需 confirm）。在指定列左侧插入新列。"""
        from models.agent_models import RowOpResponse
        import openpyxl
        resp = RowOpResponse(ok=False, table_stem=req.table_stem, sheet=req.sheet, col=req.col)
        if not req.confirm:
            resp.message = "列新增属高风险操作，需 confirm=True 才执行"
            resp.warnings.append("新增列会改变所有行的列结构，公式引用可能错位")
            return resp
        path = self._resolve_table_path(req.table_stem)
        if path is None:
            resp.message = f"表格不存在：{req.table_stem}"
            return resp
        try:
            wb = self.cli._load(path)
            ws = wb[req.sheet]
            ws.insert_cols(req.col)
            # 写列头
            if req.col_name:
                header_cell = ws.cell(row=1, column=req.col, value=req.col_name)
                if req.col_type:
                    ws.cell(row=2, column=req.col, value=f"{req.col_name}:{req.col_type}")
            from agent.excel.style_utils import copy_cell_style
            # 从相邻列复制样式
            try:
                copy_cell_style(ws.cell(row=1, column=req.col + 1), ws.cell(row=1, column=req.col))
            except Exception:
                pass
            cache_info = self.cli._save_with_cache_check(wb, path)
            try:
                self.cli._cache.clear()
            except Exception:
                pass
            resp.ok = True
            resp.message = f"已在 {req.sheet} 列{req.col} 插入新列「{req.col_name or '(空)'}」"
            resp.warnings.append("新增列后所有行的该列位置改变，公式引用已机械位移，请检查")
            if cache_info.get("needs_manual_fix"):
                resp.warnings.append(cache_info.get("cache_message", ""))
            # R26: 记录历史
            try:
                from engine.table_history import record_change
                record_change(req.table_stem, req.sheet, "column_add",
                              col=req.col, new_value=req.col_name, source="column_add")
            except Exception:
                pass
        except Exception as e:
            resp.message = f"新增列异常：{e}"
        return resp

    def delete_column(self, req: "ColumnOpRequest") -> "RowOpResponse":
        """R24: 删列（高风险，需 confirm）。"""
        from models.agent_models import RowOpResponse
        resp = RowOpResponse(ok=False, table_stem=req.table_stem, sheet=req.sheet, col=req.col)
        if not req.confirm:
            resp.message = "列删除属高风险操作，需 confirm=True 才执行"
            resp.warnings.append("删列会改变所有行的列结构，公式引用可能错位")
            return resp
        path = self._resolve_table_path(req.table_stem)
        if path is None:
            resp.message = f"表格不存在：{req.table_stem}"
            return resp
        try:
            wb = self.cli._load(path)
            ws = wb[req.sheet]
            ws.delete_cols(req.col)
            cache_info = self.cli._save_with_cache_check(wb, path)
            try:
                self.cli._cache.clear()
            except Exception:
                pass
            resp.ok = True
            resp.message = f"已删除 {req.sheet} 列{req.col}"
            resp.warnings.append("删列后右侧列左移，公式引用已机械位移，请检查")
            if cache_info.get("needs_manual_fix"):
                resp.warnings.append(cache_info.get("cache_message", ""))
            # R26: 记录历史
            try:
                from engine.table_history import record_change
                record_change(req.table_stem, req.sheet, "column_delete",
                              col=req.col, source="column_delete")
            except Exception:
                pass
        except Exception as e:
            resp.message = f"删列异常：{e}"
        return resp

    def rollback_history(self, record_id: str) -> "RowOpResponse":
        """R26: 按记录 id 回滚单次变更。"""
        from models.agent_models import RowOpResponse
        from engine.table_history import rollback_record
        info = rollback_record(record_id)
        if not info.get("ok"):
            return RowOpResponse(ok=False, message=info.get("message", "回滚失败"))
        rec = info["record"]
        rev = info["reverse_op"]
        ts = rec.get("table_stem", "")
        sh = rec.get("sheet", "")
        row = rec.get("row")
        col = rec.get("col")
        old_val = rec.get("old_value")
        resp = RowOpResponse(ok=False, table_stem=ts, sheet=sh, row=row, col=col)
        try:
            if rev == "set_value" and row and col:
                # 改回旧值
                from models.agent_models import CellUpdateRequest
                r = self.update_cell(CellUpdateRequest(
                    table_stem=ts, sheet=sh, row=row, col=col,
                    value=str(old_val) if old_val is not None else ""))
                resp.ok = r.ok
                resp.message = f"回滚：行{row} 列{col} 改回「{old_val}」" + ("" if r.ok else f"（失败：{r.error})")
            elif rev == "delete_row" and row:
                from models.agent_models import RowDeleteRequest
                r = self.delete_row(RowDeleteRequest(table_stem=ts, sheet=sh, row=row))
                resp.ok = r.ok
                resp.message = f"回滚：删除行{row}" + ("" if r.ok else f"（失败：{r.message})")
            elif rev == "insert_row" and row:
                from models.agent_models import RowInsertRequest
                extra = rec.get("extra", {})
                vals = extra.get("deleted_values", [])
                values = {str(i + 1): v for i, v in enumerate(vals) if v is not None}
                r = self.insert_row(RowInsertRequest(
                    table_stem=ts, sheet=sh, row=row, values=values))
                resp.ok = r.ok
                resp.message = f"回滚：用旧值重插行{row}" + ("" if r.ok else f"（失败：{r.message})")
            elif rev == "delete_column" and col:
                from models.agent_models import ColumnOpRequest
                r = self.delete_column(ColumnOpRequest(
                    table_stem=ts, sheet=sh, col=col, confirm=True))
                resp.ok = r.ok
                resp.message = f"回滚：删除列{col}" + ("" if r.ok else f"（失败：{r.message})")
            elif rev == "manual":
                resp.ok = False
                resp.message = info.get("message", "需手动恢复")
            else:
                resp.message = f"不支持的回滚操作：{rev}"
        except Exception as e:
            resp.message = f"回滚异常：{e}"
        return resp

    def suggest_rows(self, table_stem: str, sheet: str, value: str,
                     col: str = "", top: int = 3) -> "SuggestResponse":
        """R19: 名称定位失败时返回 top-N 相近行（FuzzyMatcher 模糊匹配）。

        col 可传列名或列号；为空时自动取含"名称"/"名字"的列，否则第一数据列。
        每条建议含行号/值/相似度/置信度/前3个非空字段摘要，供配表模式展示候选卡片。
        """
        from models.agent_models import SuggestResponse, SuggestRow
        from agent.excel.fuzzy_matcher import FuzzyMatcher

        resp = SuggestResponse(table_stem=table_stem, sheet=sheet, value=value)
        path = self._resolve_table_path(table_stem)
        if path is None:
            resp.message = f"表格不存在：{table_stem}"
            return resp

        try:
            rows = self.cli.read_sheet(path, sheet)
        except Exception as e:
            resp.message = f"读取 sheet 失败：{e}"
            return resp
        if not rows:
            resp.message = f"sheet '{sheet}' 无数据"
            return resp

        try:
            headers = self.cli.read_header(path, sheet) or []
        except Exception:
            headers = []

        start = (self.cli._resolve_data_start(path, sheet)
                 if hasattr(self.cli, "_resolve_data_start")
                 else getattr(self.cli, "data_start_row", 5))

        # 解析定位列：col 传列名或列号；为空自动取"名称"/"名字"列，否则第一列
        col_idx_0 = -1
        if col:
            if col.isdigit():
                ci = int(col) - 1
                if 0 <= ci < len(headers):
                    col_idx_0 = ci
            if col_idx_0 < 0:
                for i, h in enumerate(headers):
                    if h and col in str(h):
                        col_idx_0 = i
                        break
        if col_idx_0 < 0:
            for i, h in enumerate(headers):
                if h and ("名称" in str(h) or "名字" in str(h)):
                    col_idx_0 = i
                    break
        if col_idx_0 < 0:
            col_idx_0 = 0

        col_name = (str(headers[col_idx_0]).split(":")[0].strip()
                    if col_idx_0 < len(headers) and headers[col_idx_0]
                    else f"列{col_idx_0 + 1}")

        # 收集定位列非空值 + 行号 + 整行数据（供摘要）
        vals: list[tuple[int, str]] = []
        row_cells: dict[int, list] = {}
        for i, row in enumerate(rows):
            cell = row[col_idx_0] if col_idx_0 < len(row) else None
            if cell is None or not str(cell).strip():
                continue
            abs_row = start + i
            vals.append((abs_row, str(cell).strip()))
            row_cells[abs_row] = row

        if not vals:
            resp.message = f"定位列 '{col_name}' 无数据"
            resp.col = col_name
            resp.col_index = col_idx_0 + 1
            return resp

        # FuzzyMatcher top-N（子串+编辑距离+字符重叠三路融合）
        matcher = FuzzyMatcher(top_k=max(top, 1))
        by_val: dict[str, int] = {}
        for r, v in vals:
            by_val.setdefault(v, r)
        matched = matcher.search(value, [v for _, v in vals])

        suggestions: list[SuggestRow] = []
        for c in matched:
            r = by_val.get(c.value)
            if r is None:
                continue
            # 摘要：前 3 个非空数据列（跳过定位列）
            summary: dict[str, Any] = {}
            row = row_cells.get(r, [])
            for ci, cell in enumerate(row):
                if ci == col_idx_0 or ci >= len(headers):
                    continue
                if cell is None or not str(cell).strip():
                    continue
                hn = str(headers[ci]).split(":")[0].strip() or f"列{ci + 1}"
                summary[hn] = cell
                if len(summary) >= 3:
                    break
            suggestions.append(SuggestRow(
                row=r, value=c.value, score=c.score,
                confidence=c.confidence, summary=summary,
            ))
            if len(suggestions) >= top:
                break

        resp.col = col_name
        resp.col_index = col_idx_0 + 1
        resp.suggestions = suggestions
        resp.total = len(suggestions)
        if not suggestions:
            resp.message = (f"无相近项（定位列 '{col_name}' 共 {len(vals)} 个值，"
                            f"均低于相似度阈值）")
        return resp

    # ── Agent Chat ──

    # codemaker 错误类型 → 面向用户的具体提示映射。
    # key 与 CodemakerClient.CodemakerError.* 对应；缺省走兜底文案。
    _ERR_HINTS = {
        "serve_down": (
            "AI 服务未启动",
            "AI 底座（codemaker serve）未启动或不可达。请先运行：codemaker serve --port 8666 --hostname 0.0.0.0，再重试。"
        ),
        "auth_failed": (
            "AI 服务鉴权失败",
            "codemaker serve 返回 401/403，用户名/密码错误或无权限。请检查 .env 的 CODEMAKER_USERNAME / CODEMAKER_PASSWORD。"
        ),
        "provider_error": (
            "底层模型调用失败",
            "AI 底座转发到模型供应商失败（常见：账户余额不足、限流、模型不存在）。请检查模型供应商账户额度，或更换 .env 的 CODEMAKER_MODEL。"
        ),
        "timeout": (
            "AI 响应超时",
            "等待 AI 回复超时。模型负载高或网络抖动可能导致，请稍后重试；若持续超时请检查 codemaker serve 与模型供应商连通性。"
        ),
        "bad_request": (
            "请求格式错误",
            "codemaker serve 拒绝请求（400）。常见于 .env 的 CODEMAKER_MODEL 配置非法，请改为 provider/model 格式或留空。"
        ),
    }

    def _infer_error_type(self, msg: str) -> str:
        """从异常文本反推 error_type（异常未带 error_type 属性时的兜底）。"""
        m = (msg or "").lower()
        if "不可用" in msg or "连接失败" in msg or "urLError" in m or "connection refused" in m:
            return "serve_down"
        if "余额不足" in msg or "balance" in m or "provider" in m or "403" in m:
            return "provider_error"
        if "超时" in msg or "timeout" in m:
            return "timeout"
        if "400" in msg or "bad request" in m or "model" in m:
            return "bad_request"
        return "unknown"

    def _map_codemaker_error(self, err_type: str, raw: str) -> tuple[str, str]:
        """把 codemaker 错误映射成 (简短概括, 面向用户的中文提示)。

        raw 为原始异常文本，附在提示末尾便于定位（如 provider 错误的具体 code）。
        """
        hint = self._ERR_HINTS.get(err_type)
        if hint:
            summary, advice = hint
            # provider_error 把原始信息（含 code/message）一并展示，便于定位真因
            if err_type == "provider_error" and raw:
                advice = f"{advice}\n原始错误：{raw}"
            return summary, advice
        return "处理出错", f"抱歉，处理您的请求时出错：{raw}"

    def _compose_failure_summary(self, user_text: str, base_msg: str,
                                 failures: list[dict]) -> str:
        """汇总失败原因，并尝试调用 LLM 给出用户指令修改建议。"""
        lines = [base_msg or "操作未能全部完成", ""]
        lines.append("**失败原因：**")
        for i, f in enumerate(failures, 1):
            loc = f"{f.get('table', '') or '未知表'}/{f.get('sheet', '') or ''}".rstrip("/")
            col = f.get("col") or ""
            if col:
                loc += f" 列[{col}]"
            root = f.get("root_cause") or f.get("reason") or "未知"
            lines.append(f"{i}. `{loc}` — {root}")
            strat = f.get("attempted_strategies")
            if strat:
                lines.append(f"   - 已试：{strat}")

        # 调用 LLM 生成用户可执行的修改建议
        ai = getattr(self.agent, "_ai_enhancer", None)
        llm_sug = ""
        if ai is not None:
            try:
                failures_text = "\n".join(
                    f"- {f.get('table', '') or '未知表'}/{f.get('sheet', '') or ''} "
                    f"{f.get('col', '') or ''}：{f.get('root_cause', '') or f.get('reason', '') or '未知'}"
                    for f in failures
                )
                prompt = (
                    f"用户原始指令：「{user_text}」\n"
                    f"执行时发生以下失败：\n{failures_text}\n\n"
                    "请给出 1~2 条用户应如何修改原始指令才能成功的具体建议。"
                    '输出 JSON：{"suggestion": "..."}，只输出 JSON，无其他文字。'
                )
                raw = ai._call_llm(prompt, timeout=25)
                if raw:
                    data = ai.client.extract_json_from_response(raw)
                    if isinstance(data, dict) and data.get("suggestion"):
                        llm_sug = str(data["suggestion"]).strip()
            except Exception:
                llm_sug = ""

        # LLM 不可用时降级使用失败对象自带的 suggestion
        if not llm_sug:
            sugs = [str(f.get("suggestion", "")).strip() for f in failures if f.get("suggestion")]
            if sugs:
                llm_sug = "；".join(sugs)
        if llm_sug:
            lines.extend(["", "**修改建议：**", llm_sug])
        return "\n".join(lines)

    def _llm_judge_confirm(self, text: str) -> str:
        """LLM 判短句确认意图。返回 'affirm'/'deny'/'new'。

        opt-in（CODEMAKER_CONFIRM_LLM=1）：未命中 confirm_affirm/deny 词表的短句
        交 LLM 裁决，覆盖词表未收的措辞变体（「行」「可以啊」「别删」「算了」等）。
        复用 StepAIEnhancer._call_llm（隔离 session 免 R7）。失败/不可达返 'new'
        （保守作废 pending，不误触发执行）。
        """
        ai = getattr(self.agent, "_ai_enhancer", None)
        if ai is None:
            return "new"
        prompt = (
            f"用户上一轮被要求确认一个配表操作。用户回复：「{text}」\n"
            f"判断这是「确认执行」(affirm)、「取消」(deny) 还是「新指令/不相关」(new)？\n"
            f"仅输出 JSON: {{\"verdict\":\"affirm\"|\"deny\"|\"new\"}}"
        )
        try:
            raw = ai._call_llm(prompt, timeout=20)
        except Exception:
            return "new"
        if not raw:
            return "new"
        m = _re.search(r"\{.*\}", raw, _re.DOTALL)
        if not m:
            return "new"
        try:
            d = json.loads(m.group(0))
        except ValueError:
            return "new"
        v = str(d.get("verdict", "")).lower()
        return v if v in ("affirm", "deny", "new") else "new"

    def chat(self, text: str, session_id: str = "default",
             dry_run: bool = False, table_hint: Optional[str] = None,
             confirm_token: Optional[str] = None,
             confirm_cascade: bool = True) -> AgentChatResponse:
        """执行自然语言指令（通过 OrchestratorAgent 分诊 QA / CRUD）。"""
        if dry_run:
            return self._dry_run_chat(text, session_id, table_hint=table_hint)

        # ── 管道模式判定 ──
        # 输入含文件路径或多表关键词 → 走 7 步管道;否则走旧 CRUD 路径
        # 可配置 CODEMAKER_PIPELINE_MODE=auto|off|on(默认 auto)
        try:
            from agent.excel.pipeline.pipeline import (
                should_trigger_pipeline, extract_file_path, Pipeline
            )
            from agent.excel.cli_instrument import ToolSink  # noqa: F401
            if should_trigger_pipeline(text):
                file_path = extract_file_path(text)
                if file_path:
                    # 管道 step_sink:复用 agent 的 thinking_sink 通道推送 step 事件,
                    # chat_stream 统一经 _sink 转 SSE(step 与 thinking 共享 queue)
                    step_sink_cb = getattr(self.agent, "_agent_step_sink", None)
                    pipe = Pipeline(
                        cli=getattr(self.agent, "cli", None),
                        parser=getattr(self.agent, "parser", None),
                        thinking_sink=getattr(self.agent, "_agent_thinking_sink", None),
                        tool_sink=getattr(self.agent, "_agent_tool_sink", None),
                        writer=self.agent,
                        step_sink=step_sink_cb,
                    )
                    pipe_result = pipe.run(file_path, text=text,
                                           session_id=session_id,
                                           workspace=getattr(self.cli, "workspace", "Trunk"))
                    # PipelineResult → AgentChatResponse 兼容格式
                    from models.agent_models import AgentStepInfo
                    step_infos = [AgentStepInfo(name=s.name, ok=(s.status == "done"),
                                                detail=s.output or s.error or "")
                                  for s in pipe_result.steps]
                    return AgentChatResponse(
                        ok=pipe_result.ok or False,
                        session_id=session_id,
                        intent="pipeline",
                        reply_type="crud",
                        message=pipe_result.message,
                        steps=step_infos,
                        sub_tasks=pipe_result.sub_tasks,
                        thinking_steps=[{"phase": s.step_id, "detail": s.output or s.error or ""}
                                         for s in pipe_result.steps],
                        data={"produced": pipe_result.produced,
                              "report_path": pipe_result.report_path},
                        error="; ".join(e["error"] for e in pipe_result.errors) if pipe_result.errors else None,
                    )
        except Exception as e:
            # 管道模式异常 → 返回管道失败响应(保留管道上下文,不降级 QA 丢失进度)
            import logging as _log
            _log.getLogger(__name__).warning(
                f"管道模式执行异常: {e}", exc_info=True)
            return AgentChatResponse(
                ok=False,
                session_id=session_id,
                intent="pipeline",
                reply_type="crud",
                message=f"管道执行异常: {e}",
                error=f"pipeline_error: {type(e).__name__}: {e}",
                thinking_steps=[{"phase": "pipeline", "detail": f"执行异常: {e}"}],
            )

        # ─文字确认分支 ──
        # 多端对话式确认：上一轮留有 pending 确认时，用户直接回短句
        # （"无误/确认/是/取消/不"等）→ 映射为 confirm_token 回传/取消，
        # 不必点确认按钮。长文本视为新指令，作废旧 pending。
        if not confirm_token:
            pending = self._pending_confirms.get(session_id)
            if pending and time.time() < pending.get("expires_at", 0):
                t = (text or "").strip().lower().rstrip("。.！!？?~")
                _AFFIRM, _DENY = _load_confirm_keywords()
                if t in _AFFIRM:
                    confirm_token = pending["token"]
                    confirm_cascade = True
                elif t in _DENY:
                    self._pending_confirms.pop(session_id, None)
                    return AgentChatResponse(
                        ok=True, session_id=session_id, reply_type="qa",
                        intent="qa", message="已取消，未执行任何操作。",
                    )
                else:
                    # 未命中词表：opt-in LLM 裁决确认意图（CODEMAKER_CONFIRM_LLM=1），
                    # 覆盖词表未收的措辞变体（行/可以啊/别删/算了等）；否则保守作废 pending（新指令）
                    if os.getenv("CODEMAKER_CONFIRM_LLM", "0") == "1":
                        verdict = self._llm_judge_confirm(t)
                        if verdict == "affirm":
                            confirm_token = pending["token"]
                            confirm_cascade = True
                        elif verdict == "deny":
                            self._pending_confirms.pop(session_id, None)
                            return AgentChatResponse(
                                ok=True, session_id=session_id, reply_type="qa",
                                intent="qa", message="已取消，未执行任何操作。",
                            )
                        else:
                            self._pending_confirms.pop(session_id, None)
                    else:
                        self._pending_confirms.pop(session_id, None)
            elif pending:
                self._pending_confirms.pop(session_id, None)

        # ─回传分支 ──
        # 前端对危险操作点「确认」后，带上一轮返回的 confirm_token 重发原指令。
        # 校验令牌匹配且未超时 → 绕过 QA/CRUD 分诊，直接带 token 调 agent.run() 执行；
        # （分诊经 LangGraph 内部跑 agent 不透传 token，故此处直连 agent。）
        if confirm_token:
            pending = self._pending_confirms.pop(session_id, None)
            valid = bool(
                pending
                and pending.get("token") == confirm_token
                and time.time() < pending.get("expires_at", 0)
            )
            if not valid:
                return AgentChatResponse(
                    ok=False, session_id=session_id, reply_type="crud",
                    intent="delete",
                    message="确认已失效（超过 5 分钟或令牌不匹配），请重新发起该操作。",
                    error="confirm_token expired or mismatch",
                )
            session_history = self._sessions.setdefault(session_id, [])
            context = self._build_context(session_history)
            result = self.agent.run(text, confirm_token=confirm_token,
                                    context=context, confirm_cascade=confirm_cascade,
                                    session_id=session_id)
            if result is None:
                return AgentChatResponse(
                    ok=False, session_id=session_id,
                    message="执行失败：无返回结果", error="无返回结果",
                )
            return self._finalize_crud(result, session_id, text, session_history)

        # 表单式新增拦截：纯类别描述（如「增加灵兽」，无具体对象/字段）→
        # 返回可填写表单，不写盘。前端据 reply_type="form" 渲染表单，
        # 填好后调用 /api/tables/add-form/validate 与 /commit 完成新增。
        form = self._try_intercept_add_form(text)
        if form is not None:
            cols = [c.model_dump() if hasattr(c, "model_dump") else c.dict()
                    for c in form.columns]
            return AgentChatResponse(
                ok=True, session_id=session_id, intent="add", reply_type="form",
                message=form.message or "请填写新增信息",
                data={
                    "form": True,
                    "table_stem": form.table_stem,
                    "sheet": form.sheet,
                    "table_path": form.table_path,
                    "columns": cols,
                    "empty_row": form.empty_row,
                    "validate_url": "/api/tables/add-form/validate",
                    "commit_url": "/api/tables/add-form/commit",
                },
            )

        session_history = self._sessions.setdefault(session_id, [])

        # 构建上下文：最近N条操作摘要
        context = self._build_context(session_history)

        # 通过 OrchestratorAgent 分诊（传递 context 帮助 LLM 消解代词）
        try:
            route_result: RouteResult = self.router.chat(text, session_id, context=context)
        except Exception as e:
            err_type = getattr(e, "error_type", "") or self._infer_error_type(str(e))
            message, qa_answer = self._map_codemaker_error(err_type, str(e))
            route_result = RouteResult(
                intent="qa",
                message=message,
                qa_answer=qa_answer,
            )

        # ── QA 分支 ──
        if route_result.intent == "qa":
            op_record = {
                'id': uuid.uuid4().hex[:8],
                'text': text,
                'intent': 'qa',
                'reply_type': 'qa',
                'qa_answer': route_result.qa_answer,
                'timestamp': datetime.now().isoformat(),
            }
            session_history.append(op_record)

            return AgentChatResponse(
                ok=True,
                session_id=session_id,
                intent="qa",
                message=route_result.qa_answer,
                reply_type="qa",
                steps=[],
            )

        # ── CRUD 分支 ──
        result = route_result.crud_result
        if result is None:
            # 单一入口路由：原按 CODEMAKER_EXCEL_PIPELINE_V2（默认 "0"）再判一次 run_v2
            # 与 agent.py:5178 / orchestrator.py:135 / configuration.py:54 四处默认 "1"
            # 不一致，造成同开关两默认值 + 重复路由混叠。现统一收敛到 agent.run()：
            # run() 内部据默认 ON 的同一开关分流到 run_v2，此处不再二次猜，消除分歧。
            result = self.agent.run(text, context=context, session_id=session_id)

        if result is None:
            return AgentChatResponse(
                ok=False, session_id=session_id,
                message="执行失败：无返回结果", error="无返回结果",
            )

        return self._finalize_crud(result, session_id, text, session_history)

    # ── 真流式 chat：思考过程边执行边推送 ──
    def chat_stream(self, text: str, session_id: str = "default",
                    dry_run: bool = False, table_hint: Optional[str] = None,
                    confirm_token: Optional[str] = None,
                    confirm_cascade: bool = True,
                    cancel_event=None,
                    reply_queue=None):
        """async generator：yield (event_type, payload)。

        event_type:
          - "thinking": {"phase","detail"} —— 思考逐条实时推送
          - "tool": {name,desc,cmd,result,ok,ts} —— CLI 调用可见性(管道模式)
          - "done": AgentChatResponse.model_dump() —— 最终结果
          - "error": {"message"}

        实现：agent.run 同步放线程，on_thinking 回调推 queue，
        主循环 async 等 queue，思考实时 yield，run 完 yield done。
        """
        import asyncio
        import queue as _queue
        import threading

        loop = asyncio.get_event_loop()
        q: _queue.Queue = _queue.Queue()
        _SENTINEL = object()

        def _sink(phase: str, detail: str):
            q.put(("thinking", {"phase": phase, "detail": detail}))

        # tool sink:管道模式 instrument 层推送 (event_type, payload)
        def _tool_sink(event_type: str, payload: dict):
            q.put((event_type, payload))

        # step sink:管道模式每步完成推送 {name,ok,detail,step_id,status}
        def _step_sink(payload: dict):
            q.put(("step", payload))

        # subtask sink:Step5 子任务级进度（start/done），前端增量渲染卡片骨架
        def _subtask_sink(event_type: str, payload: dict):
            q.put((event_type, payload))

        _finished = {"v": False}  # 心跳守护退出信号：worker 结束后置 True

        def _worker():
            try:
                # 注入思考回调 + tool 回调 + step 回调 + subtask 回调到 agent 实例
                self.agent._agent_thinking_sink = _sink
                self.agent._agent_tool_sink = _tool_sink
                self.agent._agent_step_sink = _step_sink
                self.agent._agent_subtask_sink = _subtask_sink
                self.agent._cancel_event = cancel_event
                # 中断反问回调（#41：env CODEMAKER_INTERACTIVE_REPAIR=0 关闭 →
                # _ask_callback 不注入，阻断错误走原 ABORT 路径，防非交互场景卡死）
                if os.getenv("CODEMAKER_INTERACTIVE_REPAIR", "1") != "0":
                    # dry_run / 无 reply_queue 时默认"接受建议"（非 skip）：
                    # - PK 冲突（_ask_pk_conflict）：返 accept_suggest=True → validator
                    #   调 _apply_pk_to_intent 写入 suggested_id，标 _pk_resolved
                    # - 通用硬 issue（_ask_hard_issue）：返 mode=field + value=suggestion
                    #   → _apply_issue_fix_to_intent 改写 fields[col]
                    # 语义：链路完整性优先，不跳过子任务（用户要求默认接受建议）。
                    # 真 skip 仅由用户在交互 UI 主动点击产生。
                    _accept_suggestion = os.getenv(
                        "CODEMAKER_DRY_RUN_ACCEPT_SUGGEST", "1") == "1"
                    def _ask_callback(question: dict):
                        q.put(("ask", question))
                        if reply_queue is None:
                            if _accept_suggestion:
                                _sug = question.get("suggested_id")
                                if _sug is None:
                                    _sug = question.get("suggestion")
                                return {
                                    "mode": "field",
                                    "accept_suggest": True,
                                    "value": _sug,
                                    "custom_id": _sug,
                                    "text": _sug,
                                }
                            return {"mode": "skip"}
                        _ce = getattr(self.agent, "_cancel_event", None)
                        while True:
                            try:
                                return reply_queue.get(timeout=0.5)
                            except _queue.Empty:
                                if _ce is not None and _ce.is_set():
                                    return {"mode": "skip"}
                    self.agent._ask_callback = _ask_callback
                    # §4.5 交互反问通道下传 ValidatorAgent（两段式 ask_user 用）
                    _validator = getattr(self.agent, "_validator_agent", None)
                    if _validator is not None and hasattr(_validator, "set_ask_callback"):
                        _validator.set_ask_callback(_ask_callback)
                # QA / 路由路径也注入 sink：意图分类与 QAHandler 内部步骤可见
                router = getattr(self, "router", None)
                qa_handler = getattr(router, "qa_handler", None)
                if router is not None:
                    router._thinking_sink = _sink
                    router._step_sink = _step_sink
                if qa_handler is not None:
                    qa_handler._thinking_sink = _sink
                    qa_handler._step_sink = _step_sink
                    # 取消事件下传 QA 路径：qa_handler 直连 prompt + 共享 chat model
                    # （classify 节点与 QA model.invoke 同一 CodemakerChatModel 实例）
                    qa_handler._cancel_event = cancel_event
                    _qm = getattr(qa_handler, "model", None)
                    if _qm is not None:
                        _qm._cancel_event = cancel_event
                # token 级流式：LLM 同步调用期间经 /event SSE 推 llm_token
                set_token_sink(lambda kind, delta: q.put(
                    ("llm_token", {"kind": kind, "delta": delta})))
                # dry_run 标志下传 agent（Step2 _step2_validate_intents 读此，
                # dry_run 时硬 issue 复位放行不真过滤，保链路完整走通预览）。
                self.agent._dry_run_flag = bool(dry_run)
                resp = self.chat(text=text, session_id=session_id, dry_run=dry_run,
                                  table_hint=table_hint, confirm_token=confirm_token,
                                  confirm_cascade=confirm_cascade)
                q.put(("done", resp))
            except Exception as e:
                import traceback as _tb
                logger.warning("chat_stream worker 异常: %s: %s", type(e).__name__, e)
                _tb.print_exc()
                q.put(("error", {"message": str(e)}))
            finally:
                set_token_sink(None)
                self.agent._agent_thinking_sink = None
                self.agent._agent_tool_sink = None
                self.agent._agent_step_sink = None
                self.agent._agent_subtask_sink = None
                self.agent._cancel_event = None
                self.agent._ask_callback = None
                # §4.5 清理 ValidatorAgent 的 _ask_callback
                _validator = getattr(self.agent, "_validator_agent", None)
                if _validator is not None and hasattr(_validator, "set_ask_callback"):
                    _validator.set_ask_callback(None)
                router = getattr(self, "router", None)
                qa_handler = getattr(router, "qa_handler", None)
                if router is not None:
                    router._thinking_sink = None
                    router._step_sink = None
                if qa_handler is not None:
                    qa_handler._thinking_sink = None
                    qa_handler._step_sink = None
                    qa_handler._cancel_event = None
                    _qm = getattr(qa_handler, "model", None)
                    if _qm is not None:
                        _qm._cancel_event = None
                _finished["v"] = True
                q.put(_SENTINEL)

        t = threading.Thread(target=_worker, daemon=True)
        t.start()

        # ── 阶段切分：统一单一开关 CODEMAKER_EXCEL_PIPELINE_V2（默认 ON）──
        # V2 路径走 core/pipeline 契约，SSE event 自带 step_id（step1_parse/
        # step2_validate/step3_execute/step4_conclude），前端直通不翻译。
        # =0 降级到旧 run() 6 步路径，用 s1_parse/s2_partition/... 前缀映射。
        _V2 = os.getenv("CODEMAKER_EXCEL_PIPELINE_V2", "1") != "0"
        if _V2:
            _STAGE_ORDER = ["step1_parse", "step2_validate", "step3_execute",
                            "step4_conclude", "summary"]
            _STAGE_TITLES = {
                "step1_parse": "Step1 解析",
                "step2_validate": "Step2 校验",
                "step3_execute": "Step3 执行",
                "step4_conclude": "Step4 汇总",
                "summary": "总结归纳",
            }
        else:
            _STAGE_ORDER = ["s1_parse", "s2_partition", "s3_plan",
                            "s4_verify", "s5_apply", "s6_summary", "summary"]
            _STAGE_TITLES = {
                "s1_parse": "Step1 解析",
                "s2_partition": "Step2 分区",
                "s3_plan": "Step3 计划",
                "s4_verify": "Step4 校验",
                "s5_apply": "Step5 应用",
                "s6_summary": "Step6 汇总",
                "summary": "总结归纳",
            }
        _stage_total = len(_STAGE_ORDER) - 1  # V2=4,6 步=6（去 summary）

        def _stage_for_thinking(payload):
            detail = str(payload.get("detail", "") or "")
            phase = str(payload.get("phase", "") or "")
            # V2 路径：SSE event 已带 step_id（step1_parse 等），thinking 的 phase
            # 映射到对应 V2 阶段。废弃旧字符串前缀「Step1:」猜测（contracts 已禁）。
            if _V2:
                # run_v2 的 stage_start/stage_end 用 title（"Step1 解析"）与
                # step_id（"step1_parse"）作 phase 发 thinking，前端必须都能映射回
                # 对应阶段，否则 step 卡片打不开（阶段气泡缺失）。
                _sid_map = {
                    "step1_parse": "step1_parse", "step2_validate": "step2_validate",
                    "step3_execute": "step3_execute", "step4_conclude": "step4_conclude",
                    "Step1 解析": "step1_parse", "Step2 校验": "step2_validate",
                    "Step3 执行": "step3_execute", "Step4 汇总": "step4_conclude",
                }
                if phase in _sid_map:
                    return _sid_map[phase]
                if phase in ("意图分类", "问题分析", "解析", "路由",
                             "定位", "跨表探索", "计划", "分区"):
                    return "step1_parse"
                if phase == "校验":
                    return "step2_validate"
                if phase == "执行":
                    return "step3_execute"
                if phase == "汇总":
                    return "step4_conclude"
                return None
            # 6 步流程：detail 前缀显式标记所属步
            # 兼容旧 pipeline 8 步前缀：Step0 跳过、Step3/4 合并计划、Step7 归汇总
            for prefix, sid in (("Step1:", "s1_parse"), ("Step2:", "s2_partition"),
                                ("Step3:", "s3_plan"), ("Step4:", "s3_plan"),
                                ("Step5:", "s4_verify"), ("Step6:", "s5_apply"),
                                ("Step7:", "s6_summary")):
                if detail.startswith(prefix):
                    return sid
            # summary 阶段仅由 done 开启，流内事件不映射，避免总结气泡重复
            # 旧 CRUD 路径 phase 兜底（无 Step 前缀时）
            if phase in ("意图分类", "问题分析", "解析"):
                return "s1_parse"
            if phase in ("路由", "定位", "跨表探索"):
                return "s2_partition"
            if phase == "计划":
                return "s3_plan"
            if phase == "校验":
                return "s4_verify"
            if phase == "执行":
                return "s5_apply"
            if phase == "汇总":
                return "s6_summary"
            return None

        def _stage_for_step(payload):
            sid0 = str(payload.get("step_id", "") or "")
            name = str(payload.get("name", "") or "")
            # V2 路径：event 自带 step_id（step1_parse 等），直通不翻译。
            # 仅 legacy step name 兜底映射到 V2 阶段（兼容旧 SubAgent 产出）。
            if _V2:
                if sid0 in ("step1_parse", "step2_validate",
                            "step3_execute", "step4_conclude"):
                    return sid0
                if name in ("resolve_table", "resolve_sheet", "match_locator",
                            "match_target", "locate_row", "read_cell"):
                    return "step1_parse"
                if name == "add_values":
                    return "step3_execute"
                if name in ("write", "write_op", "append_row", "delete_row",
                            "delete_cell", "delete_column", "auto_sort"):
                    return "step3_execute"
                return None
            # 新 6 步精确匹配（单表 CRUD 路径，优先于旧前缀兼容）
            if name == "Step1解析": return "s1_parse"
            if name == "Step3计划": return "s3_plan"
            if name == "Step4校验": return "s4_verify"
            if name == "Step5应用": return "s5_apply"
            if name == "Step6汇总": return "s6_summary"
            tag = sid0.split("_")[0] if sid0 else ""
            # 旧 pipeline 8 步前缀兼容（tag 0 跳过，3/4 合并计划，7 归汇总）
            if tag == "0" or name.startswith("Step0"):
                return None  # 断点检查不作为前端阶段
            if tag == "1" or name.startswith("Step1"):
                return "s1_parse"
            if tag == "2" or name.startswith("Step2"):
                return "s2_partition"
            if tag == "3" or name.startswith("Step3"):
                return "s3_plan"
            if tag == "4" or name.startswith("Step4"):
                return "s3_plan"
            if tag == "5" or name.startswith("Step5"):
                return "s4_verify"
            if tag == "6" or name.startswith("Step6"):
                return "s5_apply"
            if tag == "7" or name.startswith("Step7"):
                return "s6_summary"
            # 旧 CRUD step 名兜底
            if name in ("resolve_table", "resolve_sheet", "match_locator",
                        "match_target", "locate_row", "read_cell"):
                return "s2_partition"
            if name == "add_values":
                return "s5_apply"
            if name in ("write", "write_op", "append_row", "delete_row",
                        "delete_cell", "delete_column", "auto_sort"):
                return "s5_apply"
            return None

        def _compose_stage_content(sid, buf):
            lines = []
            if buf["steps"] or buf["tools"]:
                for s in buf["steps"]:
                    # s 可能是 dict（step 事件 payload）或 AgentStep 对象（旧路径），
                    # 统一用 getattr/dict 取值兼容
                    _ok = s.get("ok") if isinstance(s, dict) else getattr(s, "ok", False)
                    _name = s.get("name", "") if isinstance(s, dict) else getattr(s, "name", "")
                    _detail = s.get("detail", "") if isinstance(s, dict) else getattr(s, "detail", "")
                    icon = "✅" if _ok else "❌"
                    lines.append(f"- {icon} **{_name}**：{_detail}")
                for tl in buf["tools"]:
                    icon = "✅" if tl.get("ok") else "❌"
                    lines.append(f"- {icon} `{tl.get('name', '')}` {tl.get('desc', '')}")
            else:
                for _ph, d in buf["thinking"]:
                    # 开发者 meta 行不进回复体（Thinking 折叠里仍可见）
                    if d.startswith("规则短路优先"):
                        continue
                    lines.append(f"- {d}")
            return "\n".join(lines)

        state = {"cur": None, "buf": None}

        def _open(sid):
            state["cur"] = sid
            state["buf"] = {"thinking": [], "steps": [], "tools": []}
            return ("stage_start", {"stage_id": sid, "title": _STAGE_TITLES[sid], "total": _stage_total})

        def _close():
            sid = state["cur"]
            if sid is None:
                return None
            content = _compose_stage_content(sid, state["buf"])
            state["cur"] = None
            state["buf"] = None
            return ("stage_end", {"stage_id": sid, "title": _STAGE_TITLES[sid],
                                  "content": content})

        # 多阶段聚合 buffer：各阶段独立累积，最终按首次出现顺序统一关闭
        # 解决多指令聚合模式下阶段回退导致气泡碎片化的问题
        stage_buffers: dict[str, dict] = {}
        stage_order: list[str] = []
        # 当前活跃阶段（用于 thinking/step 归入；tool 无阶段归属时归当前）
        active_sid: list[str | None] = [None]
        # 已 stage_end 关闭的阶段集合：done 批量收尾时跳过，避免重复事件
        closed_stages: set[str] = set()

        def _ensure_stage(sid):
            """确保阶段 buffer 存在并标记顺序，返回是否首次（需发 stage_start）。"""
            if sid not in stage_buffers:
                stage_buffers[sid] = {"thinking": [], "steps": [], "tools": []}
                stage_order.append(sid)
                return True
            return False

        # 心跳守护：15s 无新事件推一条 heartbeat thinking，防单次长 LLM 期间前端空白。
        # worker 结束（_finished）或达上限自退；q.put 线程安全，不丢消息。
        cur_subtask = {"idx": 0, "total": 0}

        def _peek_llm():
            c = getattr(self.agent, "_llm_counter", None)
            return c.peek_total() if c is not None else 0

        # F4: LLM 计数长时间不变 → 主动提示疑似 hang（避免用户被动等待无效）
        # 快赢2:心跳去重——内容(idx/total/calls)不变则不 put,避免 ask 阻塞时刷屏。
        _hb_state = {"last_calls": -1, "stale": 0, "urgent_sent": False,
                     "last_detail": ""}

        async def _heartbeat_loop():
            for _ in range(240):  # 上限 60 分钟
                await asyncio.sleep(15)
                if _finished["v"]:
                    return
                calls = _peek_llm()
                idx, total = cur_subtask["idx"], cur_subtask["total"]
                if total > 0:
                    detail = f"仍在执行：第 {idx}/{total} 子任务，已调用 {calls} 次模型"
                else:
                    # 解析/规划阶段（Step5 未启动，total=0）：不显示误导的 0/0
                    detail = f"意图解析中：已调用 {calls} 次模型"
                # 快赢2:内容不变不重发(避免 ask 阻塞时刷屏),urgent 仍放行
                if detail != _hb_state["last_detail"]:
                    _hb_state["last_detail"] = detail
                    q.put(("heartbeat", {
                        "phase": "心跳",
                        "detail": detail,
                        "llm_calls": calls, "subtask_idx": idx, "subtask_total": total,
                    }))
                # F4: LLM 计数 ≥60s(4×15s)不变且 worker 未完 → 发 heartbeat_urgent
                if calls == _hb_state["last_calls"]:
                    _hb_state["stale"] += 1
                else:
                    _hb_state["stale"] = 0
                    _hb_state["last_calls"] = calls
                if _hb_state["stale"] >= 4 and not _hb_state["urgent_sent"]:
                    _hb_state["urgent_sent"] = True
                    q.put(("heartbeat_urgent", {
                        "phase": "心跳",
                        "detail": "疑似 serve hang（LLM 调用计数长时间未变化），建议点击停止",
                        "llm_calls": calls, "subtask_idx": idx, "subtask_total": total,
                    }))

        _hb_task = asyncio.ensure_future(_heartbeat_loop())

        async def _gen():
            while True:
                item = await loop.run_in_executor(None, q.get)
                if item is _SENTINEL:
                    break
                etype, payload = item
                if etype in ("heartbeat", "heartbeat_urgent"):
                    # 心跳：直挂当前阶段气泡（thinking 流），不触发 stage 切分
                    if active_sid[0] is not None and active_sid[0] in stage_buffers:
                        stage_buffers[active_sid[0]]["thinking"].append(
                            ("心跳", payload.get("detail", "")))
                    yield etype, payload
                    continue
                if etype in ("thinking", "tool", "step", "subtask_start", "subtask_done"):
                    if etype == "thinking":
                        sid = _stage_for_thinking(payload)
                    elif etype == "step":
                        sid = _stage_for_step(payload)
                    elif etype in ("subtask_start", "subtask_done"):
                        # 子任务事件归属执行阶段（V2=step3_execute, 6步=s5_apply）
                        # 原 _4STEP 重命名漏改点修复（_4STEP→_V2）
                        sid = "step3_execute" if _V2 else "s5_apply"
                        if etype == "subtask_start":
                            cur_subtask["idx"] = payload.get("idx", cur_subtask["idx"])
                            cur_subtask["total"] = payload.get("total", cur_subtask["total"])
                    else:
                        sid = None
                    if sid is not None:
                        is_new = _ensure_stage(sid)
                        if is_new:
                            # 首次出现该阶段：先关闭当前活跃阶段（顺序收尾，
                            # 前端 step1 气泡底部 spinner 在开 step2 时即停），
                            # 再发新 stage_start 新建气泡。
                            # 多指令 stage_resume（is_new=False）路径不变 →
                            # 聚合模式气泡碎片化修复仍生效。
                            prev_sid = active_sid[0]
                            if prev_sid is not None and prev_sid != sid \
                                    and prev_sid in stage_buffers:
                                prev_buf = stage_buffers[prev_sid]
                                prev_content = _compose_stage_content(prev_sid, prev_buf)
                                yield ("stage_end", {
                                    "stage_id": prev_sid,
                                    "title": _STAGE_TITLES[prev_sid],
                                    "content": prev_content,
                                })
                                closed_stages.add(prev_sid)
                            yield ("stage_start", {"stage_id": sid, "title": _STAGE_TITLES[sid], "total": _stage_total})
                        elif active_sid[0] != sid:
                            # 阶段重新激活（多指令2的Step2 在任务1Step6后）
                            # 发 stage_resume，前端切回该阶段已有气泡（不新建）
                            yield ("stage_resume", {"stage_id": sid, "title": _STAGE_TITLES[sid]})
                        active_sid[0] = sid
                    # 归入对应阶段 buffer（sid 已知）或当前活跃阶段（tool 无阶段归属时）
                    target_sid = sid if sid is not None else active_sid[0]
                    if target_sid is not None and target_sid in stage_buffers:
                        buf = stage_buffers[target_sid]
                        if etype == "thinking":
                            buf["thinking"].append(
                                (payload.get("phase", ""), payload.get("detail", "")))
                        elif etype == "step":
                            buf["steps"].append(payload)
                        elif etype == "tool":
                            buf["tools"].append(payload)
                        # subtask 不进 buffer（前端实时渲染独立卡片），仅透传
                        yield etype, payload
                    elif etype in ("tool", "subtask_start", "subtask_done"):
                        # tool 无阶段归属且无活跃阶段：仍 yield（工具调用独立展示）
                        # subtask 兜底（sid=s5_apply 通常已建 buffer，此处防异常）
                        yield etype, payload
                elif etype in ("done", "error"):
                    # done/error：按首次出现顺序关闭所有阶段（跳过已顺序收尾的阶段）
                    for sid in stage_order:
                        if sid in closed_stages:
                            continue
                        buf = stage_buffers[sid]
                        content = _compose_stage_content(sid, buf)
                        yield ("stage_end", {"stage_id": sid, "title": _STAGE_TITLES[sid],
                                             "content": content})
                        closed_stages.add(sid)
                    msg = (getattr(payload, "message", "") or "") if etype == "done" \
                        else (payload.get("message", "") or "")
                    failures = (getattr(payload, "failures", []) or []) if etype == "done" else []
                    if etype == "done" and failures:
                        # 失败场景：总结归纳里显式列出原因 + LLM 修改建议（线程池执行避免阻塞事件循环）
                        msg = await loop.run_in_executor(
                            None, self._compose_failure_summary, text, msg, failures)
                    elif etype == "done" and getattr(payload, "result_table", None) is not None:
                        # 有结构化 result_table 时，总结回复只保留首行摘要，
                        # 明细由前端表体卡片渲染，避免 k=v 长文本与卡片重复
                        msg = msg.split("\n", 1)[0]
                    yield ("stage_start", {"stage_id": "summary",
                                           "title": _STAGE_TITLES["summary"]})
                    yield ("stage_end", {"stage_id": "summary",
                                         "title": _STAGE_TITLES["summary"],
                                         "content": msg})
                    yield etype, payload
                    # drain 剩余 sentinel
                    break
                else:
                    yield etype, payload
        return _gen()

    # D6.4: 用户纠正识别关键词（消息含其一 → 强制纠正标记，绕过 resolved 比对）
    _CORRECTION_KEYWORDS = ("不是", "搞错了", "应该是", "改回", "不对", "错了")

    def _detect_user_correction(self, result: AgentResult, session_id: str,
                                text: str) -> None:
        """跨轮用户纠正识别：本轮定位 vs 上一轮缓存。

        判定规则：
        1. 上一轮缓存存在且本轮定位了同表同 sheet 但 resolved 不同 → 纠正。
        2. 本轮消息含纠正关键词（"不是/搞错了/应该是/改回/不对/错了"）→ 强制纠正。
        3. 命中纠正 → result.user_corrected=True，corrected_to 取本轮 resolved
           （行纠正优先于列纠正，因行定位更易歧义）。

        无论是否命中纠正，本轮 evidence（table_stem/sheet/col_resolved/row_resolved）
        都写入缓存，供下一轮比对。无定位结果（needs_confirm 或 dry_run 失败）跳过缓存。

        注意：写盘在 agent.py `_run_single` finally，本方法只设字段值。
        """
        if result.needs_confirm:
            return  # 待确认态不更新缓存，等确认回传时再比

        last = self._session_last_evidence.get(session_id)
        cur_table = result.table_stem
        cur_sheet = result.table_sheet
        cur_col = result.col_evidence.get("resolved") if result.col_evidence else None
        cur_row = result.row_evidence.get("resolved") if result.row_evidence else None

        corrected = False
        if last:
            # 规则2: 纠正关键词触发，强制标记（即使 resolved 恰好相同也标记，
            # 因关键词显式表达纠正意图）
            if any(kw in text for kw in self._CORRECTION_KEYWORDS):
                corrected = True
            # 规则1: 同表同 sheet 但 resolved 不同 → 纠正
            elif (last.get("table_stem") == cur_table
                  and last.get("sheet") == cur_sheet
                  and last.get("col_resolved") and cur_col
                  and last["col_resolved"] != cur_col):
                corrected = True
            elif (last.get("table_stem") == cur_table
                  and last.get("sheet") == cur_sheet
                  and last.get("row_resolved") and cur_row
                  and last["row_resolved"] != cur_row):
                corrected = True

        if corrected:
            result.user_corrected = True
            # 行纠正优先（行定位歧义场景多），无行纠正则取列纠正
            result.corrected_to = cur_row or cur_col

        # 规则: 无定位结果不更新缓存，避免下一轮误比对
        if cur_table and (cur_col or cur_row):
            self._session_last_evidence[session_id] = {
                "table_stem": cur_table,
                "sheet": cur_sheet,
                "col_resolved": cur_col,
                "row_resolved": cur_row,
            }

    def _finalize_crud(self, result: AgentResult, session_id: str, text: str,
                       session_history: list) -> AgentChatResponse:
        """把 AgentResult 组装为 AgentChatResponse（含 needs_confirm 映射）。

        危险操作首次触发时 result.needs_confirm=True：暂存 confirm_token 到
        session 级 pending（带 TTL），响应回传 needs_confirm/confirm_token/
        confirm_message，前端据此渲染确认按钮。用户点确认后带 token 重发，
        走 chat() 的确认回传分支执行。
        """
        steps = [AgentStepInfo(name=s.name, ok=s.ok, detail=s.detail)
                 for s in result.steps]

        # D6.4: 跨轮用户纠正识别——本轮定位结果与上一轮缓存比对，判定是否纠正。
        # 命中纠正 → 回填 result.user_corrected / corrected_to（evidence 写盘已由
        # agent.py _log_evidence 在 _run_single finally 触发，这里只设字段值）。
        self._detect_user_correction(result, session_id, text)

        # T2/R15: 行定位歧义候选列表（含 summary），供前端渲染候选卡片。
        # agent.py _fill_row_evidence 已组装 alternatives（含 row/value/current/summary）。
        row_alts: list[dict] = []
        re_ev = result.row_evidence or {}
        if re_ev.get("ambiguous"):
            for alt in re_ev.get("alternatives") or []:
                row_alts.append({
                    "row": alt.get("row"),
                    "value": alt.get("value"),
                    "current": alt.get("current", False),
                    "summary": alt.get("summary") or {},
                })

        # 需二次确认：不写盘、不记历史、不拍 checkpoint，暂存待确认令牌
        if result.needs_confirm and result.confirm_token:
            self._pending_confirms[session_id] = {
                "token": result.confirm_token,
                "text": text,
                "expires_at": time.time() + self._CONFIRM_TTL_SECONDS,
            }
            return AgentChatResponse(
                ok=bool(result.ok) if result.ok is not None else False,
                session_id=session_id,
                intent=result.intent.action if result.intent else "delete",
                message=result.message,
                reply_type="crud",
                steps=steps,
                needs_confirm=True,
                confirm_token=result.confirm_token,
                confirm_message=result.message,
                confirm_kind=getattr(result, "confirm_kind", "") or "cascade",
                row_alternatives=row_alts,
                thinking_steps=self._build_thinking_steps(result),
                cross_table_candidates=getattr(result, "cross_table_candidates", []) or [],
                pending_search=getattr(result, "pending_search", None),
            )

        diff_preview = None
        checkpoint_id = None
        if result.ok and result.final and result.final.ok:
            op_record = {
                'id': uuid.uuid4().hex[:8],
                'text': text,
                'intent': result.intent.action if result.intent else '',
                'reply_type': 'crud',
                'table_hint': result.intent.table_hint if result.intent else '',
                'sheet_hint': result.intent.sheet_hint if result.intent else '',
                'target_field': result.intent.target_field if result.intent else '',
                'value': result.intent.value if result.intent else '',
                'steps': [{'name': s.name, 'ok': s.ok, 'detail': s.detail}
                          for s in result.steps],
                'timestamp': datetime.now().isoformat(),
            }
            session_history.append(op_record)
            diff_preview = self._build_diff_from_result(result)
            # 写动作成功后拍 checkpoint：本次操作完成后的 resources/ 全量快照。
            # 回退到该 checkpoint = 还原到"本次写操作完成后/下次输入前"的状态。
            checkpoint_id = self._make_checkpoint(session_id, text)

        result_table = self._build_result_table(result)

        # 复合操作：把每个子任务的步骤/结果行/定位表各自归组，前端可分段渲染。
        # 单指令时 result.sub_tasks 为空，sub_tasks 也为空，前端走原平铺路径。
        sub_tasks: list[SubTaskInfo] = []
        for sub in result.sub_tasks:
            sub_rt = self._build_result_table_from_subtask(sub)
            _sub_steps = sub.get("steps", []) if isinstance(sub, dict) else []
            sub_tasks.append(SubTaskInfo(
                index=sub.get("index", 1),
                intent_action=sub.get("intent_action", ""),
                ok=sub.get("ok", True),
                message=sub.get("message", ""),
                steps=[AgentStepInfo(
                           name=(s.get("name", "") if isinstance(s, dict)
                                 else getattr(s, "name", "")),
                           ok=(s.get("ok", False) if isinstance(s, dict)
                               else getattr(s, "ok", False)),
                           detail=(s.get("detail", "") if isinstance(s, dict)
                                   else getattr(s, "detail", "")))
                       for s in _sub_steps],
                result_table=sub_rt,
                table_stem=sub.get("table_stem", ""),
                table_sheet=sub.get("table_sheet", ""),
                needs_user_fill=sub.get("needs_user_fill", []) or [],
                partial=sub.get("partial", False),
            ))

        # D5: ok=False 时用规范化 aggregated_message（不含成功步骤文本）
        final_message = (result.aggregated_message if result.ok is False
                         else result.message)
        return AgentChatResponse(
            ok=bool(result.ok) if result.ok is not None else False,  # D2: None（未完成验证）视为失败
            session_id=session_id,
            intent=result.intent.action if result.intent else "unknown",
            message=final_message,
            reply_type="crud",
            steps=steps,
            data=result.final.data if (result.final and isinstance(result.final.data, dict)) else None,
            diff_preview=diff_preview,
            result_table=result_table,
            sub_tasks=sub_tasks,
            checkpoint_id=checkpoint_id,
            error=None if result.ok else final_message,
            needs_manual_fix=getattr(result.final, "needs_manual_fix", False) if result.final else False,
            cache_message=getattr(result.final, "cache_message", "") if result.final else "",
            row_alternatives=row_alts,
            thinking_steps=self._build_thinking_steps(result),
            multi_results=self._build_multi_results(result),
            cross_table_candidates=getattr(result, "cross_table_candidates", []) or [],
            pending_search=getattr(result, "pending_search", None),
            needs_user_fill=getattr(result, "needs_user_fill", []) or [],
            partial=getattr(result, "partial", False),
            failures=getattr(result, "failures", []) or [],
        )

    def _dry_run_chat(self, text: str, session_id: str,
                      table_hint: Optional[str] = None) -> AgentChatResponse:
        """预览模式：在临时副本上执行，返回差异预览。

        先经 OrchestratorAgent.classify 分诊：qa 类问题（"有哪些表""你好"）
        不写盘，直接返回 QA 答案；crud 类才进入临时副本执行。

        table_hint 已给定时（调用方已确知 crud 与目标表），跳过 LLM 分类与
        预解析——run() 内部会 parse_multi，预览路径由 3 次 LLM 调用降到 1 次，
        既省时又让重试预算可塞进上层 120s（E02/K02 曾因串联调用挂起超时）。
        """
        # 0. 意图分诊：table_hint 已给定→确属 crud，跳过；否则 qa 天然适配预览
        if not table_hint:
            try:
                route = self.router.classify(text, session_id)
                if route.intent == "qa":
                    return AgentChatResponse(
                        ok=True, session_id=session_id, intent="qa",
                        message=route.qa_answer, reply_type="qa", steps=[],
                    )
            except Exception:
                # 分诊失败不阻断，回退到 CRUD 解析路径（由其自行报错）
                pass

        # 1. 解析意图：table_hint 已知时用 stub 仅用于定位表（run() 内部会真正 parse_multi）
        if table_hint:
            intent = NLIntent(raw=text, action="get", table_hint=table_hint)
        else:
            try:
                intent = self.agent.parser.parse(text)
            except Exception as e:
                # O20f：parse 单意图失败时降级 parse_multi（复杂跨表指令更健壮，
                # S4 万圣狂欢 6 表 add+modify 混合指令单 parse 易超时/空响应崩）。
                # parse_multi 失败返空 list 不 raise，取首条作为定位 intent。
                # 仍无 intent 才返回错误（真正 LLM 不可用）。
                pm_intents = []
                if hasattr(self.agent.parser, "parse_multi"):
                    try:
                        pm_intents = self.agent.parser.parse_multi(text) or []
                    except Exception:
                        pm_intents = []
                if pm_intents:
                    intent = pm_intents[0]
                else:
                    err_type = getattr(e, "error_type", "") or self._infer_error_type(str(e))
                    _, advice = self._map_codemaker_error(err_type, str(e))
                    return AgentChatResponse(
                        ok=False, session_id=session_id,
                        message=advice, error=str(e),
                    )

        # 2. 定位表格文件
        path, sheet = self.agent._resolve_table(intent)
        if path is None:
            return AgentChatResponse(
                ok=False, session_id=session_id,
                message="无法定位表格文件",
                intent=intent.action,
            )

        # 3. 复制到临时目录执行
        tmp_dir = tempfile.mkdtemp()
        tmp_path = Path(tmp_dir) / path.name
        shutil.copy2(path, tmp_path)

        try:
            tmp_cli = RealCodeMakerCLI(workspace=Path(tmp_dir))
            tmp_agent = TableAgent(
                cli=tmp_cli,
                parser=self.agent.parser,
                resolver=self.agent.resolver,
                column_cfg=self.agent.column_cfg,
                row_cfg=self.agent.row_cfg,
                ctx_cfg=self.agent.ctx_cfg,
                live_index=False,
                enable_skill=self.enable_skill,
            )
            tmp_agent.enable_evidence = False  # dry-run 临时副本不写真实证据
            # 复用主 agent 的流式 sink（thinking/step/tool/subtask/ask/cancel），
            # 使 dry_run 预览路径的 4-Step ParseAgent/ValidateAgent/ExecuteAgent
            # 阶段 thinking 实时进 SSE queue，前端 stage s2/s3/s4 正常显现
            # O20h：_llm_counter 也共享主 agent 实例，使 heartbeat loop 读主 counter
            # = tmp_agent counter 实时可见（原 tmp_agent 新建独立 counter，bench llm_calls=0）。
            for _attr in ("_agent_thinking_sink", "_agent_step_sink",
                          "_agent_tool_sink", "_agent_subtask_sink",
                          "_cancel_event", "_llm_counter"):
                _v = getattr(self.agent, _attr, None)
                if _v is not None:
                    setattr(tmp_agent, _attr, _v)
            # dry_run 非交互：注入"接受建议"ask_callback（非 skip）。
            # §链路完整性：dry_run 预览应完整走通 4-Step，硬 issue 默认接受建议
            # 放行不 skip（用户要求默认接受建议，不跳过子任务）。PK 冲突 accept_suggest，
            # 通用硬 issue 返 mode=field + value=suggestion → validator 改写 fields。
            _accept_suggestion = os.getenv(
                "CODEMAKER_DRY_RUN_ACCEPT_SUGGEST", "1") == "1"
            def _dry_ask_cb(question: dict):
                if _accept_suggestion:
                    _sug = question.get("suggested_id")
                    if _sug is None:
                        _sug = question.get("suggestion")
                    return {
                        "mode": "field",
                        "accept_suggest": True,
                        "value": _sug,
                        "custom_id": _sug,
                        "text": _sug,
                    }
                return {"mode": "skip"}
            tmp_agent._ask_callback = _dry_ask_cb
            tmp_agent._dry_run_flag = True  # Step2 读此走复位放行分支
            _va = getattr(self.agent, "_validator_agent", None)
            if _va is not None:
                tmp_agent._validator_agent = _va
                # 临时覆盖共享 validator 的 ask_callback 为 accept（run 后恢复）
                _orig_va_cb = getattr(_va, "_ask_callback", None)
                if hasattr(_va, "set_ask_callback"):
                    _va.set_ask_callback(_dry_ask_cb)
            # dry_run V2 复用主 agent 的 locator/decompose agent（用主 cli 拉真实 resources
            # schema），避免 tmp_cli(workspace=临时空目录) list_tables 只1表致 decompose 产空
            _la = getattr(self.agent, "_locator_agent", None)
            _da = getattr(self.agent, "_decompose_agent", None)
            if _la is not None:
                tmp_agent._locator_agent = _la
            if _da is not None:
                tmp_agent._decompose_agent = _da
            try:
                result = tmp_agent.run(text)
            finally:
                # 恢复共享 validator 原始 ask_callback
                if _va is not None and hasattr(_va, "set_ask_callback") and _orig_va_cb is not None:
                    _va.set_ask_callback(_orig_va_cb)

            # 读取变更
            changes = self._diff_files(path, tmp_path, sheet, result)
            # 同值写入（值未变）时文件 diff 为空，用 result.final 补一条 change，
            # 使预览能反映"已执行写入"的意图（如改为与原值相同的值）。
            # 仅对写操作生效，避免 get 的 final.data 被误判为变更。
            # 用 result.intent.action（真实执行意图），避免 table_hint 预览路径
            # 传 stub intent(action=get) 误判。
            real_action = result.intent.action if result.intent else intent.action
            if (not changes and result.ok
                    and real_action in ("set", "add", "insert")
                    and result.final and isinstance(result.final.data, dict)):
                d = result.final.data
                if "row" in d and "col" in d and "value" in d:
                    changes.append(CellChange(
                        col=d.get("col", 0),
                        col_name=d.get("column") or f"col_{d.get('col', 0)}",
                        new_value=d.get("value"),
                    ))

            steps = [AgentStepInfo(name=s.name, ok=s.ok, detail=s.detail)
                     for s in result.steps]

            preview = DiffPreview(
                file=str(path.relative_to(self.resources_dir)),
                sheet=sheet or "",
                changes=changes,
            )

            op_id = uuid.uuid4().hex[:8]
            self._previews[op_id] = {
                'text': text, 'path': str(path), 'tmp_path': str(tmp_path),
                'sheet': sheet, 'changes': changes,
            }

            return AgentChatResponse(
                ok=bool(result.ok) if result.ok is not None else False,
                session_id=session_id,
                intent=result.intent.action if result.intent else intent.action,
                message=f"[预览] {result.message}",
                steps=steps,
                data=result.final.data if (result.final and isinstance(result.final.data, dict)) else None,
                diff_preview=preview,
                result_table=self._build_result_table(result),
            )
        except Exception as e:
            err_type = getattr(e, "error_type", "") or self._infer_error_type(str(e))
            _, advice = self._map_codemaker_error(err_type, str(e))
            return AgentChatResponse(
                ok=False, session_id=session_id,
                message=advice, error=str(e),
            )
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def _diff_files(self, original: Path, modified: Path,
                    sheet: str, result: AgentResult) -> List[CellChange]:
        """比较两个文件的变更。"""
        changes = []
        try:
            orig_data = self.cli.read_sheet(original, sheet) if sheet else []
            tmp_cli = RealCodeMakerCLI(workspace=modified.parent)
            mod_data = tmp_cli.read_sheet(modified, sheet) if sheet else []
        except Exception:
            return changes

        # 简单逐行比较
        max_rows = max(len(orig_data), len(mod_data))
        for r in range(max_rows):
            orig_row = orig_data[r] if r < len(orig_data) else []
            mod_row = mod_data[r] if r < len(mod_data) else []
            max_cols = max(len(orig_row), len(mod_row))
            for c in range(max_cols):
                ov = orig_row[c] if c < len(orig_row) else None
                mv = mod_row[c] if c < len(mod_row) else None
                if str(ov) != str(mv):
                    changes.append(CellChange(
                        col=c + 1, col_name=f"col_{c + 1}",
                        old_value=ov, new_value=mv,
                    ))
        return changes

    def _build_diff_from_result(self, result: AgentResult) -> Optional[DiffPreview]:
        """从 AgentResult 构建 DiffPreview（修改场景的旧值→新值预览）。

        优先用 result.result_rows（携带 old_value/col_name，准确）；
        result_rows 为空时回退到 final.data 的单格信息（补一条 new_value）。
        file/sheet/col_name/old_value 全部填充，修复旧版旧值恒为 None 的问题。
        """
        if not result.final or not result.final.ok:
            return None

        file_disp = result.table_stem or ""
        sheet_disp = result.table_sheet or ""
        row_no = None
        data = result.final.data
        if isinstance(data, dict):
            row_no = data.get('row')

        # 修改/清空场景：result_rows 已携带 old_value/new_value/col_name
        if result.result_rows:
            changes = [CellChange(
                col=r.get("col", 0),
                col_name=r.get("col_name") or "",
                old_value=r.get("old_value"),
                new_value=r.get("new_value"),
            ) for r in result.result_rows]
            return DiffPreview(
                file=file_disp, sheet=sheet_disp, row=row_no, changes=changes,
            )

        # 回退：单格写入（无 result_rows，如旧路径）
        if isinstance(data, dict) and 'row' in data and 'col' in data and 'value' in data:
            return DiffPreview(
                file=file_disp, sheet=sheet_disp,
                row=data.get('row'),
                changes=[CellChange(
                    col=data.get('col', 0),
                    col_name=data.get('column') or f"col_{data.get('col', 0)}",
                    new_value=data.get('value'),
                )],
            )
        return None

    def _build_result_table(self, result: AgentResult) -> Optional[ResultTable]:
        """从 AgentResult 构建表体结构（列名+行值），供前端直观渲染。

        各操作返回形态：
          set    → 仅变更列，旧值→新值
          add    → 生成行的列名+值
          delete → 删除行的列名+被删值（清空单格时退化为旧值→空）
          get    → 查询到的列名+值
        无 result_rows 或执行失败时返回 None。
        """
        if not result or not result.result_rows:
            return None

        action = result.intent.action if result.intent else ""
        # 清空单格：delete 意图但 new_value=="" → 用 set 形态展示旧值→空
        kind = action
        if action == "delete":
            rows = result.result_rows
            if len(rows) == 1 and rows[0].get("new_value") == "":
                kind = "set"

        row_no = None
        if result.final and isinstance(result.final.data, dict):
            row_no = result.final.data.get('row')

        file_disp = result.table_stem or ""
        sheet_disp = result.table_sheet or ""
        columns = [ResultColumn(
            col=r.get("col", 0),
            col_name=r.get("col_name") or "",
            old_value=r.get("old_value"),
            new_value=r.get("new_value"),
        ) for r in result.result_rows]

        return ResultTable(
            kind=kind or "set",
            file=file_disp,
            sheet=sheet_disp,
            row=row_no,
            columns=columns,
        )

    def _build_result_table_from_subtask(self, sub: dict) -> Optional[ResultTable]:
        """从子任务 dict 构建 ResultTable（多指令复合操作时每个子任务独立一份）。

        与 _build_result_table 逻辑一致，只是数据源从 AgentResult 换成 sub_tasks 项。
        子任务无 final.data，row 取 result_rows 里的 col 信息，行号留空（由 message 文本体现）。
        """
        rows = sub.get("result_rows") or []
        if not rows:
            return None
        action = sub.get("intent_action", "")
        kind = action
        if action == "delete" and len(rows) == 1 and rows[0].get("new_value") == "":
            kind = "set"
        file_disp = sub.get("table_stem", "")
        sheet_disp = sub.get("table_sheet", "")
        columns = [ResultColumn(
            col=r.get("col", 0),
            col_name=r.get("col_name") or "",
            old_value=r.get("old_value"),
            new_value=r.get("new_value"),
        ) for r in rows]
        return ResultTable(
            kind=kind or "set",
            file=file_disp,
            sheet=sheet_disp,
            row=None,
            columns=columns,
        )

    def _build_thinking_steps(self, result: AgentResult) -> List[Dict[str, str]]:
        """R9: 从 AgentResult 构建友好的思考过程步骤，供前端折叠展示。

        优先用 agent 核心直接填的 result.thinking_steps（细粒度，配表模式增强）；
        若为空则回退从 steps 映射（兼容旧逻辑）。
        """
        agent_thinking = getattr(result, "thinking_steps", None) or []
        if agent_thinking:
            return [{"phase": t.get("phase", ""), "detail": t.get("detail", "")}
                    for t in agent_thinking]
        thinking = []
        for s in result.steps:
            phase, detail = self._map_step_to_thinking(s)
            if phase:
                thinking.append({"phase": phase, "detail": detail})
        return thinking

    @staticmethod
    def _map_step_to_thinking(step: AgentStep) -> tuple:
        """将技术步骤名映射为友好的思考阶段名。"""
        name = step.name
        detail = step.detail
        if name == "parse_intent":
            return ("解析", f"理解意图: {detail}")
        elif name == "resolve_table":
            return ("路由", f"定位表格: {detail}")
        elif name == "resolve_sheet":
            return ("路由", f"定位Sheet: {detail}")
        elif name == "match_locator":
            return ("定位", f"匹配定位列: {detail}")
        elif name == "match_target":
            return ("定位", f"匹配目标列: {detail}")
        elif name == "locate_row":
            return ("定位", f"行定位结果: {detail}")
        elif name == "read_cell":
            return ("执行", f"读取单元格: {detail}")
        elif name == "read_row":
            return ("执行", f"读取整行: {detail}")
        elif name == "read_rows":
            return ("执行", f"读取多行: {detail}")
        elif name == "write":
            return ("执行", f"写入: {detail}")
        elif name == "write_cell":
            return ("执行", f"写入单元格: {detail}")
        elif name == "coerce_value":
            return ("校验", f"类型转换: {detail}")
        elif name == "validate_id":
            return ("校验", f"ID校验: {detail}")
        elif name == "add_values":
            return ("执行", f"新增行: {detail}")
        elif name == "id_scope":
            return ("校验", f"ID段校验: {detail}")
        return ("", "")

    def _build_multi_results(self, result: AgentResult) -> List[ResultTable]:
        """R9: 从 AgentResult.multi_rows 构建多行结果表列表，供前端渲染多行表格。"""
        multi_rows = getattr(result, "multi_rows", None) or []
        if not multi_rows:
            return []
        tables = []
        action = result.intent.action if result.intent else "get"
        file_disp = result.table_stem or ""
        sheet_disp = result.table_sheet or ""
        for mr in multi_rows:
            cols = [ResultColumn(
                col=c.get("col", 0),
                col_name=c.get("col_name") or "",
                new_value=c.get("value"),
            ) for c in (mr.get("columns") or [])]
            tables.append(ResultTable(
                kind=action,
                file=file_disp,
                sheet=sheet_disp,
                row=mr.get("row"),
                columns=cols,
            ))
        return tables

    def get_history(self, session_id: str = "default") -> List[dict]:
        """获取会话操作历史。"""
        return self._sessions.get(session_id, [])

    def _make_checkpoint(self, session_id: str, text: str) -> Optional[str]:
        """写动作成功后拍一个 checkpoint，返回 checkpoint_id。

        checkpoint = resources/ 全量快照（复用 workflow.create_snapshot_sync），
        记录 {checkpoint_id, snap_id, timestamp, text} 到 _session_checkpoints。
        一个 checkpoint 对应一次自然语言输入（可能含多个原子写操作），非原子粒度。
        失败静默返回 None，不阻断主流程。
        """
        try:
            snap_id = create_snapshot_sync(name=f"ckpt_{session_id}")
        except Exception:
            return None
        ckpt = {
            "checkpoint_id": uuid.uuid4().hex[:8],
            "snap_id": snap_id,
            "timestamp": datetime.now().isoformat(),
            "text": text,
        }
        self._session_checkpoints.setdefault(session_id, []).append(ckpt)
        return ckpt["checkpoint_id"]

    def _save_nl_checkpoint(self, session_id: str, stage: str,
                             intents: list) -> bool:
        """P27：拍 4-step NL 路径中间态 checkpoint（parse/validate 后）。

        序列化 NLIntent[]（to_checkpoint_dict）到 _nl_checkpoints[session_id][stage]，
        stall 可从中间态续跑免 Step1 重 LLM decompose。opt-in
        CODEMAKER_4STEP_CHECKPOINT=1 时写，默认 off。失败静默返 False 不阻断。
        接线（4-step 路径调 save + stall 检测 + resume）留 follow-up。
        """
        if os.getenv("CODEMAKER_4STEP_CHECKPOINT", "0") != "1":
            return False
        try:
            from agent.excel.parser.nl_parser import NLIntent
            serialized = [it.to_checkpoint_dict()
                          if isinstance(it, NLIntent) else it
                          for it in (intents or [])]
            self._nl_checkpoints.setdefault(session_id, {})[stage] = {
                "intents": serialized,
                "timestamp": datetime.now().isoformat(),
                "stage": stage,
            }
            return True
        except Exception:
            return False

    def _load_nl_checkpoint(self, session_id: str, stage: str):
        """P27：加载 NL checkpoint（_save_nl_checkpoint 的逆）。

        返回反序列化的 NLIntent[] 或 None（无 checkpoint）。供 stall 续跑：
        跳过已成功 Step5 op + 从未完成的中间态继续。
        """
        ckpt = self._nl_checkpoints.get(session_id, {}).get(stage)
        if not ckpt:
            return None
        try:
            from agent.excel.parser.nl_parser import NLIntent
            return [NLIntent.from_checkpoint_dict(d)
                    for d in ckpt.get("intents") or []]
        except Exception:
            return None

    def list_checkpoints(self, session_id: str = "default") -> List[dict]:
        """列出该会话的所有 checkpoint（按时间顺序）。

        每项 {checkpoint_id, timestamp, text, label}，label 供前端展示。
        """
        ckpts = self._session_checkpoints.get(session_id, [])
        out = []
        for i, c in enumerate(ckpts):
            preview = c["text"] if c["text"] else ""
            if len(preview) > 30:
                preview = preview[:30] + "…"
            out.append({
                "checkpoint_id": c["checkpoint_id"],
                "timestamp": c["timestamp"],
                "text": c["text"],
                "label": f"#{i + 1} {preview}",
            })
        return out

    def rollback_to_checkpoint(self, session_id: str = "default",
                               checkpoint_id: Optional[str] = None) -> dict:
        """把表格回退到指定 checkpoint（= 某次写操作完成后的状态）。

        Args:
            session_id: 会话 id。
            checkpoint_id: 目标 checkpoint id；None 时回退到最近一个 checkpoint
                           （即"最后一次写操作完成后的状态"，等于"撤销最近一次输入的影响"）。

        回退成功后，丢弃比该 checkpoint 更晚的所有 checkpoint（时间线截断），
        保留该 checkpoint 及之前的。复用 workflow.restore_snapshot_sync：
        恢复前自动备份回退本身可逆；并清 CLI 缓存 + 刷新索引。

        Returns:
            {ok, message, session_id, checkpoint_id, rolled_back_count}
        """
        ckpts = self._session_checkpoints.setdefault(session_id, [])
        if not ckpts:
            return {
                "ok": False, "session_id": session_id,
                "message": f"会话 '{session_id}' 无可用 checkpoint，无法回退",
            }

        # 定位目标 checkpoint
        if checkpoint_id is None:
            idx = len(ckpts) - 1
        else:
            idx = next((i for i, c in enumerate(ckpts)
                        if c["checkpoint_id"] == checkpoint_id), -1)
            if idx < 0:
                return {
                    "ok": False, "session_id": session_id,
                    "message": f"checkpoint '{checkpoint_id}' 不存在",
                }

        target = ckpts[idx]
        # 丢弃比目标更晚的 checkpoint（时间线截断）
        dropped = ckpts[idx + 1:]
        ckpts[:] = ckpts[:idx + 1]

        try:
            r = restore_snapshot_sync(target["snap_id"])
        except FileNotFoundError:
            # 快照文件已不在磁盘：清理失效 checkpoint，提示用户
            ckpts[:] = ckpts[:idx] if idx > 0 else []
            return {
                "ok": False, "session_id": session_id,
                "checkpoint_id": target["checkpoint_id"],
                "message": f"checkpoint '{target['checkpoint_id']}' 对应的快照已不存在，无法回退",
            }

        # 同步裁剪会话历史：丢弃比目标 checkpoint 晚的操作记录，
        # 保持 history 与 checkpoint 时间线一致。
        # checkpoint idx 对应第 (idx+1) 次写操作，历史中写操作记录也应截断到 idx+1 条。
        hist = self._sessions.get(session_id, [])
        if len(hist) > idx + 1:
            hist[:] = hist[:idx + 1]

        return {
            "ok": True, "session_id": session_id,
            "checkpoint_id": target["checkpoint_id"],
            "message": f"已回退到 checkpoint #{idx + 1}（{r.get('message', '')}）",
            "rolled_back_count": len(dropped),
        }

    def _on_file_changed(self, result) -> None:
        """TableFileWatcher 回调：索引已由 watcher 刷新，此处只需清 openpyxl 缓存。

        避免 _load 命中旧 Workbook 对象导致读取到改表前的内容。
        T9 D9.2: 顺带跑一次 skill 衰减扫描（轻量，扫 runtime yaml 的 last_seen），
        并对已删除的表标记 cascade 规则 stale（D9.4）。
        """
        try:
            self.cli._cache.clear()
        except Exception:
            pass
        # D9.4: 失效标记 — 删除的表对应的 cascade 规则标 stale
        try:
            from agent.excel.skill_updater import get_skill_updater
            su = get_skill_updater()
            removed = getattr(result, "removed", []) or []
            if removed:
                su.detect_stale_rules(removed)
        except Exception:
            pass
        # D9.2: 衰减扫描（轻量）
        try:
            from agent.excel.skill_updater import get_skill_updater
            get_skill_updater().decay_scan()
        except Exception:
            pass

    def invalidate_cache(self) -> None:
        """清除 CLI 工作簿缓存并增量刷新表格索引。

        快照恢复/外部文件变更后调用，保证后续读取与磁盘实时一致
        （openpyxl 工作簿缓存 + 行级倒排索引都会随之更新）。
        """
        try:
            self.cli._cache.clear()
        except Exception:
            pass
        try:
            from agent.table_index import refresh_if_changed
            result = refresh_if_changed(self.resources_dir)
            # D9.4: 手动 refresh 也检测 removed → 标 stale cascade 规则
            removed = getattr(result, "removed", []) or []
            if removed:
                try:
                    from agent.excel.skill_updater import get_skill_updater
                    get_skill_updater().detect_stale_rules(removed)
                except Exception:
                    pass
        except Exception:
            pass

    # ── 批量操作 ──

    def batch(self, messages: List[str], session_id: str = "default",
              stop_on_error: bool = True) -> BatchResponse:
        """批量执行多条指令。"""
        results = []
        success = 0
        fail = 0

        for i, msg in enumerate(messages):
            resp = self.chat(msg, session_id=session_id)
            item = BatchItemResult(
                index=i, message=msg, ok=resp.ok,
                result_message=resp.message,
                error=resp.error,
            )
            results.append(item)
            if resp.ok:
                success += 1
            else:
                fail += 1
                if stop_on_error:
                    break

        # ok 语义：全部成功为 True；stop_on_error=False 且至少一条成功时也视为 True
        # （批量容错场景，部分失败可接受，由 success_count/fail_count 体现明细）。
        ok = (fail == 0) or (not stop_on_error and success > 0)
        return BatchResponse(
            ok=ok,
            results=results,
            success_count=success,
            fail_count=fail,
        )

    # ── 数据验证 ──

    def validate(self, tables: List[str] = None,
                 check_types: List[str] = None) -> ValidateResponse:
        """执行数据一致性校验。"""
        if check_types is None:
            check_types = ["referential", "uniqueness", "range"]

        issues = []
        index = load_index()
        all_tables = self.cli.list_tables()

        # 过滤要检查的表
        target_tables = all_tables
        if tables:
            target_tables = [t for t in all_tables if t.stem in tables]

        for path in target_tables:
            # 找到对应索引
            tmeta = None
            for t in index:
                if t.stem == path.stem:
                    tmeta = t
                    break
            if tmeta is None:
                continue

            for s in tmeta.sheets:
                try:
                    sheet_issues = self._validate_sheet(
                        path, s.name, tmeta, check_types)
                    issues.extend(sheet_issues)
                except Exception:
                    continue

        # 构建摘要
        errors = [i for i in issues if i.severity == "error"]
        warnings = [i for i in issues if i.severity == "warning"]
        ok = len(errors) == 0
        summary = f"检查完成：{len(errors)} 个错误，{len(warnings)} 个警告"

        return ValidateResponse(ok=ok, issues=issues, summary=summary)

    def _validate_sheet(self, path: Path, sheet: str, tmeta: TableMeta,
                        check_types: List[str]) -> List[ValidationIssue]:
        """检查单个 sheet 的数据一致性。"""
        issues = []

        try:
            headers = self.cli.read_header(path, sheet)
            rows = self.cli.read_sheet(path, sheet)
        except Exception:
            return issues

        if not rows:
            return issues

        # 1. 唯一性检查
        if "uniqueness" in check_types:
            # 找 ID/名称类列
            id_cols = []
            for i, h in enumerate(headers):
                hl = str(h).lower()
                if any(kw in hl for kw in ("id", "编号", "名称")):
                    # 排除描述类列
                    if not any(bad in hl for bad in ("描述", "图标", "icon")):
                        id_cols.append(i)

            for ci in id_cols:
                seen = {}
                for ri, row in enumerate(rows):
                    val = str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ""
                    if val and val in seen:
                        issues.append(ValidationIssue(
                            severity="warning",
                            table=path.stem, sheet=sheet,
                            row=ri + 1, col=ci + 1,
                            col_name=headers[ci] if ci < len(headers) else "",
                            message=f"重复值 '{val}'（首次出现在行 {seen[val] + 1}）",
                        ))
                    elif val:
                        seen[val] = ri

        # 2. 范围检查
        if "range" in check_types:
            for ci, h in enumerate(headers):
                hl = str(h).lower()
                # 概率类列
                if any(kw in hl for kw in ("概率", "rate", "ratio", "chance")):
                    for ri, row in enumerate(rows):
                        val = row[ci] if ci < len(row) else None
                        if val is not None:
                            try:
                                fv = float(val)
                                if fv < 0 or fv > 100:
                                    issues.append(ValidationIssue(
                                        severity="warning",
                                        table=path.stem, sheet=sheet,
                                        row=ri + 1, col=ci + 1,
                                        col_name=h,
                                        message=f"概率值 {fv} 超出 [0, 100] 范围",
                                    ))
                            except (ValueError, TypeError):
                                pass

        # 3. 引用完整性检查
        if "referential" in check_types:
            issues.extend(self._check_referential(path, sheet, headers, rows))

        return issues

    def _check_referential(self, path: Path, sheet: str,
                           headers: List[str], rows: List[list]) -> List[ValidationIssue]:
        """检查引用完整性：ID 列是否在目标表中存在。"""
        issues = []
        # 常见的外键映射：{列名关键词: 目标表 stem}
        FK_MAP = {
            "灵兽id": "pet", "宠物id": "pet",
            "物品编号": "item", "物品id": "item", "道具id": "item",
            "神通id": "school_ability", "技能id": "spell",
            "仙友id": "assistant", "人物id": "hero", "英雄id": "hero",
            "配方id": "item_recipe", "门派id": "school",
            "活动id": "activity", "任务id": "quest",
            "buff": "spell", "法宝id": "fabao",
        }

        for ci, h in enumerate(headers):
            hl = str(h).lower()
            for fk_kw, target_stem in FK_MAP.items():
                if fk_kw in hl:
                    # 找到目标表并检查值是否存在
                    target_path = None
                    for p in self.resources_dir.rglob(f"{target_stem}.xlsx"):
                        target_path = p
                        break
                    if target_path is None:
                        continue

                    # 读取目标表的所有 ID 值
                    try:
                        target_rows = self.cli.read_sheet(target_path, sheet)
                        target_headers = self.cli.read_header(target_path, sheet)
                    except Exception:
                        continue

                    # 找目标表的 ID/名称列
                    target_col = 0  # 默认第一列
                    for ti, th in enumerate(target_headers):
                        thl = str(th).lower()
                        if "id" in thl or "编号" in thl:
                            target_col = ti
                            break

                    valid_ids = set()
                    for tr in target_rows:
                        if target_col < len(tr) and tr[target_col] is not None:
                            valid_ids.add(str(tr[target_col]).strip())

                    # 检查源表的引用值
                    for ri, row in enumerate(rows):
                        val = str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ""
                        if val and val not in valid_ids and val != "0" and val != "":
                            issues.append(ValidationIssue(
                                severity="warning",
                                table=path.stem, sheet=sheet,
                                row=ri + 1, col=ci + 1,
                                col_name=h,
                                message=f"引用的 {fk_kw}={val} 在 {target_stem} 表中不存在",
                            ))

        return issues

    # ── AI 辅助冲突解决 ──

    def _get_llm(self):
        """懒加载合并建议 LLM（CodemakerChatModel），复用 router 的 codemaker client。

        serve 不可用或构造失败返回 None，调用方回退规则建议。
        """
        if self._llm is not None:
            return self._llm
        try:
            from agent.llm import CodemakerChatModel
            self._llm = CodemakerChatModel(client=self.router.client)
        except Exception:
            self._llm = None
        return self._llm

    def _suggest_merge_via_llm(self, table_stem: str, sheet: str, col_name: str,
                               row_key: str, base_value: Any,
                               versions: dict, version_meta: dict,
                               base_file: str = "") -> Optional[dict]:
        """调 LLM 生成建议。失败/不可用返回 None，调用方回退规则。

        prompt 含各版本 SVN rev/date（时间先后）+ 单元格值 + 列名/表名，
        要求 LLM 只输出 JSON {suggested_version, reasoning, confidence}。
        base_file 仅在 prompt 标注为参考，校验时排除（不建议采纳基准）。
        """
        llm = self._get_llm()
        if llm is None:
            return None
        try:
            from langchain_core.messages import SystemMessage, HumanMessage
        except Exception:
            return None

        lines = []
        for fn, val in versions.items():
            meta = version_meta.get(fn, {}) if version_meta else {}
            rev = meta.get("rev", "")
            date = meta.get("date", "")
            tag = []
            if rev: tag.append(f"rev={rev}")
            if date: tag.append(f"date={date}")
            tagstr = "|".join(tag)
            base_mark = " [基准·仅参考]" if fn == base_file else ""
            lines.append(f"{fn}{'|'+tagstr if tagstr else ''}{base_mark}: {val!r}")
        versions_block = "\n".join(lines) if lines else "（无）"

        user_prompt = (
            f"表={table_stem} sheet={sheet} 列={col_name} 行键={row_key}\n"
            f"基准值={base_value!r}\n各版本（含SVN修订）：\n{versions_block}\n\n"
            f"建议采纳哪个版本。suggested_version 必须是上面文件名之一。"
        )
        try:
            resp = llm.invoke([
                SystemMessage(content=MERGE_SUGGEST_SYSTEM_PROMPT),
                HumanMessage(content=user_prompt),
            ])
            text = resp.content if hasattr(resp, "content") else str(resp)
            if not isinstance(text, str):
                text = str(text)
            m = _re.search(r"\{.*\}", text, _re.S)
            if not m:
                return None
            obj = json.loads(m.group(0))
            sv = str(obj.get("suggested_version", "")).strip()
            # 排除 base（基准仅参考）+ 必须是输入版本之一
            if not sv or sv not in versions or sv == base_file:
                return None
            try:
                conf = float(obj.get("confidence", 0.5))
            except (TypeError, ValueError):
                conf = 0.5
            # 融合规则因子：rev 领先度 + 值域合理性
            other_revs = [
                (version_meta or {}).get(fn, {}).get("rev")
                for fn in versions if fn != sv
            ]
            rev_conf = _rev_confidence(
                (version_meta or {}).get(sv, {}).get("rev"), other_revs)
            value_conf = _value_confidence(versions.get(sv), col_name)
            fused_conf = _fuse_confidence(conf, rev_conf, value_conf)
            return {
                "suggested_version": sv,
                "suggestion": str(versions.get(sv, "")),
                "reasoning": str(obj.get("reasoning", "")),
                "confidence": fused_conf,
            }
        except Exception:
            return None

    def _suggest_merge_batch_via_llm(self, table_stem: str, sheet: str,
                                     base_file: str, version_meta: dict,
                                     items: list) -> Optional[list]:
        """一次 LLM 调用为所有冲突格生成建议，返回与 items 顺序对应的结果列表。

        将全部冲突格拼入单个 prompt，LLM 一次返回 JSON 数组，大幅减少网络往返。
        失败返回 None，调用方回退逐格并行。
        """
        llm = self._get_llm()
        if llm is None:
            return None
        try:
            from langchain_core.messages import SystemMessage, HumanMessage
        except Exception:
            return None

        parts = []
        for idx, it in enumerate(items):
            col_name = it.get("col_name", "")
            row_key = it.get("row_key", "")
            base_value = it.get("base_value")
            versions = it.get("versions", {})
            lines = []
            for fn, val in versions.items():
                meta = version_meta.get(fn, {}) if version_meta else {}
                rev = meta.get("rev", "")
                date = meta.get("date", "")
                tag = []
                if rev: tag.append(f"rev={rev}")
                if date: tag.append(f"date={date}")
                tagstr = "|".join(tag)
                base_mark = " [基准·仅参考]" if fn == base_file else ""
                lines.append(f"  {fn}{'|'+tagstr if tagstr else ''}{base_mark}: {val!r}")
            versions_block = "\n".join(lines) if lines else "  （无）"
            parts.append(
                f"[{idx}] 列={col_name} 行键={row_key}\n"
                f"  基准值={base_value!r}\n各版本：\n{versions_block}"
            )
        all_parts = "\n\n".join(parts)

        user_prompt = (
            f"表={table_stem} sheet={sheet}\n"
            f"为以下 {len(items)} 个冲突单元格分别建议采纳哪个版本：\n\n"
            f"{all_parts}\n\n"
            f"输出 JSON 数组，每个元素按输入序号 [0]~[{len(items)-1}] 顺序返回。"
            f"suggested_version 必须是非基准衍生文件名之一。"
        )
        try:
            resp = llm.invoke([
                SystemMessage(content=MERGE_SUGGEST_BATCH_SYSTEM_PROMPT),
                HumanMessage(content=user_prompt),
            ])
            text = resp.content if hasattr(resp, "content") else str(resp)
            if not isinstance(text, str):
                text = str(text)
            m = _re.search(r"\[.*\]", text, _re.S)
            if not m:
                return None
            arr = json.loads(m.group(0))
            if not isinstance(arr, list):
                return None

            results = [None] * len(items)
            for obj in arr:
                idx = int(obj.get("index", -1))
                if idx < 0 or idx >= len(items):
                    continue
                sv = str(obj.get("suggested_version", "")).strip()
                versions = items[idx].get("versions", {})
                if not sv or sv not in versions or sv == base_file:
                    continue
                try:
                    conf = float(obj.get("confidence", 0.5))
                except (TypeError, ValueError):
                    conf = 0.5
                other_revs = [
                    (version_meta or {}).get(fn, {}).get("rev")
                    for fn in versions if fn != sv
                ]
                rev_conf = _rev_confidence(
                    (version_meta or {}).get(sv, {}).get("rev"), other_revs)
                value_conf = _value_confidence(
                    versions.get(sv), items[idx].get("col_name", ""))
                results[idx] = {
                    "suggested_version": sv,
                    "suggestion": str(versions.get(sv, "")),
                    "reasoning": str(obj.get("reasoning", "")),
                    "confidence": _fuse_confidence(conf, rev_conf, value_conf),
                }
            return results
        except Exception:
            return None

    def _suggest_cache_get(self, cache_key: tuple):
        """带 TTL 的缓存读：命中且未过期则 move_to_end（LRU）并返回值，否则返回 None。"""
        ent = self._suggest_cache.get(cache_key)
        if ent is None:
            return None
        val, ts = ent
        if time.time() - ts > self._suggest_cache_ttl:
            try:
                del self._suggest_cache[cache_key]
            except KeyError:
                pass
            return None
        self._suggest_cache.move_to_end(cache_key)
        return val

    def _suggest_cache_put(self, cache_key: tuple, val) -> None:
        """写缓存：超限时按 LRU 淘汰最旧条目，保证有界。"""
        self._suggest_cache[cache_key] = (val, time.time())
        self._suggest_cache.move_to_end(cache_key)
        while len(self._suggest_cache) > self._suggest_cache_max:
            self._suggest_cache.popitem(last=False)

    def suggest_merge_batch(self, table_stem: str, sheet: str,
                            version_meta: dict, items: list,
                            base_file: str = "") -> dict:
        """批量 AI 建议：优先一次 LLM 调用覆盖全部冲突格（秒级返回），失败回退逐格并行。

        一次 LLM 调用的 prompt 含所有冲突格信息，LLM 一次返回 JSON 数组，
        将 N 次网络往返（每格 15-45s）压缩为 1 次（约 30-60s），总时间大幅缩减。
        失败则回退原有逐格并发策略。
        """
        results: Dict[str, dict] = {}

        # 1. 优先走缓存
        uncached_items = []
        for it in items:
            ri = it.get("ri", 0)
            ci = it.get("ci", 0)
            key = f"{ri}-{ci}"
            versions = it.get("versions", {})
            cache_key = (table_stem, sheet, it.get("col_name", ""), _json_hash(versions))
            cached = self._suggest_cache_get(cache_key)
            if cached is not None:
                results[key] = cached
            else:
                uncached_items.append(it)

        if not uncached_items:
            return {"results": results}

        # 2. 尝试批量 LLM（一次调用覆盖全部未缓存格）
        if version_meta and len(uncached_items) >= 2:
            batch_results = self._suggest_merge_batch_via_llm(
                table_stem, sheet, base_file, version_meta, uncached_items,
            )
            if batch_results is not None:
                for i, it in enumerate(uncached_items):
                    ri = it.get("ri", 0)
                    ci = it.get("ci", 0)
                    key = f"{ri}-{ci}"
                    r = batch_results[i]
                    if r is None:
                        # 单格回退规则
                        r = self.suggest_merge(
                            table_stem=table_stem, sheet=sheet,
                            col_name=it.get("col_name", ""),
                            row_key=it.get("row_key", ""),
                            base_value=it.get("base_value"),
                            versions=it.get("versions", {}),
                            version_meta=version_meta,
                            base_file=base_file,
                        )
                    results[key] = r
                    # 写缓存
                    versions = it.get("versions", {})
                    cache_key = (table_stem, sheet, it.get("col_name", ""), _json_hash(versions))
                    self._suggest_cache_put(cache_key, r)
                return {"results": results}

        # 3. 回退：逐格并行（原有逻辑）
        from concurrent.futures import ThreadPoolExecutor

        def _one(it: dict) -> tuple:
            ri = it.get("ri", 0)
            ci = it.get("ci", 0)
            key = f"{ri}-{ci}"
            try:
                res = self.suggest_merge(
                    table_stem=table_stem, sheet=sheet,
                    col_name=it.get("col_name", ""),
                    row_key=it.get("row_key", ""),
                    base_value=it.get("base_value"),
                    versions=it.get("versions", {}),
                    version_meta=version_meta,
                    base_file=base_file,
                )
                # 写缓存
                versions = it.get("versions", {})
                cache_key = (table_stem, sheet, it.get("col_name", ""), _json_hash(versions))
                self._suggest_cache_put(cache_key, res)
                return key, res
            except Exception:
                return key, {
                    "suggested_version": "",
                    "suggestion": "",
                    "reasoning": "建议生成失败",
                    "confidence": 0.0,
                }

        max_workers = min(8, max(1, len(uncached_items)))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            for key, res in ex.map(_one, uncached_items):
                results[key] = res
        return {"results": results}

    def suggest_merge(self, table_stem: str, sheet: str, col_name: str,
                      row_key: str, base_value: Any,
                      versions: dict, version_meta: Optional[dict] = None,
                      base_file: str = "") -> dict:
        """AI 辅助冲突解决建议。

        优先调 LLM（基于 SVN 修订时间先后 + 内容），失败或无 version_meta 时回退规则。
        base_file 仅作参考，LLM/规则均不建议采纳基准。
        规则策略（按优先级）：
        1. ID/编号列 → 选较新版本（非 base）
        2. 数值列 → 选合理范围内的值
        3. 描述/文本列 → 选较长的（通常更完整）
        4. 兜底 → 选非 base 版本中最新修改的
        """
        if not versions:
            return {
                "suggested_version": "",
                "suggestion": "",
                "reasoning": "无可用版本",
                "confidence": 0.0,
            }

        # 优先调 LLM（需 version_meta 提供 SVN 修订时间）；失败回退下方规则
        if version_meta:
            llm_res = self._suggest_merge_via_llm(
                table_stem, sheet, col_name, row_key, base_value, versions, version_meta, base_file,
            )
            if llm_res is not None:
                return llm_res

        base_str = str(base_value) if base_value is not None else ""
        non_base = {k: v for k, v in versions.items() if str(v) != base_str}

        if not non_base:
            return {
                "suggested_version": list(versions.keys())[0],
                "suggestion": str(list(versions.values())[0]),
                "reasoning": "所有版本值相同",
                "confidence": 1.0,
            }

        col_lower = col_name.lower()

        # 策略 1：ID/编号类列 → 选非 base 中最大的
        if any(kw in col_lower for kw in ("id", "编号", "等级", "level")):
            best_ver = max(non_base.keys(), key=lambda k: (str(non_base[k]) if non_base[k] is not None else ""))
            return {
                "suggested_version": best_ver,
                "suggestion": str(non_base[best_ver]),
                "reasoning": f"ID/编号类列，建议采用非基准版本 '{best_ver}'（通常基准版本较旧）",
                "confidence": 0.75,
            }

        # 策略 2：数值列 → 选非 base 中值最大的
        try:
            numeric_vals = {}
            for k, v in non_base.items():
                if v is not None:
                    numeric_vals[k] = float(v)
            if numeric_vals:
                # 排除异常值（NaN, Inf）
                valid = {k: v for k, v in numeric_vals.items() if v == v and v != float('inf')}
                if valid:
                    best_ver = max(valid.keys(), key=lambda k: valid[k])
                    return {
                        "suggested_version": best_ver,
                        "suggestion": str(non_base[best_ver]),
                        "reasoning": f"数值列，建议采用最大值版本 '{best_ver}'（值={valid[best_ver]}）",
                        "confidence": 0.6,
                    }
        except (ValueError, TypeError):
            pass

        # 策略 3：文本/描述列 → 选最长的
        text_versions = {}
        for k, v in non_base.items():
            if v is not None:
                text_versions[k] = len(str(v))
        if text_versions:
            best_ver = max(text_versions.keys(), key=lambda k: text_versions[k])
            return {
                "suggested_version": best_ver,
                "suggestion": str(non_base[best_ver]),
                "reasoning": f"文本列，建议采用最长版本 '{best_ver}'（{text_versions[best_ver]} 字符，通常描述更完整）",
                "confidence": 0.55,
            }

        # 策略 4：兜底 → 选第一个非 base 版本
        first_ver = list(non_base.keys())[0]
        return {
            "suggested_version": first_ver,
            "suggestion": str(non_base[first_ver]),
            "reasoning": f"默认建议采用版本 '{first_ver}'",
            "confidence": 0.3,
        }


# 单例
_agent_service: Optional[AgentService] = None


def get_agent_service() -> AgentService:
    """获取 AgentService 单例。"""
    global _agent_service
    if _agent_service is None:
        _agent_service = AgentService()
    return _agent_service


def init_agent_service(resources_dir: Path = None) -> AgentService:
    """初始化 AgentService（可指定资源目录）。"""
    global _agent_service
    _agent_service = AgentService(resources_dir)
    return _agent_service
