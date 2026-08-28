"""模式一：跨分支合并 API（absorb / merge_back）。

基于 SVN copyfrom 反查真实 merge-base，三方对比 source/target 分支最新版本，
复用 compare_sheet + commit_authors（同作者自动合并）+ formula_notice 规则。

- direction=absorb：source 的改动合到 target 分支（落点 target）
- direction=merge_back：分支合回 trunk（落点 trunk，版本化命名）

compare 阶段两者处理一致，区别在 apply 落点（任务5.3）。

数据源统一为真实 SVN 工作副本（merge/svn/demo_svn/wc：trunk + branches/*），
版本标识为真实 SVN revision 号，merge-base 通过 `svn log --stop-on-copy` 反查
copyfrom 得到，不再依赖文件夹模拟快照（该方式已下线）。
"""
import asyncio
import json
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
import uuid
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

from fastapi import APIRouter, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from config import MERGE_DIR, SVN_DEMO_WC_DIR
from engine.models import CompareResponse, MergeRequest, SheetMergeData
from engine.parallel_compare import parallel_map_tables
from routers.merge_stages import (
    _build_group, _unresolved_conflicts, _sheets_stats,
    _collect_changes, _append_audit, _validate_apply_refs,
)
from engine.fast_apply import collect_disk_sheet_pks
from routers.precommit_hold import preflight_row_manifest, record_hold_audit
from routers.svn_history import _resolve_branch_path, _run_soft, _resolve_branch_point
from routers.structural import compute_structural_changes, build_display_group, build_display_sheet
# 4.4 worker 隔离:不再模块级 import services.agent_service(触发 agent/__init__ 重依赖,
# ProcessPool spawn 子进程导入本模块会死锁)。_prefetch_ai_suggestions 内已改局部 import。

# AI 建议预取后台线程池（fire-and-forget，绝不阻塞 compare 响应）
_AI_PREFETCH_POOL = ThreadPoolExecutor(max_workers=2, thread_name_prefix="ai-prefetch")
# 预取上限：大合并多表多冲突时控制 LLM 配额，只预取冲突最多的前 N 个 sheet
_PREFETCH_SHEET_LIMIT = 5
# 在途 sheet 去重：避免用户点格与预取同 sheet 重复发起 LLM 调用
_prefetch_inflight: set = set()


def _prefetch_ai_suggestions(result: CompareResponse) -> None:
    """compare 返回后，后台预取冲突最多的前 N 个 sheet 的 AI 建议填缓存。

    成本控制：
    - 仅当 AgentService 已初始化时预取（避免后台线程触发 60+ 表索引重建 + 文件监听）；
    - 只预取冲突最多的前 _PREFETCH_SHEET_LIMIT 个 sheet，大合并不会耗光 LLM 配额；
    - 在途 sheet 去重（_prefetch_inflight），避免用户点格与预取重复调用 LLM。
    异常全吞，绝不影响 compare 主流程。
    """
    from services.agent_service import _agent_service
    if _agent_service is None:
        return  # 未初始化（LLM 后端未启动），跳过预取避免重初始化
    service = _agent_service

    # 按冲突数降序取前 N 个 sheet
    sheet_jobs = []
    for gn, fg in result.groups.items():
        base_file = fg.base_file or ""
        vm = fg.version_meta or {}
        for sn, sd in fg.sheets.items():
            conflicts = []
            for row in sd.rows:
                rk = str(row.key)
                for cell in row.cells:
                    if cell.conflict and cell.versions:
                        conflicts.append({
                            "ri": 0, "ci": cell.col,
                            "col_name": cell.col_letter,
                            "row_key": rk,
                            "base_value": cell.versions.get(base_file),
                            "versions": cell.versions,
                        })
            if conflicts:
                sheet_jobs.append((len(conflicts), gn, sn, base_file, vm, conflicts))
    sheet_jobs.sort(key=lambda x: x[0], reverse=True)
    sheet_jobs = sheet_jobs[:_PREFETCH_SHEET_LIMIT]

    def _worker():
        for _, gn, sn, base_file, vm, conflicts in sheet_jobs:
            key = (gn, sn)
            if key in _prefetch_inflight:
                continue  # 已在途（用户刚点过同 sheet），跳过避免重复 LLM 调用
            _prefetch_inflight.add(key)
            try:
                service.suggest_merge_batch(
                    table_stem=gn, sheet=sn,
                    items=conflicts, version_meta=vm, base_file=base_file,
                )
            except Exception:
                pass  # 单 sheet 预取失败不影响其余
            finally:
                _prefetch_inflight.discard(key)

    _AI_PREFETCH_POOL.submit(_worker)
from routers.diff import (
    _apply_edits_and_save, _save_with_formula_cache, get_next_merge_version,
)
from engine.fast_apply import collect_disk_sheet_pks

router = APIRouter(prefix="/api/merge/branch", tags=["merge-branch"])


# ── 真实 SVN 工作副本支持（merge/svn/demo_svn/wc：trunk + branches/*）──
# 版本标识为真实 SVN revision 号，merge-base 通过 svn log --stop-on-copy
# 反查 copyfrom 得到，不再使用文件夹模拟快照。


# 请求级 svn log 缓存：同一 compare/apply 内多张表多文件共享，避免 N 表×多文件
# 各跑一次 svn log -l 1（M9 性能）。key=文件路径字符串，value=(rev, author, date)。
# 由调用方传入空 dict，跨表/跨文件复用，整次请求结束后随 bp_cache 一起释放。
_rev_cache: Dict[str, tuple] = {}


def _fmt_rev(rev) -> str:
    """SVN revision → 'r' + 五位零填充字符串（如 3 → 'r00003'）。

    统一版本号展示格式：preview-base 返回的 source_rev/target_rev/base_rev、
    /log 下拉选项 value、version_meta[...]["rev"] 均用此格式，前端 select
    v-model 值与 option value 字符串一致才能回显。
    """
    try:
        n = int(rev)
    except (TypeError, ValueError):
        return ""
    return f"r{n:05d}"


def _rev_info_cached(target: Path, rev_cache: Optional[Dict[str, tuple]] = None) -> tuple:
    """带缓存的 svn log -l 1：取文件最新 (rev, author, date)。

    rev_cache 为空时退化为无缓存（兼容 preview-base 等非批量场景）。
    key=路径 resolve 规范化字符串，与 _prefill_rev_cache 的 svn info 输出对齐，
    避免绝对路径 vs 相对路径 key 不匹配。
    """
    key = str(Path(target).resolve())
    if rev_cache is not None and key in rev_cache:
        return rev_cache[key]
    # -r HEAD:1：工作副本未 update 时，svn log 默认查 BASE:1 会返回旧 rev 的 author/date，
    # 导致 commit_authors（同作者自动合并）误判。强制从 repo HEAD 倒查取最新提交信息。
    out, err = _run_soft(["svn", "log", "--xml", "-r", "HEAD:1", "-l", "1", key])
    info = ("", "", "")
    if err is None and out:
        try:
            root = ET.fromstring(out)
            le = root.find("logentry")
            if le is not None:
                info = (
                    (le.get("revision") or "").strip(),
                    (le.findtext("author") or "").strip(),
                    (le.findtext("date") or "").strip(),
                )
        except ET.ParseError:
            pass
    if rev_cache is not None:
        rev_cache[key] = info
    return info


def _prefill_rev_cache(dir_path: Path, rev_cache: Optional[Dict[str, tuple]] = None) -> None:
    """一次性 svn info -R --xml 取目录所有文件的 (rev, author, date) 批量填入 rev_cache。

    替代 N 表×多文件各跑一次 svn log -l 1（74 表 ~148 次 subprocess ~190s），
    一次 svn info -R 递归取全目录 ~0.1s。_rev_info_cached 命中预填后直接返回，跳过 subprocess。
    svn 不可用或解析失败静默降级（后续 _rev_info_cached 仍逐文件兜底）。
    """
    if rev_cache is None or not dir_path.is_dir():
        return
    out, err = _run_soft(["svn", "info", "-R", "--xml", str(dir_path)])
    if err is not None or not out:
        return
    try:
        root = ET.fromstring(out)
    except ET.ParseError:
        return
    for entry in root.findall("entry"):
        path = entry.get("path")
        if not path:
            continue
        commit = entry.find("commit")
        if commit is None:
            continue
        rev = (commit.get("revision") or "").strip()
        author = (commit.findtext("author") or "").strip()
        date = (commit.findtext("date") or "").strip()
        if rev:
            # entry path 规范化为绝对路径，与 _rev_info_cached 的 str(Path(target).resolve()) key 对齐
            rev_cache[str(Path(path).resolve())] = (rev, author, date)


def _copyfrom_chain_svn(wc_path: Path,
                        bp_cache: Optional[Dict[str, dict]] = None) -> List[dict]:
    """SVN 真实模式：沿 svn copyfrom 上溯，返回祖先链（从子到父，不含 inferred）。

    每条 {wc_path, copyfrom_path, copyfrom_rev}。遇 inferred（纯新建无 copyfrom）
    或 svn 不可用终止——inferred 的 copyfrom_path 是路径自身，非真 fork 点，
    不参与 LCA 计算。环保护：访问集合去重。对应 openspec merge-svn-dual-mode。

    示例：A 从 trunk@r1 fork → 链 [{A, trunk, r1}]（trunk inferred 终止不收）。
    """
    chain: List[dict] = []
    seen = set()
    cur = wc_path
    while cur.is_dir() and str(cur) not in seen:
        seen.add(str(cur))
        bp = _bp_cached(str(cur), bp_cache)
        if not bp.get("ok") or bp.get("inferred"):
            break
        cf = bp.get("copyfrom_path", "")
        cf_rev = bp.get("copyfrom_rev")
        if not cf or cf_rev is None:
            break
        chain.append({"wc_path": str(cur), "copyfrom_path": cf, "copyfrom_rev": cf_rev})
        wc_root = _wc_root_for(cur)
        if wc_root is None:
            break
        nxt = _repo_path_to_wc(cf, wc_root)
        if not nxt.is_dir() or nxt == cur:
            break
        cur = nxt
    return chain


def _lca_svn(src_dir: Path, tgt_dir: Path,
             bp_cache: Optional[Dict[str, dict]] = None) -> Optional[dict]:
    """SVN 真实模式：双方 copyfrom 链交叉求最近共同祖先 rev。

    返回 {copyfrom_path, copyfrom_rev} 或 None。按 copyfrom_path 匹配，rev 取
    双方该路径最早值——例如 A copyfrom=trunk@r1、B copyfrom=trunk@r2，LCA=trunk@r1
    （双方都包含的 trunk 内容最早版本）。无公共祖先返回 None。
    """
    src_chain = _copyfrom_chain_svn(src_dir, bp_cache)
    tgt_by_path: Dict[str, dict] = {}
    for e in _copyfrom_chain_svn(tgt_dir, bp_cache):
        p = e["copyfrom_path"]
        if not p:
            continue
        if p not in tgt_by_path or e["copyfrom_rev"] < tgt_by_path[p]["copyfrom_rev"]:
            tgt_by_path[p] = e
    for e in src_chain:
        p = e["copyfrom_path"]
        if p in tgt_by_path:
            tgt_e = tgt_by_path[p]
            lca_rev = min(e["copyfrom_rev"], tgt_e["copyfrom_rev"])
            return {"copyfrom_path": p, "copyfrom_rev": lca_rev}
    return None


class BranchCompareRequest(BaseModel):
    direction: str                            # absorb | merge_back
    source_branch: str                        # 相对 MERGE_DIR 的分支目录（如 branches/A）
    target_branch: str                        # 相对 MERGE_DIR 的目标分支目录（absorb 落点）
    group_names: Optional[List[str]] = None   # 要比对的表名列表，None=自动取两分支交集全部比对
    merge_base_override: Optional[str] = None  # 基准文件路径，覆盖自动反查
    source_rev: Optional[int] = None          # source 指定 SVN revision（None=HEAD）
    target_rev: Optional[int] = None          # target 指定 SVN revision（None=HEAD）
    base_rev: Optional[int] = None            # 共同祖先 base 指定 revision（None=自动 LCA）


# /dirs 结果缓存：merge 下目录扫描（rglob + _collect_table_keys）在文件多时耗数百
# 毫秒，产出新版本目录后主动失效。
_DIRS_CACHE: Dict[str, Any] = {"ts": 0.0, "data": None}
_DIRS_CACHE_TTL = 30.0   # 秒

# /log 结果缓存：svn log 全量历史在长生命周期分支上 subprocess 阻塞数秒，进页面选
# From/To/Base 各调一次。历史不常变（仅 apply 产出新提交时新增一条），TTL 内复用；
# apply 后随 /dirs 一起失效。key=前端传入的 path 字符串。
_LOG_CACHE: Dict[str, Any] = {}
_LOG_CACHE_TTL = 30.0   # 秒
# 限量取最近 N 条历史，足够下拉选择且大幅缩短 svn log 输出（长生命周期分支全量可达数千条）
_LOG_LIMIT = 200

def invalidate_dirs_cache():
    """apply 成功后调，让下次 /dirs 重新扫描（产出新版本目录能立即出现在下拉）。"""
    _DIRS_CACHE["ts"] = 0.0
    _DIRS_CACHE["data"] = None
    # apply 产出新提交，分支历史列表变化，/log 缓存一并失效
    _LOG_CACHE.clear()
    try:
        from routers import merge_subdir
        merge_subdir.invalidate_dirs_cache()
    except Exception:
        pass

@router.get("/dirs")
def list_branch_dirs():
    """列出可选分支目录（真实 SVN 工作副本 + merge 下目录），供前端下拉选择。

    返回 [{name, type, tables, copyfrom, copyfrom_rev, rev, note}]：
      - rev：自身版本号（svn/demo_svn wc 从目录名 _rN 解析；真实 SVN revision 由前端
        调 /log 取 HEAD 显示，此处留空）
      - copyfrom/copyfrom_rev：fork 来源（真实 SVN 模式下由 svn log --stop-on-copy
        在 compare/preview-base 阶段反查，列表阶段不填）
    tables 为该分支包含的表名前缀集合。
    """
    now = time.time()
    if _DIRS_CACHE["data"] is not None and now - _DIRS_CACHE["ts"] < _DIRS_CACHE_TTL:
        return _DIRS_CACHE["data"]
    dirs = []

    # 1. merge 下其他目录（branches/A 等 wc 目录）
    for d in sorted(MERGE_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("."):
            continue
        # 跳过已知非分支目录（svn/ 由第 2 步单独收集真实 SVN 工作副本）
        if d.name in ("trunk", "src", "svn", "history", "mergebase", "samples", "legacy", "scripts", "_seed_data", "demo"):
            continue
        # 分支子目录（如 branches/A）
        if d.name == "branches" and d.is_dir():
            for b in sorted(d.iterdir()):
                if b.is_dir() and not b.name.startswith("."):
                    tables = sorted(_collect_table_keys(b))
                    dirs.append({"name": f"branches/{b.name}", "type": "wc",
                                 "tables": tables, "copyfrom": "", "note": ""})
            continue
        tables = sorted(_collect_table_keys(d))
        if tables:
            dirs.append({"name": d.name, "type": "wc", "tables": tables,
                         "copyfrom": "", "note": ""})

    # 2. 真实 SVN 工作副本（merge/svn/demo_svn/wc：trunk + branches/*，版本标识为 SVN revision 号）
    if SVN_DEMO_WC_DIR.is_dir():
        trunk_wc = SVN_DEMO_WC_DIR / "wc" / "trunk"
        if trunk_wc.is_dir():
            tables = sorted(_collect_table_keys(trunk_wc))
            if tables:
                dirs.append({"name": "svn/demo_svn/wc/trunk", "type": "wc",
                             "tables": tables, "copyfrom": "", "note": "SVN rev 号工作副本"})
        branches_wc = SVN_DEMO_WC_DIR / "wc" / "branches"
        if branches_wc.is_dir():
            for b in sorted(branches_wc.iterdir()):
                if b.is_dir() and not b.name.startswith("."):
                    tables = sorted(_collect_table_keys(b))
                    dirs.append({"name": f"svn/demo_svn/wc/branches/{b.name}", "type": "wc",
                                 "tables": tables, "copyfrom": "", "note": "SVN rev 号工作副本"})

    data = {"dirs": dirs}
    _DIRS_CACHE["ts"] = time.time()
    _DIRS_CACHE["data"] = data
    return data


class PreviewBaseRequest(BaseModel):
    """选完两文件后预览共同祖先 base 版本。"""
    source_branch: str
    target_branch: str
    base_rev: Optional[int] = None      # 用户手动指定 base rev（None=自动 LCA）


@router.post("/preview-base")
def preview_base(req: PreviewBaseRequest):
    """选完 From/To 后预览三方版本信息：source_rev / target_rev / base_rev + base 来源。

    真实 SVN：rev 取 HEAD（svn log -l 1），base 用 _lca_svn 反查 copyfrom 最近共同祖先
    rev，或 base_rev 手动指定。LCA 失败时按 fork 逻辑回退（merge_back 的 trunk inferred、
    目录合并的子目录 inferred），取单侧 copyfrom_rev 作 base。base_source 返回工作副本相对
    路径（如 svn/demo_svn/wc/trunk），供前端 /log 加载 base rev 下拉选项。
    """
    src_dir = _resolve_demo_or_wc(req.source_branch)
    tgt_dir = _resolve_demo_or_wc(req.target_branch)
    if not src_dir.is_dir() or not tgt_dir.is_dir():
        raise HTTPException(400, "source_branch / target_branch 必须为目录")

    bp_cache: Dict[str, dict] = {}
    rev_cache: Dict[str, tuple] = {}
    # source/target rev 取该分支目录下所有文件的最新 revision（max），而非单一表的 rev：
    # SVN 每文件 rev 独立，分支"最新版本"应是树内文件的最大 rev（含最新提交）。
    # src/tgt 两个目录并行跑（各 74 文件 svn log），串行会翻倍耗时。
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=2) as ex:
        src_rev_f = ex.submit(_dir_head_rev, src_dir, rev_cache)
        tgt_rev_f = ex.submit(_dir_head_rev, tgt_dir, rev_cache)
        src_rev = src_rev_f.result()
        tgt_rev = tgt_rev_f.result()
    if req.base_rev is not None:
        # 手动指定 base_rev：base_source 仍解析真实工作副本路径供前端 /log 加载历史，
        # 避免 base_source="手动指定" 时前端 /log?path=手动指定 返回 400。
        _, auto_source = _resolve_base_preview(src_dir, tgt_dir, bp_cache)
        base_rev = _fmt_rev(req.base_rev)
        base_source = auto_source
    else:
        base_rev, base_source = _resolve_base_preview(src_dir, tgt_dir, bp_cache)
    return {
        "mode": "svn",
        "source_rev": src_rev,
        "target_rev": tgt_rev,
        "base_rev": base_rev,
        "base_source": base_source,
        "source_copyfrom": "",
        "target_copyfrom": "",
    }


def _repo_path_to_wc_rel(repo_path: str, src_dir: Path, tgt_dir: Path) -> str:
    """把 SVN 仓库内路径（如 /trunk）转为前端可选的工作副本相对路径。

    前端 /log 需要相对 MERGE_DIR 的路径（如 svn/demo_svn/wc/trunk）才能调后端取历史，
    仓库路径 /trunk 无法直接用。通过 wc_root 反推 svn/demo_svn/wc 前缀 + repo_path 的
    trunk/branches 段拼出完整相对路径。
    """
    wc_root = _wc_root_for(src_dir) or _wc_root_for(tgt_dir)
    if wc_root is None:
        return repo_path
    try:
        rel = wc_root.resolve().relative_to(MERGE_DIR.resolve()).as_posix()
    except ValueError:
        return repo_path
    return f"{rel}/{(repo_path or '').lstrip('/')}"


def _dir_head_rev(d: Path, rev_cache: Optional[Dict[str, tuple]] = None) -> str:
    """取目录树的最新 SVN revision（分支树内最新一次提交的 rev，即 repo HEAD 视角）。

    `svn log` 对工作副本路径默认查 BASE:1（工作副本停的旧 rev 之前），工作副本未 update
    时返回 fork rev 而非真实 HEAD。必须加 `-r HEAD:1` 强制从 repo HEAD 倒查，取树内最新一条
    logentry。默认递归，一条 logentry 的 revision = 树内文件 rev 最大值。
    一次 subprocess 替代旧实现的 N 文件并行 svn log（74 文件 ~12s → ~0.3s）。
    """
    if not d.is_dir():
        return ""
    # 目录下无 xlsx 则无意义（避免 svn log 空目录）
    has_xlsx = any(
        not fp.name.startswith("~$") and not fp.name.startswith("_")
        for fp in d.rglob("*.xlsx")
    )
    if not has_xlsx:
        return ""
    dir_key = str(d)
    if rev_cache is not None and dir_key in rev_cache:
        info = rev_cache[dir_key]
        rev = info[0] if isinstance(info, tuple) else ""
        return _fmt_rev(rev) if rev else ""
    # -r HEAD:1：从 repo HEAD 倒查（非工作副本 BASE），-l 1 取最新一条
    out, err = _run_soft(["svn", "log", "--xml", "-r", "HEAD:1", "-l", "1", dir_key])
    rev = ""
    if err is None and out:
        try:
            root = ET.fromstring(out)
            le = root.find("logentry")
            if le is not None:
                rev = (le.get("revision") or "").strip()
        except ET.ParseError:
            pass
    if rev_cache is not None and rev:
        rev_cache[dir_key] = (rev, "", "")
    return _fmt_rev(rev) if rev else ""


def _resolve_base_preview(src_dir: Path, tgt_dir: Path,
                          bp_cache: Dict[str, dict]) -> tuple:
    """preview-base 的 base 解析：LCA 优先，子目录场景次之，失败走 fork 回退。

    返回 (base_rev_str, base_source)。base_source 为工作副本相对路径
    （供前端 /log 加载 base rev 选项）或"手动指定"占位。
    """
    # 1. 双方 LCA
    lca = _lca_svn(src_dir, tgt_dir, bp_cache)
    if lca is not None:
        return _fmt_rev(lca["copyfrom_rev"]), _repo_path_to_wc_rel(
            lca["copyfrom_path"], src_dir, tgt_dir)

    # 2. 子目录场景：一侧是另一侧的子目录（目录合并），base = 父目录 @ 子目录 copyfrom_rev
    #    子目录创建时刻继承父目录版本，base_source 取父目录工作副本路径供前端 /log 加载。
    for child, parent in ((src_dir, tgt_dir), (tgt_dir, src_dir)):
        try:
            child.resolve().relative_to(parent.resolve())
        except ValueError:
            continue
        bp_child = _bp_cached(str(child), bp_cache)
        if bp_child.get("ok"):
            rev = bp_child.get("copyfrom_rev")
            if rev is not None:
                try:
                    rel = parent.resolve().relative_to(MERGE_DIR.resolve()).as_posix()
                except ValueError:
                    rel = str(parent)
                return _fmt_rev(rev), rel

    # 3. 单侧 inferred（merge_back 的 trunk）→ 沿另一侧链找 fork rev
    bp_t = _bp_cached(str(tgt_dir), bp_cache)
    bp_s = _bp_cached(str(src_dir), bp_cache)
    if (bp_t.get("ok") and bp_t.get("inferred")) or (bp_s.get("ok") and bp_s.get("inferred")):
        for infer_dir, other_dir in ((tgt_dir, src_dir), (src_dir, tgt_dir)):
            infer_bp = bp_t if infer_dir == tgt_dir else bp_s
            if not infer_bp.get("ok") or not infer_bp.get("inferred"):
                continue
            wc_root = _wc_root_for(infer_dir) or _wc_root_for(other_dir)
            if wc_root is None:
                continue
            try:
                infer_repo = infer_dir.resolve().relative_to(wc_root).as_posix()
            except ValueError:
                continue
            if not infer_repo:
                continue
            best = None
            for e in _copyfrom_chain_svn(other_dir, bp_cache):
                e_repo = (e["copyfrom_path"] or "").strip("/")
                if e_repo == infer_repo and e["copyfrom_rev"] is not None:
                    if best is None or e["copyfrom_rev"] < best["copyfrom_rev"]:
                        best = e
            if best is not None:
                return _fmt_rev(best["copyfrom_rev"]), _repo_path_to_wc_rel(
                    best["copyfrom_path"], src_dir, tgt_dir)

    # 4. 单侧 fork 兜底（target 优先，merge_back 的 trunk 无 copyfrom 时取 source）
    bp = _bp_cached(str(tgt_dir), bp_cache)
    use_dir = tgt_dir
    if not bp.get("ok") or bp.get("inferred"):
        bp_src = _bp_cached(str(src_dir), bp_cache)
        if bp_src.get("ok") and not bp_src.get("inferred"):
            bp = bp_src
            use_dir = src_dir
    if bp.get("ok") and not bp.get("inferred"):
        return _fmt_rev(bp["copyfrom_rev"]), _repo_path_to_wc_rel(
            bp["copyfrom_path"], src_dir, tgt_dir)

    return "", ""


def _any_table(d: Path) -> str:
    """取目录下任意一个表名（用于取 HEAD revision 信息）。"""
    keys = _collect_table_keys(d)
    return next(iter(keys), "") if keys else ""


@router.get("/log")
def list_branch_history(path: str):
    """返回某分支/目录的历史版本列表，供 From/To/Base rev 下拉选择。

    真实 SVN：svn log --xml 取该路径最近 _LOG_LIMIT 条历史提交 → [{rev, author, date, msg}, ...]。
    TTL 缓存：同路径 TTL 内复用（apply 产出新提交后由 invalidate_dirs_cache 主动失效）。
    """
    p = _resolve_demo_or_wc(path)
    if not p.is_dir():
        raise HTTPException(400, f"路径不存在或不是目录: {path}")

    now = time.time()
    cached = _LOG_CACHE.get(path)
    if cached and now - cached["ts"] < _LOG_CACHE_TTL:
        return cached["data"]

    try:
        # -r HEAD:1：工作副本未 update 时，svn log 默认查 BASE:1 只返回旧 rev 之前的，
        # 下拉看不到真实 HEAD。强制从 repo HEAD 倒查。-l _LOG_LIMIT 限量取最近 N 条，
        # 避免长生命周期分支全量历史（数千条）导致 subprocess 阻塞。
        result = subprocess.run(
            ["svn", "log", "--xml", "-r", "HEAD:1", "-l", str(_LOG_LIMIT), str(p)],
            capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30,
        )
        if result.returncode != 0:
            raise HTTPException(400, f"svn log 失败: {result.stderr}")
        root = ET.fromstring(result.stdout)
        entries = []
        for entry in root.findall("./logentry"):
            entries.append({
                "rev": _fmt_rev(entry.get("revision")),
                "author": entry.findtext("author", ""),
                "date": entry.findtext("date", ""),
                "msg": entry.findtext("msg", "").strip(),
            })
        data = {"mode": "svn", "path": path, "entries": entries}
        _LOG_CACHE[path] = {"ts": now, "data": data}
        return data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"获取 SVN 历史失败: {e}")


@router.get("/tables")
def list_common_tables(source_branch: str, target_branch: str):
    """返回 source/target 两分支的交集表名，供前端下拉选择。

    支持嵌套子文件夹：表名为相对分支目录的 posix 路径（如 'config/skill'），
    扁平表名为文件名 stem（如 'ability'，保持旧兼容）。
    """
    src_dir = _resolve_demo_or_wc(source_branch)
    tgt_dir = _resolve_demo_or_wc(target_branch)
    if not src_dir.is_dir() or not tgt_dir.is_dir():
        raise HTTPException(400, "source_branch / target_branch 必须为目录")
    src_tables = _collect_table_keys(src_dir)
    tgt_tables = _collect_table_keys(tgt_dir)
    common = sorted(src_tables & tgt_tables)
    return {"common_tables": common, "source_only": sorted(src_tables - tgt_tables),
            "target_only": sorted(tgt_tables - src_tables)}


_TABLE_KEYS_CACHE: Dict[Tuple[str, float], set] = {}
_TABLE_KEYS_LRU = 256


def _collect_table_keys(base_dir: Path) -> set:
    """收集目录下所有 xlsx 的表名 key（相对 base_dir 的 posix 路径，去 .xlsx 后缀）。

    支持嵌套：config/skill.xlsx → 'config/skill'；扁平 ability.xlsx → 'ability'。
    跳过 ~$ 临时文件、_ 前缀元数据文件。

    方案 ④ P0-2: (dir, max_mtime) 缓存——svn up 改 mtime → 自动失效；
    冷 /dirs 列 380 分支 + 同分支多端点访问免重复 rglob+keys 构造。LRU 256 防涨内存。
    单次 rglob 同时取 mtime 与 keys，避免命中/未命中双 rglob。
    """
    if not base_dir.is_dir():
        return set()
    files = [fp for fp in base_dir.rglob("*.xlsx")
             if not fp.name.startswith("~$") and not fp.name.startswith("_")]
    try:
        m = max((fp.stat().st_mtime for fp in files), default=0.0)
    except OSError:
        m = 0.0
    key = (str(base_dir), m)
    cached = _TABLE_KEYS_CACHE.get(key)
    if cached is not None:
        return cached
    keys = {fp.relative_to(base_dir).with_suffix("").as_posix() for fp in files}
    if len(_TABLE_KEYS_CACHE) >= _TABLE_KEYS_LRU:
        _TABLE_KEYS_CACHE.pop(next(iter(_TABLE_KEYS_CACHE)))
    _TABLE_KEYS_CACHE[key] = keys
    return keys


def _resolve_demo_or_wc(branch: str) -> Path:
    """分支名解析为绝对路径：相对路径拼到 MERGE_DIR，绝对路径直接用（经越界校验）。

    前端下拉选项的 name 如 'svn/demo_svn/wc/branches/dev1'、'branches/A'，
    均相对 MERGE_DIR 解析；_resolve_branch_path 内部做越界校验。
    保留旧函数名以兼容 merge_subdir 的 import。
    """
    if not branch:
        raise HTTPException(400, "branch 不能为空")
    return _resolve_branch_path(branch)


def _list_branch_files(branch_dir: Path, group_name: str) -> List[Path]:
    """列出分支目录下匹配 group_name 的 xlsx 文件（wc 最新版，通常唯一）。

    支持嵌套子文件夹：group_name 为相对路径（posix，如 'config/skill'），
    在 branch_dir 下按相对路径定位 {group}.xlsx；也兼容旧扁平表名（如 'ability'）。
    """
    files = _locate_group_files(branch_dir, group_name)
    if not files:
        raise HTTPException(400, f"分支目录 {branch_dir} 下未找到匹配 {group_name} 的 xlsx")
    return files


def _locate_group_files(base_dir: Path, group_name: str) -> List[Path]:
    """在 base_dir 下按 group_name（相对路径 posix，如 'config/skill'）定位 xlsx。

    先尝试规范名 {group}.xlsx，再尝试带时间点后缀 {group}_sN.xlsx（目录合并场景），
    最后兼容旧平铺前缀匹配 {group}*.xlsx（用于 wc 目录下 item_merged_devbranch1 等中间名）。
    返回匹配文件列表（按文件名排序）。group_name 含路径分隔符时只走精确路径定位。
    """
    rel = (group_name or "").strip().lstrip("/")
    if not rel:
        return []
    # 嵌套路径：group_name 含分隔符 → 按相对路径精确文件名定位
    if "/" in rel or "\\" in rel:
        parts = rel.replace("\\", "/").split("/")
        rel_path = Path(*parts)
        direct = base_dir / rel_path.parent / f"{rel_path.name}.xlsx"
        if direct.is_file() and not direct.name.startswith("~$"):
            return [direct]
        # 时间点后缀变体：{leaf}_sN.xlsx（目录合并嵌套场景）
        parent_dir = (base_dir / rel_path.parent).resolve() if str(rel_path.parent) != "." else base_dir
        leaf = rel_path.name
        s_versions = []
        if parent_dir.is_dir():
            for p in parent_dir.glob("*.xlsx"):
                if p.name.startswith("~$") or p.name.startswith("_"):
                    continue
                m = re.match(rf"^{re.escape(leaf)}_s(\d+)$", p.stem)
                if m:
                    s_versions.append(p)
        return sorted(s_versions)
    # 扁平表名：优先精确匹配 {group}.xlsx；不存在再兼容旧前缀匹配 {group}*.xlsx
    #（wc 目录下可能有 _merged 等中间名）。避免 "space" 同时把 space_landmark 等
    # 同名前缀表卷进同一 compare group，导致前端出现多个 From/To 列。
    exact = base_dir / f"{rel}.xlsx"
    if exact.is_file() and not exact.name.startswith("~$"):
        return [exact]
    return [fp for fp in sorted(base_dir.glob(f"{rel}*.xlsx"))
            if not fp.name.startswith("~$") and not fp.name.startswith("_")]


def _svn_rel_target(target: Path) -> Tuple[List[str], Optional[str]]:
    """SVN 命令参数 + cwd 归一：target 在工作副本内时转相对路径 + cwd=wc 根，
    规避 Windows 上 TortoiseSVN/unisvn 对绝对路径大小写解析偶发 E720005；
    非工作副本回退绝对路径 + 无 cwd。返回 (argv 追加项, cwd 或 None)。
    """
    wc_root = _find_wc_root(target)
    if wc_root is not None:
        try:
            rel = target.resolve().relative_to(wc_root).as_posix()
            return [rel], str(wc_root)
        except ValueError:
            pass
    return [str(target)], None


def _svn_cat(target: Path, rev: Optional[int], dest: Path) -> None:
    """svn cat -r rev 取文件内容写到 dest；rev=None 取最新。"""
    cmd = ["svn", "cat"]
    if rev is not None:
        cmd += ["-r", str(rev)]
    argv, cwd = _svn_rel_target(target)
    cmd += argv
    try:
        r = subprocess.run(cmd, capture_output=True, timeout=60, cwd=cwd)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        raise HTTPException(500, "svn cat 失败：svn 不可用或超时")
    if r.returncode != 0:
        err = r.stderr.decode("utf-8", errors="replace").strip()
        raise HTTPException(500, f"svn cat {target}@{rev} 失败: {err}")
    dest.write_bytes(r.stdout)


def _latest_author(target: Path) -> str:
    """svn log -l 1 取文件最新提交作者。

    批量比对场景应优先用 _rev_info_cached 共享 rev_cache，避免每文件各跑一次
    subprocess（见 _compare_one_table 改造）。本函数保留供单次调用兜底。
    """
    out, err = _run_soft(["svn", "log", "--xml", "-l", "1", str(target)])
    if err is not None or not out:
        return ""
    try:
        root = ET.fromstring(out)
    except ET.ParseError:
        return ""
    le = root.find("logentry")
    return (le.findtext("author") or "").strip() if le is not None else ""


def _latest_revision_info(target: Path, rev_cache: Optional[Dict[str, tuple]] = None) -> Dict[str, str]:
    """svn log -l 1 取文件最新提交的 {rev, author, date}。

    rev_cache 传入时复用缓存（同一 compare 内 N 表×多文件共享，M9 性能优化），
    不传则单次调用。rev 统一为 _fmt_rev 五位零填充 'r00003' 格式。
    """
    rev, author, date = _rev_info_cached(target, rev_cache)
    if not rev:
        return {}
    return {"rev": _fmt_rev(rev), "author": author, "date": date}


def _wc_root_for(target: Path) -> Optional[Path]:
    """从工作副本路径推断 wc 根（路径里 trunk/branches 的父目录）。"""
    parts = target.parts
    for i, p in enumerate(parts):
        if p in ("trunk", "branches") and i > 0:
            return Path(*parts[:i]).resolve()
    return None

def _find_wc_root(p: Path) -> Optional[Path]:
    """返回 p 所属 SVN 工作副本根目录（含 .svn 的最近祖先），非工作副本返回 None。"""
    cur = p if p.is_dir() else p.parent
    while cur != cur.parent:
        if (cur / ".svn").is_dir():
            return cur
        cur = cur.parent
    return None


def _svn_commit_apply(base_dir: Path, commit_files: List[Path], commit_name: str) -> dict:
    """apply 完成后把合并产物提交到目标 SVN 工作副本，返回新 HEAD revision 与状态。

    base_dir 不在 SVN 工作副本内（demo 快照目录）时不做任何提交。
    以 cwd=工作副本根 + 相对路径执行 svn add/commit/info，规避 Windows 上
    TortoiseSVN/unisvn 对绝对路径大小写解析偶发的 E720005。提交失败不回滚文件，
    只回传 error 供调用方透出告警（文件已落盘，用户可手动 svn commit）。
    """
    wc_root = _find_wc_root(base_dir)
    if wc_root is None:
        return {"is_svn": False, "committed": False, "revision": None, "error": None}
    rels: List[str] = []
    for f in commit_files:
        try:
            rels.append(f.resolve().relative_to(wc_root).as_posix())
        except ValueError:
            pass
    if not rels:
        return {"is_svn": True, "committed": False, "revision": None, "error": "无可提交文件"}
    try:
        # svn add 对已版本化文件是 no-op，仅登记新文件（含父目录自动 --parents）
        subprocess.run(["svn", "add", "--parents", *rels],
                       cwd=str(wc_root), capture_output=True, timeout=120)
        # svn add 对已版本化文件是 no-op，不校验其返回码（新文件已登记即可）
        msg = (commit_name or "").strip() or "merge apply"
        r = subprocess.run(["svn", "commit", "-m", msg, *rels],
                           cwd=str(wc_root), capture_output=True, timeout=300)
        if r.returncode != 0:
            err = r.stderr.decode("utf-8", "replace").strip() if isinstance(r.stderr, bytes) else str(r.stderr).strip()
            return {"is_svn": True, "committed": False, "revision": None, "error": err or "svn commit 失败"}
        # 用 commit 输出解析实际 revision（"Committed revision 132."），
        # 比 svn info --show-item revision（多 wc 根时取不到本次提交 rev）更可靠
        rev = None
        out = r.stdout.decode("utf-8", "replace") if isinstance(r.stdout, bytes) else str(r.stdout)
        m = re.search(r"Committed revision (\d+)", out)
        if m:
            rev = int(m.group(1))
        if rev is None:
            info = subprocess.run(["svn", "info", "--show-item", "revision"],
                                  cwd=str(wc_root), capture_output=True, timeout=60)
            if info.returncode == 0:
                try:
                    rev = int(info.stdout.decode("utf-8", "replace").strip())
                except ValueError:
                    rev = None
        return {"is_svn": True, "committed": True, "revision": rev, "error": None}
    except FileNotFoundError:
        return {"is_svn": True, "committed": False, "revision": None, "error": "未找到 svn 命令"}
    except Exception as e:
        return {"is_svn": True, "committed": False, "revision": None, "error": str(e)}


def _git_commit_apply(repo_root: Path, commit_files: List[Path], commit_name: str) -> dict:
    """apply 成功后把本次改动 git 提交到项目仓库，返回 {committed, revision, error}。

    提交范围：仓库内所有已跟踪文件的修改（含 _append_audit 更新的审计日志与
    本次代码改动）。merge/svn/demo_svn 下的合并产物由 SVN 管理且被 .gitignore
    排除，不进 git。失败不抛异常，只回传 error（SVN 已落盘，git 失败不影响结果）。
    """
    if not (repo_root / ".git").exists():
        return {"committed": False, "revision": None, "error": "非 git 仓库"}
    msg = (commit_name or "").strip() or "merge apply"
    try:
        # 只暂存已跟踪文件的修改（git add -u），避免把 gitignore 的 SVN 产物拖进来
        subprocess.run(["git", "-C", str(repo_root), "add", "-u"],
                       capture_output=True, timeout=60)
        r = subprocess.run(["git", "-C", str(repo_root), "commit", "-m", msg],
                           capture_output=True, timeout=120)
        if r.returncode != 0:
            err = r.stderr.decode("utf-8", "replace").strip() if isinstance(r.stderr, bytes) else str(r.stderr).strip()
            if "nothing to commit" in (r.stdout.decode("utf-8", "replace") if isinstance(r.stdout, bytes) else str(r.stdout)) or \
               "nothing to commit" in err:
                return {"committed": False, "revision": None, "error": None}
            return {"committed": False, "revision": None, "error": err or "git commit 失败"}
        rev = None
        out = subprocess.run(["git", "-C", str(repo_root), "rev-parse", "--short", "HEAD"],
                             capture_output=True, timeout=30)
        if out.returncode == 0:
            rev = out.stdout.decode("utf-8", "replace").strip() or None
        return {"committed": True, "revision": rev, "error": None}
    except FileNotFoundError:
        return {"committed": False, "revision": None, "error": "未找到 git 命令"}
    except Exception as e:
        return {"committed": False, "revision": None, "error": str(e)}



def _repo_path_to_wc(repo_path: str, wc_root: Path) -> Path:
    """仓库内路径 /trunk/x → wc_root/trunk/x。"""
    rel = (repo_path or "").lstrip("/")
    return (wc_root / rel).resolve()


def _flat_group_name(group_name: str) -> str:
    """把嵌套表名（如 'config/skill'）扁平化为安全文件名片段（'config__skill'）。

    用于 tmp 内中间文件命名，避免 group_name 含路径分隔符时拼出不存在的子目录。
    扁平表名（如 'ability'）原样返回。
    """
    return (group_name or "").replace("/", "__").replace("\\", "__").strip("/")


def _group_path_in_dir(base_dir: Path, group_name: str) -> Path:
    """在 base_dir 下按 group_name（相对路径 posix 或扁平名）定位 xlsx 文件路径。

    group_name='config/skill' → base_dir/config/skill.xlsx；
    group_name='ability' → base_dir/ability.xlsx；
    目录合并场景 group_name 可带 _sN 后缀语义（如 'match_stat' 匹配 match_stat_s1.xlsx 的
    规范前缀），但本函数取规范名 {group}.xlsx；带后缀变体由 _locate_group_files 处理。
    """
    rel = (group_name or "").strip().lstrip("/")
    parts = rel.replace("\\", "/").split("/")
    return base_dir.joinpath(*parts).with_suffix(".xlsx")


def _bp_cached(path: str, bp_cache: Optional[Dict[str, dict]] = None) -> dict:
    """请求级缓存 _resolve_branch_point 结果（M9）：同一次 compare/apply 内 src_dir/tgt_dir
    不变，N 张表原会触发 N 次 svn log --stop-on-copy（每次 subprocess + 60s 超时）。
    传入 bp_cache dict 复用首次结果，将 N 次 svn 调用降为 1-2 次。
    """
    if bp_cache is not None and path in bp_cache:
        return bp_cache[path]
    result = _resolve_branch_point(path=path)
    if bp_cache is not None:
        bp_cache[path] = result
    return result


def _svn_export_dir(repo_wc_path: Path, rev: Optional[int], dest_dir: Path) -> bool:
    """svn export -r rev 整个目录到 dest_dir（一次性导出，避免 N 次 svn cat）。

    用于 base 批量导出：compare 入口确定 base 的 (copyfrom_path, rev) 后，一次性
    导出整个 trunk@rev，N 张表从本地副本拷贝，0 次 per-table svn cat。
    rev=None 取 HEAD。返回成功与否（失败回退到 per-table svn cat）。
    """
    cmd = ["svn", "export"]
    if rev is not None:
        cmd += ["-r", str(rev)]
    argv, cwd = _svn_rel_target(repo_wc_path)
    cmd += argv
    cmd += ["--force", str(dest_dir)]
    try:
        r = subprocess.run(cmd, capture_output=True, timeout=120, cwd=cwd)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False
    return r.returncode == 0 and dest_dir.is_dir()


# 方案 ② P0-3: base export 按 svn rev 磁盘复用——SVN rev 不可变，同 rev 必同内容，
# 连续比对同分支同 base_rev 第二次起跳 svn export（直击 ~120s 主因）。
# MERGE_BASE_EXPORT_CACHE 可改根目录；MERGE_BASE_EXPORT_KEEP 控保留份数。
_BASE_EXPORT_KEEP = max(1, int(os.environ.get("MERGE_BASE_EXPORT_KEEP", "5") or "5"))


def _base_export_dest(rev: int, tmp: Optional[Path] = None) -> Path:
    """base export 落点：优先 MERGE_DIR/.base_export_cache 共享磁盘缓存；
    tmp 仅在缓存禁用时回退（旧行为）。每次调用读 MERGE_DIR，便于测试 monkeypatch。"""
    env = os.environ.get("MERGE_BASE_EXPORT_CACHE", "").strip()
    if env and env.lower() not in ("off", "0"):
        root = Path(env)
    else:
        root = MERGE_DIR / ".base_export_cache"
    return root / f"base_export_r{rev}"


def _base_export_cache_hit(dest: Path) -> bool:
    """命中判断：目录存在且至少含一个 .xlsx（导出过且非空）。"""
    return dest.is_dir() and any(dest.rglob("*.xlsx"))


def _maybe_prune_base_export_cache() -> None:
    """保留最近 _BASE_EXPORT_KEEP 份 base export（按 mtime 降序），删更老。
    删除失败静默跳过，不阻断 compare 主流程。"""
    env = os.environ.get("MERGE_BASE_EXPORT_CACHE", "").strip()
    if env and env.lower() not in ("off", "0"):
        root = Path(env)
    else:
        root = MERGE_DIR / ".base_export_cache"
    if not root.exists():
        return
    try:
        dirs = [p for p in root.iterdir() if p.is_dir()]
    except OSError:
        return
    dirs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for old in dirs[_BASE_EXPORT_KEEP:]:
        try:
            shutil.rmtree(old)
        except OSError:
            pass


def _base_export_or_cache(wc: Path, rev: Optional[int], tmp: Path) -> Optional[Path]:
    """svn export -r rev 整目录到磁盘缓存，命中则零 subprocess 复用。

    方案 ②：SVN rev 不可变 → 同 rev 必同内容。命中缓存直接返回；未命中走 svn export
    后落缓存并 prune 老目录。返回 dest 路径或 None（导出失败）。
    """
    if rev is None:
        return None
    dest = _base_export_dest(rev, tmp)
    if _base_export_cache_hit(dest):
        return dest
    if _svn_export_dir(wc, rev, dest):
        _maybe_prune_base_export_cache()
        return dest
    return None


def _prepare_base_export(src_dir: Path, tgt_dir: Path, base_rev: Optional[int],
                         tmp: Path, bp_cache: Dict[str, dict]) -> Optional[Path]:
    """确定 base 的 (copyfrom_path, rev) 后一次性 svn export 整个目录到 tmp。

    branch_compare 入口调用，导出的目录供 _resolve_base_for_compare 从本地拷贝，
    跳过 per-table svn cat。三种 base 来源统一在此解析：
      - base_rev 指定 → 取 src/tgt 侧 wc 目录作导出源
      - 双方 LCA → copyfrom_path + copyfrom_rev
      - 单侧 inferred（merge_back 的 trunk）→ 沿另一侧链找指向 inferred 侧的 fork rev
    任一方式确定 (wc_path, rev) 后 svn export；解析不出或导出失败返回 None。
    """
    # 1. base_rev 指定 → 用 src/tgt 侧目录（两者同名 base），取首个存在
    if base_rev is not None:
        for d in (src_dir, tgt_dir):
            if d.is_dir():
                dest = _base_export_or_cache(d, base_rev, tmp)
                if dest is not None:
                    return dest
        return None

    # 2. 自动 LCA
    lca = _lca_svn(src_dir, tgt_dir, bp_cache)
    if lca is not None:
        copyfrom_path = lca["copyfrom_path"]
        copyfrom_rev = lca["copyfrom_rev"]
        wc_root = _wc_root_for(src_dir) or _wc_root_for(tgt_dir)
        if wc_root is not None:
            base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
            if base_wc.is_dir():
                dest = _base_export_or_cache(base_wc, copyfrom_rev, tmp)
                if dest is not None:
                    return dest

    # 3. 单侧 inferred（merge_back 的 trunk）→ 沿另一侧链找 fork rev
    bp_t = _bp_cached(str(tgt_dir), bp_cache)
    bp_s = _bp_cached(str(src_dir), bp_cache)
    if (bp_t.get("ok") and bp_t.get("inferred")) or (bp_s.get("ok") and bp_s.get("inferred")):
        for infer_dir, other_dir in ((tgt_dir, src_dir), (src_dir, tgt_dir)):
            infer_bp = bp_t if infer_dir == tgt_dir else bp_s
            if not infer_bp.get("ok") or not infer_bp.get("inferred"):
                continue
            wc_root = _wc_root_for(infer_dir) or _wc_root_for(other_dir)
            if wc_root is None:
                continue
            try:
                infer_repo = infer_dir.resolve().relative_to(wc_root).as_posix()
            except ValueError:
                continue
            if not infer_repo:
                continue
            best = None
            for e in _copyfrom_chain_svn(other_dir, bp_cache):
                e_repo = (e["copyfrom_path"] or "").strip("/")
                if e_repo == infer_repo and e["copyfrom_rev"] is not None:
                    if best is None or e["copyfrom_rev"] < best["copyfrom_rev"]:
                        best = e
            if best is not None:
                copyfrom_path = best["copyfrom_path"]
                copyfrom_rev = best["copyfrom_rev"]
                base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
                if base_wc.is_dir():
                    dest = tmp / f"base_export_r{copyfrom_rev}"
                    if _svn_export_dir(base_wc, copyfrom_rev, dest):
                        return dest

    # 4. 单侧 fork 兜底（target 优先）
    bp = _bp_cached(str(tgt_dir), bp_cache)
    use_dir = tgt_dir
    if not bp.get("ok") or bp.get("inferred"):
        bp_src = _bp_cached(str(src_dir), bp_cache)
        if bp_src.get("ok") and not bp_src.get("inferred"):
            bp = bp_src
            use_dir = src_dir
    if bp.get("ok") and not bp.get("inferred"):
        copyfrom_path = bp["copyfrom_path"]
        copyfrom_rev = bp["copyfrom_rev"]
        wc_root = _wc_root_for(use_dir)
        if wc_root is not None:
            base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
            if base_wc.is_dir():
                dest = _base_export_or_cache(base_wc, copyfrom_rev, tmp)
                if dest is not None:
                    return dest

    return None


def _resolve_base_for_compare(src_dir: Path, tgt_dir: Path, group_name: str,
                              override: Optional[str], tmp: Path,
                              bp_cache: Optional[Dict[str, dict]] = None,
                              base_rev: Optional[int] = None,
                              base_export_dir: Optional[Path] = None) -> Path:
    """定位 source/target 的共同 merge-base 文件，基于 SVN 真实 LCA。

    base_rev 指定时直接 svn cat -r base_rev 取共同祖先版本；
    否则双方 copyfrom 链交叉求最近共同祖先 rev（_lca_svn），svn cat -r rev 取 base。
    无公共祖先时回退单侧 copyfrom（fork-point）。
    group_name 支持嵌套相对路径（如 'config/skill'）：按相对路径定位 base 文件，
    tmp 内中间文件名用扁平化片段（config__skill）避免子目录问题。

    base_export_dir（批量导出优化）：compare 入口已一次性 svn export 整个 base 目录
    到本地，本函数优先从该目录拷贝 {group}.xlsx，跳过 per-table svn cat（N 次 → 0 次）。
    """
    flat = _flat_group_name(group_name)

    def _base_from_export(rev: int) -> Optional[Path]:
        """从批量导出的 base_export_dir 取本地文件，失败返回 None 回退 svn cat。"""
        if base_export_dir is None:
            return None
        local = _group_path_in_dir(base_export_dir, group_name)
        if not local.is_file():
            return None
        dest = tmp / f"{flat}_base_r{rev}.xlsx"
        try:
            shutil.copy2(local, dest)
            return dest
        except OSError:
            return None

    if override:
        base_src = _resolve_branch_path(override)
        if not base_src.is_file():
            raise HTTPException(400, f"merge_base_override 文件不存在: {base_src}")
        base_path = tmp / f"{flat}_base.xlsx"
        shutil.copy2(base_src, base_path)
        return base_path

    # 1a. base_rev 显式指定 → 直接 svn cat -r base_rev 取 base（不查 LCA）
    if base_rev is not None:
        # 优先用批量导出的本地文件（compare 入口已 svn export 整个目录）
        cached = _base_from_export(base_rev)
        if cached is not None:
            return cached
        wc_root = _wc_root_for(src_dir) or _wc_root_for(tgt_dir)
        if wc_root is not None:
            # base_rev 指定时，copyfrom_path 未知，取 src_dir/tgt_dir 侧同名文件作 cat 目标
            # （约定：base 是双方共同祖先版本，文件路径与 src/tgt 同名）
            for d in (src_dir, tgt_dir):
                base_wc = _group_path_in_dir(d, group_name)
                if base_wc.exists():
                    base_path = tmp / f"{flat}_base_r{base_rev}.xlsx"
                    try:
                        _svn_cat(base_wc, base_rev, base_path)
                        return base_path
                    except HTTPException:
                        pass
        raise HTTPException(
            400,
            f"base_rev={base_rev} svn cat 失败或路径不可用，请检查 revision 或用 merge_base_override"
        )

    # 1b. 自动 LCA：双方 copyfrom 链交叉求最近共同祖先 rev
    lca = _lca_svn(src_dir, tgt_dir, bp_cache)
    if lca is not None:
        copyfrom_path = lca["copyfrom_path"]
        copyfrom_rev = lca["copyfrom_rev"]
        # 优先用批量导出的本地文件
        cached = _base_from_export(copyfrom_rev)
        if cached is not None:
            return cached
        wc_root = _wc_root_for(src_dir) or _wc_root_for(tgt_dir)
        if wc_root is not None:
            base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
            if base_wc.is_dir():
                base_wc = _group_path_in_dir(base_wc, group_name)
            base_path = tmp / f"{flat}_base_r{copyfrom_rev}.xlsx"
            try:
                _svn_cat(base_wc, copyfrom_rev, base_path)
                return base_path
            except HTTPException:
                pass  # svn cat 失败，回退单侧 fork-point

    # LCA 失败回退：一侧 inferred（如 merge_back 的 trunk）时，沿另一侧 copyfrom 链
    # 找指向 inferred 侧路径的最早 rev —— 该 rev 即 fork 点（共同祖先）。
    # 例：B←A@r12←trunk@r10，merge_back B→trunk；trunk inferred 链空，LCA 交叉落空，
    # 沿 B 链找指向 trunk 的条目 → trunk@r10（而非 B 直接 copyfrom A@r12，会丢 trunk@r10→r12 段）。
    bp_t = _bp_cached(str(tgt_dir), bp_cache)
    bp_s = _bp_cached(str(src_dir), bp_cache)
    if (bp_t.get("ok") and bp_t.get("inferred")) or (bp_s.get("ok") and bp_s.get("inferred")):
        for infer_dir, other_dir in ((tgt_dir, src_dir), (src_dir, tgt_dir)):
            infer_bp = bp_t if infer_dir == tgt_dir else bp_s
            if not infer_bp.get("ok") or not infer_bp.get("inferred"):
                continue
            wc_root = _wc_root_for(infer_dir) or _wc_root_for(other_dir)
            if wc_root is None:
                continue
            try:
                infer_repo = infer_dir.resolve().relative_to(wc_root).as_posix()
            except ValueError:
                continue
            if not infer_repo:
                continue
            best = None
            for e in _copyfrom_chain_svn(other_dir, bp_cache):
                e_repo = (e["copyfrom_path"] or "").strip("/")
                if e_repo == infer_repo and e["copyfrom_rev"] is not None:
                    if best is None or e["copyfrom_rev"] < best["copyfrom_rev"]:
                        best = e
            if best is not None:
                copyfrom_path = best["copyfrom_path"]
                copyfrom_rev = best["copyfrom_rev"]
                base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
                if base_wc.is_dir():
                    base_wc = _group_path_in_dir(base_wc, group_name)
                base_path = tmp / f"{flat}_base_r{copyfrom_rev}.xlsx"
                try:
                    _svn_cat(base_wc, copyfrom_rev, base_path)
                    return base_path
                except HTTPException:
                    pass  # cat 失败继续走单侧 fork 兜底

    # 单侧 fork 兜底（target 优先，merge_back 的 trunk 无 copyfrom 时取 source）
    bp = _bp_cached(str(tgt_dir), bp_cache)
    use_dir = tgt_dir
    if not bp.get("ok") or bp.get("inferred"):
        bp_src = _bp_cached(str(src_dir), bp_cache)
        if bp_src.get("ok") and not bp_src.get("inferred"):
            bp = bp_src
            use_dir = src_dir
    if bp.get("ok") and not bp.get("inferred"):
        copyfrom_path = bp["copyfrom_path"]
        copyfrom_rev = bp["copyfrom_rev"]
        wc_root = _wc_root_for(use_dir)
        if wc_root is not None:
            base_wc = _repo_path_to_wc(copyfrom_path, wc_root)
            if base_wc.is_dir():
                base_wc = _group_path_in_dir(base_wc, group_name)
            base_path = tmp / f"{flat}_base_r{copyfrom_rev}.xlsx"
            _svn_cat(base_wc, copyfrom_rev, base_path)
            return base_path

    raise HTTPException(400, "无法自动定位 merge-base（SVN copyfrom 反查失败），请通过 merge_base_override 手工指定基准")


def branch_compare(req: BranchCompareRequest, progress_cb=None):
    """跨分支三方对比：base(LCA 共同 base) + source 最新版 + target 最新版。

    一次性比对多个表（group_names 为空时自动取两分支交集全部表），
    返回 CompareResponse.groups 含每个表的 FileGroup。
    HTTP 入口见 branch_compare_http（手动序列化绕开慢编码器）。

    progress_cb(phase, done=None, total=None, detail=""):可选进度回调,供 SSE 上报。
    """
    if req.direction not in ("absorb", "merge_back"):
        raise HTTPException(400, "direction 必须为 absorb 或 merge_back")

    src_dir = _resolve_demo_or_wc(req.source_branch)
    tgt_dir = _resolve_demo_or_wc(req.target_branch)
    if not src_dir.is_dir() or not tgt_dir.is_dir():
        raise HTTPException(400, "source_branch / target_branch 必须为目录")

    if req.group_names:
        # group_names 可能含单侧独有的表（结构增删），_compare_one_table 对 src/tgt
        # 均不存在的表会抛错。取与 src∩tgt 的交集做单元格比对，单侧独有的交给
        # structural_changes 标注（compute_structural_changes 识别 source_added 等）。
        src_tables = _collect_table_keys(src_dir)
        tgt_tables = _collect_table_keys(tgt_dir)
        table_names = [gn for gn in req.group_names if gn in src_tables and gn in tgt_tables]
    else:
        src_tables = _collect_table_keys(src_dir)
        tgt_tables = _collect_table_keys(tgt_dir)
        table_names = sorted(src_tables & tgt_tables)
    if not table_names:
        raise HTTPException(400, "两分支无交集表名，无法比对")

    tmp = Path(tempfile.mkdtemp(prefix="merge_branch_"))
    try:
        groups = {}
        # M9: 同一次 compare 内 src_dir/tgt_dir 不变，缓存 _resolve_branch_point 结果
        # 避免每张表重复 svn log（N 表 → 1-2 次 svn 调用）。
        bp_cache: Dict[str, dict] = {}
        # rev_cache：缓存 svn log -l 1 的 (rev, author, date)，N 表×多文件共享，
        # 避免每文件各跑一次 subprocess（M9 性能优化）。
        rev_cache: Dict[str, tuple] = {}

        # base 批量导出优化：先确定 base 的 (copyfrom_path, rev)，一次性 svn export
        # 整个 base 目录到本地，N 张表从本地副本拷贝，避免 N 次 per-table svn cat
        # （74 表 ~190s → 数秒）。失败回退到 per-table svn cat（_resolve_base_for_compare 兜底）。
        base_export_dir: Optional[Path] = None
        if not req.merge_base_override:
            try:
                base_export_dir = _prepare_base_export(
                    src_dir, tgt_dir, req.base_rev, tmp, bp_cache)
            except Exception:
                base_export_dir = None  # 导出失败不阻断，回退 per-table svn cat

        # 预填充 bp_cache：src/tgt 两分支的 branch_point 各跑一次 svn log --stop-on-copy，
        # 避免表级并行时 N 表并发各跑一次（重复 subprocess）。base_export_dir 已用首查结果。
        _bp_cached(str(src_dir), bp_cache)
        _bp_cached(str(tgt_dir), bp_cache)
        # 批量预填充 rev_cache：一次 svn info -R 取 src/tgt 目录所有文件 (rev,author,date)，
        # 替代 N 表×多文件逐个 svn log -l 1（~190s → ~0.2s）。_rev_info_cached 命中后跳过 subprocess。
        _prefill_rev_cache(src_dir, rev_cache)
        _prefill_rev_cache(tgt_dir, rev_cache)
        if progress_cb:
            progress_cb("resolve_branches", detail="分支基准解析完成")

        # 表级并行：_compare_one_table 内部 read_group_files(lxml 释放 GIL) + svn subprocess
        # 均释放 GIL，多表并发可压缩总耗时（N 表串行 → 并行，CPU-bound 的 compare_sheet
        # 单表内已 sheet 级并行，表级再并行对 IO 段有收益）。rev_cache 多线程重复 miss
        # 只产生冗余 svn log（结果一致），不破坏正确性。
        # #32: 表数 >= 4 时切 ProcessPool（compare_sheet 真并行，GIL 下 ThreadPool 对纯 CPU
        # 无收益）；失败自动回退 ThreadPool。worker_fn 为 module-level _compare_one_table_proc。
        _table_worker = partial(
            _compare_one_table_proc,
            src_dir=src_dir, tgt_dir=tgt_dir, override=req.merge_base_override,
            tmp=tmp, bp_cache=bp_cache, source_rev=req.source_rev,
            target_rev=req.target_rev, base_rev=req.base_rev,
            rev_cache=rev_cache, base_export_dir=base_export_dir)
        _tbl_progress = (lambda d, t: progress_cb("compare_tables", d, t)) if progress_cb else None
        for gn, fg in parallel_map_tables(_table_worker, table_names, progress_cb=_tbl_progress):
            groups[gn] = fg
        # 结构增删标注：base/src/tgt 三方 {group: set(sheet)} 对比，识别表/sheet 增删
        try:
            base_sheets = _dir_sheet_sets(base_export_dir) if base_export_dir else {}
            src_sheets = _dir_sheet_sets(src_dir)
            tgt_sheets = _dir_sheet_sets(tgt_dir)
            structural = compute_structural_changes(base_sheets, src_sheets, tgt_sheets)
        except Exception:
            structural = []
        if progress_cb:
            progress_cb("done", 1, 1, "比对完成")
        return CompareResponse(
            groups=groups,
            session_id="",
            conflict_origin=f"branch_{req.direction}",
            structural_changes=structural,
        )
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@router.post("/compare")
async def branch_compare_http(req: BranchCompareRequest):
    """HTTP 入口：大响应（10w 行级）手动序列化——FastAPI jsonable_encoder 对嵌套
    CellData 模型逐字段编码 ~5s/50MB，model_dump_json 仅 ~0.3s，故绕开编码器。
    内部 branch_compare 仍返回 CompareResponse（供测试/复用直接调用）。

    异步化 + 全链路异常捕获：避免同步函数内未处理异常导致空 body 引发前端
    "Unexpected end of JSON input"。
    """
    import asyncio
    import traceback as _tb
    try:
        result = await asyncio.to_thread(branch_compare, req)
    except HTTPException:
        raise
    except Exception as e:
        _tb.print_exc()
        raise HTTPException(500, f"比对内部错误：{e}")

    try:
        content = result.model_dump_json()
    except Exception as e:
        _tb.print_exc()
        raise HTTPException(500, f"序列化响应失败（数据量可能过大，请尝试指定 group_names 缩小范围）：{e}")

    # AI 建议后台预取：compare 返回后 fire-and-forget 遍历冲突格填 _suggest_cache，
    # 前端点 🎯 时命中缓存零延迟返回。不阻塞响应，失败静默。
    _prefetch_ai_suggestions(result)

    return Response(content=content, media_type="application/json")


# ── compare 任务队列 + SSE 进度上报(3.5) ──────────────────────────
# 大表 compare 耗时高,同步 HTTP 阻塞无进度。新增异步入口:POST /compare/async
# 立即返回 task_id,GET /compare/progress/{task_id} 以 SSE 推送进度事件至完成。
# 既有 POST /compare 同步入口保留不变(向后兼容)。
_COMPARE_TASKS: Dict[str, dict] = {}
_COMPARE_TASKS_LOCK = threading.Lock()


def _compare_task_emit(task_id: str, event: dict) -> None:
    """追加事件到任务事件流,done/error 置终态。"""
    with _COMPARE_TASKS_LOCK:
        t = _COMPARE_TASKS.get(task_id)
        if t is None:
            return
        t["events"].append(event)
        et = event.get("type")
        if et == "done":
            t["status"] = "done"
            t["result"] = event.get("result")
        elif et == "error":
            t["status"] = "error"
            t["error"] = event.get("error", "")


@router.post("/compare/async")
async def branch_compare_async(req: BranchCompareRequest):
    """异步 compare:入队后台执行,立即返回 task_id。"""
    task_id = uuid.uuid4().hex
    with _COMPARE_TASKS_LOCK:
        _COMPARE_TASKS[task_id] = {"status": "running", "events": [], "result": None, "error": ""}

    def _progress(phase, done=None, total=None, detail=""):
        ev: dict = {"type": "progress", "phase": phase, "detail": detail}
        if done is not None:
            ev["done"] = done
        if total is not None:
            ev["total"] = total
        _compare_task_emit(task_id, ev)

    def _run():
        try:
            result = branch_compare(req, progress_cb=_progress)
            _compare_task_emit(task_id, {"type": "done",
                                         "result": json.loads(result.model_dump_json())})
        except HTTPException as e:
            _compare_task_emit(task_id, {"type": "error",
                                         "error": f"{e.status_code}: {e.detail}"})
        except Exception as e:
            _compare_task_emit(task_id, {"type": "error", "error": str(e)})

    threading.Thread(target=_run, daemon=True, name=f"compare-{task_id[:8]}").start()
    return {"task_id": task_id}


@router.get("/compare/progress/{task_id}")
async def branch_compare_progress(task_id: str):
    """SSE:推送 compare 进度事件至 done/error。"""
    async def _stream():
        idx = 0
        # 未知 task_id → 立即报错关闭
        with _COMPARE_TASKS_LOCK:
            if task_id not in _COMPARE_TASKS:
                yield f"event: error\ndata: {json.dumps({'error': '未知 task_id'}, ensure_ascii=False)}\n\n"
                return
        while True:
            with _COMPARE_TASKS_LOCK:
                t = _COMPARE_TASKS.get(task_id)
                if t is None:
                    break
                events = t["events"][idx:]
                idx = len(t["events"])
                status = t["status"]
            for ev in events:
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
            if status in ("done", "error"):
                # 终态:清理任务,关闭流
                with _COMPARE_TASKS_LOCK:
                    _COMPARE_TASKS.pop(task_id, None)
                return
            await asyncio.sleep(0.2)

    return StreamingResponse(_stream(), media_type="text/event-stream")


def _sheet_names(fp: Path) -> set:
    """读 xlsx 的 sheet 名集合（轻量读取）。

    优先 python-calamine（Rust，解析 10w 行文件 <0.1s，openpyxl read_only 需 0.3-0.5s/文件），
    不可用时回退 openpyxl。read_only 工作簿持有文件句柄直至 close()——漏关会锁死 Windows 上
    被扫描的文件（重建目录时 rmtree 报 WinError 32），故回退路径必须显式关闭。
    """
    try:
        from python_calamine import CalamineWorkbook
        return set(CalamineWorkbook.from_path(fp).sheet_names)
    except Exception:
        pass
    wb = None
    try:
        wb = load_workbook(fp, read_only=True, data_only=True)
        return set(wb.sheetnames)
    except Exception:
        return set()
    finally:
        if wb is not None:
            wb.close()


def _dir_sheet_sets(d: Path) -> dict:
    """{table_key: {sheet_names}} for xlsx in dir（分支按相对路径 key 分组，支持嵌套）。

    table_key 为相对目录的 posix 路径去 .xlsx 后缀（如 'config/skill'；扁平 'ability'）。
    """
    out: Dict[str, set] = {}
    if not d or not d.is_dir():
        return out
    for p in d.rglob("*.xlsx"):
        if p.name.startswith("~$") or p.name.startswith("_"):
            continue
        key = p.relative_to(d).with_suffix("").as_posix()
        out.setdefault(key, set()).update(_sheet_names(p))
    return out


def _branch_display_file(src_dir: Path, tgt_dir: Path, base_dir: Optional[Path], group: str, status: str) -> Optional[Path]:
    """结构增删展示内容来源：新增/删除方取尚存一侧的文件；both_deleted 取 base(LCA) 快照。

    group 为相对路径 key（如 'config/skill'），按 _group_path_in_dir 定位 {group}.xlsx。
    """
    if status in ("source_added", "target_deleted"):
        f = _group_path_in_dir(src_dir, group)
    elif status in ("target_added", "source_deleted"):
        f = _group_path_in_dir(tgt_dir, group)
    else:  # both_deleted
        f = _group_path_in_dir(base_dir, group) if base_dir else None
    return f if f and f.is_file() else None


# 单侧展示表文件大小上限：超此值跳过全量读入（read_excel + 构 CellData），
# 避免 structural 的 target_added/source_deleted 大表（如 10w 行 big_data 2.5MB）
# 全量塞进 compare 响应致 JSON 爆 50-100MB → 前端 V8 崩 STATUS_BREAKPOINT。
# 阈值 1MB：ability/item/match_stat/config 等正常表（<200KB）放行，big_data（2.5MB）跳过。
_DISPLAY_FILE_SIZE_LIMIT = 1 * 1024 * 1024


def _attach_display_entries(groups: dict, structural: list, src_dir: Path, tgt_dir: Path,
                            base_dir: Optional[Path]) -> None:
    """把"仅一侧存在"的表/sheet 以只读展示分组挂进 groups，前端可直接点击查看内容。

    - 表级：source_added/target_added/source_deleted/target_deleted/both_deleted → 单侧展示 FileGroup。
    - sheet 级：common/both_added 表内仅一侧有的 sheet → 单侧展示 SheetDiff 补进该表。
    both_added（双方各新增，有真实比对数据）跳过。
    超过 _DISPLAY_FILE_SIZE_LIMIT 的单侧大表跳过全量读入（structural_changes 徽标仍保留）。
    """
    from engine.parser import read_excel
    for c in structural:
        if c["status"] in ("common", "both_added"):
            continue
        fp = _branch_display_file(src_dir, tgt_dir, base_dir, c["table"], c["status"])
        if fp is None:
            continue
        if fp.stat().st_size > _DISPLAY_FILE_SIZE_LIMIT:
            continue
        if c["kind"] == "table":
            if c["table"] not in groups:
                groups[c["table"]] = build_display_group(c["table"], fp, c["status"], c["origin"])
        else:
            g = groups.get(c["table"])
            if g is None or c["sheet"] in g.sheets:
                continue
            rows = read_excel(str(fp)).get(c["sheet"])
            if rows is not None:
                g.sheets[c["sheet"]] = build_display_sheet(fp, c["sheet"], rows, c["status"], c["origin"])


def _compare_one_table(src_dir: Path, tgt_dir: Path, group_name: str,
                        override: Optional[str], tmp: Path,
                        bp_cache: Optional[Dict[str, dict]] = None,
                        source_rev: Optional[int] = None,
                        target_rev: Optional[int] = None,
                        base_rev: Optional[int] = None,
                        rev_cache: Optional[Dict[str, tuple]] = None,
                        base_export_dir: Optional[Path] = None) -> FileGroup:
    """单个表的三方比对，返回 FileGroup。

    SVN 真实模式下 source_rev/target_rev 指定 src/tgt 取哪个 revision（None=HEAD，
    svn cat -r rev 取历史版本）；base_rev 指定共同祖先 revision（None=自动 LCA）。
    """
    src_files = _list_branch_files(src_dir, group_name)
    tgt_files = _list_branch_files(tgt_dir, group_name)
    if not src_files or not tgt_files:
        raise HTTPException(400, f"分支 {src_dir.name}/{tgt_dir.name} 下未找到 {group_name}.xlsx")

    base_path = _resolve_base_for_compare(
        src_dir, tgt_dir, group_name, override, tmp, bp_cache,
        base_rev=base_rev, base_export_dir=base_export_dir)
    base_name = base_path.name
    paths: List[str] = [str(base_path)]
    commit_authors: Dict[str, str] = {}
    version_meta: Dict[str, Dict[str, Any]] = {}

    # base 版本元信息：SVN 模式从 base 文件名 _base_r<rev>.xlsx 解析 copyfrom_rev
    m = re.search(r"_base_r(\d+)\.xlsx$", base_name)
    if m:
        version_meta[base_name] = {"rev": _fmt_rev(m.group(1)), "author": "", "date": ""}

    flat = _flat_group_name(group_name)
    rev_map = {"src": source_rev, "tgt": target_rev}
    for tag, files in (("src", src_files), ("tgt", tgt_files)):
        rev = rev_map[tag]
        for fp in files:
            stem = fp.stem
            leaf = group_name.rsplit("/", 1)[-1]
            if stem.startswith(leaf):
                suffix = stem[len(leaf):].lstrip("_")
                new_name = (f"{flat}_{tag}_{suffix}.xlsx" if suffix
                            else f"{flat}_{tag}.xlsx")
            else:
                new_name = f"{stem}_{tag}.xlsx"
            dest = tmp / new_name
            if rev is not None:
                # SVN 真实模式 + 指定 revision：svn cat -r rev 取历史版本
                _svn_cat(fp, rev, dest)
            else:
                shutil.copy2(fp, dest)
            paths.append(str(dest))
            # 指定 rev 时 version_meta 用该 rev（author/date 取 HEAD 的作近似，复用 rev_cache）
            if rev is not None:
                _, author, date = _rev_info_cached(fp, rev_cache)
                version_meta[new_name] = {"rev": _fmt_rev(rev), "author": author, "date": date}
                commit_authors[new_name] = author
            else:
                info = _latest_revision_info(fp, rev_cache)
                version_meta[new_name] = info
                commit_authors[new_name] = info.get("author", "")

    # demo_svn 是单用户演示仓库（file:// 协议，author 恒为系统用户），
    # 同作者自动合并会让双方改动互相吸收而看不到冲突。演示场景禁用同作者合并，
    # 真实多用户仓库（svnserve/http）author 不同，不受影响。
    use_authors = commit_authors
    if "demo_svn" in str(src_dir) or "demo_svn" in str(tgt_dir):
        use_authors = {}

    return _build_group(
        paths,
        base_name=base_name,
        group_name=group_name,
        merge_base_file=base_name,
        commit_authors=use_authors,
        version_meta=version_meta,
    )


def _compare_one_table_proc(group_name, src_dir, tgt_dir, override, tmp,
                            bp_cache, source_rev, target_rev, base_rev,
                            rev_cache, base_export_dir):
    """#32 ProcessPool worker：单表 compare 全在子进程跑（svn cat + read + compare_sheet）。
    module-level 以保证可 pickle；共享缓存（rev_cache/bp_cache）经 partial 只读绑定，
    _prefill_rev_cache 已预填，子进程命中即跳过 subprocess。返回 (group_name, FileGroup)。"""
    return group_name, _compare_one_table(
        src_dir, tgt_dir, group_name, override, tmp, bp_cache,
        source_rev=source_rev, target_rev=target_rev, base_rev=base_rev,
        rev_cache=rev_cache, base_export_dir=base_export_dir)


class TableApplyItem(BaseModel):
    """单个表的 apply 数据。"""
    group_name: str
    sheets: List[SheetMergeData]  # 前端编辑后的 sheet 数据


class BranchApplyRequest(BaseModel):
    """branch apply 请求：多表批量应用。"""
    direction: str
    source_branch: str
    target_branch: str
    tables: List[TableApplyItem]
    apply_mode: str = "new_version"   # "new_version"(absorb 产新版本目录 / merge_back 产新版本文件) | "overwrite"(覆盖目标最新)
    merge_base_override: Optional[str] = None
    commit_name: str = ""             # 本次合并的提交名称（可空，审计/历史记录展示用）
    source_rev: Optional[int] = None  # source 指定 SVN revision（None=HEAD）；与 compare 保持一致
    target_rev: Optional[int] = None  # target 指定 SVN revision（None=HEAD）；与 compare 保持一致
    base_rev: Optional[int] = None    # 共同祖先 base 指定 revision（None=自动 LCA）；与 compare 保持一致
    force: bool = False               # 方法 B：pre_commit_hold override（漏行预检命中时确认放行）


def _branch_head_file(tgt_dir: Path, group_name: str) -> Optional[Path]:
    """目标分支 HEAD 文件：时间点后缀 _sN 最大者（如 match_stat_s2）；无则规范名 {group}.xlsx。

    与 subdir 版 _target_head_file 同款逻辑。带 _sN 后缀的分支（目录合并吸收产物）
    必须取 max(_sN) 写到最新版，否则会写到最旧的 _s1 造成"应用成功但数据没落 HEAD"。
    """
    files = _locate_group_files(tgt_dir, group_name)
    if not files:
        return None
    leaf = group_name.rsplit("/", 1)[-1]
    head = None
    head_n = -1
    for p in files:
        m = re.match(rf"^{re.escape(leaf)}_s(\d+)$", p.stem)
        if m and int(m.group(1)) > head_n:
            head_n = int(m.group(1))
            head = p
    if head is not None:
        return head
    # 无 _sN 后缀变体：规范名优先（_locate_group_files 已保证精确名在前）
    return files[0]


def _locate_target_for(tgt_dir: Path, group_name: str, direction: str):
    """定位单表的 target 落点，返回 (ours_path, dest_dir, dest_name, mode_tag)。

    仅用于「无法确定版本目录规则」时的兜底命名（merge_back 落 trunk 版本化文件，
    或 target 非 `{name}_r{N}` 命名的分支/工作副本）。absorb 场景下若 target 匹配
    版本目录命名规则，改由 `_next_branch_revision_dir` 产出全新版本目录，见 branch_apply。

    支持嵌套 group_name（如 'config/skill'）：dest_dir 为 group_name 所在子目录，
    dest_name 用 leaf 名（'skill'）+ 版本号，避免路径分隔符出现在文件名里。
    """
    head = _branch_head_file(tgt_dir, group_name)
    if head is None:
        raise HTTPException(400, f"target_branch 下未找到 {group_name} 文件")
    ours_path = head
    leaf = group_name.rsplit("/", 1)[-1]
    # 嵌套表：dest_dir 到 group_name 所在子目录；扁平表 dest_dir = tgt_dir
    sub_rel = group_name.rsplit("/", 1)[0] if "/" in group_name else ""
    dest_dir = (tgt_dir / sub_rel).resolve() if sub_rel else tgt_dir

    if direction == "merge_back":
        # 落点 = 用户选定的 target（最新 trunk）。原实现硬编码 TRUNK_DIR，在工作副本
        # 目标（如 trunk_rN 或 svn/fixture 工作副本）下会错位（文件写到 merge/trunk 而非
        # 所选 trunk），此处改为写回 tgt_dir，与 merge_back"目标=trunk"的语义一致。
        dest_name = get_next_merge_version(dest_dir, leaf)
        mode_tag = "branch_merge_back"
    else:
        existing = {p.name for p in dest_dir.glob(f"{leaf}_*.xlsx")}
        n = 1
        while f"{leaf}_{n}.xlsx" in existing:
            n += 1
        dest_name = f"{leaf}_{n}.xlsx"
        mode_tag = "branch_absorb"
    return ours_path, dest_dir, dest_name, mode_tag


def _rollback_written(written: List[Tuple[Path, Optional[Path]]]) -> None:
    """apply 失败时按写入记账还原目标目录：已存在的文件恢复备份，新建的文件删除。

    written 元素为 (dest, backup)：backup 非空表示该 dest 在本次写入前已存在，
    需用备份覆盖回原文件；backup 为空表示本次新建，删除即可避免残留半成品。
    """
    for dest, backup in written:
        try:
            if backup is not None and backup.is_file():
                shutil.copy2(backup, dest)
            elif dest.exists():
                dest.unlink()
        except OSError:
            pass


_REV_DIR_RE = re.compile(r"^(.*)_r(\d+)$")


def _next_branch_revision_dir(tgt_dir: Path) -> Optional[Path]:
    """absorb 落点若匹配版本目录命名 `{name}_r{N}`（如 B_r2），返回下一版本目录路径
    `{name}_r{N+1}`（如 B_r3；若已存在则继续 +1 直到空位）。

    不匹配该命名规则（如真实 SVN 工作副本 branches/A）时返回 None，
    调用方应回退到 `_locate_target_for` 的旧版内命名。
    """
    m = _REV_DIR_RE.match(tgt_dir.name)
    if not m:
        return None
    prefix, rev = m.group(1), int(m.group(2))
    parent = tgt_dir.parent
    n = rev + 1
    while (parent / f"{prefix}_r{n}").exists():
        n += 1
    return parent / f"{prefix}_r{n}"


@router.post("/apply")
def branch_apply(req: BranchApplyRequest):
    """应用人工解决结果（多表批量），写入落点。

    apply_mode（与目录合并一致，两种落点）：
      - "overwrite"：覆盖目标分支最新表格（dest = 目标 HEAD 同名，内容为合并结果）。
      - "new_version"（默认）：absorb 且 target 匹配版本目录命名 `{name}_r{N}` 时产出
        全新兄弟版本目录 `{name}_r{N+1}`（整份快照 + 合并结果）；
        否则（merge_back 或真实工作副本）版本化命名 {group}_{N+1}.xlsx。
    ours = target 最新版；未决冲突阻断写回（预检查全部表格后才创建版本目录，避免半成品）；
    一次合并（多表）只写一条审计记录。
    """
    src_dir = _resolve_demo_or_wc(req.source_branch)
    tgt_dir = _resolve_demo_or_wc(req.target_branch)
    if not src_dir.is_dir() or not tgt_dir.is_dir():
        raise HTTPException(400, "source_branch / target_branch 必须为目录")
    if req.apply_mode not in ("overwrite", "new_version"):
        raise HTTPException(400, "apply_mode 必须为 overwrite 或 new_version")

    from engine.models import MergeRequest as _MR

    # 1) 预检查：全部表格零未决冲突才继续（避免创建半成品版本目录）
    mrs = {}
    for item in req.tables:
        mr = _MR(group_name=item.group_name, sheets=item.sheets)
        unresolved = _unresolved_conflicts(mr)
        if unresolved:
            raise HTTPException(
                400,
                f"表 {item.group_name} 存在 {len(unresolved)} 处未决冲突："
                + json.dumps(unresolved, ensure_ascii=False),
            )
        mrs[item.group_name] = mr

    # 2) absorb + new_version 且 target 匹配版本目录命名 → 产出新兄弟版本目录（整份快照 + 合并结果）
    new_branch_dir: Optional[Path] = None
    try:
        if req.direction == "absorb" and req.apply_mode == "new_version":
            new_branch_dir = _next_branch_revision_dir(tgt_dir)
            if new_branch_dir is not None:
                new_branch_dir.mkdir(parents=True, exist_ok=True)
                # 递归拷贝整份快照（含嵌套子文件夹），跳过 ~$ 临时文件
                for fp in tgt_dir.rglob("*"):
                    rel = fp.relative_to(tgt_dir)
                    dst = new_branch_dir / rel
                    if fp.is_dir():
                        dst.mkdir(parents=True, exist_ok=True)
                    elif fp.is_file() and not fp.name.startswith("~$"):
                        dst.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(fp, dst)
    except HTTPException:
        raise
    except Exception as e:
        if new_branch_dir is not None and new_branch_dir.exists():
            shutil.rmtree(new_branch_dir, ignore_errors=True)
        raise HTTPException(500, f"创建新版本目录失败：{e}")

    results = []
    audit_tables = []          # 一次合并（多表）只写一条审计记录，表级明细收进 tables
    tmp = Path(tempfile.mkdtemp(prefix="merge_branch_apply_"))
    # 回滚记账：记录每张表写入前 dest 的存在性与备份路径，失败时还原 tgt_dir 状态
    written: List[Tuple[Path, Optional[Path]]] = []
    try:
        bp_cache: Dict[str, dict] = {}  # M9: 复用 svn log 结果，避免每表重复反查
        for item in req.tables:
            mr = mrs[item.group_name]

            if req.apply_mode == "overwrite":
                # 覆盖目标分支最新表格（同名，内容为合并结果）
                ours_path, dest_dir, dest_name, mode_tag = _locate_target_for(
                    tgt_dir, item.group_name, req.direction)
                dest_name = ours_path.name
                mode_tag = f"branch_{req.direction}_overwrite"
            elif new_branch_dir is not None:
                ours_path, sub_dest_dir, _, _ = _locate_target_for(
                    tgt_dir, item.group_name, req.direction)
                # 嵌套表：在新版本目录下重建相对子目录结构
                sub_rel = item.group_name.rsplit("/", 1)[0] if "/" in item.group_name else ""
                dest_dir = (new_branch_dir / sub_rel).resolve() if sub_rel else new_branch_dir
                dest_name = ours_path.name
                mode_tag = "branch_absorb"
            else:
                ours_path, dest_dir, dest_name, mode_tag = _locate_target_for(
                    tgt_dir, item.group_name, req.direction)

            base_path = _resolve_base_for_compare(
                src_dir, tgt_dir, item.group_name, req.merge_base_override, tmp, bp_cache,
                base_rev=req.base_rev)
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / dest_name

            # 写入前预记账：dest 已存在则备份到 tmp，便于失败时恢复原文件
            backup: Optional[Path] = None
            if dest.exists() and dest.is_file():
                # group_name 可能含路径分隔符（嵌套表如 assistant/assistant），
                # 备份名须扁平化，否则 tmp/{group}_backup_... 会拼出不存在的子目录
                # 导致 shutil.copy2 报 WinError 3
                backup = tmp / f"{_flat_group_name(item.group_name)}_backup_{dest_name}"
                shutil.copy2(dest, backup)
            written.append((dest, backup))

            # 方法 B：pre-flight hold 漏行预检（写盘前，CODEMAKER_PREFLIGHT_HOLD 默认 off 零回归）
            # ca-overview §2.3.1：apply 前预检"合并此 patch 将丢失哪些 base id"，命中阻断 + audit 留痕
            hold_mode = os.environ.get("CODEMAKER_PREFLIGHT_HOLD", "off").lower()
            preflight = None
            if hold_mode != "off":
                base_pks = collect_disk_sheet_pks(ours_path)
                preflight = preflight_row_manifest(mr, base_pks)
                if preflight.will_silently_drop:
                    if hold_mode == "on" and not req.force:
                        raise HTTPException(status_code=409, detail={
                            "preflight": preflight.to_dict(),
                            "message": "漏行预检命中，需 force=true override 或补回漏行",
                        })
                    # audit 模式 or force override → 不阻断，holds 记入 audit_tables

            cache_info = _apply_edits_and_save(ours_path, dest, mr)
            # 前端只传差异行，跨表主键集合以落盘表为准（calamine 快读）
            ref_report = _validate_apply_refs(mr, extra_pks=collect_disk_sheet_pks(ours_path))

            stats = _sheets_stats(item.sheets)
            changes, truncated = _collect_changes(item.sheets, ours_path.name)
            audit_tables.append({
                "group": item.group_name,
                "base": ours_path.name,
                "output": dest_name,
                "stats": stats,
                "changes": changes,
                "changes_truncated": truncated,
                "cache_message": cache_info["cache_message"],
                "preflight_holds": [h.to_dict() for h in preflight.holds]
                                   if (preflight and preflight.will_silently_drop) else [],
            })

            results.append({
                "group": item.group_name,
                "ok": True,
                "output": dest_name,
                "path": str(dest),
                "version_dir": new_branch_dir.name if new_branch_dir else "",
                "needs_manual_fix": cache_info["needs_manual_fix"],
                "cache_message": cache_info["cache_message"],
                "ref_integrity": ref_report,
            })

        # 一次合并一条审计记录（多表明细收进 tables；group/output 汇总供列表展示）
        audit_mode = (f"branch_{req.direction}_overwrite" if req.apply_mode == "overwrite"
                      else f"branch_{req.direction}")
        _append_audit({
            "time": datetime.now().isoformat(timespec="seconds"),
            "mode": audit_mode,
            "source_branch": req.source_branch,
            "target_branch": req.target_branch,
            "version_dir": new_branch_dir.name if new_branch_dir else "",
            "commit_name": req.commit_name,
            "group": ", ".join(t["group"] for t in audit_tables),
            "output": ", ".join(t["output"] for t in audit_tables),
            "tables": audit_tables,
        })

        invalidate_dirs_cache()

        # 真实 SVN 工作副本：把本次写出的产物 svn commit，产生新 revision；
        # 然后 git add/commit 项目仓库（失败不阻断，仅回传告警）
        svn_report = _svn_commit_apply(tgt_dir, [Path(r["path"]) for r in results], req.commit_name)
        git_report = {"committed": False, "revision": None, "error": None}
        if svn_report.get("committed"):
            git_report = _git_commit_apply(MERGE_DIR.parent, [], req.commit_name)

        return {
            "ok": True,
            "results": results,
            "new_version_dir": new_branch_dir.name if new_branch_dir else "",
            "ref_warning": "；".join(r["ref_integrity"]["warning"] for r in results if r["ref_integrity"]["warning"]),
            "svn": svn_report,
            "git": git_report,
        }
    except HTTPException:
        # 写入过程中失败：清理半成品新版本目录；还原已写入 tgt脏数据
        if new_branch_dir is not None and new_branch_dir.exists():
            shutil.rmtree(new_branch_dir, ignore_errors=True)
        _rollback_written(written)
        raise
    except Exception as e:
        if new_branch_dir is not None and new_branch_dir.exists():
            shutil.rmtree(new_branch_dir, ignore_errors=True)
        _rollback_written(written)
        raise HTTPException(500, f"应用合并失败：{e}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
