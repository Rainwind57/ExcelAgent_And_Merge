import openpyxl, os, json
base = r"C:\Users\wuzhixian\Desktop\Excel-Agent-And-Merge\merge\_seed_data\trunk"
out = {}
def dump(name, sheet=None):
    p = os.path.join(base,name)
    wb = openpyxl.load_workbook(p, data_only=True)
    sheets = [sh for sh in wb.sheetnames if sh!="CONFIG"] if sheet is None else [sheet]
    for sh in sheets:
        ws = wb[sh]
        rows=[]
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if row is None or all(c is None for c in row): break
            rows.append(list(row))
        out[name+"::"+sh]=rows
dump("entity_prefab.xlsx","Base")
dump("interaction.xlsx")
with open(r"C:\Users\wuzhixian\Desktop\Excel-Agent-And-Merge\_dump2.json","w",encoding="utf-8") as f:
    json.dump(out,f,ensure_ascii=False,indent=1)
print("done")
