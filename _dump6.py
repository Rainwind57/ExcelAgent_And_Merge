import openpyxl, os, json
base = r"C:\Users\wuzhixian\Desktop\Excel-Agent-And-Merge\merge\_seed_data\trunk"
out = {}
for name in ["id_mgr.xlsx","model_prefab.xlsx","combat/pve_combat_npc.xlsx","combat/combat.xlsx","reward.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(base,name), data_only=True)
    for sh in wb.sheetnames:
        if sh=="CONFIG": continue
        ws = wb[sh]
        rows=[]
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if row is None or all(c is None for c in row): break
            rows.append(list(row))
        out[name+"::"+sh]=rows
with open(r"C:\Users\wuzhixian\Desktop\Excel-Agent-And-Merge\_dump_idmgr.json","w",encoding="utf-8") as f:
    json.dump(out,f,ensure_ascii=False,indent=1)
print("done")
